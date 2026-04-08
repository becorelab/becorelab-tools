"""
2026년 3월 매출 데이터를 Firestore에 업로드
어제 정산한 채널별 결과 엑셀 → Firestore settlements/{uid}/months/2026-03
"""
import os
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from google.cloud import firestore
import firebase_admin
from firebase_admin import credentials, auth

# Firebase 초기화
KEY_PATH = os.path.join(os.path.dirname(__file__), '..', 'sourcing', 'analyzer',
    'becorelab-tools-firebase-adminsdk-fbsvc-c665234c8b.json')
cred = credentials.Certificate(KEY_PATH)
if not firebase_admin._apps:
    firebase_admin.initialize_app(cred)
db = firestore.Client.from_service_account_json(KEY_PATH)

USER_EMAIL = 'becorelab@gmail.com'
try:
    user = auth.get_user_by_email(USER_EMAIL)
    USER_UID = user.uid
    print(f'[AUTH] {USER_EMAIL} → UID: {USER_UID}')
except Exception as e:
    print(f'[AUTH] 실패: {e}')
    sys.exit(1)

BASE = os.path.join(os.path.dirname(__file__), '★월별매출정산 with Claude Cowork', '3월 매출 정산')

# 채널별 결과 파일 → 앱 채널명 매핑
CHANNEL_FILES = {
    '카페24': '03월 카페24/카페24_정산결과_becorelab2_20260403_356_5195.xlsx',
    '스마트스토어': '03월 스마트스토어/03월 스마트스토어.xlsx',
    '로켓배송': '03월 로켓배송/03월 로켓배송.xlsx',
    '지마켓': '03월 지마켓/03월 지마켓.xlsx',
    '11번가': '03월 11번가/03월 11번가.xlsx',
    '옥션': '03월 옥션/03월 옥션.xlsx',
    '카카오선물': '03월 카카오 선물하기/03월 카카오선물하기.xlsx',
    '신세계': '03월 신세계/03월 신세계.xlsx',
    '오늘의집': '03월 오늘의집/03월 오늘의집.xlsx',
    '에드가': '03월 에드가/03월 에드가.xlsx',
    '두버': '03월 두버/03월 두버.xlsx',
    '지엠홀딩스': '03월 지엠홀딩스/03월 지엠홀딩스.xlsx',
}

def safe_int(v):
    if v is None or v == '' or isinstance(v, str):
        return 0
    try:
        return int(round(float(v)))
    except:
        return 0

def safe_float(v):
    if v is None or v == '' or isinstance(v, str):
        return 0.0
    try:
        return float(v)
    except:
        return 0.0

print(f'\n[PARSE] 3월 정산 결과 로드 중...')

channels_data = {}
skip_names = {'총합계', '(비어 있음)', '행 레이블', '배송비', ''}

for ch_name, fpath in CHANNEL_FILES.items():
    full_path = os.path.join(BASE, fpath)
    if not os.path.exists(full_path):
        print(f'  ⏭️  {ch_name}: 파일 없음')
        continue

    wb = openpyxl.load_workbook(full_path, data_only=True)
    # 피벗 시트 찾기
    ws = None
    for sn in wb.sheetnames:
        if '피벗' in sn:
            ws = wb[sn]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    products = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if name is None or str(name).strip() in skip_names:
            continue
        name = str(name).strip()
        # 헤더/타이틀 행 스킵
        try:
            qty = int(row[1]) if row[1] and not isinstance(row[1], str) else 0
        except:
            continue

        settle = safe_int(row[2]) if len(row) > 2 else 0
        ship = safe_int(row[3]) if len(row) > 3 and row[3] and not isinstance(row[3], str) else 0

        # 매출액 = 정산금 (채널마다 의미 다르지만 앱에서는 settlement으로 통일)
        products.append({
            'name': name,
            'qty': qty,
            'settlement': settle,
            'revenue': settle,
            'unitCost': 0,
            'totalCost': 0,
            'profit': 0,
            'profitRate': 0,
            'mapped': True,
        })

    if products:
        channels_data[ch_name] = products
        total = sum(p['settlement'] for p in products)
        print(f'  ✅ {ch_name}: {len(products)}개 상품, {total:,.0f}원')
    else:
        print(f'  ⚠️  {ch_name}: 데이터 없음')

# Firestore 업로드
print(f'\n[UPLOAD] Firestore 업로드...')
base_ref = db.collection('settlements').document(USER_UID).collection('months')
doc_id = '2026-03'

channel_doc = {}
total_qty = total_sett = total_cost = total_profit = 0

for ch_name, products in channels_data.items():
    ch_qty = sum(p['qty'] for p in products)
    ch_sett = sum(p['settlement'] for p in products)

    channel_doc[ch_name] = {
        'aggregated': products,
        'summary': {
            'qty': ch_qty,
            'settlement': ch_sett,
            'cost': 0,
            'profit': 0,
            'profitRate': 0,
        }
    }
    total_qty += ch_qty
    total_sett += ch_sett

doc_data = {
    'savedAt': firestore.SERVER_TIMESTAMP,
    'channels': channel_doc,
    'totalSummary': {
        'qty': total_qty,
        'settlement': total_sett,
        'cost': total_cost,
        'profit': total_profit,
        'profitRate': 0,
    }
}

base_ref.document(doc_id).set(doc_data)
print(f'\n  ✅ {doc_id}: {len(channels_data)}개 채널, {total_qty:,}개 상품, 정산 {total_sett:,.0f}원')
print(f'\n[DONE] 3월 데이터 업로드 완료! 🎉')
print(f'  → 매출앱에서 2026년 3월 선택하면 데이터가 보입니다.')
