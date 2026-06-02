# -*- coding: utf-8 -*-
"""로켓그로스 빠른정산 일별(고객결제일) 상세 - 스크린샷 18일치(5/14~5/31) 수기 입력"""
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 고객결제일: [판매액, 판매수수료, 할인쿠폰, 정산대상액, 상품광고비, 풀필먼트비용, 빠른정산수수료, 최종지급액(로켓그로스분)]
DAILY = {
 '2026-05-14': [407100, 32223, 31690, 343187, 129833, 0,      758,   178278],
 '2026-05-15': [252720, 19702, 23180, 209838, 70475,  0,      480,   117899],
 '2026-05-16': [1071920,88132, 44600, 939188, 122396, 0,      2690,  720183],
 '2026-05-17': [1097910,89065, 59750, 949095, 134207, 0,      2609,  717370],
 '2026-05-18': [1578600,131006,51690, 1395904,123136, 0,      4498,  1128679],
 '2026-05-19': [1564160,130194,46690, 1387276,132315, 0,      4321,  1111913],
 '2026-05-20': [1162440,97208, 29530, 1035702,115579, 0,      3080,  813473],
 '2026-05-21': [603900, 50696, 13000, 540204, 85681,  0,      1477,  399026],
 '2026-05-22': [751100, 62932, 17580, 670588, 95993,  0,      1815,  505721],
 '2026-05-23': [969230, 81179, 23080, 864971, 109122, 189394, 1718,  478239],
 '2026-05-24': [1617520,136137,30740, 1450643,127429, 189169, 3361,  985620],
 '2026-05-25': [979010, 82656, 15430, 880924, 123375, 310697, 1339,  357421],
 '2026-05-26': [602240, 50079, 18580, 533581, 71411,  202133, 0,     206678],
 '2026-05-27': [255750, 20559, 16150, 219041, 61156,  98612,  871,   268142],  # 전날잔금 포함된 최종
 '2026-05-28': [298160, 24052, 17870, 256238, 46947,  78339,  362,   104966],
 '2026-05-29': [184410, 14595, 14290, 155525, 13290,  63890,  209,   62583],
 '2026-05-30': [218980, 17838, 11000, 190142, 58011,  28282,  272,   84562],
 '2026-05-31': [972950, 82163, 15000, 875787, 148005, 79323,  1736,  559144],
}

cols = ['판매액','판매수수료','할인쿠폰','정산대상액','상품광고비','풀필먼트비용','빠른정산수수료','최종지급액']
tot = [0]*8
for d,v in DAILY.items():
    for i in range(8): tot[i]+=v[i]

print('=== 빠른정산 일별 합계 (5/14~5/31, 18일) ===')
for i,c in enumerate(cols):
    print('  %-12s %12s' % (c, format(tot[i],',')))

json.dump(DAILY, open('/tmp/fastpay_daily.json','w'), ensure_ascii=False)

# 엑셀에 시트 추가
path='/Users/macmini_ky/Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브/앱/매출 정산/26.05/26.05 로켓그로스 정산_채움컴퍼니.xlsx'
wb=load_workbook(path)
if '빠른정산 일별(수금)' in wb.sheetnames: del wb['빠른정산 일별(수금)']
ws=wb.create_sheet('빠른정산 일별(수금)')
thin=Side(style='thin',color='CCCCCC'); border=Border(left=thin,right=thin,top=thin,bottom=thin)
hdrfill=PatternFill('solid',fgColor='4472C4'); hdrfont=Font(color='FFFFFF',bold=True,size=10)
totfill=PatternFill('solid',fgColor='FFF2CC')
ws['A1']='로켓그로스 빠른정산 일별 상세 (고객결제일 기준 = 실수금) — Wing 화면 캡처'
ws['A1'].font=Font(bold=True,size=12)
hdr=['고객결제일']+cols
ws.append([]); ws.append(hdr)
for c in range(1,len(hdr)+1):
    cell=ws.cell(row=3,column=c); cell.fill=hdrfill; cell.font=hdrfont
    cell.alignment=Alignment(horizontal='center'); cell.border=border
r=4
for d in sorted(DAILY):
    ws.append([d]+DAILY[d]); r+=1
ws.append(['합계']+tot)
for c in range(1,len(hdr)+1):
    cell=ws.cell(row=r,column=c); cell.fill=totfill; cell.font=Font(bold=True)
for row in ws.iter_rows(min_row=4,max_row=r,min_col=2,max_col=9):
    for cell in row: cell.number_format='#,##0'
ws.column_dimensions['A'].width=13
for col in 'BCDEFGHI': ws.column_dimensions[col].width=12
# 메모
ws.cell(row=r+2,column=1,value='※ 5/27은 전날 빠른정산 잔금 포함된 최종지급액. 풀필먼트비용은 5/23부터 발생(배송완료 기준).')
ws.cell(row=r+3,column=1,value='※ 상품광고비/풀필먼트비용 = Wing 실제 차감액(정확). 5/1~5/13은 추가 캡처 필요.')
wb.save(path)
print('\n엑셀 시트 추가 완료: 빠른정산 일별(수금)')
