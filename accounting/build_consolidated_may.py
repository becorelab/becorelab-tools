# -*- coding: utf-8 -*-
"""통합 정산 5월본 — 채널별 5월 데이터를 월별 온라인 매출정산 매트릭스에 채움"""
import openpyxl, json, re, os
from collections import defaultdict
from openpyxl.styles import Font, PatternFill

DST="/Users/macmini_ky/Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브/앱/매출 정산/26.05/2026. 05 온라인 매출정산.xlsx"
B26="/Users/macmini_ky/Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브/앱/매출 정산/26.05"
NAME_MAP=json.load(open('cafe24_name_map_temp.json'))

def norm(s): return re.sub(r'\s+','',str(s)).strip()

# ── 1. 채널별 5월 데이터: 품목 -> [수량, 매출액, 배송비] ──
data=defaultdict(lambda: defaultdict(lambda:[0,0,0]))  # matrix채널명 -> 품목 -> [qty,sales,ship]
extra_rows=[]  # (matrix채널명, 라벨, 매출액) 차감/특수행

# 1-a) 카페24 (raw에서)
raw=f'{B26}/05월 카페24/카페24 05월 로우데이터.xlsx'
wb=openpyxl.load_workbook(raw,data_only=True,read_only=True); ws=wb[wb.sheetnames[0]]
hdr=[c.value for c in next(ws.iter_rows(max_row=1))]
def ci(n):
    for i,h in enumerate(hdr):
        if h and n in str(h): return i
    return None
iN=ci('주문상품명(옵션포함)');iO=ci('옵션+판매가');iQ=ci('수량');iS=ci('총 배송비')
iOrd=ci('주문번호');iRef=ci('실제 환불금액');iCpn=ci('주문서 쿠폰 할인금액');iApp=ci('앱 상품할인 금액(최종)');iMil=ci('사용한 적립금액(최종)')
def f(v):
    try: return float(str(v).replace(',','')) if v not in (None,'','None') else 0.0
    except: return 0.0
CH='카페24 / 네이버페이'
oship=set();oref=set();ocpn=set();omil=set();ref=cpn=app=mil=0.0
for row in ws.iter_rows(min_row=2,values_only=True):
    if row[iQ] in (None,'') and row[iO] in (None,''): continue
    std=NAME_MAP.get(str(row[iN] or '').strip()); o=str(row[iOrd] or '').strip()
    if std:
        q=int(f(row[iQ])); data[CH][std][0]+=q; data[CH][std][1]+=f(row[iO])*q
        if o and o not in oship: data[CH][std][2]+=f(row[iS]); oship.add(o)
    if o and o not in oref:
        rv=f(row[iRef])
        if rv>0: ref+=rv; oref.add(o)
    if o and o not in ocpn: cpn+=f(row[iCpn]); ocpn.add(o)
    app+=f(row[iApp])
    if o and o not in omil:
        mv=f(row[iMil])
        if mv>0: mil+=mv; omil.add(o)
extra_rows += [(CH,'환불(차감)',-ref),(CH,'쿠폰(차감)',-cpn),(CH,'앱상품할인(차감)',-app),(CH,'사용적립금(차감)',-mil)]

# 1-b) 나머지 채널 (출력파일에서): 매출액=정산금액, 배송비=0
CHMAP={'스마트스토어':'스마트스토어','카카오선물하기':'카카오선물하기','쿠팡로켓':'쿠팡(로켓배송)',
'옥션':'옥션','지마켓':'지마켓','11번가':'11번가','쿠팡윙':'쿠팡','SSG':'신세계몰\n(에스에스지닷컴)',
'오늘의집':'오늘의집','GS샵':'GS샵','에이블리':'에이블리'}
import glob
for d in os.listdir(B26):
    import unicodedata
    dn=unicodedata.normalize('NFC',d)
    if not dn.startswith('05월') or not os.path.isdir(os.path.join(B26,d)): continue
    chshort=dn.replace('05월 ','').replace(' ','')
    # 매칭
    mch=None
    for k,v in CHMAP.items():
        if k.replace(' ','')==chshort or k==chshort: mch=v; break
    if not mch: continue
    files=[f2 for f2 in os.listdir(os.path.join(B26,d)) if unicodedata.normalize('NFC',f2).endswith('정산.xlsx')]
    if not files: continue
    fp=os.path.join(B26,d,files[0])
    wb2=openpyxl.load_workbook(fp,data_only=True); ws2=wb2.active
    h2=[c.value for c in ws2[1]]
    def c2(n):
        for i,hh in enumerate(h2):
            if hh and n in str(hh): return i
        return None
    inm=c2('품명');iqt=c2('수량');ist=c2('정산')
    if None in (inm,iqt,ist): continue
    for r in range(2,ws2.max_row+1):
        nm=ws2.cell(r,inm+1).value
        nms=str(nm).strip() if nm else ''
        if nm is None or nms in ('합계','총합계','소계'): continue
        if '배송비' in nms:
            try: extra_rows.append((mch,'배송비',float(ws2.cell(r,ist+1).value or 0)))
            except: pass
            continue
        def tn(v):
            try: return float(str(v).replace(',','')) if v not in (None,'') else 0.0
            except: return None
        q=tn(ws2.cell(r,iqt+1).value); s=tn(ws2.cell(r,ist+1).value)
        if q is None or s is None: continue
        data[mch][str(nm).strip()][0]+=q; data[mch][str(nm).strip()][1]+=s

# ── 2. 매트릭스 채우기 ──
wb=openpyxl.load_workbook(DST, data_only=False)
ws=wb['월별 온라인 매출정산']
# 채널 블록 행범위 (data_only로 채널명 읽기)
wbv=openpyxl.load_workbook(DST, data_only=True); wsv=wbv['월별 온라인 매출정산']
rowsv=list(wsv.iter_rows(values_only=True))
blocks={}; cur=None; start=None
for i in range(9,len(rowsv)):
    ch=rowsv[i][1] if len(rowsv[i])>1 else None
    if ch:
        if ch!=cur:
            if cur is not None: blocks.setdefault(cur,[start+1,i])
            cur=ch; start=i
if cur is not None: blocks.setdefault(cur,[start+1,len(rowsv)])

def set_row(R, qty, sales, ship):
    ws.cell(R,42).value=qty; ws.cell(R,43).value=sales; ws.cell(R,44).value=ship
    ws.cell(R,45).value=f'=AQ{R}+AR{R}'
    ws.cell(R,46).value=f'=IFERROR(VLOOKUP($E{R},원가!$B:$E,4,FALSE)," ")'
    ws.cell(R,47).value=f'=IFERROR(AQ{R}-($AT{R}*$AP{R}),"")'
    ws.cell(R,48).value=f'=IFERROR(AU{R}/AS{R},"")'

filled=defaultdict(int); appended=[]
for mch, prods in data.items():
    if mch not in blocks:
        print(f'⚠️ 매트릭스에 채널 없음: {repr(mch)}'); continue
    s,e=blocks[mch]
    # 매트릭스 품목 -> 행 (정규화 매칭)
    row_of={}
    for R in range(s,e+1):
        pm=wsv.cell(R,5).value
        if pm: row_of.setdefault(norm(pm),R)
    used=set()
    for pm,(q,sl,sh) in prods.items():
        if q==0 and sl==0: continue
        R=row_of.get(norm(pm))
        if R and R not in used:
            set_row(R,q,sl,sh); used.add(R); filled[mch]+=1
        else:
            appended.append((mch,pm,q,sl,sh))
    # 차감/특수행은 append로
# extra(차감) + unmatched append: 맨 아래
last=ws.max_row
ar=last+2
for mch,label,amt in extra_rows:
    ar+=1; ws.cell(ar,2).value=mch; ws.cell(ar,5).value=label
    set_row(ar,'',amt,'')
for mch,pm,q,sl,sh in appended:
    ar+=1; ws.cell(ar,2).value=mch; ws.cell(ar,5).value=pm
    set_row(ar,q,sl,sh)

wb.save(DST)
print('=== 통합 5월 채우기 완료 ===')
for mch in data: print(f'  {repr(mch)[:26]:<28} 매칭채움 {filled[mch]} / 신규추가 {sum(1 for a in appended if a[0]==mch)}')
print(f'차감행(카페24) {len(extra_rows)}개 추가')
print(f'\n채널별 5월 매출(=정산) 합계 검증:')
for mch,prods in data.items():
    tot=sum(v[1]+v[2] for v in prods.values())
    print(f'  {repr(mch)[:24]:<26} {tot:>13,.0f}')
