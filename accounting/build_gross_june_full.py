# -*- coding: utf-8 -*-
"""로켓그로스(채움컴퍼니) 2026년 6월 정산표 — 전체 (매출인식 6/1~6/30)
새 포맷(CATEGORY_TR / WAREHOUSING_SHIPPING / STORAGE_FEE) + 셀러월렛 일별 + 주정산 실값.
build_gross_june.py(6/1~14)의 6월 완결판. 경로 = 26.06 월별 폴더 체계."""
import openpyxl, glob, os, json, re, unicodedata
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = "/Users/macmini_ky/Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브/Claude AI work space/02. 매출 정산/26.06/그로스/채움컴퍼니"
OUT = "/Users/macmini_ky/Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브/Claude AI work space/02. 매출 정산/26.06/그로스/26.06 로켓그로스 정산_채움컴퍼니_최종.xlsx"
AD  = "/Users/macmini_ky/ClaudeAITeam/marketing/coupang_data"
P1, P2 = "2026-06-01", "2026-06-30"

# 옵션ID -> 원가 (세트단가, VAT별도) — 2026-07-07 원가표(건조기시트 4,190) 반영
COST = {
    '95304424363': 4190, '95363500368': 8380, '95363500367': 12570,   # 바이올렛 1/2/3
    '95602285926': 4190,                                              # 바이올렛 1 (신규 옵션ID)
    '95377476928': 4190, '95377476939': 8380, '95377476938': 12570,   # 베이비크림 1/2/3
    '95304888338': 3199, '95304888360': 6398, '95304888356': 9597,    # 얼룩 1/2/3
    '95304888355': 19194,                                             # 얼룩 6
    '95060453340': 2200, '95060453345': 4400, '95060453346': 6600,    # 입테 1/2/3
    '95060453344': 8800, '95060453347': 11000, '95060453343': 13200,  # 입테 4/5/6
}
def bucket(name):
    n = name or ''
    if '바이올렛' in n: return '바이올렛'
    if '베이비' in n: return '베이비크림'
    if '얼룩' in n or '목때' in n: return '얼룩제거제'
    if '입벌림' in n or '오모모' in n: return '입테이프'
    return '기타'

# ---- 1) 매출/정산대상액: CATEGORY_TR ----
prod = defaultdict(lambda: {'옵션명':'', '수량':0, '판매액':0, '정산대상액':0})
cyc  = defaultdict(lambda: {'판매액':0, '정산대상액':0, '수량':0})
os.chdir(SRC)
for f in glob.glob('*CATEGORY_TR*'):
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True); ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None: continue
        rec = str(row[4])[:10]
        if not (P1 <= rec <= P2): continue
        end = str(row[1])[:10]
        oid = str(row[11]).split('.')[0]
        try: qty=float(row[16] or 0); pa=float(row[17] or 0); st=float(row[23] or 0)
        except: continue
        p = prod[oid]; p['옵션명'] = (str(row[13]) + ' / ' + str(row[14]))[:60]
        p['수량']+=qty; p['판매액']+=pa; p['정산대상액']+=st
        cyc[end]['판매액']+=pa; cyc[end]['정산대상액']+=st; cyc[end]['수량']+=qty
    wb.close()
unknown = [o for o in prod if o not in COST]
assert not unknown, f"원가 미매핑 옵션ID: {unknown}"
JUNE_CYC = sorted(cyc.keys())

# ---- 2) 물류비: 정산주기별 최종비용(VAT포함) ----
def logi_by_cycle(pat, sheet):
    seen=set(); d=defaultdict(float)
    for f in glob.glob(pat):
        wb=openpyxl.load_workbook(f,read_only=True,data_only=True)
        if sheet not in wb.sheetnames: wb.close(); continue
        for row in wb[sheet].iter_rows(min_row=3, values_only=True):
            if not row or row[0] is None: continue
            end=str(row[0])[:10]
            if end in seen: continue
            seen.add(end)
            try: d[end]+=float(row[3] or 0)
            except: pass
        wb.close()
    return d
io=logi_by_cycle('*WAREHOUSING_SHIPPING*','입출고비')
ship=logi_by_cycle('*WAREHOUSING_SHIPPING*','배송비')
stor=logi_by_cycle('*STORAGE_FEE*','보관비')

# ---- 3) 광고비: 광고센터 공식 XLSX (6월 전체, 지출 발생일 기준) ----
# 2026-07-07 대표님 제공. JSON 일별수집은 6/1~7 +105,137 / 6/29~30 +8,897 누락 확인 → XLSX가 정답.
AD_XLSX = [f for f in os.listdir(SRC) if 'pa_daily_keyword' in f and f.endswith('.xlsx')]
assert len(AD_XLSX)==1, f"광고 XLSX 1개여야 함: {AD_XLSX}"
CYC_DAYS = {'2026-06-07':(1,7), '2026-06-14':(8,14), '2026-06-21':(15,21), '2026-06-28':(22,28), '2026-06-30':(29,30)}
ad_daily=defaultdict(float); ad_prod=defaultdict(float)
_awb=openpyxl.load_workbook(os.path.join(SRC,AD_XLSX[0]),read_only=True,data_only=True)
_aws=_awb.active; _rows=_aws.iter_rows(values_only=True); _hdr=None
for _row in _rows:
    if _row and any(v=='날짜' for v in _row if isinstance(v,str)): _hdr=list(_row); break
_iD=_hdr.index('날짜'); _iC=_hdr.index('광고비'); _iP=_hdr.index('광고집행 상품명')
for _row in _rows:
    if not _row or _row[_iD] is None: continue
    m=re.match(r"(\d{4})(\d{2})(\d{2})", str(_row[_iD]))
    if not m: continue
    d=f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    try: v=float(_row[_iC] or 0)
    except: continue
    ad_daily[d]+=v; ad_prod[bucket(str(_row[_iP] or ''))]+=v
_awb.close()
ad_cyc={end: sum(ad_daily.get(f"2026-06-{d:02d}",0) for d in range(d1,d2+1)) for end,(d1,d2) in CYC_DAYS.items()}

# ---- 4) 제품별 물류 실부과 (SKU 상세, VAT별도 -> x1.1) ----
def agg_logi(sheet, tag, cA, cAB, cQ):
    seen=set(); res=defaultdict(lambda:{'수량':0,'발생':0,'실부과':0})
    for f in glob.glob('*WAREHOUSING_SHIPPING*'):
        wb=openpyxl.load_workbook(f,read_only=True,data_only=True)
        if sheet not in wb.sheetnames: wb.close(); continue
        for r in wb[sheet].iter_rows(values_only=True):
            if not r or r[5]!=tag: continue
            cy=str(r[1])[:10]
            if not cy.startswith('2026-06'): continue
            key=(cy,r[6],r[7],r[10])
            if key in seen: continue
            seen.add(key)
            try: A=float(r[cA] or 0); AB=float(r[cAB] or 0); q=float(r[cQ] or 0)
            except: continue
            d=res[bucket(str(r[12]))]; d['수량']+=q; d['발생']+=A; d['실부과']+=AB
        wb.close()
    return res
LIO=agg_logi('입출고비','입출고비 정산',22,24,19)
LSP=agg_logi('배송비','배송비 정산',21,23,20)

# ---- 5) 셀러월렛 일별 (로켓그로스 열만, 결제일 6/1~6/30) ----
def won(v):
    if v in (None,''): return 0
    return int(str(v).replace('원','').replace(',','').strip() or 0)
wallet=[]
wf=[f for f in os.listdir(SRC) if '셀러월렛' in unicodedata.normalize('NFC',f)][0]
wwb=openpyxl.load_workbook(os.path.join(SRC,wf),read_only=True,data_only=True)
wws=wwb[wwb.sheetnames[0]]
wrows=list(wws.iter_rows(values_only=True))
ncol=max(len(r) for r in wrows)
def cell(r,c): return wrows[r-1][c-1] if len(wrows[r-1])>=c else None
for j in range(1,ncol+1):
    title=str(cell(1,j) or '')
    if '로켓그로스' not in title: continue
    m=re.search(r'(20\d{2}-\d{2}-\d{2})', str(cell(2,j) or ''))
    if not m: continue
    d=m.group(1)
    if not ('2026-06-01' <= d <= '2026-06-30'): continue
    ratio=re.search(r'지급비율\s*(\d+)', str(cell(17,j) or ''))
    wallet.append({'결제일':d, '판매액':won(cell(6,j)), '취소액':won(cell(8,j)),
        '판매수수료':won(cell(12,j)), '할인쿠폰':won(cell(14,j)), '정산대상액':won(cell(16,j)),
        '지급비율':int(ratio.group(1)) if ratio else None, '빠른정산대상액':won(cell(18,j)),
        '상품광고비':won(cell(22,j)), '풀필먼트':won(cell(24,j)), '입금액':won(cell(26,j)),
        '빠른정산수수료':won(cell(28,j)), '최종지급액':won(cell(30,j))})
wwb.close()
wallet.sort(key=lambda x:x['결제일'])
assert len(wallet)==30, f"셀러월렛 6월 일수 {len(wallet)} != 30"

# ---- 6) 주정산 수금 (어드민 정산현황 실값, 2026-07-07 대표님 제공) ----
SETTLE = [
    ('2026-06-07','6/1~6/7',  '2026-07-03',390047,'2026-08-03',167148),
    ('2026-06-14','6/8~6/14', '2026-07-10',305791,'2026-08-03',133909),
    ('2026-06-21','6/15~6/21','2026-07-20',392785,'2026-08-03',168327),
    ('2026-06-28','6/22~6/28','2026-07-27',0,     '2026-08-03',0),
    ('2026-06-30','6/29~6/30','2026-08-03',66432, '미발행',None),
]

# ================= 엑셀 작성 =================
thin=Side(style='thin',color='CCCCCC'); border=Border(thin,thin,thin,thin)
hf=PatternFill('solid',fgColor='4472C4'); hfont=Font(color='FFFFFF',bold=True,size=10)
tot=PatternFill('solid',fgColor='FFF2CC')
def hdr(ws,r,n):
    for c in range(1,n+1):
        cell=ws.cell(row=r,column=c); cell.fill=hf; cell.font=hfont
        cell.alignment=Alignment(horizontal='center',vertical='center'); cell.border=border
def fmt(ws,r1,r2,c1,c2,f='#,##0'):
    for row in ws.iter_rows(min_row=r1,max_row=r2,min_col=c1,max_col=c2):
        for cell in row:
            if isinstance(cell.value,(int,float)): cell.number_format=f
def pct(ws,r1,r2,c):
    for row in ws.iter_rows(min_row=r1,max_row=r2,min_col=c,max_col=c):
        for cell in row:
            if isinstance(cell.value,(int,float)): cell.number_format='0.0%'
def totrow(ws,r,n):
    for c in range(1,n+1):
        cell=ws.cell(row=r,column=c); cell.fill=tot; cell.font=Font(bold=True)

wb=Workbook()

# Sheet1: 제품별 정산
ws=wb.active; ws.title='제품별 정산'
ws['A1']='로켓그로스(채움컴퍼니) 6월 제품별 정산 — 매출인식 6/1~6/30'; ws['A1'].font=Font(bold=True,size=13)
H=['옵션ID','상품/옵션','수량','판매액','정산대상액','원가(단가)','원가합계','매출이익(정산-원가)','이익률']
ws.append([]); ws.append(H); hdr(ws,3,len(H))
r=4; T=defaultdict(float)
for oid in sorted(prod,key=lambda x:-prod[x]['판매액']):
    p=prod[oid]; uc=COST[oid]; cc=uc*p['수량']
    prof=p['정산대상액']-cc
    ws.append([oid,p['옵션명'],p['수량'],p['판매액'],p['정산대상액'],uc,cc,prof,prof/p['판매액'] if p['판매액'] else 0])
    T['수량']+=p['수량']; T['판매액']+=p['판매액']; T['정산대상액']+=p['정산대상액']; T['원가합계']+=cc; T['이익']+=prof
    r+=1
ws.append(['','합계',T['수량'],T['판매액'],T['정산대상액'],'',T['원가합계'],T['이익'],T['이익']/T['판매액']])
totrow(ws,r,9); fmt(ws,4,r,3,8); pct(ws,4,r,9)
for col,w in zip('ABCDEFGHI',[13,44,6,12,12,11,11,15,8]): ws.column_dimensions[col].width=w

# Sheet2: 월 요약·정산주기
ws2=wb.create_sheet('월 요약·정산주기')
ws2['A1']='6월 정산주기별 요약 (정산대상액 = 판매액 − 판매수수료·쿠폰 차감 후)'; ws2['A1'].font=Font(bold=True,size=12)
H2=['정산주기(종료일)','매출인식','수량','판매액','공제(수수료·쿠폰)','정산대상액','입출고비','배송비','보관비','물류계','광고비','실수령(정산-물류-광고)','실수령률']
ws2.append([]); ws2.append(H2); hdr(ws2,3,len(H2))
RANGE={'2026-06-07':'6/1~6/7','2026-06-14':'6/8~6/14','2026-06-21':'6/15~6/21','2026-06-28':'6/22~6/28','2026-06-30':'6/29~6/30'}
rr=4; G=defaultdict(float)
for end in JUNE_CYC:
    c=cyc[end]; pa=c['판매액']; st=c['정산대상액']; fee=pa-st
    a,b,s2=io.get(end,0),ship.get(end,0),stor.get(end,0); lg=a+b+s2; adc=ad_cyc.get(end,0)
    net=st-lg-adc
    ws2.append([end,RANGE.get(end,''),c['수량'],pa,fee,st,a,b,s2,lg,adc,net,net/pa if pa else 0])
    for k,v in [('수량',c['수량']),('판매액',pa),('수수료',fee),('정산',st),('입출고',a),('배송',b),('보관',s2),('물류',lg),('광고',adc),('실수령',net)]: G[k]+=v
    rr+=1
ws2.append(['합계','',G['수량'],G['판매액'],G['수수료'],G['정산'],G['입출고'],G['배송'],G['보관'],G['물류'],G['광고'],G['실수령'],G['실수령']/G['판매액']])
totrow(ws2,rr,13); fmt(ws2,4,rr,3,12); pct(ws2,4,rr,13)
for col,w in zip('ABCDEFGHIJKLM',[15,10,7,12,13,12,10,10,8,10,11,18,9]): ws2.column_dimensions[col].width=w

# Sheet3: 제품군 순마진 (광고비 배분 포함)
ws5=wb.create_sheet('제품군 순마진')
ws5['A1']='6월 제품군별 순마진 — 정산대상액 − 원가 − 물류실부과(VAT포함) − 광고비'; ws5['A1'].font=Font(bold=True,size=12)
H5=['제품군','수량','판매액','정산대상액','원가','물류실부과(VAT포함)','광고비','순이익','순마진율(판매액대비)']
ws5.append([]); ws5.append(H5); hdr(ws5,3,len(H5))
pg=defaultdict(lambda: defaultdict(float))
for oid,p in prod.items():
    b=bucket(p['옵션명']); pg[b]['수량']+=p['수량']; pg[b]['판매액']+=p['판매액']; pg[b]['정산']+=p['정산대상액']; pg[b]['원가']+=COST[oid]*p['수량']
for b in set(LIO)|set(LSP):
    pg[b]['물류']+= (LIO[b]['실부과']+LSP[b]['실부과'])*1.1
for b,v in ad_prod.items():
    if v: pg[b]['광고']+=v
r5=4; T5=defaultdict(float)
for b in sorted(pg,key=lambda x:-pg[x]['판매액']):
    d=pg[b]; net=d['정산']-d['원가']-d['물류']-d['광고']
    ws5.append([b,d['수량'],d['판매액'],d['정산'],d['원가'],d['물류'],d['광고'],net,net/d['판매액'] if d['판매액'] else 0])
    for k in ['수량','판매액','정산','원가','물류','광고']: T5[k]+=d[k]
    T5['순이익']+=net; r5+=1
ws5.append(['합계',T5['수량'],T5['판매액'],T5['정산'],T5['원가'],T5['물류'],T5['광고'],T5['순이익'],T5['순이익']/T5['판매액']])
totrow(ws5,r5,9); fmt(ws5,4,r5,2,8); pct(ws5,4,r5,9)
ws5.cell(row=r5+2,column=1,value='※ 광고비 배분 = 광고 리포트 "광고집행 상품명" 기준. 물류실부과 = SKU상세 할인적용가×1.1 (주기합계와 미세 반올림 차이 가능).')
for col,w in zip('ABCDEFGHI',[12,7,12,12,11,17,11,12,10]): ws5.column_dimensions[col].width=w

# Sheet4: 제품별 물류비 (발생 vs 실부과 vs 면제)
wsl=wb.create_sheet('제품별 물류비')
wsl['A1']='6월 제품별 물류비 — 발생비용(할인前) vs 실부과 vs 면제 (6월 전체, VAT별도)'; wsl['A1'].font=Font(bold=True,size=12)
HL=['제품','입출고 발생','입출고 실부과','배송 발생','배송 실부과','발생합계','실부과합계','면제액','면제율','실부과(VAT포함)']
wsl.append([]); wsl.append(HL); hdr(wsl,3,len(HL))
rl=4; LT=defaultdict(float)
for p in sorted(set(LIO)|set(LSP), key=lambda x:-(LIO[x]['실부과']+LSP[x]['실부과'])):
    a1=LIO[p]['발생']; b1=LIO[p]['실부과']; a2=LSP[p]['발생']; b2=LSP[p]['실부과']
    ta=a1+a2; tb=b1+b2
    wsl.append([p,a1,b1,a2,b2,ta,tb,ta-tb,(ta-tb)/ta if ta else 0,round(tb*1.1)])
    for k,v in [('a1',a1),('b1',b1),('a2',a2),('b2',b2),('ta',ta),('tb',tb)]: LT[k]+=v
    rl+=1
wsl.append(['합계',LT['a1'],LT['b1'],LT['a2'],LT['b2'],LT['ta'],LT['tb'],LT['ta']-LT['tb'],(LT['ta']-LT['tb'])/LT['ta'] if LT['ta'] else 0,round(LT['tb']*1.1)])
totrow(wsl,rl,10); fmt(wsl,4,rl,2,8); fmt(wsl,4,rl,10,10); pct(wsl,4,rl,9)
wsl.cell(row=rl+2,column=1,value='※ 면제분 = 저가상품 할인(14,000원 미만). 2027.1.31 종료 시 발생비용 전액 부과 리스크.')
for col,w in zip('ABCDEFGHIJ',[12,11,12,10,11,11,12,11,8,14]): wsl.column_dimensions[col].width=w

# Sheet5: 빠른정산 일별(수금) — 셀러월렛
ws6=wb.create_sheet('빠른정산 일별(수금)')
ws6['A1']='로켓그로스 셀러월렛(빠른정산) 일별 — 고객결제일 6/1~6/30'; ws6['A1'].font=Font(bold=True,size=12)
H6=['고객결제일','판매액','취소액','판매수수료','할인쿠폰','정산대상액','지급비율%','빠른정산대상액','상품광고비','풀필먼트비용','입금액','빠른정산수수료','최종지급액']
ws6.append([]); ws6.append(H6); hdr(ws6,3,len(H6))
r6=4; W=defaultdict(float)
for w_ in wallet:
    ws6.append([w_['결제일'],w_['판매액'],w_['취소액'],w_['판매수수료'],w_['할인쿠폰'],w_['정산대상액'],w_['지급비율'],w_['빠른정산대상액'],w_['상품광고비'],w_['풀필먼트'],w_['입금액'],w_['빠른정산수수료'],w_['최종지급액']])
    for k in ['판매액','취소액','판매수수료','할인쿠폰','정산대상액','빠른정산대상액','상품광고비','풀필먼트','입금액','빠른정산수수료','최종지급액']: W[k]+=w_[k]
    r6+=1
ws6.append(['합계',W['판매액'],W['취소액'],W['판매수수료'],W['할인쿠폰'],W['정산대상액'],'',W['빠른정산대상액'],W['상품광고비'],W['풀필먼트'],W['입금액'],W['빠른정산수수료'],W['최종지급액']])
totrow(ws6,r6,13); fmt(ws6,4,r6,2,13)
flip=[w_['결제일'] for w_ in wallet if w_['지급비율']==100]
ws6.cell(row=r6+2,column=1,value=f"※ 지급비율 90%→100% 전환: {flip[0] if flip else '6월 중 없음'} 결제분부터. 셀러월렛 선지급 후 주정산에서 회수 상계 (6/22~28 주기부터 주정산 최종지급 0원 = 정상).")
for col,w in zip('ABCDEFGHIJKLM',[12,11,9,11,10,12,9,13,11,12,11,12,12]): ws6.column_dimensions[col].width=w

# Sheet6: 수금일정 (주정산)
ws3=wb.create_sheet('수금일정')
ws3['A1']='6월 매출분 주정산 수금일정 (어드민 정산현황 실값 2026-07-07 기준 / 셀러월렛 선지급 별도)'; ws3['A1'].font=Font(bold=True,size=12)
H3=['정산주기','매출인식','1차 70% 지급일','1차 금액','2차 30% 지급일','2차 금액']
ws3.append([]); ws3.append(H3); hdr(ws3,3,len(H3))
r3=4
for end,rng,d1,a1,d2,a2 in SETTLE:
    ws3.append([end,rng,d1,a1,d2,a2 if a2 is not None else '미발행']); r3+=1
ws3.append(['','','합계 1차',sum(s[3] for s in SETTLE),'2차',sum(s[5] or 0 for s in SETTLE)])
totrow(ws3,r3,6); fmt(ws3,4,r3,4,4); fmt(ws3,4,r3,6,6)
ws3.cell(row=r3+2,column=1,value='※ 6/22~28부터 0원 = 셀러월렛 100% 지급 전환으로 주정산 잔여 소멸 (선지급 회수 상계). 매출 정상.')
for col,w in zip('ABCDEF',[14,12,15,12,15,12]): ws3.column_dimensions[col].width=w

# Sheet7: 정산 메모
ws4=wb.create_sheet('정산 메모')
notes=[
 '[로켓그로스 채움컴퍼니 6월 정산 메모 / 작성 2026-07-07]','',
 '[범위] 매출인식일 2026-06-01 ~ 06-30 전체 (정산주기 5개: ~6/7, ~6/14, ~6/21, ~6/28, ~6/30).',
 '[기준] 정산대상액 = CATEGORY_TR col23 (판매수수료·할인쿠폰 차감 후). 물류비 = 주기별 최종비용(VAT포함).',
 '',
 '[교차검증 — 2026-07-07 수행]',
 ' - 6/1~14: 판매 11,403,920 / 정산 10,035,610 / 물류 2,376,926 = 6/18 기존 정산과 0원 일치.',
 ' - 6/22~28: 판매 8,012,460 / 정산 7,084,473 / 물류 1,568,955 = Wing 어드민 상세내역과 1원 단위 일치.',
 ' - 광고비: 공식 XLSX(지출 발생일) 6/22~28 1,034,559 vs 어드민 상계 1,093,843 (−59,284 = 청구시점 차이).',
 '',
 '[광고비] 광고센터 공식 XLSX(A00940134_pa_daily_keyword_20260601_20260630) 합산 = 3,600,061.',
 ' - JSON 일별수집 대비 +114,032 (6/1~7 +105,137, 6/29~30 +8,897 소급 반영). 제품군 배분 = "광고집행 상품명".',
 ' - 기타 31,071 = 캡슐표백제(과탄산 캡슐) 광고 — 6월 그로스 판매품목 아님, 전계정 합산 방침대로 포함.',
 '[수금] ①셀러월렛(빠른정산): 결제일별 익일 선지급, 6/22 결제분부터 지급비율 90→100% ②주정산: 어드민 실값.',
 ' - 6/22~28 주기부터 주정산 최종지급 0원 = 셀러월렛 선지급 회수 상계 (매출 이상 없음, 지급 경로만 변경).',
 '[원가] 6/26 대표님 원가표(VAT별도). 신규 옵션: 베이비크림 2/3개입, 바이올렛 신ID(95602285926), 얼룩 6개입.',
]
for i,n in enumerate(notes,1): ws4.cell(row=i,column=1,value=n)
ws4.cell(row=1,column=1).font=Font(bold=True,size=12); ws4.column_dimensions['A'].width=110

wb.save(OUT)

# ---- 콘솔 요약 + 검증 ----
tpa=G['판매액']; tst=G['정산']; tlg=G['물류']; tad=G['광고']; tcost=T['원가합계']; tprof=tst-tcost-tlg-tad
print('저장:', OUT); print()
print('=== 6월 그로스 정산 요약 (6/1~6/30 전체) ===')
print(f"판매액        {int(tpa):>12,}")
print(f"공제(수수료·쿠폰) {int(tpa-tst):>10,}")
print(f"정산대상액    {int(tst):>12,}")
print(f"물류비        {int(tlg):>12,}  (입출고 {int(G['입출고']):,} + 배송 {int(G['배송']):,} + 보관 {int(G['보관']):,})")
print(f"광고비        {int(tad):>12,}")
print(f"실수령(정산-물류-광고) {int(tst-tlg-tad):>12,}  ({(tst-tlg-tad)/tpa*100:.1f}%)")
print(f"원가          {int(tcost):>12,}")
print(f"순이익(정산-원가-물류-광고) {int(tprof):>10,}  ({tprof/tpa*100:.1f}%)")
print()
print('=== 검증 ===')
plsum=(LT['tb'])*1.1
print(f"물류 제품별합(실부과×1.1) {plsum:,.0f} vs 주기별 최종비용합 {tlg:,.0f} (차이 {plsum-tlg:+,.0f})")
print(f"셀러월렛 6월 결제분: 판매 {W['판매액']:,.0f} / 정산대상 {W['정산대상액']:,.0f} / 최종지급 {W['최종지급액']:,.0f}")
print(f"광고 제품군 배분합 {sum(ad_prod.values()):,.0f} vs 주기합 {tad:,.0f}")
print(f"지급비율 100% 첫날: {flip[0] if flip else '없음'}")
print('\n=== 제품군 순마진 ===')
for b in sorted(pg,key=lambda x:-pg[x]['판매액']):
    d=pg[b]; net=d['정산']-d['원가']-d['물류']-d['광고']
    print(f"{b:8s} 판매 {d['판매액']:>11,.0f} 순이익 {net:>10,.0f} ({net/d['판매액']*100 if d['판매액'] else 0:5.1f}%)")
