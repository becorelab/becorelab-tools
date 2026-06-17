# -*- coding: utf-8 -*-
"""6월 SKU별 순마진 표 → 엑셀 + 구글시트 추가"""
import openpyxl, glob, json, os
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import gspread
from google.oauth2.service_account import Credentials

SRC="/Users/macmini_ky/Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브/앱/매출 정산/26.05/그로스"
XLSX="/Users/macmini_ky/Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브/앱/매출 정산/26.05/26.06 로켓그로스 정산_채움컴퍼니.xlsx"
KEY="/Users/macmini_ky/ClaudeAITeam/sourcing/analyzer/becorelab-tools-firebase-adminsdk-fbsvc-4af6f0c5ac.json"
SID="1W2gWtqcUnJxYMNKC74AuMFRsz-5yWzLousRf4Hrn8Fc"
AD="/Users/macmini_ky/ClaudeAITeam/marketing/coupang_data"
NAME={'95304424363':'바이올렛 1개','95363500368':'바이올렛 2개','95363500367':'바이올렛 3개',
 '95377476928':'베이비크림 1개','95304888338':'얼룩제거제 1개','95304888360':'얼룩제거제 2개','95304888356':'얼룩제거제 3개',
 '95060453340':'입테이프 1개','95060453345':'입테이프 2개','95060453346':'입테이프 3개',
 '95060453344':'입테이프 4개','95060453347':'입테이프 5개','95060453343':'입테이프 6개'}
COST={'95304424363':3932,'95363500368':7864,'95363500367':11796,'95377476928':3932,
 '95304888338':3199,'95304888360':6398,'95304888356':9597,
 '95060453340':2200,'95060453345':4400,'95060453346':6600,'95060453344':8800,'95060453347':11000,'95060453343':13200}
def fam(nm):
    if '바이올렛' in nm: return '바이올렛'
    if '베이비' in nm: return '베이비크림'
    if '얼룩' in nm or '목때' in nm: return '얼룩제거제'
    if '입테이프' in nm or '입벌림' in nm or '오모모' in nm or '코골이' in nm: return '입테이프'
    if '표백제' in nm: return '표백제'
    return nm

os.chdir(SRC)
S=defaultdict(lambda:{'수량':0,'정산':0,'판매':0})
for f in glob.glob('*CATEGORY_TR*'):
    wb=openpyxl.load_workbook(f,read_only=True,data_only=True)
    for r in wb.active.iter_rows(min_row=2,values_only=True):
        if not r or r[0] is None or not ('2026-06-01'<=str(r[4])[:10]<='2026-06-14'): continue
        o=str(r[11]).split('.')[0]
        try: S[o]['수량']+=float(r[16] or 0); S[o]['정산']+=float(r[23] or 0); S[o]['판매']+=float(r[17] or 0)
        except: pass
    wb.close()
def logi(sheet,tag,cAB):
    seen=set(); d=defaultdict(float)
    for f in glob.glob('*WAREHOUSING_SHIPPING*'):
        wb=openpyxl.load_workbook(f,read_only=True,data_only=True)
        if sheet not in wb.sheetnames: wb.close(); continue
        for r in wb[sheet].iter_rows(values_only=True):
            if not r or r[5]!=tag or not str(r[1])[:10].startswith('2026-06'): continue
            key=(str(r[1])[:10],r[6],r[7],r[10])
            if key in seen: continue
            seen.add(key)
            try: d[str(r[10]).split('.')[0]]+=float(r[cAB] or 0)
            except: pass
        wb.close()
    return d
io=logi('입출고비','입출고비 정산',24); sp=logi('배송비','배송비 정산',23)
LOGI={o:(io.get(o,0)+sp.get(o,0))*1.1 for o in set(io)|set(sp)}
adfam=defaultdict(float)
for day in range(1,15):
    try: rows=json.load(open(f"{AD}/A00940134_pa_daily_keyword_202606{day:02d}_202606{day:02d}.json"))
    except: continue
    for r in rows: adfam[fam(str(r.get('광고집행 상품명','')))]+=float(r.get('광고비',0) or 0)
famst=defaultdict(float)
for o in S: famst[fam(NAME.get(o,o))]+=S[o]['정산']

rows=[]
for o in sorted(S,key=lambda x:-S[x]['정산']):
    nm=NAME.get(o,o); f=fam(nm); q=S[o]['수량']; pa=S[o]['판매']; st=S[o]['정산']; cc=COST.get(o,0)*q
    lg=LOGI.get(o,0); ad=adfam.get(f,0)*(st/famst[f]) if famst[f] else 0
    rows.append([nm,q,pa,st,cc,lg,ad,st-cc-lg-ad])
T=[sum(r[i] for r in rows) for i in range(1,8)]

# ===== 엑셀: SKU별 순마진 시트 =====
wb=openpyxl.load_workbook(XLSX)
if 'SKU별 순마진' in wb.sheetnames: del wb['SKU별 순마진']
ws=wb.create_sheet('SKU별 순마진', 1)
thin=Side(style='thin',color='CCCCCC'); bd=Border(thin,thin,thin,thin)
hf=PatternFill('solid',fgColor='4472C4'); hfont=Font(color='FFFFFF',bold=True,size=10); totf=PatternFill('solid',fgColor='FFF2CC')
ws['A1']='6월 SKU별 순마진 (6/1~6/14) — 순마진=정산대상-원가-물류-광고(제품군 배분)'; ws['A1'].font=Font(bold=True,size=12)
H=['SKU','수량','판매액','정산대상','원가','물류(VAT포함)','광고(배분)','순마진','마진율']
ws.append([]); ws.append(H)
for c in range(1,len(H)+1):
    cell=ws.cell(row=3,column=c); cell.fill=hf; cell.font=hfont; cell.alignment=Alignment(horizontal='center'); cell.border=bd
for r in rows: ws.append(r[:1]+r[1:]+[r[7]/r[3] if r[3] else 0])
ws.append(['합계']+T+[T[6]/T[2] if T[2] else 0])
last=ws.max_row
for c in range(1,10):
    cell=ws.cell(row=last,column=c); cell.fill=totf; cell.font=Font(bold=True)
for row in ws.iter_rows(min_row=4,max_row=last,min_col=2,max_col=8):
    for cell in row:
        if isinstance(cell.value,(int,float)): cell.number_format='#,##0'
for row in ws.iter_rows(min_row=4,max_row=last,min_col=9,max_col=9):
    for cell in row:
        if isinstance(cell.value,(int,float)): cell.number_format='0.0%'
ws.cell(row=last+2,column=1,value='※ 물류=옵션ID별 실부과(A-B)×1.1. 광고=제품군 총광고를 정산대상 비중으로 배분(추정). 표백제 광고 11,190원은 6월 그로스 매출 없어 제외.')
ws.cell(row=last+3,column=1,value='※ 베이비크림: 광고 12.6만 집행했으나 그로스 매출 1개 → 적자. 광고 효율/재고 점검 필요. 입테이프: 광고>물류로 마진율 11%대.')
for col,w in zip('ABCDEFGHI',[14,6,12,12,11,13,11,12,8]): ws.column_dimensions[col].width=w
wb.save(XLSX)
print('엑셀 SKU별 순마진 시트 추가 완료')

# ===== 구글시트: 6월 SKU 마진 탭 =====
def cm(n): return f"{int(round(n)):,}"
cr=Credentials.from_service_account_file(KEY,scopes=["https://www.googleapis.com/auth/spreadsheets","https://www.googleapis.com/auth/drive"])
sh=gspread.authorize(cr).open_by_key(SID)
t='6월 SKU 마진'
try: g=sh.worksheet(t); g.clear()
except gspread.WorksheetNotFound: g=sh.add_worksheet(title=t,rows=30,cols=9)
out=[['로켓그로스 6월(1~14일) SKU별 순마진'],[],
 ['SKU','수량','판매액','정산대상','원가','물류','광고','순마진','마진율']]
for r in rows: out.append([r[0],cm(r[1]),cm(r[2]),cm(r[3]),cm(r[4]),cm(r[5]),cm(r[6]),cm(r[7]),f"{r[7]/r[3]*100:.1f}%"])
out.append(['합계',cm(T[0]),cm(T[1]),cm(T[2]),cm(T[3]),cm(T[4]),cm(T[5]),cm(T[6]),f"{T[6]/T[2]*100:.1f}%"])
out.append([])
out.append(['※ 광고=제품군 총광고를 정산대상 비중으로 배분(추정). 표백제 11,190원 제외(그로스 매출無).'])
out.append(['※ 베이비크림 적자(광고 과다·매출1개), 입테이프 광고>물류로 마진율 11%대 — 광고 점검 필요.'])
g.update(out,'A1')
print('구글시트 [6월 SKU 마진] 탭 추가 완료 / 합계 순마진', cm(T[6]),f"({T[6]/T[2]*100:.1f}%)")
