# -*- coding: utf-8 -*-
"""그로스 수금일정 정정 — 실제 정산 리포트(주정산) 기준으로 재작성"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

# 실제 정산 리포트 (정산일, 지급비율, 매출인식일, 최종지급액)
R=[
('2026-05-12','30%','03/30~03/31',14964),('2026-05-12','30%','04/01~04/05',72772),
('2026-05-12','70%','04/06~04/12',186905),('2026-06-01','30%','04/06~04/12',80120),
('2026-05-19','70%','04/13~04/19',229452),('2026-06-01','30%','04/13~04/19',126883),
('2026-05-27','70%','04/20~04/26',304126),('2026-06-01','30%','04/20~04/26',130359),
('2026-06-02','70%','04/27~04/30',482602),('2026-06-10','30%','04/27~04/30',206855),
('2026-06-02','70%','05/01~05/03',1078922),('2026-06-10','30%','05/01~05/03',462393),
('2026-06-09','70%','05/04~05/10',2864040),('2026-07-01','30%','05/04~05/10',1246436),
('2026-06-16','70%','05/11~05/17',696975),('2026-07-01','30%','05/11~05/17',298680),
('2026-06-23','70%','05/18~05/24',514174),('2026-07-01','30%','05/18~05/24',220355),
('2026-06-29','70%','05/25~05/31',257799),('2026-07-01','30%','05/25~05/31',93336),
('2026-07-03','70%','06/01~06/07',562524),
]
TODAY='2026-06-05'
빠른정산=8799897  # 5/14~31 선지급 (캡처 기준, 이미 수령)

path='/Users/macmini_ky/Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브/앱/매출 정산/26.05/26.05 로켓그로스 정산_채움컴퍼니.xlsx'
wb=load_workbook(path)
if '수금일정' in wb.sheetnames: del wb['수금일정']
ws=wb.create_sheet('수금일정',1)
thin=Side(style='thin',color='CCCCCC'); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
hf=PatternFill('solid',fgColor='4472C4'); hfont=Font(color='FFFFFF',bold=True,size=10)
tf=PatternFill('solid',fgColor='FFF2CC'); doneF=PatternFill('solid',fgColor='E2EFDA')

ws['A1']='로켓그로스 채움컴퍼니 — 수금일정 (실제 정산 리포트 기준)'; ws['A1'].font=Font(bold=True,size=13)
ws['A2']='주정산(이 표) + 빠른정산(5/14~31 선지급 8,799,897 이미 수령). 기준일 2026-06-05'; ws['A2'].font=Font(size=9,italic=True)

# 정산일별 합계
byday=defaultdict(int)
for d,r,m,a in R: byday[d]+=a
ws.append([]); ws.append(['정산일','입금액','상태','(구성)'])
for c in range(1,5):
    cell=ws.cell(row=4,column=c); cell.fill=hf; cell.font=hfont; cell.alignment=Alignment(horizontal='center'); cell.border=bd
row=5; fut=0
for d in sorted(byday):
    comp=' / '.join(f"{m}({r})" for dd,r,m,a in R if dd==d)
    status='입금완료' if d<=TODAY else '예정'
    if d>TODAY: fut+=byday[d]
    ws.cell(row,1,d); ws.cell(row,2,byday[d]); ws.cell(row,3,status); ws.cell(row,4,comp[:55])
    if d<=TODAY:
        for c in range(1,4): ws.cell(row,c).fill=doneF
    row+=1
ws.cell(row,1,'주정산 합계'); ws.cell(row,2,sum(byday.values()))
for c in range(1,3): ws.cell(row,c).fill=tf; ws.cell(row,c).font=Font(bold=True)
row+=2

# 요약
ws.cell(row,1,'[ 수금 요약 ]').font=Font(bold=True,size=11); row+=1
summ=[('빠른정산 선지급 (5/14~31, 이미 수령)',빠른정산),
      ('주정산 — 이미 받은 분 (~6/5)',sum(byday[d] for d in byday if d<=TODAY)),
      ('주정산 — 앞으로 받을 분 (6/5 이후)',fut),
      ('★ 6/5 이후 그로스 추가 입금 (주정산)',fut)]
for nm,v in summ:
    ws.cell(row,1,nm); ws.cell(row,2,v)
    if '★' in nm:
        for c in range(1,3): ws.cell(row,c).fill=tf; ws.cell(row,c).font=Font(bold=True)
    row+=1
row+=1
for note in ['※ 빠른정산=구매확정의 ~90% 선지급(전일 매일). 주정산=정식정산으로 잔여분+2차를 며칠 뒤 추가지급.',
 '※ 5/1~13 주차는 빠른정산 전이라 주정산 풀(70% 108만/286만). 5/18~31은 빠른정산 후라 주정산은 잔액만.',
 '※ 이전 추정치(218만 등)는 폐기 — 본 표가 실제 Wing 정산 리포트 기준.']:
    ws.cell(row,1,note).font=Font(size=9); row+=1

for col in ws.iter_rows(min_row=5,max_row=row,min_col=2,max_col=2):
    for cell in col:
        if isinstance(cell.value,(int,float)): cell.number_format='#,##0'
ws.column_dimensions['A'].width=34; ws.column_dimensions['B'].width=13
ws.column_dimensions['C'].width=11; ws.column_dimensions['D'].width=58
wb.save(path)
print('✅ 수금일정 시트 정정 완료 (실제 정산 리포트 기준)\n')
print('정산일별 입금:')
for d in sorted(byday): print(f'  {d}: {byday[d]:>10,}  {"(완료)" if d<=TODAY else "(예정)"}')
print(f'\n주정산 총액: {sum(byday.values()):,}')
print(f'6/5 이후 추가 입금(주정산): {fut:,}')
