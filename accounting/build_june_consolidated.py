# -*- coding: utf-8 -*-
"""6월 통합 정산 종합 시안 v3 — 배송비/원가 원본 직접 (2026-07-10)"""
import openpyxl, os, unicodedata, json
from collections import defaultdict, OrderedDict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
B="/Users/macmini_ky/Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브/Claude AI work space/02. 매출 정산/26.06"
def NM(s): return unicodedata.normalize("NFC",s)
def num(v):
    if v in (None,'','None'): return 0.0
    try: return float(str(v).replace(',','').replace('원',''))
    except: return 0.0
# data[ch][품명] = [수량, 매출, 이익, 원가];  ship_det[ch][품명]=배송비
data=defaultdict(lambda: defaultdict(lambda:[0.,0.,0.,0.]))
ship_det=defaultdict(lambda: defaultdict(float))
def find(kw):
    for root,dirs,files in os.walk(B):
        if "_archive" in root or "중복" in root: continue
        for f in files:
            n=NM(f)
            if kw in n and "정산" in n and n.endswith(".xlsx") and n.startswith("06월") and "라이플로우_" not in n:
                return os.path.join(root,f)
def read_std(path,ch):
    wb=openpyxl.load_workbook(path,read_only=True,data_only=True); ws=wb.active
    rows=list(ws.iter_rows(values_only=True)); hdr=rows[0]
    ipr=next((i for i,h in enumerate(hdr) if h and '이익' in str(h)),-1)
    ico=next((i for i,h in enumerate(hdr) if h and '원가합계' in str(h)), next((i for i,h in enumerate(hdr) if h and '원가' in str(h)),-1))
    irev=next((i for i,h in enumerate(hdr) if h and ('정산금액' in str(h) or '공급가매출' in str(h))),2)
    for r in rows[1:]:
        nm=str(r[0] or '').strip()
        if not nm or nm in ('합계','총합계','상품 소계','최종 (정산)'): continue
        q=num(r[1]); rev=num(r[irev]); prof=num(r[ipr]); co=num(r[ico])
        data[ch][nm][0]+=q; data[ch][nm][1]+=rev; data[ch][nm][2]+=prof; data[ch][nm][3]+=co
    wb.close()
STD={'스마트스토어':'스마트스토어','쿠팡(로켓배송)':'쿠팡로켓','지마켓':'지마켓','옥션':'옥션','11번가':'11번가',
'GS샵':'GS','SSG':'SSG','오늘의집':'오늘의집','에이블리':'에이블리','카카오선물하기':'카카오선물하기','카카오쇼핑하기':'카카오 쇼핑하기',
'PG컴퍼니':'PG컴퍼니','두버':'두버','에드가':'에드가'}
for ch,kw in STD.items():
    p=find(kw)
    if p: read_std(p,ch)
# 카페24 (매출액/배송비 별도, 원가·이익 원본)
p=find('카페24'); wb=openpyxl.load_workbook(p,read_only=True,data_only=True); ws=wb.active
for r in ws.iter_rows(min_row=4,values_only=True):
    nm=str(r[0] or '').strip()
    if not nm or nm in ('상품 소계','최종 (정산)'): continue
    if '차감' in nm:
        data['카페24'][nm][1]+=num(r[2]); data['카페24'][nm][2]+=num(r[5]); continue
    data['카페24'][nm][0]+=num(r[1]); data['카페24'][nm][1]+=num(r[2]); data['카페24'][nm][2]+=num(r[5]); data['카페24'][nm][3]+=num(r[4])
    ship_det['카페24'][nm]+=num(r[3])
wb.close()
# 그로스 채움 (제품별정산) + WAREHOUSING 옵션ID별 물류 실부과
import glob as _glob
GLOGI=defaultdict(lambda:[0.,0.])  # nm -> [입출고VAT포함, 배송VAT포함]
p=f"{B}/그로스/26.06 로켓그로스 정산_채움컴퍼니_최종.xlsx"
wb=openpyxl.load_workbook(p,read_only=True,data_only=True); ws=wb['제품별 정산']
OID2NM={}
for r in ws.iter_rows(min_row=4,values_only=True):
    nm=str(r[1] or '').strip()
    if not nm or nm=='합계': continue
    oid=str(r[0]).split('.')[0]; OID2NM[oid]=nm
    data['쿠팡(로켓그로스)'][nm][0]+=num(r[2]); data['쿠팡(로켓그로스)'][nm][1]+=num(r[4]); data['쿠팡(로켓그로스)'][nm][2]+=num(r[7]); data['쿠팡(로켓그로스)'][nm][3]+=num(r[6])
wb.close()
GDIR=f"{B}/그로스/채움컴퍼니"
def _agg(sheet,tag,cAB,seen):
    for f in _glob.glob(GDIR+"/*WAREHOUSING_SHIPPING*"):
        w=openpyxl.load_workbook(f,read_only=True,data_only=True)
        if sheet not in w.sheetnames: w.close(); continue
        for r in w[sheet].iter_rows(values_only=True):
            if not r or r[5]!=tag: continue
            cy=str(r[1])[:10]
            if not cy.startswith('2026-06'): continue
            key=(cy,r[6],r[7],r[10])
            if key in seen: continue
            seen.add(key)
            oid=str(r[10]).split('.')[0]; nm=OID2NM.get(oid)
            if nm: 
                idx=0 if '입출고' in tag else 1
                try: GLOGI[nm][idx]+=float(r[cAB] or 0)*1.1
                except: pass
        w.close()
_agg('입출고비','입출고비 정산',24,set()); _agg('배송비','배송비 정산',23,set())
# 그로스 이익 = 정산대상 − 원가 − 물류
for nm,v in data['쿠팡(로켓그로스)'].items():
    io,sp=GLOGI.get(nm,[0,0]); v[2]=v[1]-v[3]-(io+sp)
# 라이플로우 (일비아+클린햇)
cost=json.load(open('/Users/macmini_ky/ClaudeAITeam/accounting/cost_master_2026.json'))
CMAP={'하트 식기세척기 세제':'하트식세기','캡슐 표백제':'캡슐표백제','건조기시트':'건조기시트','캡슐세제':'캡슐세제','섬유탈취제 400ml':'섬유탈취제 400','섬유탈취제 100ml':'섬유탈취제 100','얼룩제거제 100ml':'얼룩제거제100','수세미':'수세미'}
CHC={'변기세정제 케이스형 2개입':4180,'변기세정제 코인형 12개입':4180,'숯 제습제 10개입':6380}
P="/Users/macmini_ky/Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브/Claude AI work space/02. 매출 정산/26.06/06월 라이플로우/06월 라이플로우_260610.xlsx"
wb=openpyxl.load_workbook(P,read_only=True,data_only=True)
g=lambda sn:[wb[s] for s in wb.sheetnames if NM(s)==sn][0]
for sn,ri in [('피벗',2),('피벗(6월 2차 정산)',3)]:
    for r in g(sn).iter_rows(min_row=4,values_only=True):
        nm=str(r[0] or '').strip()
        if not nm or nm=='총합계': continue
        q=num(r[1]); rev=num(r[ri]); c=cost.get(CMAP.get(nm,''),0)
        data['라이플로우'][nm][0]+=q; data['라이플로우'][nm][1]+=rev; data['라이플로우'][nm][2]+=rev-c*q; data['라이플로우'][nm][3]+=c*q
for r in g('피벗 클린햇').iter_rows(min_row=4,values_only=True):
    nm=str(r[0] or '').strip()
    if not nm or nm=='총합계': continue
    q=num(r[1]); rev=num(r[3]); c=CHC.get(nm,0)
    data['라이플로우'][nm][0]+=q; data['라이플로우'][nm][1]+=rev; data['라이플로우'][nm][2]+=rev-c*q; data['라이플로우'][nm][3]+=c*q
wb.close()

GROUPS=OrderedDict([("자사몰",["카페24"]),("네이버",["스마트스토어"]),("쿠팡",["쿠팡(로켓배송)","쿠팡(로켓그로스)"]),
 ("오픈마켓",["지마켓","옥션","11번가"]),("종합몰",["SSG","GS샵","오늘의집","에이블리","카카오선물하기","카카오쇼핑하기"]),
 ("수동발주",["라이플로우","PG컴퍼니","두버","에드가"])])
SHIP={"카페24":1326000,"스마트스토어":1578000,"지마켓":37713,"라이플로우":537000,"PG컴퍼니":57000,"두버":153000,"에드가":6000}
# 광고비 (2026-07-14 마감 확정, 대표님 제공): 로켓 14,263,512 / 메타 13,030,612 / 스스 = 쏘핑수수료 8,300,292 + 애드부스트 1,292,283 + 검색광고 1,280,511
AD={"카페24":13030612,"쿠팡(로켓배송)":14263512,"쿠팡(로켓그로스)":3600061,"스마트스토어":8300292+1292283+1280511,"11번가":157409}
# 5월 확정값 (2026-07-06 마감: 카페24 차감4행·사은품·샘플키트·마케팅허수 / PG 수량 292 정정 반영)
MAY5={'카페24':(28487162,17773951),'스마트스토어':(12666380,9216508.3),'카카오선물하기':(105920,73184),'쿠팡(로켓배송)':(70192906,35768629.2),'옥션':(696962,461422.3),'지마켓':(2561213,1683646.4),'11번가':(1724963,1063669.4),'SSG':(122185,90832.8),'오늘의집':(125172,95497.1),'GS샵':(66720,40770),'에이블리':(21606,0),'에드가':(117840,54450),'두버':(908420,335208),'PG컴퍼니':(1324720,109164),'라이플로우':(845335,309672.8),'쿠팡(로켓그로스)':(19303923,10110963),'카카오쇼핑하기':(0,0)}
rev_=lambda ch:sum(v[1] for v in data.get(ch,{}).values())
# 이익 = 채널파일 이익 + 배송수익 (총매출과 대칭, 2026-07-14 대표님 확정. 카페24는 파일 이익에 배송 기포함이라 제외)
prof_=lambda ch:sum(v[2] for v in data.get(ch,{}).values())+(SHIP.get(ch,0) if ch!='카페24' else 0)
cost_=lambda ch:sum(v[3] for v in data.get(ch,{}).values())
ship_=lambda ch:SHIP.get(ch,0)
tot_=lambda ch:rev_(ch)+ship_(ch)   # 총매출=매출+배송
GTr=sum(tot_(ch) for chs in GROUPS.values() for ch in chs)
GTp=sum(prof_(ch) for chs in GROUPS.values() for ch in chs)

print(f"{'채널':16s} {'매출':>12s} {'배송비':>9s} {'총매출':>12s} {'이익':>11s}")
for chs in GROUPS.values():
    for ch in chs:
        print(f"{ch:16s} {rev_(ch):>12,.0f} {ship_(ch):>9,.0f} {tot_(ch):>12,.0f} {prof_(ch):>11,.0f}")
print(f"{'종합':16s} {'':12s} {'':9s} {GTr:>12,.0f} {GTp:>11,.0f}")

# ===== 워크북 =====
HF=PatternFill('solid',fgColor='1B2A4A'); HFONT=Font(color='FFFFFF',bold=True,size=10)
SUB=PatternFill('solid',fgColor='DCEDC8'); TOT=PatternFill('solid',fgColor='FFF2CC'); AS_=PatternFill('solid',fgColor='FFF9C4'); MF=PatternFill('solid',fgColor='FCE4D6')
bd=Border(bottom=Side(style='thin',color='DDDDDD'))
def m(c):c.number_format='#,##0'
def pc(c):c.number_format='0.0%'
wb=openpyxl.Workbook()
# 시트2
ws2=wb.active; ws2.title='채널별 세부'; ws2['A1']='2026년 6월 채널별 세부'; ws2['A1'].font=Font(bold=True,size=13)
ws2.append([]); ws2.append(['그룹','채널','품명','수량','매출(정산)','입출고비','배송비','원가','이익','이익률'])
for c in range(1,11): ws2.cell(row=3,column=c).fill=HF; ws2.cell(row=3,column=c).font=HFONT
r=4
for gp,chs in GROUPS.items():
    for ch in chs:
        isg=(ch=='쿠팡(로켓그로스)')
        for nm,(q,rev,prof,co) in sorted(data.get(ch,{}).items(),key=lambda x:-x[1][1]):
            if isg:
                io,sp=GLOGI.get(nm,[0,0]); base=rev  # 매출=정산대상
                ws2.append([gp,ch,nm,q or '',rev,round(io),round(sp),co,prof,(prof/rev if rev else 0)])
            else:
                sd=ship_det.get(ch,{}).get(nm,0)
                ws2.append([gp,ch,nm,q or '',rev,'',sd or '',co,prof,(prof/(rev+sd) if (rev+sd) else 0)])
            for c in [4,5,6,7,8,9]: m(ws2.cell(row=r,column=c))
            pc(ws2.cell(row=r,column=10))
            if '차감' in nm or rev<0:
                for c in range(1,11): ws2.cell(row=r,column=c).fill=MF
            r+=1
        cr=rev_(ch); cp=prof_(ch); cc=cost_(ch)
        if isg:
            tio=sum(GLOGI.get(n,[0,0])[0] for n in data[ch]); tsp=sum(GLOGI.get(n,[0,0])[1] for n in data[ch])
            ws2.append(['',ch+' 합계','','',cr,round(tio),round(tsp),cc,cp,(cp/cr if cr else 0)])
        else:
            ws2.append(['',ch+' 합계','','',cr,'',ship_(ch),cc,cp,(cp/(cr+ship_(ch)) if (cr+ship_(ch)) else 0)])
        for c in range(1,11): ws2.cell(row=r,column=c).fill=SUB; ws2.cell(row=r,column=c).font=Font(bold=True)
        for c in [5,6,7,8,9]: m(ws2.cell(row=r,column=c))
        pc(ws2.cell(row=r,column=10)); r+=1
ws2.append(['','■ 비코어랩 종합','','',sum(rev_(c) for chs in GROUPS.values() for c in chs),'',sum(SHIP.values()),sum(cost_(c) for chs in GROUPS.values() for c in chs),GTp,(GTp/GTr)])
for c in range(1,11): ws2.cell(row=r,column=c).fill=TOT; ws2.cell(row=r,column=c).font=Font(bold=True,size=11)
for c in [5,6,7,8,9]: m(ws2.cell(row=r,column=c))
pc(ws2.cell(row=r,column=10)); r+=1
ws2.append(['','   총매출 (매출+배송)','','','','',GTr,'','',''])
ws2.cell(row=r,column=2).font=Font(bold=True); m(ws2.cell(row=r,column=7)); ws2.cell(row=r,column=7).font=Font(bold=True); r+=2
ws2.cell(row=r,column=1,value='◆ 전체 이익 차감 (물류 운임 — 대표님 기입)').font=Font(bold=True,size=11); r+=1
ws2.cell(row=r,column=2,value='항목'); ws2.cell(row=r,column=3,value='금액(직접입력)')
for c in [2,3]: ws2.cell(row=r,column=c).fill=HF; ws2.cell(row=r,column=c).font=HFONT
r+=1; ts=r
# 6월 운임 확정 (대표님 2026-07-14): 바이피엘 9,475,160+140,250 / 밀크런 311,600 / 스타배송 149,820
FREIGHT={'바이피엘 운임':9615410,'밀크런 운임':311600,'스타배송 운임':149820,'기타':None}
for it,fv in FREIGHT.items():
    ws2.cell(row=r,column=2,value=it); ws2.cell(row=r,column=3,value=fv); ws2.cell(row=r,column=3).fill=AS_; m(ws2.cell(row=r,column=3)); r+=1
ws2.cell(row=r,column=2,value='운임 차감 합계').font=Font(bold=True); ws2.cell(row=r,column=3,value=f'=SUM(C{ts}:C{r-1})').font=Font(bold=True); m(ws2.cell(row=r,column=3)); r+=1
ws2.cell(row=r,column=2,value='최종 이익 (종합−운임)').font=Font(bold=True); ws2.cell(row=r,column=3,value=f'={GTp:.0f}-C{r-1}').font=Font(bold=True); m(ws2.cell(row=r,column=3))
for col,w in zip('ABCDEFGHIJ',[9,17,32,6,12,10,12,11,11,7]): ws2.column_dimensions[col].width=w
ws2.freeze_panes='A4'
# 시트3
ws3=wb.create_sheet('채널별 매출이익'); ws3['A1']='2026년 6월 채널별 매출·이익 (광고비 전)'; ws3['A1'].font=Font(bold=True,size=13)
ws3.append([]); ws3.append(['그룹','채널','매출+배송','원가','이익','이익률'])
for c in range(1,7): ws3.cell(row=3,column=c).fill=HF; ws3.cell(row=3,column=c).font=HFONT
r=4
for gp,chs in GROUPS.items():
    gs=r
    for ch in chs:
        ct=tot_(ch); cp=prof_(ch)
        ws3.append([gp if ch==chs[0] else '',ch,ct,ct-cp,cp,(cp/ct if ct else 0)])
        for c in [3,4,5]: m(ws3.cell(row=r,column=c))
        pc(ws3.cell(row=r,column=6))
        for c in range(1,7): ws3.cell(row=r,column=c).border=bd
        r+=1
    if len(chs)>1: ws3.merge_cells(start_row=gs,start_column=1,end_row=r-1,end_column=1)
    ws3.cell(row=gs,column=1).alignment=Alignment(vertical='center',horizontal='center'); ws3.cell(row=gs,column=1).font=Font(bold=True)
    gt=sum(tot_(c) for c in chs); gpp=sum(prof_(c) for c in chs)
    ws3.append(['',gp+' 소계',gt,gt-gpp,gpp,(gpp/gt if gt else 0)])
    for c in range(1,7): ws3.cell(row=r,column=c).fill=SUB; ws3.cell(row=r,column=c).font=Font(bold=True)
    for c in [3,4,5]: m(ws3.cell(row=r,column=c))
    pc(ws3.cell(row=r,column=6)); r+=1
ws3.append(['','■ 종합',GTr,GTr-GTp,GTp,(GTp/GTr)])
for c in range(1,7): ws3.cell(row=r,column=c).fill=TOT; ws3.cell(row=r,column=c).font=Font(bold=True,size=11)
for c in [3,4,5]: m(ws3.cell(row=r,column=c))
pc(ws3.cell(row=r,column=6))
for col,w in zip('ABCDEF',[10,20,13,12,13,8]): ws3.column_dimensions[col].width=w
# 시트4
ws4=wb.create_sheet('채널별 광고후이익'); ws4['A1']='2026년 6월 광고비 적용 후 이익 (노란셀=광고비)'; ws4['A1'].font=Font(bold=True,size=13)
ws4.append([]); ws4.append(['채널','매출','매출이익','광고비','광고비중','광고후이익','광고후이익률'])
for c in range(1,8): ws4.cell(row=3,column=c).fill=HF; ws4.cell(row=3,column=c).font=HFONT
r=4; adT=0
for gp,chs in GROUPS.items():
    for ch in chs:
        t=tot_(ch); pf=prof_(ch); ad=AD.get(ch,0); adT+=ad
        ws4.append([ch,t,pf,ad,(ad/t if t else 0),pf-ad,((pf-ad)/t if t else 0)])
        for c in [2,3,4,6]: m(ws4.cell(row=r,column=c))
        pc(ws4.cell(row=r,column=5)); pc(ws4.cell(row=r,column=7))
        if ad: ws4.cell(row=r,column=4).fill=AS_
        r+=1
ws4.append(['■ 종합',GTr,GTp,adT,(adT/GTr),GTp-adT,((GTp-adT)/GTr)])
for c in range(1,8): ws4.cell(row=r,column=c).fill=TOT; ws4.cell(row=r,column=c).font=Font(bold=True)
for c in [2,3,4,6]: m(ws4.cell(row=r,column=c))
pc(ws4.cell(row=r,column=5)); pc(ws4.cell(row=r,column=7))
ws4.cell(row=r+2,column=1,value='※ 스스 광고비 = 쏘핑 공구 수수료 8,300,292(대표님 지시로 광고비 처리) + 애드부스트 1,292,283 + 검색광고 1,280,511. 11번가 = 후불광고비. 로켓 14,263,512 = 대표님 확정(7/14).')
for col,w in zip('ABCDEFG',[18,13,12,12,8,13,10]): ws4.column_dimensions[col].width=w
# 시트5 증감: 6월 매출 내림차순 정렬 + 그래프 2종
ws5=wb.create_sheet('채널별 증감'); ws5['A1']='채널별 5월 → 6월 증감'; ws5['A1'].font=Font(bold=True,size=13)
ws5.append([]); ws5.append(['채널','5월 매출','6월 매출','매출증감','증감율','5월 이익','6월 이익','이익증감'])
for c in range(1,9): ws5.cell(row=3,column=c).fill=HF; ws5.cell(row=3,column=c).font=HFONT
allch=[ch for chs in GROUPS.values() for ch in chs]
allch.sort(key=lambda c:-tot_(c))  # 6월 매출 내림차순
GREEN=Font(color='1B7A3D',bold=True); RED=Font(color='C0392B',bold=True)
r=4
for ch in allch:
    m5,p5=MAY5.get(ch,(0,0)); m6=tot_(ch); p6=prof_(ch); dm=m6-m5
    ws5.append([ch,m5,m6,dm,(dm/m5 if m5 else 0),p5,p6,p6-p5])
    for c in [2,3,4,6,7,8]: m(ws5.cell(row=r,column=c))
    pc(ws5.cell(row=r,column=5))
    ws5.cell(row=r,column=4).font=GREEN if dm>=0 else RED
    ws5.cell(row=r,column=5).font=GREEN if dm>=0 else RED
    ws5.cell(row=r,column=8).font=GREEN if (p6-p5)>=0 else RED
    r+=1
tm5=sum(MAY5.get(c,(0,0))[0] for c in allch); tp5=sum(MAY5.get(c,(0,0))[1] for c in allch)
ws5.append(['■ 종합',tm5,GTr,GTr-tm5,((GTr-tm5)/tm5),tp5,GTp,GTp-tp5])
for c in range(1,9): ws5.cell(row=r,column=c).fill=TOT; ws5.cell(row=r,column=c).font=Font(bold=True)
for c in [2,3,4,6,7,8]: m(ws5.cell(row=r,column=c))
pc(ws5.cell(row=r,column=5))
last=r
# 상위 8개만 그래프(스케일 큰 채널) — 정렬돼 있으니 4~11행
TOPN=min(8,len(allch)); te=3+TOPN
from openpyxl.chart import BarChart, Reference, Series
c1=BarChart(); c1.type='col'; c1.grouping='clustered'; c1.title='채널별 5월 vs 6월 매출 (상위 8)'
c1.height=8.5; c1.width=20; c1.style=10; c1.y_axis.numFmt='#,##0'; c1.y_axis.majorGridlines=None
c1.add_data(Reference(ws5,min_col=2,max_col=3,min_row=3,max_row=te),titles_from_data=True)
c1.set_categories(Reference(ws5,min_col=1,min_row=4,max_row=te))
c1.gapWidth=60; ws5.add_chart(c1,'J3')
c2=BarChart(); c2.type='bar'; c2.title='매출 증감액 (6월−5월, 상위 8)'
c2.height=8.5; c2.width=20; c2.style=12; c2.x_axis.numFmt='#,##0'; c2.legend=None
c2.add_data(Reference(ws5,min_col=4,max_col=4,min_row=3,max_row=te),titles_from_data=True)
c2.set_categories(Reference(ws5,min_col=1,min_row=4,max_row=te))
ws5.add_chart(c2,'J20')
for col,w in zip('ABCDEFGH',[16,12,12,12,8,12,12,12]): ws5.column_dimensions[col].width=w

# 시트1 대시보드
ws1=wb.create_sheet('대시보드',0); ws1['A1']='2026년 6월 매출 종합 대시보드'; ws1['A1'].font=Font(bold=True,size=14)
ws1.append([]); ws1.append(['그룹','매출(총)','이익','이익률','매출비중'])
for c in range(1,6): ws1.cell(row=3,column=c).fill=HF; ws1.cell(row=3,column=c).font=HFONT
r=4
for gp,chs in GROUPS.items():
    gt=sum(tot_(c) for c in chs); gpp=sum(prof_(c) for c in chs)
    ws1.append([gp,gt,gpp,(gpp/gt if gt else 0),(gt/GTr)])
    for c in [2,3]: m(ws1.cell(row=r,column=c))
    pc(ws1.cell(row=r,column=4)); pc(ws1.cell(row=r,column=5)); r+=1
ws1.append(['■ 종합',GTr,GTp,(GTp/GTr),1])
for c in range(1,6): ws1.cell(row=r,column=c).fill=TOT; ws1.cell(row=r,column=c).font=Font(bold=True)
for c in [2,3]: m(ws1.cell(row=r,column=c))
pc(ws1.cell(row=r,column=4)); pc(ws1.cell(row=r,column=5))
for col,w in zip('ABCDE',[14,14,13,8,9]): ws1.column_dimensions[col].width=w

# 시트6 수수료·정산일정 (노란셀=대표님 확인/수정)
ws6=wb.create_sheet('수수료·정산일정')
ws6['A1']='채널별 수수료율 · 정산 일정 (노란셀 확인·수정)'; ws6['A1'].font=Font(bold=True,size=13)
ws6.append([]); ws6.append(['채널','수수료율','산출근거','정산주기','지급일정','비고'])
for c in range(1,7): ws6.cell(row=3,column=c).fill=HF; ws6.cell(row=3,column=c).font=HFONT
FEE=[
 ('카페24','3.4%','추정(PG)','수시(PG)','승인+영업일','자사몰 PG수수료'),
 ('스마트스토어','4.56%','실측(정산데이터)','빠른정산','구매확정 익일','Npay+매출연동 실효'),
 ('쿠팡(로켓배송)','—','직매입','월정산','매입 후 익월','수수료 없음'),
 ('쿠팡(로켓그로스)','11.67%','실측(판매vs정산)','빠른정산+주정산','익일90/100%·주70·30','+물류·광고 별도'),
 ('지마켓','추정~13%','추정','구매확정','확정+영업일',''),
 ('옥션','추정~13%','추정','구매확정','확정+영업일',''),
 ('11번가','추정~13%','추정','구매확정','확정+영업일','후불광고비 별도'),
 ('SSG','추정','추정','월마감','익월','위수탁'),
 ('GS샵','추정','추정','월정산','익월',''),
 ('오늘의집','추정~25%','추정','월정산','익월',''),
 ('에이블리','추정','추정','정산주기','',''),
 ('카카오선물하기','추정~10%','추정','정산주기','',''),
 ('카카오쇼핑하기','추정~10%','추정','정산주기','',''),
 ('라이플로우','0%','공급가직거래','수동/협의','발주 후 협의','수동발주'),
 ('PG컴퍼니','0%','공급가직거래','수동/협의','',''),
 ('두버','0%','공급가직거래','수동/협의','',''),
 ('에드가','0%','공급가직거래','수동/협의','',''),
]
r=4
for row in FEE:
    ws6.append(list(row))
    if '추정' in str(row[2]):
        for c in [2,3,4,5]: ws6.cell(row=r,column=c).fill=AS_
    r+=1
ws6.cell(row=r+1,column=1,value='※ 수수료율·정산일정은 대표님 확인값으로 수정. 그로스 물류·광고는 별도(시트2/4).')
for col,w in zip('ABCDEF',[18,12,16,16,20,20]): ws6.column_dimensions[col].width=w

OUT=f"{B}/[시안] 2026.06 온라인 매출정산_종합.xlsx"; wb.save(OUT); print("\n✅ 저장 v5")
