# -*- coding: utf-8 -*-
"""로켓그로스(세이랩컴퍼니 A01707416) 2026년 6월 정산표 — 매출인식 6/1~6/30
build_gross_june_full.py(채움)와 동일 방법론. 광고비 = 광고센터 공식 XLSX."""
import openpyxl, glob, os, re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = "/Users/macmini_ky/Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브/Claude AI work space/02. 매출 정산/26.06/세이랩컴퍼니"
OUT = "/Users/macmini_ky/Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브/Claude AI work space/02. 매출 정산/26.06/26.06 로켓그로스 정산_세이랩컴퍼니.xlsx"
P1, P2 = "2026-06-01", "2026-06-30"

# 옵션ID -> 원가 (세트단가, VAT별도) — 2026-06-26 대표님 원가표 (식세기 3,304 / 한장수세미 1,860)
COST = {
    '95510001821': 3304,  '95556504824': 6608,  '95556504822': 9912,  '95556504823': 13216,  # 식세기 1/2/3/4
    '95463771120': 1860,  '95607833748': 1860,                                               # 수세미 1 (ID 2종)
    '95463791496': 3720,  '95463791493': 5580,  '95463791494': 7440,                          # 수세미 2/3/4
    '95463791492': 9300,  '95463791495': 11160,                                              # 수세미 5/6
}
def bucket(name):
    n = name or ''
    if '식기세척기' in n or '식세기' in n: return '식기세척기세제'
    if '수세미' in n: return '한장수세미'
    return '기타'

# ---- 1) 매출/정산대상액: CATEGORY_TR ----
prod = defaultdict(lambda: {'옵션명':'', '수량':0, '판매액':0, '정산대상액':0})
cyc  = defaultdict(lambda: {'판매액':0, '정산대상액':0, '수량':0})
os.chdir(SRC)
for f in glob.glob('*CATEGORY_TR*'):
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True); ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None: continue
        rec = str(row[4])[:10]
        if not (P1 <= rec <= P2): continue
        end = str(row[1])[:10]
        oid = str(row[11]).split('.')[0]
        try: qty=float(row[16] or 0); pa=float(row[17] or 0); st=float(row[23] or 0)
        except: continue
        p = prod[oid]; p['옵션명'] = (str(row[13]) + ' / ' + str(row[14]))[:60]
        p['수량']+=qty; p['판매액']+=pa; p['정산대상액']+=st
        cyc[end]['판매액']+=pa; cyc[end]['정산대상액']+=st; cyc[end]['수량']+=qty
    wb.close()
unknown = [o for o in prod if o not in COST]
assert not unknown, f"원가 미매핑 옵션ID: {unknown}"
JUNE_CYC = sorted(cyc.keys())

# ---- 2) 물류비: 정산주기별 최종비용(VAT포함) ----
def logi_by_cycle(pat, sheet):
    seen=set(); d=defaultdict(float)
    for f in glob.glob(pat):
        wb=openpyxl.load_workbook(f,read_only=True,data_only=True)
        if sheet not in wb.sheetnames: wb.close(); continue
        for row in wb[sheet].iter_rows(min_row=3, values_only=True):
            if not row or row[0] is None: continue
            end=str(row[0])[:10]
            if end in seen: continue
            seen.add(end)
            try: d[end]+=float(row[3] or 0)
            except: pass
        wb.close()
    return d
io=logi_by_cycle('*WAREHOUSING_SHIPPING*','입출고비')
ship=logi_by_cycle('*WAREHOUSING_SHIPPING*','배송비')
stor=logi_by_cycle('*STORAGE_FEE*','보관비')

# ---- 3) 광고비: 광고센터 공식 XLSX ----
AD_XLSX=[f for f in os.listdir(SRC) if 'pa_daily_keyword' in f and f.endswith('.xlsx')]
assert len(AD_XLSX)==1, f"광고 XLSX 1개여야 함: {AD_XLSX}"
CYC_DAYS={'2026-06-07':(1,7),'2026-06-14':(8,14),'2026-06-21':(15,21),'2026-06-28':(22,28),'2026-06-30':(29,30)}
ad_daily=defaultdict(float); ad_prod=defaultdict(float)
_awb=openpyxl.load_workbook(os.path.join(SRC,AD_XLSX[0]),read_only=True,data_only=True)
_aws=_awb.active; _rows=_aws.iter_rows(values_only=True); _hdr=None
for _row in _rows:
    if _row and any(v=='날짜' for v in _row if isinstance(v,str)): _hdr=list(_row); break
_iD=_hdr.index('날짜'); _iC=_hdr.index('광고비'); _iP=_hdr.index('광고집행 상품명')
for _row in _rows:
    if not _row or _row[_iD] is None: continue
    m=re.match(r"(\d{4})(\d{2})(\d{2})", str(_row[_iD]))
    if not m: continue
    d=f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    try: v=float(_row[_iC] or 0)
    except: continue
    ad_daily[d]+=v; ad_prod[bucket(str(_row[_iP] or ''))]+=v
_awb.close()
ad_cyc={end: sum(ad_daily.get(f"2026-06-{d:02d}",0) for d in range(d1,d2+1)) for end,(d1,d2) in CYC_DAYS.items()}

# ---- 4) 제품별 물류 실부과 ----
def agg_logi(sheet, tag, cA, cAB, cQ):
    seen=set(); res=defaultdict(lambda:{'수량':0,'발생':0,'실부과':0})
    for f in glob.glob('*WAREHOUSING_SHIPPING*'):
        wb=openpyxl.load_workbook(f,read_only=True,data_only=True)
        if sheet not in wb.sheetnames: wb.close(); continue
        for r in wb[sheet].iter_rows(values_only=True):
            if not r or r[5]!=tag: continue
            cy=str(r[1])[:10]
            if not cy.startswith('2026-06'): continue
            key=(cy,r[6],r[7],r[10])
            if key in seen: continue
            seen.add(key)
            try: A=float(r[cA] or 0); AB=float(r[cAB] or 0); q=float(r[cQ] or 0)
            except: continue
            d=res[bucket(str(r[12]))]; d['수량']+=q; d['발생']+=A; d['실부과']+=AB
        wb.close()
    return res
LIO=agg_logi('입출고비','입출고비 정산',22,24,19)
LSP=agg_logi('배송비','배송비 정산',21,23,20)

# ---- 5) 기타비용 + 주정산 수금 (어드민 실값, 2026-07-07 대표님 제공) ----
MILKRUN = 267080  # 밀크런(입고운송) 2건×133,540 — ①6/1~7 주정산 차감(7/3 지급분, 상세내역 검증) ②6/30 빠른정산 차감(스크린샷 검증). 대표님 결정: 6월 비용 귀속
SETTLE = [
    ('2026-06-07','6/1~6/7',  '2026-07-03',643073,'2026-08-03',332753),
    ('2026-06-14','6/8~6/14', '2026-07-10',696080,'2026-08-03',298269),
    ('2026-06-21','6/15~6/21','2026-07-20',274436,'2026-08-03',126918),
    ('2026-06-28','6/22~6/28','2026-07-27',0,     '2026-08-03',0),
    ('2026-06-30','6/29~6/30','2026-08-03',146425,'미발행',None),
]

# ================= 엑셀 작성 =================
thin=Side(style='thin',color='CCCCCC'); border=Border(thin,thin,thin,thin)
hf=PatternFill('solid',fgColor='4472C4'); hfont=Font(color='FFFFFF',bold=True,size=10)
tot=PatternFill('solid',fgColor='FFF2CC')
def hdr(ws,r,n):
    for c in range(1,n+1):
        cell=ws.cell(row=r,column=c); cell.fill=hf; cell.font=hfont
        cell.alignment=Alignment(horizontal='center',vertical='center'); cell.border=border
def fmt(ws,r1,r2,c1,c2,f='#,##0'):
    for row in ws.iter_rows(min_row=r1,max_row=r2,min_col=c1,max_col=c2):
        for cell in row:
            if isinstance(cell.value,(int,float)): cell.number_format=f
def pct(ws,r1,r2,c):
    for row in ws.iter_rows(min_row=r1,max_row=r2,min_col=c,max_col=c):
        for cell in row:
            if isinstance(cell.value,(int,float)): cell.number_format='0.0%'
def totrow(ws,r,n):
    for c in range(1,n+1):
        cell=ws.cell(row=r,column=c); cell.fill=tot; cell.font=Font(bold=True)

wb=Workbook()
ws=wb.active; ws.title='제품별 정산'
ws['A1']='로켓그로스(세이랩컴퍼니) 6월 제품별 정산 — 매출인식 6/1~6/30'; ws['A1'].font=Font(bold=True,size=13)
H=['옵션ID','상품/옵션','수량','판매액','정산대상액','원가(단가)','원가합계','매출이익(정산-원가)','이익률']
ws.append([]); ws.append(H); hdr(ws,3,len(H))
r=4; T=defaultdict(float)
for oid in sorted(prod,key=lambda x:-prod[x]['판매액']):
    p=prod[oid]; uc=COST[oid]; cc=uc*p['수량']; prof=p['정산대상액']-cc
    ws.append([oid,p['옵션명'],p['수량'],p['판매액'],p['정산대상액'],uc,cc,prof,prof/p['판매액'] if p['판매액'] else 0])
    T['수량']+=p['수량']; T['판매액']+=p['판매액']; T['정산대상액']+=p['정산대상액']; T['원가합계']+=cc; T['이익']+=prof
    r+=1
ws.append(['','합계',T['수량'],T['판매액'],T['정산대상액'],'',T['원가합계'],T['이익'],T['이익']/T['판매액']])
totrow(ws,r,9); fmt(ws,4,r,3,8); pct(ws,4,r,9)
for col,w in zip('ABCDEFGHI',[13,46,6,12,12,11,11,15,8]): ws.column_dimensions[col].width=w

ws2=wb.create_sheet('월 요약·정산주기')
ws2['A1']='6월 정산주기별 요약'; ws2['A1'].font=Font(bold=True,size=12)
H2=['정산주기(종료일)','매출인식','수량','판매액','공제(수수료·쿠폰)','정산대상액','입출고비','배송비','보관비','물류계','광고비','실수령(정산-물류-광고)','실수령률']
ws2.append([]); ws2.append(H2); hdr(ws2,3,len(H2))
RANGE={'2026-06-07':'6/1~6/7','2026-06-14':'6/8~6/14','2026-06-21':'6/15~6/21','2026-06-28':'6/22~6/28','2026-06-30':'6/29~6/30'}
rr=4; G=defaultdict(float)
for end in JUNE_CYC:
    c=cyc[end]; pa=c['판매액']; st=c['정산대상액']; fee=pa-st
    a,b,s2=io.get(end,0),ship.get(end,0),stor.get(end,0); lg=a+b+s2; adc=ad_cyc.get(end,0)
    net=st-lg-adc
    ws2.append([end,RANGE.get(end,''),c['수량'],pa,fee,st,a,b,s2,lg,adc,net,net/pa if pa else 0])
    for k,v in [('수량',c['수량']),('판매액',pa),('수수료',fee),('정산',st),('입출고',a),('배송',b),('보관',s2),('물류',lg),('광고',adc),('실수령',net)]: G[k]+=v
    rr+=1
ws2.append(['합계','',G['수량'],G['판매액'],G['수수료'],G['정산'],G['입출고'],G['배송'],G['보관'],G['물류'],G['광고'],G['실수령'],G['실수령']/G['판매액']])
totrow(ws2,rr,13); fmt(ws2,4,rr,3,12); pct(ws2,4,rr,13)
for col,w in zip('ABCDEFGHIJKLM',[15,10,7,12,13,12,10,10,8,10,11,18,9]): ws2.column_dimensions[col].width=w

ws5=wb.create_sheet('제품군 순마진')
ws5['A1']='6월 제품군별 순마진 — 정산대상액 − 원가 − 물류실부과(VAT포함) − 광고비'; ws5['A1'].font=Font(bold=True,size=12)
H5=['제품군','수량','판매액','정산대상액','원가','물류실부과(VAT포함)','광고비','순이익','순마진율(판매액대비)']
ws5.append([]); ws5.append(H5); hdr(ws5,3,len(H5))
pg=defaultdict(lambda: defaultdict(float))
for oid,p in prod.items():
    b=bucket(p['옵션명']); pg[b]['수량']+=p['수량']; pg[b]['판매액']+=p['판매액']; pg[b]['정산']+=p['정산대상액']; pg[b]['원가']+=COST[oid]*p['수량']
for b in set(LIO)|set(LSP): pg[b]['물류']+=(LIO[b]['실부과']+LSP[b]['실부과'])*1.1
for b,v in ad_prod.items():
    if v: pg[b]['광고']+=v
r5=4; T5=defaultdict(float)
for b in sorted(pg,key=lambda x:-pg[x]['판매액']):
    d=pg[b]; net=d['정산']-d['원가']-d['물류']-d['광고']
    ws5.append([b,d['수량'],d['판매액'],d['정산'],d['원가'],d['물류'],d['광고'],net,net/d['판매액'] if d['판매액'] else 0])
    for k in ['수량','판매액','정산','원가','물류','광고']: T5[k]+=d[k]
    T5['순이익']+=net; r5+=1
ws5.append(['소계',T5['수량'],T5['판매액'],T5['정산'],T5['원가'],T5['물류'],T5['광고'],T5['순이익'],T5['순이익']/T5['판매액']])
totrow(ws5,r5,9); r5+=1
ws5.append(['밀크런(입고운송, 제품 미배분)','','','','','','',-MILKRUN,''])
r5+=1
ws5.append(['총 순이익 (밀크런 반영)','','','','','','',T5['순이익']-MILKRUN,(T5['순이익']-MILKRUN)/T5['판매액']])
totrow(ws5,r5,9)
fmt(ws5,4,r5,2,8); pct(ws5,4,r5,9)
for col,w in zip('ABCDEFGHI',[24,7,12,12,11,17,11,12,10]): ws5.column_dimensions[col].width=w

ws3=wb.create_sheet('수금일정')
ws3['A1']='6월 매출분 주정산 수금일정 (어드민 정산현황 실값 2026-07-07 / 셀러월렛 선지급 별도)'; ws3['A1'].font=Font(bold=True,size=12)
H3=['정산주기','매출인식','1차 70% 지급일','1차 금액','2차 30% 지급일','2차 금액']
ws3.append([]); ws3.append(H3); hdr(ws3,3,len(H3))
r3=4
for end,rng,d1,a1,d2,a2 in SETTLE:
    ws3.append([end,rng,d1,a1,d2,a2 if a2 is not None else '미발행']); r3+=1
ws3.append(['','','합계 1차',sum(s[3] for s in SETTLE),'2차',sum(s[5] or 0 for s in SETTLE)])
totrow(ws3,r3,6); fmt(ws3,4,r3,4,4); fmt(ws3,4,r3,6,6)
ws3.cell(row=r3+2,column=1,value='※ 6/22~28 주기 0원 = 빠른정산(셀러월렛) 선지급 회수 상계 — 채움과 동일 패턴, 매출 정상. 빠른정산은 6월 하순 가입(정확 일자 미상), 확인된 셀러월렛 지급: 6/30 결제분 703,470(밀크런 133,540 차감 포함).')
for col,w in zip('ABCDEF',[14,12,15,12,15,12]): ws3.column_dimensions[col].width=w

ws4=wb.create_sheet('정산 메모')
notes=[
 '[로켓그로스 세이랩컴퍼니(A01707416) 6월 정산 메모 / 작성 2026-07-07]','',
 '[범위] 매출인식일 2026-06-01 ~ 06-30 (정산주기 5개). 5/26~5/31분은 6/26 첫 정산에서 처리 완료 — 본 표 미포함.',
 '[기준] 채움컴퍼니와 동일 방법론. 정산대상액 = CATEGORY_TR col23. 물류 = 주기별 최종비용(VAT포함).',
 '[특이] 6/15 매출인식 0건 (주기 6/21 파일에 행 없음 — 윙 화면 확인 대기).',
 '[광고비] 광고센터 공식 XLSX(A01707416_pa_daily_keyword_20260601_20260630) 지출 발생일 기준.',
 '[물류비] 신규 계정 비용제로(입고 90일) 면제 구간 — 사실상 0원. 면제 소진 시 채움 수준(판매액의 ~20%) 각오.',
 '[원가] 식기세척기세제 3,304 / 한장수세미 1,860 (VAT별도, 2026-06-26 대표님 원가표).',
 '[밀크런] 267,080 = 2건×133,540 (①6/1~7 주정산 차감 ②6/30 빠른정산 차감 — 각 최종지급액 검산 일치, 실차감 확정).',
 ' - 6월 비용 귀속 (대표님 결정 2026-07-07). 동일 금액 2건 = 정액요금 입고 2회 추정. 이중청구 의심 시 쿠팡 문의 여지.',
 ' - 6/30 빠른정산에 상품광고비 -6,662 (환급, 무효클릭 조정 추정) 확인.',
 '[6/15] 판매 0건 확정 — 재다운로드 파일이 기존과 데이터 완전 동일 (원본에 행 없음).',
 '[수금] 주정산 = 어드민 실값(2026-07-07). 빠른정산 = 6월 하순 가입(일자 미상), 6/22~28 주기부터 상계 0원 패턴.',
]
for i,n in enumerate(notes,1): ws4.cell(row=i,column=1,value=n)
ws4.cell(row=1,column=1).font=Font(bold=True,size=12); ws4.column_dimensions['A'].width=110

wb.save(OUT)
tpa=G['판매액']; tst=G['정산']; tlg=G['물류']; tad=G['광고']; tcost=T['원가합계']; tprof=tst-tcost-tlg-tad-MILKRUN
print('저장:',OUT); print()
print('=== 세이랩 6월 정산 요약 (6/1~6/30) ===')
print(f"판매액        {int(tpa):>11,}")
print(f"정산대상액    {int(tst):>11,}")
print(f"물류비        {int(tlg):>11,}")
print(f"광고비        {int(tad):>11,}")
print(f"원가          {int(tcost):>11,}")
print(f"밀크런(6월귀속) {MILKRUN:>11,}")
print(f"실수령(정산-물류-광고) {int(tst-tlg-tad):>11,} ({(tst-tlg-tad)/tpa*100:.1f}%)")
print(f"순이익(밀크런 반영) {int(tprof):>11,} ({tprof/tpa*100:.1f}%)")
print()
print("광고 배분합:", f"{sum(ad_prod.values()):,.0f}", "/ 주기합:", f"{tad:,.0f}")
print("광고 일자 범위:", min(ad_daily), "~", max(ad_daily))
print('=== 제품군 순마진 ===')
for b in sorted(pg,key=lambda x:-pg[x]['판매액']):
    d=pg[b]; net=d['정산']-d['원가']-d['물류']-d['광고']
    print(f"{b:10s} 판매 {d['판매액']:>10,.0f} 광고 {d['광고']:>9,.0f} 순이익 {net:>9,.0f} ({net/d['판매액']*100 if d['판매액'] else 0:5.1f}%)")
EOF_MARKER_NOT_USED = True
