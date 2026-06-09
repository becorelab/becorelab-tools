"""그로스 옵션별 물류비 단가표 생성
정산 원본(WAREHOUSING_SHIPPING) 전체 집계 → 옵션ID별 개당 입출고비/배송비 정가(VAT별도).
대표님 통찰: 사이즈별로 거의 고정값 → 단가표화. 실부과는 프로모션 룰로 별도 적용.
"""
import openpyxl, glob, json, os
from collections import defaultdict, Counter

GROSS_DIR = "/Users/macmini_ky/Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브/앱/매출 정산/26.05/그로스"
OUT = "/Users/macmini_ky/ClaudeAITeam/accounting/gross_logistics_rates.json"

def per_unit_rates():
    # 옵션ID -> {입출고비 개당단가 후보, 배송비 개당단가 후보, 옵션명, 사이즈}
    wh = defaultdict(list)   # 입출고비 개당
    ship = defaultdict(list) # 배송비 개당
    meta = {}                # 옵션ID -> (옵션명, 사이즈)
    files = glob.glob(os.path.join(GROSS_DIR, "*WAREHOUSING_SHIPPING*.xlsx"))
    seen_rows = set()  # 중복파일 방지 (주문ID+배송ID+거래유형)
    for f in files:
        wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
        # 입출고비: col10 옵션ID, col13 옵션명, col17 사이즈, col19 판매수량, col22 비용
        for sht, qty_c, fee_c, bucket in [("입출고비",19,22,wh), ("배송비",20,21,ship)]:
            if sht not in wb.sheetnames: continue
            ws = wb[sht]
            for r in ws.iter_rows(min_row=9, values_only=True):
                oid, qty, fee = r[10], r[qty_c], r[fee_c]
                if oid is None or not isinstance(fee,(int,float)) or not isinstance(qty,(int,float)) or qty<=0:
                    continue
                if fee <= 0:  # 환불(-)/면제(0) 제외, 정가만
                    continue
                key = (sht, r[6], r[7], r[3])  # 시트+주문ID+배송ID+발생일
                if key in seen_rows: continue
                seen_rows.add(key)
                bucket[str(int(oid))].append(round(fee/qty))
                if str(int(oid)) not in meta:
                    meta[str(int(oid))] = {"옵션명": r[13], "사이즈": r[17]}
        wb.close()
    # 옵션별 최빈 단가
    rates = {}
    allids = set(wh)|set(ship)
    for oid in allids:
        w = Counter(wh.get(oid,[])).most_common(1)
        s = Counter(ship.get(oid,[])).most_common(1)
        rates[oid] = {
            "옵션명": meta.get(oid,{}).get("옵션명"),
            "사이즈": meta.get(oid,{}).get("사이즈"),
            "입출고비_개당": w[0][0] if w else None,
            "배송비_개당": s[0][0] if s else None,
            "표본_입출고": len(wh.get(oid,[])), "표본_배송": len(ship.get(oid,[])),
        }
    return rates

if __name__ == "__main__":
    rates = per_unit_rates()
    with open(OUT,"w",encoding="utf-8") as f:
        json.dump(rates, f, ensure_ascii=False, indent=2)
    print(f"옵션 {len(rates)}개 단가표 → {OUT}\n")
    for oid, r in sorted(rates.items(), key=lambda x:-(x[1]['표본_입출고'])):
        print(f"  {oid} | {str(r['옵션명'])[:22]:22} [{r['사이즈']}] 입출고 {r['입출고비_개당']}/배송 {r['배송비_개당']} (n={r['표본_입출고']}/{r['표본_배송']})")
