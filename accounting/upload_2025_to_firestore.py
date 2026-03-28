"""
2025년 전체 매출 데이터를 Firestore에 업로드
엑셀 파일: 2025. 12 온라인 매출정산.xlsx → Firestore settlements/{uid}/months/2025-01 ~ 2025-12
"""
import os
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from google.cloud import firestore
import firebase_admin
from firebase_admin import credentials, auth

# Firebase 초기화
KEY_PATH = os.path.join(os.path.dirname(__file__), '..', 'sourcing', 'analyzer',
    'becorelab-tools-firebase-adminsdk-fbsvc-c665234c8b.json')
cred = credentials.Certificate(KEY_PATH)
if not firebase_admin._apps:
    firebase_admin.initialize_app(cred)
db = firestore.Client.from_service_account_json(KEY_PATH)

# 사용자 UID 찾기 (becorelab@gmail.com)
USER_EMAIL = 'becorelab@gmail.com'
try:
    user = auth.get_user_by_email(USER_EMAIL)
    USER_UID = user.uid
    print(f'[AUTH] {USER_EMAIL} → UID: {USER_UID}')
except Exception as e:
    print(f'[AUTH] 사용자를 찾을 수 없습니다: {e}')
    print('  → Firebase Console에서 로그인한 적 있는 계정의 UID를 직접 입력하세요.')
    sys.exit(1)

# 엑셀 로드
XLSX_PATH = os.path.join(os.path.dirname(__file__), '2025. 12 온라인 매출정산.xlsx')
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
ws = wb[wb.sheetnames[1]]  # 채널 판매별 상품현황

# 월별 시작 컬럼 (0-indexed, row 6에서 확인)
MONTH_START_COLS = {
    1: 5, 2: 14, 3: 23, 4: 32, 5: 41, 6: 50,
    7: 59, 8: 68, 9: 77, 10: 86, 11: 95, 12: 104
}
# 각 월 블록 내 오프셋: 0=판매수량, 1=매출액, 2=배송비, 3=매출+배송비, 4=단가원가, 5=원가, 6=마진율, 7=정산금액, 8=원가합계

# 채널명 매핑 (엑셀 채널명 → 앱 채널명)
CHANNEL_MAP = {
    '카페24 / 네이버스토어': '카페24',
    '스마트스토어': '스마트스토어',
    '쿠팡(로켓배송)': '로켓배송',
    '쿠팡': '로켓배송',
    '지마켓': '지마켓',
    '옥션': '옥션',
    '11번가': '11번가',
    '오늘의집': '오늘의집',
    '카카오쇼핑라이브': '카카오쇼핑',
    '카카오선물하기': '카카오선물',
    '두버': '두버',
    'GS샵': 'GS',
    '에이블리': '에이블리',
    '굿트리': '굿트리',
    '고그로우': '고그로우',
    '에드가': '에드가',
    '신세계': '신세계',
    '홈해살림': '지엠홀딩스',
}

def safe_num(v, default=0):
    if v is None or v == '' or v == '#DIV/0!':
        return default
    try:
        return float(v)
    except:
        return default

def safe_int(v, default=0):
    return int(safe_num(v, default))


# 상품 데이터 파싱
print('\n[PARSE] 엑셀 파싱 중...')

# {month: {channel: [products]}}
month_data = {m: {} for m in range(1, 13)}

current_channel = None
for r in range(10, ws.max_row + 1):
    # 채널명 (col 1, 0-indexed)
    ch_cell = ws.cell(r, 2).value
    if ch_cell:
        current_channel = ch_cell.strip()

    if not current_channel:
        continue

    # 상품명 (col 4, 0-indexed col 3)
    product_name = ws.cell(r, 5).value
    if not product_name:
        continue
    product_name = str(product_name).strip()

    # 매핑된 채널명
    mapped_ch = CHANNEL_MAP.get(current_channel, current_channel)

    # 월별 데이터 추출
    for month, start_col in MONTH_START_COLS.items():
        qty = safe_int(ws.cell(r, start_col + 1).value)
        revenue = safe_num(ws.cell(r, start_col + 2).value)  # 매출액
        settlement = safe_num(ws.cell(r, start_col + 8).value)  # 정산금액
        unit_cost = safe_num(ws.cell(r, start_col + 5).value)  # 단가원가
        total_cost = safe_num(ws.cell(r, start_col + 6).value)  # 원가
        profit_rate = safe_num(ws.cell(r, start_col + 7).value)  # 마진율

        if qty == 0 and revenue == 0 and settlement == 0:
            continue

        profit = settlement - total_cost if total_cost > 0 else 0

        if mapped_ch not in month_data[month]:
            month_data[month][mapped_ch] = []

        month_data[month][mapped_ch].append({
            'name': product_name,
            'qty': qty,
            'settlement': safe_int(settlement),
            'revenue': safe_int(revenue),
            'unitCost': safe_int(unit_cost),
            'totalCost': safe_int(total_cost),
            'profit': safe_int(profit),
            'profitRate': round(profit_rate, 4) if isinstance(profit_rate, float) else 0,
            'mapped': True,
        })

# Firestore 업로드
print('\n[UPLOAD] Firestore 업로드 시작...')
base_ref = db.collection('settlements').document(USER_UID).collection('months')

for month in range(1, 13):
    doc_id = f'2025-{month:02d}'
    channels = month_data[month]

    if not channels:
        print(f'  {doc_id}: 데이터 없음, 스킵')
        continue

    channel_doc = {}
    total_qty = 0
    total_sett = 0
    total_cost = 0
    total_profit = 0

    for ch_name, products in channels.items():
        ch_qty = sum(p['qty'] for p in products)
        ch_sett = sum(p['settlement'] for p in products)
        ch_cost = sum(p['totalCost'] for p in products)
        ch_profit = sum(p['profit'] for p in products)

        channel_doc[ch_name] = {
            'aggregated': products,
            'summary': {
                'qty': ch_qty,
                'settlement': ch_sett,
                'cost': ch_cost,
                'profit': ch_profit,
                'profitRate': ch_profit / ch_sett if ch_sett > 0 else 0,
            }
        }
        total_qty += ch_qty
        total_sett += ch_sett
        total_cost += ch_cost
        total_profit += ch_profit

    doc_data = {
        'savedAt': firestore.SERVER_TIMESTAMP,
        'channels': channel_doc,
        'totalSummary': {
            'qty': total_qty,
            'settlement': total_sett,
            'cost': total_cost,
            'profit': total_profit,
            'profitRate': total_profit / total_sett if total_sett > 0 else 0,
        }
    }

    base_ref.document(doc_id).set(doc_data)
    ch_count = len(channels)
    prod_count = sum(len(v) for v in channels.values())
    print(f'  ✅ {doc_id}: {ch_count}개 채널, {prod_count}개 상품, 정산 {total_sett:,.0f}원')

print('\n[DONE] 2025년 1~12월 전체 업로드 완료! 🎉')
print('  → 매출정산 앱에서 월 선택하면 데이터가 자동으로 나타납니다.')
