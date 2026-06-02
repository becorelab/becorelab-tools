# -*- coding: utf-8 -*-
"""로켓그로스 5월 수금일정 (2-track: 일반정산 5/1~13 + 빠른정산 5/14~31)"""
import json, datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 일반정산분 (5/1~13): 정산주기 -> (정산대상액, 광고비추정, 종료일)
GENERAL = {
    '~5/3 (5/1~3)':   (1541315, 210847, '2026-05-03'),
    '~5/10 (5/4~10)': (4154806, 721953, '2026-05-10'),
    '~5/17 (5/11~13)':(1965140, 306462, '2026-05-13'),  # 5/14~17은 빠른정산이라 제외
}
fastpay = json.load(open('/tmp/fastpay_daily.json'))

def biz_plus(dstr, days=28):
    return (datetime.date.fromisoformat(dstr) + datetime.timedelta(days=days)).isoformat()

path='/Users/macmini_ky/Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브/앱/매출 정산/26.05/26.05 로켓그로스 정산_채움컴퍼니.xlsx'
wb=load_workbook(path)
if '수금일정' in wb.sheetnames: del wb['수금일정']
ws=wb.create_sheet('수금일정', 1)

thin=Side(style='thin',color='CCCCCC'); border=Border(left=thin,right=thin,top=thin,bottom=thin)
hdrfill=PatternFill('solid',fgColor='4472C4'); hdrfont=Font(color='FFFFFF',bold=True,size=10)
secfill=PatternFill('solid',fgColor='D9E1F2'); totfill=PatternFill('solid',fgColor='FFF2CC')
def hdr_row(row, vals):
    for i,v in enumerate(vals,1):
        c=ws.cell(row=row,column=i,value=v); c.fill=hdrfill; c.font=hdrfont
        c.alignment=Alignment(horizontal='center'); c.border=border

TODAY='2026-06-02'
ws['A1']='로켓그로스 채움컴퍼니 — 2026년 5월 수금일정'; ws['A1'].font=Font(bold=True,size=13)
ws['A2']='빠른정산은 5/14 고객결제분부터 적용. 이전(5/1~13)은 일반 주정산 일정.'; ws['A2'].font=Font(size=9,italic=True)

# ── Track A: 일반 주정산 ──
ws.cell(row=4,column=1,value='[ Track A ] 일반 주정산 — 5/1~13 판매분').font=Font(bold=True,size=11)
hdr_row(5, ['정산주기','정산대상액','광고비(추정)','1차 70%(광고차감후)','1차 입금예정','2차 30%','2차 입금예정'])
r=6; tA1=0; tA2=0
for k,(settle,ad,end) in GENERAL.items():
    s1=settle*0.7-ad; s2=settle*0.3
    d1=biz_plus(end); d1m=d1+(' (입금완료)' if d1<=TODAY else ' (예정)')
    ws.cell(row=r,column=1,value=k); ws.cell(row=r,column=2,value=settle)
    ws.cell(row=r,column=3,value=ad); ws.cell(row=r,column=4,value=round(s1))
    ws.cell(row=r,column=5,value=d1m); ws.cell(row=r,column=6,value=round(s2))
    ws.cell(row=r,column=7,value='2026-07-15 (예정)')
    tA1+=s1; tA2+=s2; r+=1
ws.cell(row=r,column=1,value='소계'); ws.cell(row=r,column=4,value=round(tA1)); ws.cell(row=r,column=6,value=round(tA2))
for c in range(1,8):
    ws.cell(row=r,column=c).fill=totfill; ws.cell(row=r,column=c).font=Font(bold=True)
rA=r; r+=2

# ── Track B: 빠른정산 ──
ws.cell(row=r,column=1,value='[ Track B ] 빠른정산 — 5/14~31 판매분 (고객결제일 익일 매일 입금)').font=Font(bold=True,size=11)
r+=1
hdr_row(r, ['고객결제일','정산대상액','상품광고비','풀필먼트(물류)','빠른정산수수료','실수금(입금액)','입금일(익일)'])
r+=1; start=r; tB=0; tBad=0; tBful=0
for d in sorted(fastpay):
    v=fastpay[d]  # [판매,수수료,쿠폰,정산대상,광고,풀필,빠수,최종]
    indt=(datetime.date.fromisoformat(d)+datetime.timedelta(days=1)).isoformat()
    ws.cell(row=r,column=1,value=d); ws.cell(row=r,column=2,value=v[3])
    ws.cell(row=r,column=3,value=v[4]); ws.cell(row=r,column=4,value=v[5])
    ws.cell(row=r,column=5,value=v[6]); ws.cell(row=r,column=6,value=v[7])
    ws.cell(row=r,column=7,value=indt+' (완료)')
    tB+=v[7]; tBad+=v[4]; tBful+=v[5]; r+=1
ws.cell(row=r,column=1,value='소계'); ws.cell(row=r,column=3,value=tBad); ws.cell(row=r,column=4,value=tBful); ws.cell(row=r,column=6,value=tB)
for c in range(1,8):
    ws.cell(row=r,column=c).fill=totfill; ws.cell(row=r,column=c).font=Font(bold=True)
rB=r; r+=2

# ── 월 수금 요약 ──
ws.cell(row=r,column=1,value='[ 5월 판매분 수금 요약 ]').font=Font(bold=True,size=11); r+=1
summary=[
    ('빠른정산 실수금 (5/14~31, 입금완료)', round(tB)),
    ('일반정산 1차 (5/1~13, 6월 입금)', round(tA1)),
    ('일반정산 2차 (5/1~13, 7월 입금)', round(tA2)),
    ('합계 (5월 판매분 총 수금 추정)', round(tB+tA1+tA2)),
]
for name,val in summary:
    ws.cell(row=r,column=1,value=name); ws.cell(row=r,column=2,value=val)
    if '합계' in name:
        for c in range(1,3): ws.cell(row=r,column=c).fill=totfill; ws.cell(row=r,column=c).font=Font(bold=True)
    r+=1
r+=1
for note in ['※ 광고비(추정)=쿠팡 광고리포트 기준. 빠른정산분 광고비/풀필먼트는 Wing 실제 차감액(정확).',
             '※ 일반정산 물류비는 0원(비용제로 90일 혜택). 2차 입금일은 판매 익익월=7월(영업일 따라 변동).',
             '※ 빠른정산분은 고객결제일 익일 입금되어 6/1까지 전액 수금 완료.']:
    ws.cell(row=r,column=1,value=note).font=Font(size=9); r+=1

# 포맷
for row in ws.iter_rows(min_row=6,max_row=rB,min_col=2,max_col=6):
    for cell in row:
        if isinstance(cell.value,(int,float)): cell.number_format='#,##0'
for row in ws.iter_rows(min_row=r-10,max_row=r,min_col=2,max_col=2):
    for cell in row:
        if isinstance(cell.value,(int,float)): cell.number_format='#,##0'
ws.column_dimensions['A'].width=34
for col in 'BCDEFG': ws.column_dimensions[col].width=15

wb.save(path)
print('수금일정 시트 저장 완료')
print()
print('=== 5월 판매분 수금 요약 ===')
for name,val in summary:
    print('  %-40s %12s' % (name, format(val,',')))
