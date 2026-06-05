# -*- coding: utf-8 -*-
"""통합 5월 블록합계(매출금액/이익합계)를 SUMIFS(셀참조)로 변경 + 전 채널 검증"""
import openpyxl
DST="/Users/macmini_ky/Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브/앱/매출 정산/26.05/2026. 05 온라인 매출정산.xlsx"
wb=openpyxl.load_workbook(DST, data_only=False); ws=wb['월별 온라인 매출정산']
n=0
for R in range(1, ws.max_row+1):
    for col,colL in [(49,'AS'),(50,'AU')]:
        c=ws.cell(R,col)
        if isinstance(c.value,str) and c.value.startswith(f'=SUM(${colL}'.replace('$','')) and ws.cell(R,2).value:
            c.value=f'=SUMIFS(${colL}:${colL},$B:$B,$B{R})'; n+=1
        elif isinstance(c.value,str) and c.value.startswith(f'=SUM({colL}') and ws.cell(R,2).value:
            c.value=f'=SUMIFS(${colL}:${colL},$B:$B,$B{R})'; n+=1
wb.save(DST); print(f'✅ 블록합계 {n}개 → SUMIFS(셀참조)')

# 검증
wbv=openpyxl.load_workbook(DST, data_only=True); wsv=wbv['월별 온라인 매출정산']
expect={'카페24 / 네이버페이':28487162,'스마트스토어':12666380,'카카오선물하기':105920,
'쿠팡(로켓배송)':70192906,'옥션':696962,'지마켓':2561213,'11번가':1724963,'쿠팡':35106,
'신세계몰\n(에스에스지닷컴)':122185,'오늘의집':125172,'GS샵':66720,'에이블리':21606}
got={}
for R in range(2, wsv.max_row+1):
    ch=wsv.cell(R,2).value
    if ch in expect:
        a=wsv.cell(R,43).value or 0; b=wsv.cell(R,44).value or 0
        try: got[ch]=got.get(ch,0)+float(a)+float(b)
        except: pass
print(f"\n{'채널':<22}{'매트릭스':>13}{'기대값':>13}{'차이':>9}")
allok=True
for ch,exp in expect.items():
    g=got.get(ch,0); d=g-exp; mark='✅' if abs(d)<1 else '❌'
    if abs(d)>=1: allok=False
    print(f"{repr(ch)[:21]:<22}{g:>13,.0f}{exp:>13,.0f}{d:>9,.0f} {mark}")
print('\n'+('🎉 전 채널 일치! 오류 없음' if allok else '⚠️ 불일치 있음'))
