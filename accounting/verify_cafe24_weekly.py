"""
카페24 14개월 검증 - 주간 정산 방식
2025.01 ~ 2026.02 자사몰 실제 매출/이익 정확 계산
"""
import sys, io, os, json, re
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

BASE = 'N:/개인/Becorelab/03. 영업/20. 월별 매출정산/기타/자사몰 검증'
COST_FILE = f'{BASE}/상품명 및 원가 파일 전체 통합.xlsx'

# 1. 원가 맵 로드
print('[1/3] 원가 맵 로드...')
wb_cost = openpyxl.load_workbook(COST_FILE, data_only=True)
ws_cost = wb_cost['원가']
cost_map = {}
cost_map_norm = {}

def norm(s):
    return re.sub(r'\s+', '', s)

# Row 2~4 특수 항목
for r in [2, 3, 4]:
    name = str(ws_cost.cell(row=r, column=1).value or '').strip()
    cost = ws_cost.cell(row=r, column=3).value
    if name and cost:
        try:
            c = float(cost)
            cost_map[name] = c
            cost_map_norm[norm(name)] = c
        except: pass

# Row 7부터 본 데이터
for row in ws_cost.iter_rows(min_row=7, values_only=True):
    name = str(row[1] or '').strip()
    cost = row[4]  # E열 = 판관비포함원가(VAT포함)
    if name and cost:
        try:
            c = float(cost)
            cost_map[name] = c
            cost_map_norm[norm(name)] = c
        except: pass

print(f'  원가 맵: {len(cost_map)}개')

# 2. 품명 매핑
print('[2/3] 품명 매핑 로드...')
with open('cafe24_name_map_temp.json', 'r', encoding='utf-8') as f:
    name_map = json.load(f)
print(f'  품명 매핑: {len(name_map)}개')

def sf(v):
    if v is None or v == '': return 0.0
    try: return float(str(v).replace(',', ''))
    except: return 0.0

def process_month(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['원본']
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]

    def col_idx(kw):
        for i, h in enumerate(headers):
            if h and kw in str(h): return i
        return None

    IDX_ORDER = col_idx('주문번호')
    IDX_QTY = col_idx('수량')
    IDX_OPT_PRICE = col_idx('옵션+판매가')
    IDX_NAME = col_idx('주문상품명(옵션포함)')
    IDX_SHIP = col_idx('총 배송비(KRW)')
    IDX_COUPON = col_idx('쿠폰 할인금액(최초)')
    IDX_REFUND = col_idx('실제 환불금액')
    IDX_POINT = col_idx('사용한 적립금액(최종)')
    IDX_RANK = col_idx('회원등급 추가할인금액')
    IDX_APP = col_idx('앱 상품할인 금액(최종)')
    IDX_EXTRA = col_idx('상품별 추가할인금액')

    orders_seen = set()
    sum_gross = 0
    sum_ship = 0
    sum_coupon = 0
    sum_refund = 0
    sum_point = 0
    sum_rank = 0
    sum_app = 0
    sum_extra = 0
    total_payment = 0
    total_cost = 0
    total_qty = 0
    no_cost_count = 0
    no_cost_value = 0
    unmapped_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[IDX_ORDER]: continue
        order = str(row[IDX_ORDER]).strip()
        qty = int(sf(row[IDX_QTY]))
        opt_price = sf(row[IDX_OPT_PRICE])
        gross = qty * opt_price

        is_first = order not in orders_seen
        if is_first:
            orders_seen.add(order)

        coupon = sf(row[IDX_COUPON]) if is_first else 0
        point = sf(row[IDX_POINT]) if is_first else 0
        ship_first = sf(row[IDX_SHIP]) if is_first else 0
        refund = sf(row[IDX_REFUND])
        rank = sf(row[IDX_RANK])
        app = sf(row[IDX_APP])
        extra = sf(row[IDX_EXTRA])

        payment = gross - coupon - refund - point - rank - app - extra

        # 원가
        raw_name = str(row[IDX_NAME] or '').strip()
        final = name_map.get(raw_name, '')
        unit_cost = cost_map.get(final, 0) if final else 0
        if unit_cost == 0 and final:
            unit_cost = cost_map_norm.get(norm(final), 0)

        if not final:
            unmapped_count += 1
        elif unit_cost == 0:
            no_cost_count += 1
            no_cost_value += gross

        cost = unit_cost * qty

        total_payment += payment
        total_cost += cost
        total_qty += qty

        sum_gross += gross
        sum_coupon += coupon
        sum_refund += refund
        sum_point += point
        sum_rank += rank
        sum_app += app
        sum_extra += extra
        sum_ship += ship_first

    return {
        'gross': sum_gross,
        'ship': sum_ship,
        'coupon': sum_coupon,
        'refund': sum_refund,
        'point': sum_point,
        'rank': sum_rank,
        'app': sum_app,
        'extra': sum_extra,
        'payment': total_payment,
        'payment_with_ship': total_payment + sum_ship,
        'cost': total_cost,
        'profit': total_payment - total_cost,
        'margin_rate': (total_payment - total_cost) / total_payment if total_payment > 0 else 0,
        'qty': total_qty,
        'unmapped': unmapped_count,
        'no_cost': no_cost_count,
        'no_cost_value': no_cost_value,
    }

# 3. 14개월 처리
print('[3/3] 14개월 처리...\n')
months = []
for f in sorted(os.listdir(BASE)):
    if '카페24 원본' in f and f.endswith('.xlsx'):
        m = re.match(r'(\d{2})\. (\d{2}) 카페24', f)
        if m:
            year = '20' + m.group(1)
            month = m.group(2)
            months.append((f'{year}-{month}', f'{BASE}/{f}'))

results = {}
for label, fp in months:
    print(f'  {label} 처리 중...')
    results[label] = process_month(fp)

# 결과 출력
print()
print('=' * 110)
print(f"  카페24 14개월 — 주간 정산 방식 검증")
print('=' * 110)
print(f"{'월':<10} {'매출(gross)':>13} {'-쿠폰':>10} {'-환불':>10} {'-적립':>10} {'-기타':>10} {'+배송':>10} {'최종매출':>13} {'원가':>12} {'마진':>12} {'마진율':>7}")
print('-' * 130)

total_g = total_pay = total_pwship = total_co = total_pr = 0
for label, r in results.items():
    others = r['rank'] + r['app'] + r['extra']
    print(f"{label:<10} {r['gross']:>13,.0f} {r['coupon']:>10,.0f} {r['refund']:>10,.0f} {r['point']:>10,.0f} {others:>10,.0f} {r['ship']:>10,.0f} {r['payment_with_ship']:>13,.0f} {r['cost']:>12,.0f} {r['profit']:>12,.0f} {r['margin_rate']*100:>6.1f}%")
    total_g += r['gross']
    total_pay += r['payment']
    total_pwship += r['payment_with_ship']
    total_co += r['cost']
    total_pr += r['profit']

print('-' * 130)
total_margin = total_pr / total_pay if total_pay > 0 else 0
print(f"{'합계':<10} {total_g:>13,.0f} {'':>10} {'':>10} {'':>10} {'':>10} {'':>10} {total_pwship:>13,.0f} {total_co:>12,.0f} {total_pr:>12,.0f} {total_margin*100:>6.1f}%")

print()
print('※ 원가 매핑 안된 건수 (월별):')
for label, r in results.items():
    if r['no_cost'] > 0 or r['unmapped'] > 0:
        print(f"  {label}: 미매핑 {r['unmapped']}, 원가없음 {r['no_cost']} (₩{r['no_cost_value']:,.0f})")
