# -*- coding: utf-8 -*-
"""로켓그로스 3~5월 원가·이익·이익률 계산 + 월별요약 시트 업데이트"""
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

costm = json.load(open('/tmp/cost_by_month.json'))      # month -> 정산대상액,원가,미상수량
agg = json.load(open('/tmp/grossagg2.json'))
ad5 = json.load(open('/tmp/ad_week.json'))
ad34 = json.load(open('/tmp/ad_week34.json'))
fastpay = json.load(open('/tmp/fastpay_daily.json'))

# 월별 판매/정산/물류
def mtot(key):
    t = {'2026-03':0,'2026-04':0,'2026-05':0}
    for p,a in agg.items():
        t[p[:7]] += a[key]
    return t
sales = mtot('판매액'); settle = mtot('정산대상액')
logi = {m:0 for m in ['2026-03','2026-04','2026-05']}
for p,a in agg.items():
    logi[p[:7]] += a['입출고비']+a['배송비']+a['보관비']

# 광고비: 5월 = 5/1~13 추정 + 5/14~31 빠른정산 실제 / 4월 = 추정 / 3월 0
ad_5_1to13 = ad5['~5/03'] + ad5['~5/10'] + 306462  # 5/11~13
ad_5_14to31 = sum(fastpay[d][4] for d in fastpay)   # 빠른정산 상품광고비 실제
ad = {
    '2026-03': 0,
    '2026-04': sum(ad34.get(k,0) for k in ad34),
    '2026-05': ad_5_1to13 + ad_5_14to31,
}
MN = {'2026-03':'3월','2026-04':'4월','2026-05':'5월'}

print('=== 로켓그로스 3~5월 손익 ===')
print('%-5s %12s %12s %12s %10s %12s %12s %7s' % ('월','판매액','정산대상액','원가','물류비','광고비','이익','이익률'))
TS=TST=TC=TL=TA=TP=0
rows=[]
for m in ['2026-03','2026-04','2026-05']:
    s=sales[m]; st=settle[m]; c=costm[m]['원가']; l=logi[m]; a=ad[m]
    profit = st - c - l - a
    rate = profit/s if s else 0
    rows.append((m,s,st,c,l,a,profit,rate))
    print('%-5s %12s %12s %12s %10s %12s %12s %6.1f%%' % (
        MN[m], format(int(s),','), format(int(st),','), format(int(c),','),
        format(int(l),','), format(int(a),','), format(int(profit),','), rate*100))
    TS+=s; TST+=st; TC+=c; TL+=l; TA+=a; TP+=profit
print('%-5s %12s %12s %12s %10s %12s %12s %6.1f%%' % (
    '합계', format(int(TS),','), format(int(TST),','), format(int(TC),','),
    format(int(TL),','), format(int(TA),','), format(int(TP),','), TP/TS*100))

# ── 월별 요약 시트 재작성 (이익 포함) ──
path='/Users/macmini_ky/Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브/앱/매출 정산/26.05/26.05 로켓그로스 정산_채움컴퍼니.xlsx'
wb=load_workbook(path)
if '월별 손익' in wb.sheetnames: del wb['월별 손익']
ws=wb.create_sheet('월별 손익',0)
thin=Side(style='thin',color='CCCCCC'); border=Border(left=thin,right=thin,top=thin,bottom=thin)
hdrfill=PatternFill('solid',fgColor='4472C4'); hdrfont=Font(color='FFFFFF',bold=True,size=10)
totfill=PatternFill('solid',fgColor='FFF2CC'); pf=PatternFill('solid',fgColor='E2EFDA')
ws['A1']='로켓그로스(채움컴퍼니) 3~5월 손익 — 이익·이익률'; ws['A1'].font=Font(bold=True,size=13)
hdr=['월','판매액','정산대상액','원가','물류비','광고비','이익','이익률(판매)','이익률(정산대상)']
ws.append([]); ws.append(hdr)
for c in range(1,len(hdr)+1):
    cell=ws.cell(row=3,column=c); cell.fill=hdrfill; cell.font=hdrfont
    cell.alignment=Alignment(horizontal='center'); cell.border=border
r=4
for (m,s,st,c,l,a,profit,rate) in rows:
    ws.append([MN[m],s,st,c,l,a,round(profit),rate, profit/st if st else 0]); r+=1
ws.append(['합계',TS,TST,TC,TL,TA,round(TP),TP/TS, TP/TST])
for c in range(1,len(hdr)+1):
    ws.cell(row=r,column=c).fill=totfill; ws.cell(row=r,column=c).font=Font(bold=True)
for row in ws.iter_rows(min_row=4,max_row=r,min_col=2,max_col=7):
    for cell in row: cell.number_format='#,##0'
for row in ws.iter_rows(min_row=4,max_row=r,min_col=8,max_col=9):
    for cell in row: cell.number_format='0.0%'
# 이익 강조
for rr in range(4,r+1):
    ws.cell(row=rr,column=7).fill=pf
for col,w in zip('ABCDEFGHI',[6,13,13,12,10,12,12,12,14]): ws.column_dimensions[col].width=w
# 메모
ws.cell(row=r+2,column=1,value='※ 이익 = 정산대상액 − 원가 − 물류비 − 광고비 (쿠팡 그로스 영업이익. 본사 간접비/인건비 제외)')
ws.cell(row=r+3,column=1,value='※ 광고비: 5/14~31 빠른정산 실제값 + 5/1~13·4월 광고리포트 추정. 빠른정산수수료(~3만) 미반영')
ws.cell(row=r+4,column=1,value='※ 원가 미상 소량(집게5·와이셔츠1개) 제외, 영향 무시 수준. 물류비 3~4월 0원(비용제로 90일)')
wb.save(path)
print('\n월별 손익 시트 추가 완료')
