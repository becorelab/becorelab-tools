#!/usr/bin/env python3
"""
정산 검증 게이트 (validate_settlement.py)
==========================================
월별 통합정산 xlsx를 받아, 마무리 전에 "상품명 불일치로 이익이 누락되는" 문제를 자동으로 잡는다.
2026-06 5월 정산에서 '식세기/섬유탈취제 100ml' 등 이름이 원가시트와 안 맞아 54건/13.9M(이익 ~550만)이
누락됐던 사고의 재발 방지용. (배경: 메모리 reference_consolidated_settlement, 인수인계 2026-06-30)

핵심 원리:
- 원가시트 = 상품명 마스터(SSOT). 모든 판매 품명은 원가시트에 존재해야 VLOOKUP으로 이익이 잡힌다.
- 매출>0 인데 이익이 빈칸 = 원가 미매칭 = 이름 불일치(또는 진짜 신규). → 마무리 전에 잡는다.

사용법:
    python3 validate_settlement.py "<통합정산.xlsx>"
    python3 validate_settlement.py "<통합정산.xlsx>" --month 5     # 특정 월만 (기본: 전월 자동탐지)

종료코드: 문제 0건이면 0, 있으면 1 (CI/자동화에서 게이트로 사용 가능)
"""
import sys, re, argparse
from collections import defaultdict
import openpyxl

SHEET = '월별 온라인 매출정산'
COST_SHEET = '원가'
# 월별 블록: 시작열(1-based). 1월=F(6), 이후 9칸 간격. 블록 내 오프셋: 수량0 매출액1 배송2 매출+배송3 원가4 이익5
MONTH_BLOCK_START = {m: 6 + 9 * (m - 1) for m in range(1, 13)}


def col_letter(idx):
    s = ''
    while idx > 0:
        idx, r = divmod(idx - 1, 26)
        s = chr(65 + r) + s
    return s


def load_cost_names(wb):
    """원가시트 품명 집합 (B열=이름). 공백제거 정규화 키도 같이."""
    cs = wb[COST_SHEET]
    names, norm = set(), set()
    for r in range(1, cs.max_row + 1):
        v = cs.cell(r, 2).value  # B열
        if isinstance(v, str) and v.strip():
            names.add(v.strip())
            norm.add(v.strip().replace(' ', ''))
    return names, norm


def detect_month(ws):
    """데이터가 있는 가장 마지막 월 자동탐지 (매출액 합>0)."""
    last = None
    for m in range(1, 13):
        c = MONTH_BLOCK_START[m] + 1  # 매출액
        s = sum(ws.cell(r, c).value for r in range(10, ws.max_row + 1)
                if isinstance(ws.cell(r, c).value, (int, float)))
        if s:
            last = m
    return last


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('file')
    ap.add_argument('--month', type=int, default=None)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.file, data_only=True)
    if SHEET not in wb.sheetnames:
        print(f'❌ "{SHEET}" 시트 없음'); sys.exit(2)
    ws = wb[SHEET]
    cost_names, cost_norm = load_cost_names(wb)
    month = args.month or detect_month(ws)
    if not month:
        print('❌ 데이터 있는 월을 못 찾음'); sys.exit(2)

    bs = MONTH_BLOCK_START[month]
    C_QTY, C_AMT, C_SHIP, C_PROFIT = bs, bs + 1, bs + 2, bs + 5
    print(f'=== 정산 검증: {month}월 ({SHEET}) ===\n')

    # 캐시 유무 감지: 매출>0 행 중 이익이 숫자인 비율. 낮으면 openpyxl 저장 등으로 캐시 비워진 상태.
    sales_rows = [r for r in range(10, ws.max_row + 1)
                  if isinstance(ws.cell(r, C_AMT).value, (int, float)) and ws.cell(r, C_AMT).value > 0]
    num_profit = sum(1 for r in sales_rows if isinstance(ws.cell(r, C_PROFIT).value, (int, float)))
    cache_ok = sales_rows and (num_profit / len(sales_rows) > 0.3)
    if not cache_ok:
        print('ℹ️  이익 캐시 없음(엑셀 미재계산 파일) → 품명↔원가시트 대조만 수행.'
              ' 정확한 이익빈칸 확인은 엑셀에서 한 번 열어 저장 후 재실행.\n')

    # PRIMARY(캐시 무관): 매출>0 인데 품명이 원가시트에 없음 = 이름 불일치/신규 → 이익 누락 예정
    # SECONDARY(캐시 있을 때만): 매출>0 인데 이익 빈칸 = 실제 미계산
    bad = []
    by_ch = defaultdict(lambda: [0, 0.0])
    for r in sales_rows:
        b = ws.cell(r, 2).value
        e = ws.cell(r, 5).value
        amt = ws.cell(r, C_AMT).value
        prof = ws.cell(r, C_PROFIT).value
        en = (e.strip() if isinstance(e, str) else '')
        in_cost = en in cost_names or en.replace(' ', '') in cost_norm
        prof_blank = cache_ok and not isinstance(prof, (int, float))
        if (not in_cost) or prof_blank:
            ch = (str(b).strip() if isinstance(b, str) else '?')
            reason = []
            if not in_cost:
                reason.append('원가시트에없음')
            if prof_blank:
                reason.append('이익빈칸')
            bad.append((r, ch, en, amt, '+'.join(reason)))
            by_ch[ch][0] += 1
            by_ch[ch][1] += amt

    grand = sum(ws.cell(r, C_AMT).value for r in range(10, ws.max_row + 1)
                if isinstance(ws.cell(r, C_AMT).value, (int, float)))
    if not bad:
        print('✅ 통과: 매출 있는 모든 품명이 원가시트에 존재 + 이익 계산됨.')
        sys.exit(0)

    bad_amt = sum(x[3] for x in bad)
    print(f'🔴 문제 {len(bad)}건 / 매출 {bad_amt:,.0f}원' + (f' (전체의 {bad_amt/grand*100:.1f}%)' if grand else ''))
    print('   (매출>0 인데 품명이 원가시트에 없거나 이익이 빈칸 → 이름 불일치로 이익 누락 위험)\n')
    print('채널별:')
    for ch, (n, s) in sorted(by_ch.items(), key=lambda x: -x[1][1]):
        print(f'   {ch:18} {n:>3}건  매출 {s:>12,.0f}')
    print('\n상세 (매출 큰 순):')
    for r, ch, e, amt, why in sorted(bad, key=lambda x: -x[3])[:50]:
        print(f'   행{r:>5} [{ch[:12]:12}] {e[:30]:30} 매출 {amt:>11,.0f}  [{why}]')

    print(f'\n→ 조치: 품명을 원가시트({COST_SHEET}) 정식명으로 교체하면 이익 자동 계산.')
    print('   "하트" 규칙: 원본에 "하트" 있으면 하트식세기, 없으면 일반 식기세척기.')
    print('   원가시트에 아예 없는 진짜 신규는 원가부터 등록.')
    sys.exit(1)


if __name__ == '__main__':
    main()
