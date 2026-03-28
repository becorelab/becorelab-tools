"""
2023~2024년 매출 데이터를 Firestore에 업로드
"""
import os, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl
from google.cloud import firestore
import firebase_admin
from firebase_admin import credentials, auth

# Firebase 초기화
KEY_PATH = os.path.join(os.path.dirname(__file__), '..', 'sourcing', 'analyzer',
    'becorelab-tools-firebase-adminsdk-fbsvc-c665234c8b.json')
cred = credentials.Certificate(KEY_PATH)
if not firebase_admin._apps:
    firebase_admin.initialize_app(cred)
db = firestore.Client.from_service_account_json(KEY_PATH)

user = auth.get_user_by_email('becorelab@gmail.com')
USER_UID = user.uid
print(f'[AUTH] UID: {USER_UID}')

CHANNEL_MAP = {
    '카페24 / 네이버페이': '카페24',
    '카페24 / 네이버스토어': '카페24',
    '스마트스토어': '스마트스토어',
    '쿠팡(로켓배송)': '로켓배송',
    '쿠팡': '로켓배송',
    '지마켓': '지마켓',
    '옥션': '옥션',
    '11번가': '11번가',
    '오늘의집': '오늘의집',
    '카카오쇼핑라이브': '카카오쇼핑',
    '카카오선물하기': '카카오선물',
    '두버': '두버',
    'GS샵': 'GS',
    '에이블리': '에이블리',
    '굿트리': '굿트리',
    '고그로우': '고그로우',
    '에드가': '에드가',
    '신세계': '신세계',
    '홈해살림': '지엠홀딩스',
}

def safe_num(v, default=0):
    if v is None or v == '' or isinstance(v, str):
        return default
    try:
        return float(v)
    except:
        return default

def safe_int(v, default=0):
    return int(safe_num(v, default))


def process_file(year, filename, sheet_index, data_start_row, block_size):
    filepath = os.path.join(os.path.dirname(__file__), filename)
    print(f'\n{"="*50}')
    print(f'[{year}년] {filename}')
    print(f'{"="*50}')

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[sheet_index]]
    print(f'시트: {wb.sheetnames[sheet_index]} (rows={ws.max_row})')

    # 월별 시작 컬럼 찾기 (row 4, 5, 6 순서로 탐색)
    month_start = {}
    for check_row in [4, 5, 6]:
        for c in range(ws.max_column):
            v = ws.cell(check_row, c + 1).value
            if v and '월' in str(v):
                try:
                    m = int(str(v).replace('월', '').strip().lstrip('0'))
                    if 1 <= m <= 12:
                        month_start[m] = c  # 0-indexed
                except:
                    pass
        if month_start:
            break

    print(f'월별 시작 컬럼: {month_start}')

    # 블록 내 오프셋 결정
    # 2023/2024: 0=qty, 1=revenue, 2=shipping, 3=total, 4=unitcost, 5=profit, 6=settlement, 7=profitTotal
    # settlement은 offset 6 (매출금액/정산금액)

    month_data = {m: {} for m in range(1, 13)}
    current_channel = None

    for r in range(data_start_row, ws.max_row + 1):
        ch_cell = ws.cell(r, 2).value  # col B = 채널명
        if ch_cell:
            current_channel = str(ch_cell).strip()

        if not current_channel:
            continue

        product_name = ws.cell(r, 5).value  # col E = 상품명
        if not product_name:
            continue
        product_name = str(product_name).strip()

        mapped_ch = CHANNEL_MAP.get(current_channel, current_channel)

        for month, start_col in month_start.items():
            if month > 12:
                continue
            # 0-indexed → cell은 1-indexed
            qty = safe_int(ws.cell(r, start_col + 1).value)        # 판매수량
            revenue = safe_num(ws.cell(r, start_col + 2).value)    # 매출액
            settlement = safe_num(ws.cell(r, start_col + 7).value) # 정산금액/매출금액 (offset 6)
            unit_cost = safe_num(ws.cell(r, start_col + 5).value)  # 개별원가
            profit = safe_num(ws.cell(r, start_col + 6).value)     # 이익
            profit_total = safe_num(ws.cell(r, start_col + 8).value)  # 이익합계 (offset 7)

            if qty == 0 and revenue == 0 and settlement == 0:
                continue

            if mapped_ch not in month_data[month]:
                month_data[month][mapped_ch] = []

            month_data[month][mapped_ch].append({
                'name': product_name,
                'qty': qty,
                'settlement': safe_int(settlement),
                'revenue': safe_int(revenue),
                'unitCost': safe_int(unit_cost),
                'totalCost': safe_int(unit_cost * qty) if unit_cost > 0 else 0,
                'profit': safe_int(profit_total) if profit_total != 0 else safe_int(profit),
                'profitRate': round(profit / settlement, 4) if settlement > 0 and isinstance(profit, (int, float)) else 0,
                'mapped': True,
            })

    # Firestore 업로드
    print(f'\n[UPLOAD] Firestore 업로드...')
    base_ref = db.collection('settlements').document(USER_UID).collection('months')

    for month in range(1, 13):
        doc_id = f'{year}-{month:02d}'
        channels = month_data[month]

        if not channels:
            print(f'  {doc_id}: 데이터 없음')
            continue

        channel_doc = {}
        t_qty = t_sett = t_cost = t_profit = 0

        for ch_name, products in channels.items():
            ch_qty = sum(p['qty'] for p in products)
            ch_sett = sum(p['settlement'] for p in products)
            ch_cost = sum(p['totalCost'] for p in products)
            ch_profit = sum(p['profit'] for p in products)

            channel_doc[ch_name] = {
                'aggregated': products,
                'summary': {
                    'qty': ch_qty, 'settlement': ch_sett, 'cost': ch_cost,
                    'profit': ch_profit,
                    'profitRate': ch_profit / ch_sett if ch_sett > 0 else 0,
                }
            }
            t_qty += ch_qty; t_sett += ch_sett; t_cost += ch_cost; t_profit += ch_profit

        doc_data = {
            'savedAt': firestore.SERVER_TIMESTAMP,
            'channels': channel_doc,
            'totalSummary': {
                'qty': t_qty, 'settlement': t_sett, 'cost': t_cost, 'profit': t_profit,
                'profitRate': t_profit / t_sett if t_sett > 0 else 0,
            }
        }
        base_ref.document(doc_id).set(doc_data)
        ch_count = len(channels)
        prod_count = sum(len(v) for v in channels.values())
        print(f'  {doc_id}: {ch_count}개 채널, {prod_count}개 상품, 정산 {t_sett:,.0f}원')

    print(f'\n[DONE] {year}년 업로드 완료!')


# 2024년: sheet_index=1, data_start=10, block=8
process_file(2024, '2024. 12 온라인 매출정산 -원가수정.xlsx', 1, 10, 8)

# 2023년: sheet_index=2 (월별 온라인 매출정산), data_start=8, block=8
process_file(2023, '2023. 12 온라인 매출정산.xlsx', 2, 8, 8)

print('\n\n=== 전체 완료! 2023~2024년 Firestore 업로드 성공 ===')
