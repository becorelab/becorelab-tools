# -*- coding: utf-8 -*-
"""로켓그로스(채움컴퍼니) 2026년 6월 정산표 (6/1~6/14 매출인식분)
새 포맷(CATEGORY_TR / WAREHOUSING_SHIPPING / STORAGE_FEE) 기반."""
import openpyxl, glob, os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = "/Users/macmini_ky/Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브/앱/매출 정산/26.05/그로스"
OUT = "/Users/macmini_ky/Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브/앱/매출 정산/26.05/26.06 로켓그로스 정산_채움컴퍼니.xlsx"
P1, P2 = "2026-06-01", "2026-06-14"   # 6월 매출인식 범위

# 채움컴퍼니 원가표 (옵션ID -> 원가, 세트단가 / VAT별도) — 가이드 + 신규 얼룩 멀티팩
COST = {
    '95304424363': 3932, '95363500368': 7864, '95363500367': 11796,   # 바이올렛 1/2/3
    '95377476928': 3932,                                              # 베이비크림 1
    '95304888338': 3199, '95304888360': 6398, '95304888356': 9597,    # 얼룩 1/2/3
    '95060453340': 2200, '95060453345': 4400, '95060453346': 6600,    # 입테 1/2/3
    '95060453344': 8800, '95060453347': 11000, '95060453343': 13200,  # 입테 4/5/6
}

# ---- 1) 매출/정산대상액: CATEGORY_TR, 매출인식일 6/1~6/14 ----
prod = defaultdict(lambda: {'옵션명':'', '수량':0, '판매액':0, '정산대상액':0})
cyc  = defaultdict(lambda: {'판매액':0, '정산대상액':0})
os.chdir(SRC)
for f in glob.glob('*CATEGORY_TR*'):
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True); ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None: continue
        rec = str(row[4])[:10]
        if not (P1 <= rec <= P2): continue
        end = str(row[1])[:10]
        oid = str(row[11]).split('.')[0]
        try: qty=float(row[16] or 0); pa=float(row[17] or 0); st=float(row[23] or 0)
        except: continue
        p = prod[oid]; p['옵션명'] = (str(row[13]) + ' / ' + str(row[14]))[:50]
        p['수량']+=qty; p['판매액']+=pa; p['정산대상액']+=st
        cyc[end]['판매액']+=pa; cyc[end]['정산대상액']+=st
    wb.close()

# ---- 2) 물류비: WAREHOUSING_SHIPPING(입출고비/배송비) + STORAGE_FEE(보관비), 정산주기별 최종비용 ----
def logi_by_cycle(pat, sheet):
    seen=set(); d=defaultdict(float)
    for f in glob.glob(pat):
        wb=openpyxl.load_workbook(f,read_only=True,data_only=True)
        if sheet not in wb.sheetnames: wb.close(); continue
        for row in wb[sheet].iter_rows(min_row=3, values_only=True):
            if not row or row[0] is None: continue
            end=str(row[0])[:10]
            if end in seen: continue
            seen.add(end)
            try: d[end]+=float(row[3] or 0)
            except: pass
        wb.close()
    return d
io=logi_by_cycle('*WAREHOUSING_SHIPPING*','입출고비')
ship=logi_by_cycle('*WAREHOUSING_SHIPPING*','배송비')
stor=logi_by_cycle('*STORAGE_FEE*','보관비')
JUNE_CYC = sorted(cyc.keys())   # 6월 정산주기 종료일들

# ---- 3) 광고비 (쿠팡 광고 리포트 JSON, 6/1~6/14, 정산주기별) ----
import json
AD = "/Users/macmini_ky/ClaudeAITeam/marketing/coupang_data"
def adsum(d1,d2):
    t=0
    for day in range(d1,d2+1):
        fn=f"{AD}/A00940134_pa_daily_keyword_202606{day:02d}_202606{day:02d}.json"
        try: t+=sum(float(r.get('광고비',0) or 0) for r in json.load(open(fn)))
        except: pass
    return t
ad_cyc = {'2026-06-07': adsum(1,7), '2026-06-14': adsum(8,14)}

# ---- 주정산 수금 (정산 리포트 목록, 정확값) ----
# (정산주기, 매출인식범위, 1차70%지급일, 1차금액, 2차30%지급일, 2차금액)
SETTLE = [
    ('2026-06-07','2026-06-01~06-07','2026-07-03',390047,'미발행',None),
    ('2026-06-14','2026-06-08~06-14','2026-07-10',305791,'미발행',None),
]

# ================= 엑셀 작성 =================
thin=Side(style='thin',color='CCCCCC'); border=Border(thin,thin,thin,thin)
hf=PatternFill('solid',fgColor='4472C4'); hfont=Font(color='FFFFFF',bold=True,size=10)
tot=PatternFill('solid',fgColor='FFF2CC')
def hdr(ws,r,n):
    for c in range(1,n+1):
        cell=ws.cell(row=r,column=c); cell.fill=hf; cell.font=hfont
        cell.alignment=Alignment(horizontal='center',vertical='center'); cell.border=border
def fmt(ws,r1,r2,c1,c2,f='#,##0'):
    for row in ws.iter_rows(min_row=r1,max_row=r2,min_col=c1,max_col=c2):
        for cell in row:
            if isinstance(cell.value,(int,float)): cell.number_format=f

wb=Workbook()

# Sheet1: 제품별 정산
ws=wb.active; ws.title='제품별 정산'
ws['A1']='로켓그로스(채움컴퍼니) 6월 제품별 정산 — 매출인식 6/1~6/14'; ws['A1'].font=Font(bold=True,size=13)
H=['옵션ID','상품/옵션','수량','판매액','정산대상액','원가(단가)','원가합계','매출이익(정산-원가)','이익률']
ws.append([]); ws.append(H); hdr(ws,3,len(H))
r=4; T=defaultdict(float)
for oid in sorted(prod,key=lambda x:-prod[x]['판매액']):
    p=prod[oid]; uc=COST.get(oid); cc=uc*p['수량'] if uc else None
    prof=(p['정산대상액']-cc) if cc is not None else None
    pr=(prof/p['판매액']) if (prof is not None and p['판매액']) else None
    ws.append([oid,p['옵션명'],p['수량'],p['판매액'],p['정산대상액'],uc,cc,prof,pr])
    T['수량']+=p['수량']; T['판매액']+=p['판매액']; T['정산대상액']+=p['정산대상액']
    if cc is not None: T['원가합계']+=cc
    if prof is not None: T['이익']+=prof
    r+=1
ws.append(['','합계',T['수량'],T['판매액'],T['정산대상액'],'',T['원가합계'],T['이익'],T['이익']/T['판매액']])
for c in range(1,10):
    cell=ws.cell(row=r,column=c); cell.fill=tot; cell.font=Font(bold=True)
fmt(ws,4,r,3,8);
for row in ws.iter_rows(min_row=4,max_row=r,min_col=9,max_col=9):
    for cell in row:
        if isinstance(cell.value,(int,float)): cell.number_format='0.0%'
for col,w in zip('ABCDEFGHI',[13,40,6,12,12,11,11,15,8]): ws.column_dimensions[col].width=w

# Sheet2: 월 요약 (정산주기별 + 합계)
ws2=wb.create_sheet('월 요약·정산주기')
ws2['A1']='6월 정산주기별 요약 (정산대상액 = 판매액 - 판매수수료 차감 후)'; ws2['A1'].font=Font(bold=True,size=12)
H2=['정산주기(종료일)','매출인식','판매액','판매수수료','정산대상액','입출고비','배송비','보관비','물류계','광고비','실수령(정산-물류-광고)','실수령률']
ws2.append([]); ws2.append(H2); hdr(ws2,3,len(H2))
rr=4; G=defaultdict(float)
RANGE={'2026-06-07':'6/1~6/7','2026-06-14':'6/8~6/14'}
for end in JUNE_CYC:
    pa=cyc[end]['판매액']; st=cyc[end]['정산대상액']; fee=pa-st
    a,b,c=io.get(end,0),ship.get(end,0),stor.get(end,0); lg=a+b+c; adc=ad_cyc.get(end,0)
    net=st-lg-adc
    ws2.append([end,RANGE.get(end,''),pa,fee,st,a,b,c,lg,adc,net,net/pa if pa else 0])
    for k,v in [('판매액',pa),('수수료',fee),('정산대상액',st),('입출고',a),('배송',b),('보관',c),('물류',lg),('광고',adc),('실수령',net)]: G[k]+=v
    rr+=1
ws2.append(['합계','',G['판매액'],G['수수료'],G['정산대상액'],G['입출고'],G['배송'],G['보관'],G['물류'],G['광고'],G['실수령'],G['실수령']/G['판매액']])
for c in range(1,13):
    cell=ws2.cell(row=rr,column=c); cell.fill=tot; cell.font=Font(bold=True)
fmt(ws2,4,rr,3,11)
for row in ws2.iter_rows(min_row=4,max_row=rr,min_col=12,max_col=12):
    for cell in row:
        if isinstance(cell.value,(int,float)): cell.number_format='0.0%'
for col,w in zip('ABCDEFGHIJKL',[16,10,12,11,12,10,10,8,10,11,18,9]): ws2.column_dimensions[col].width=w

# Sheet 제품별 물류비 (발생 vs 실부과 vs 면제)
def short(name):
    n=name or ''
    if '바이올렛' in n: return '바이올렛'
    if '베이비' in n: return '베이비크림'
    if '얼룩' in n or '목때' in n: return '얼룩제거제'
    if '입벌림' in n or '오모모' in n: return '입테이프'
    return n[:10]
def agg_logi(sheet, tag, cA, cAB, cQ):
    seen=set(); res=defaultdict(lambda:{'수량':0,'발생':0,'실부과':0})
    for f in glob.glob('*WAREHOUSING_SHIPPING*'):
        wb=openpyxl.load_workbook(f,read_only=True,data_only=True)
        if sheet not in wb.sheetnames: wb.close(); continue
        for r in wb[sheet].iter_rows(values_only=True):
            if not r or r[5]!=tag: continue
            cy=str(r[1])[:10]
            if not cy.startswith('2026-06'): continue
            key=(cy,r[6],r[7],r[10])
            if key in seen: continue
            seen.add(key)
            try: A=float(r[cA] or 0); AB=float(r[cAB] or 0); q=float(r[cQ] or 0)
            except: continue
            d=res[short(r[12])]; d['수량']+=q; d['발생']+=A; d['실부과']+=AB
        wb.close()
    return res
LIO=agg_logi('입출고비','입출고비 정산',22,24,19)
LSP=agg_logi('배송비','배송비 정산',21,23,20)
wsl=wb.create_sheet('제품별 물류비')
wsl['A1']='6월 제품별 물류비 — 발생비용(할인前) vs 실부과 vs 면제 (6/1~6/14, VAT별도)'; wsl['A1'].font=Font(bold=True,size=12)
HL=['제품','입출고 발생','입출고 실부과','배송 발생','배송 실부과','발생합계','실부과합계','면제액','면제율','실부과(VAT포함)']
wsl.append([]); wsl.append(HL); hdr(wsl,3,len(HL))
rl=4; LT=defaultdict(float)
for p in sorted(set(LIO)|set(LSP), key=lambda x:-(LIO[x]['실부과']+LSP[x]['실부과'])):
    a1=LIO[p]['발생']; b1=LIO[p]['실부과']; a2=LSP[p]['발생']; b2=LSP[p]['실부과']
    ta=a1+a2; tb=b1+b2; ex=ta-tb; exr=ex/ta if ta else 0
    wsl.append([p,a1,b1,a2,b2,ta,tb,ex,exr,round(tb*1.1)])
    for k,v in [('a1',a1),('b1',b1),('a2',a2),('b2',b2),('ta',ta),('tb',tb),('ex',ex)]: LT[k]+=v
    rl+=1
tta=LT['ta']; ttb=LT['tb']
wsl.append(['합계',LT['a1'],LT['b1'],LT['a2'],LT['b2'],tta,ttb,tta-ttb,(tta-ttb)/tta,round(ttb*1.1)])
for c in range(1,11):
    cell=wsl.cell(row=rl,column=c); cell.fill=tot; cell.font=Font(bold=True)
fmt(wsl,4,rl,2,8);
for row in wsl.iter_rows(min_row=4,max_row=rl,min_col=10,max_col=10):
    for cell in row:
        if isinstance(cell.value,(int,float)): cell.number_format='#,##0'
for row in wsl.iter_rows(min_row=4,max_row=rl,min_col=9,max_col=9):
    for cell in row:
        if isinstance(cell.value,(int,float)): cell.number_format='0.0%'
note_r=rl+2
wsl.cell(row=note_r,column=1,value='※ 요율 고정: 입출고 극소형 개당 1,650원 / 배송 극소형 개당 2,200원 (5월말·6월 동일). "실부과"는 비용제로(입고90일) 만료분. 보관비 0원(세이버 면제).')
wsl.cell(row=note_r+1,column=1,value='※ 면제분(41%)은 저가상품 할인(14,000원 미만, ~35-40%). 이 할인 2027.1.31 종료 시 발생비용 전액 부과 → 물류비 약 +70% 증가 리스크.')
for col,w in zip('ABCDEFGHIJ',[12,11,12,10,11,11,12,11,8,14]): wsl.column_dimensions[col].width=w

# Sheet3: 수금일정 (주정산)
ws3=wb.create_sheet('수금일정')
ws3['A1']='6월 매출분 주정산 수금일정 (빠른정산은 익일 선지급 별도)'; ws3['A1'].font=Font(bold=True,size=12)
H3=['정산주기','매출인식','1차 70% 지급일','1차 금액','2차 30% 지급일','2차 금액']
ws3.append([]); ws3.append(H3); hdr(ws3,3,len(H3))
r3=4
for end,rng,d1,a1,d2,a2 in SETTLE:
    ws3.append([end,rng,d1,a1,d2,a2 if a2 else '미발행(7월말~8월 예정)']); r3+=1
ws3.append(['','','합계 1차',sum(s[3] for s in SETTLE),'',''])
ws3.cell(row=r3,column=4).font=Font(bold=True)
fmt(ws3,4,r3,4,4); ws3.cell(row=r3,column=6).number_format='General'
for col,w in zip('ABCDEF',[14,14,15,12,17,22]): ws3.column_dimensions[col].width=w

# Sheet4: 메모
ws4=wb.create_sheet('정산 메모')
notes=[
 '[로켓그로스 채움컴퍼니 6월 정산 메모 / 작성 2026-06-18]','',
 '[범위] 매출인식일 2026-06-01 ~ 06-14 (어드민 자료 한계, 6/15~ 정산주기 미마감).',
 '[기준] 정산주기 종료일 기준. 정산대상액 = 판매액 - 판매수수료(약8.6%+VAT) 차감 후 (새 포맷 col23).',
 '',
 '[물류비] WAREHOUSING_SHIPPING(입출고비/배송비 시트) + STORAGE_FEE(보관비) 각 최종비용(VAT포함).',
 ' - 보관비 0원 (세이버 혜택). 정산주기 종료일별 합계 사용 (SKU 분해 불가, 새 포맷은 주기합계만 제공).',
 '',
 '[광고비] 쿠팡 광고 리포트 JSON(계정 A00940134) 일별 광고비 합산. 6/1~7=651,272 / 6/8~14=718,831.',
 ' - 광고비(지출)는 소급반영 영향 없어 JSON 사용 OK (전환/ROAS와 무관).',
 ' - Wing 빠른정산 일별 차감 상품광고비와 1~5% 오차 가능.',
 '',
 '[수금] 주정산 1차 70%: 6/1~7분 7/3(390,047), 6/8~14분 7/10(305,791) — 정산 리포트 목록 실값.',
 ' - 2차 30%은 미발행(7월말~8월 예정). 빠른정산(셀러월렛)은 익일 90% 선지급으로 별도 수령 완료.',
 ' - 1차 지급액이 작은 이유: 빠른정산으로 이미 선지급된 금액이 1차에서 차감되기 때문 (이중수령 아님).',
 '',
 '[원가] 가이드 채움 원가표(세트단가, VAT별도). 얼룩 2개입=3,199x2, 3개입=3,199x3 적용.',
]
for i,n in enumerate(notes,1): ws4.cell(row=i,column=1,value=n)
ws4.cell(row=1,column=1).font=Font(bold=True,size=12); ws4.column_dimensions['A'].width=108

wb.save(OUT)
# 콘솔 요약
print('저장:',OUT); print()
tpa=G['판매액']; tst=G['정산대상액']; tlg=G['물류']; tad=G['광고']; tcost=T['원가합계']; tprof=tst-tcost-tlg-tad
print('=== 6월 그로스 정산 요약 (6/1~6/14) ===')
print(f"판매액      {int(tpa):>12,}")
print(f"판매수수료  {int(tpa-tst):>12,}")
print(f"정산대상액  {int(tst):>12,}")
print(f"물류비      {int(tlg):>12,}  (입출고 {int(G['입출고']):,} + 배송 {int(G['배송']):,} + 보관 {int(G['보관']):,})")
print(f"광고비      {int(tad):>12,}")
print(f"실수령(정산-물류-광고) {int(tst-tlg-tad):>12,}  ({(tst-tlg-tad)/tpa*100:.1f}%)")
print(f"원가        {int(tcost):>12,}")
print(f"순이익(정산-원가-물류-광고) {int(tprof):>12,}  ({tprof/tpa*100:.1f}%)")
