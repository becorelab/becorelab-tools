# -*- coding: utf-8 -*-
"""로켓그로스(채움컴퍼니) 3~5월 정산표 생성"""
import json, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

agg = json.load(open('/tmp/grossagg2.json'))
ad5 = json.load(open('/tmp/ad_week.json'))
ad34 = json.load(open('/tmp/ad_week34.json'))

admap = {
    '2026-05-03': ad5['~5/03'], '2026-05-10': ad5['~5/10'], '2026-05-17': ad5['~5/17'],
    '2026-05-24': ad5['~5/24'], '2026-05-31': ad5['~5/31'],
    '2026-04-05': ad34.get('~4/05', 0), '2026-04-12': ad34.get('~4/12', 0),
    '2026-04-19': ad34.get('~4/19', 0), '2026-04-26': ad34.get('~4/26', 0), '2026-04-30': ad34.get('~4/30', 0),
}
periods = sorted(agg.keys())

def payout1(end):
    return (datetime.date.fromisoformat(end) + datetime.timedelta(days=28)).isoformat()
def payout2(end):
    d = datetime.date.fromisoformat(end); m = d.month + 2; y = d.year
    if m > 12: m -= 12; y += 1
    return datetime.date(y, m, 15).isoformat()

wb = Workbook()
thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdrfill = PatternFill('solid', fgColor='4472C4'); hdrfont = Font(color='FFFFFF', bold=True, size=10)
totfill = PatternFill('solid', fgColor='FFF2CC')

def style_hdr(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c); cell.fill = hdrfill; cell.font = hdrfont
        cell.alignment = Alignment(horizontal='center', vertical='center'); cell.border = border

MN = {'2026-03': '3월', '2026-04': '4월', '2026-05': '5월'}

# Sheet1 월별 요약
ws = wb.active; ws.title = '월별 요약'
ws['A1'] = '로켓그로스(채움컴퍼니) 정산 요약 - 2026년 3~5월 (정산주기 기준)'
ws['A1'].font = Font(bold=True, size=13)
hdr = ['월', '판매액', '판매수수료', '정산대상액', '입출고비', '배송비', '보관비', '물류비계', '광고비', '실수령(추정)', '실수령률']
ws.append([]); ws.append(hdr); style_hdr(ws, 3, len(hdr))
mt = {}
for p in periods:
    a = agg[p]; m = p[:7]
    mt.setdefault(m, {'판매액': 0, '정산대상액': 0, '입출고비': 0, '배송비': 0, '보관비': 0, '광고비': 0})
    for k in ['판매액', '정산대상액', '입출고비', '배송비', '보관비']:
        mt[m][k] += a[k]
    mt[m]['광고비'] += admap.get(p, 0)
r = 4
for m in sorted(mt):
    t = mt[m]; fee = t['판매액'] - t['정산대상액']; logi = t['입출고비'] + t['배송비'] + t['보관비']
    net = t['정산대상액'] - logi - t['광고비']; rate = net / t['판매액'] if t['판매액'] else 0
    ws.append([MN[m], t['판매액'], fee, t['정산대상액'], t['입출고비'], t['배송비'], t['보관비'], logi, t['광고비'], net, rate])
    r += 1
allt = {k: sum(mt[m][k] for m in mt) for k in ['판매액', '정산대상액', '입출고비', '배송비', '보관비', '광고비']}
fee = allt['판매액'] - allt['정산대상액']; logi = allt['입출고비'] + allt['배송비'] + allt['보관비']
net = allt['정산대상액'] - logi - allt['광고비']
ws.append(['합계', allt['판매액'], fee, allt['정산대상액'], allt['입출고비'], allt['배송비'], allt['보관비'], logi, allt['광고비'], net, net / allt['판매액']])
for c in range(1, 12):
    cell = ws.cell(row=r, column=c); cell.fill = totfill; cell.font = Font(bold=True)
for row in ws.iter_rows(min_row=4, max_row=r, min_col=2, max_col=10):
    for cell in row: cell.number_format = '#,##0'
for row in ws.iter_rows(min_row=4, max_row=r, min_col=11, max_col=11):
    for cell in row: cell.number_format = '0.0%'
for col, w in zip('ABCDEFGHIJK', [7, 13, 12, 13, 10, 10, 8, 11, 12, 13, 9]):
    ws.column_dimensions[col].width = w

# Sheet2 주차별
ws2 = wb.create_sheet('주차별 정산·수금')
ws2['A1'] = '주차별 정산 + 수금 예정일 (1차 70% / 2차 30%)'; ws2['A1'].font = Font(bold=True, size=12)
hdr2 = ['정산주기(종료일)', '판매액', '정산대상액', '물류비', '광고비', '실수령(추정)', '1차 70%', '1차 입금예정', '2차 30%', '2차 입금예정']
ws2.append([]); ws2.append(hdr2); style_hdr(ws2, 3, len(hdr2))
rr = 4
for p in periods:
    a = agg[p]; logi = a['입출고비'] + a['배송비'] + a['보관비']; adc = admap.get(p, 0)
    net = a['정산대상액'] - logi - adc
    ws2.append([p, a['판매액'], a['정산대상액'], logi, adc, net, a['정산대상액'] * 0.7, payout1(p), a['정산대상액'] * 0.3, payout2(p)])
    rr += 1
tp = {k: sum(agg[p][k] for p in periods) for k in ['판매액', '정산대상액', '입출고비', '배송비', '보관비']}
tadc = sum(admap.get(p, 0) for p in periods); tlogi = tp['입출고비'] + tp['배송비'] + tp['보관비']
tnet = tp['정산대상액'] - tlogi - tadc
ws2.append(['합계', tp['판매액'], tp['정산대상액'], tlogi, tadc, tnet, tp['정산대상액'] * 0.7, '', tp['정산대상액'] * 0.3, ''])
for c in range(1, 11):
    cell = ws2.cell(row=rr, column=c); cell.fill = totfill; cell.font = Font(bold=True)
for row in ws2.iter_rows(min_row=4, max_row=rr, min_col=2, max_col=10):
    for cell in row:
        if isinstance(cell.value, (int, float)): cell.number_format = '#,##0'
for col, w in zip('ABCDEFGHIJ', [16, 12, 12, 10, 11, 13, 12, 13, 11, 13]):
    ws2.column_dimensions[col].width = w

# Sheet3 메모
ws3 = wb.create_sheet('정산 메모')
notes = [
    '[로켓그로스 채움컴퍼니 정산 메모 / 2026-06-02]',
    '',
    '[기준] 정산주기 종료일 기준 (다른 채널과 동일). 정산대상액 = 판매액 - 판매수수료(약8.58%+VAT) 차감 후.',
    '',
    '[물류비] CFS 입출고비(파일 시트1) + 배송비(파일 시트2) 각 최종비용(VAT포함). 보관비는 세이버 혜택으로 0원.',
    ' - 비용제로(입고 90일) 혜택으로 5/17 주차까지 물류비 0원, 5/24 주차부터 부과 시작.',
    '',
    '[광고비] 쿠팡 광고 리포트(계정 A00940134, 전 계정) 기준. 이 계정은 그로스 상품 광고 전용.',
    ' - 키워드+캠페인 일자별 광고비 합산. 검증: 5/1~24 = 2,174,416원, 기존 마진분석 2,095,706원 (오차 4%).',
    ' - 정확한 차감액은 Wing 광고정산서와 1~5% 오차 가능. 빠른정산 이용액은 미반영.',
    '',
    '[수금일정] 1차 70%: 정산주기 종료 +약20영업일(약4주). 2차 30%: 판매 익익월 15일경. (예상일, 영업일 기준 실제와 다를 수 있음)',
    '',
    '[확인필요] 빠른정산(셀러월렛) 이용액 차감분 미반영 - Wing 정산서에서 확인 필요.',
    '[참고] 3월은 광고 미집행(상품 신규). 4월 광고비는 4/19 주차부터 발생.',
]
for i, n in enumerate(notes, 1):
    ws3.cell(row=i, column=1, value=n)
ws3.cell(row=1, column=1).font = Font(bold=True, size=12)
ws3.column_dimensions['A'].width = 105

out = '/Users/macmini_ky/Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브/앱/매출 정산/26.05/26.05 로켓그로스 정산_채움컴퍼니.xlsx'
wb.save(out)
print('저장 완료:', out)
print()
print('=== 월별 요약 ===')
for m in sorted(mt):
    t = mt[m]; fee = t['판매액'] - t['정산대상액']; logi = t['입출고비'] + t['배송비'] + t['보관비']
    net = t['정산대상액'] - logi - t['광고비']
    print('%s: 판매 %10s | 수수료 %9s | 물류 %9s | 광고 %9s | 실수령 %11s (%.0f%%)' % (
        MN[m], format(int(t['판매액']), ','), format(int(fee), ','), format(int(logi), ','),
        format(int(t['광고비']), ','), format(int(net), ','), net / t['판매액'] * 100))
