"""
키노 에이전트 - 카페24 14개월 매출/이익 계산
방식: aggregate-then-deduct (월간 정산시트 방식)
  - 모든 행의 gross를 먼저 합산
  - 6개 차감 항목(쿠폰, 환불, 적립금, 회원등급할인, 앱할인, 추가할인)도 별도 합산
  - 마지막에 차감 적용
"""
import sys
import io
import os
import json
import tempfile
import shutil
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl

SRC_DIR = r'N:/개인/Becorelab/03. 영업/20. 월별 매출정산/기타/자사몰 검증'
OUT_DIR = r'C:/Users/info/ClaudeAITeam/accounting/multi_agent_verify'
MAPPING_PATH = os.path.join(OUT_DIR, 'master_mapping.json')
RESULT_PATH = os.path.join(OUT_DIR, '03_kino_result.json')

# 포함 주문 상태
INCLUDE_STATUSES = {'배송 완료', '구매 확정', '배송중', '배송 준비중'}

# 14개월 파일 리스트 (파일명, YYYY-MM)
MONTHS = []
for y, m in [(25, i) for i in range(1, 13)] + [(26, 1), (26, 2)]:
    fn = f'{y:02d}. {m:02d} 카페24 원본 데이터.xlsx'
    ym = f'20{y:02d}-{m:02d}'
    MONTHS.append((fn, ym))


def load_mapping():
    with open(MAPPING_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)


def to_num(v):
    if v is None or v == '':
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        s = str(v).replace(',', '').strip()
        if not s:
            return 0.0
        try:
            return float(s)
        except ValueError:
            return 0.0


def find_col(headers, *candidates, prefix=False):
    """헤더 이름으로 컬럼 인덱스 찾기. prefix=True면 시작 문자열 매칭."""
    for i, h in enumerate(headers):
        if h is None:
            continue
        hs = str(h).strip()
        for c in candidates:
            if prefix:
                if hs.startswith(c):
                    return i
            else:
                if hs == c:
                    return i
    return -1


def get_cost(standard, ym, cost_map, cost_map_monthly):
    if not standard:
        return None
    if standard in cost_map:
        return cost_map[standard]
    mm = cost_map_monthly.get(ym, {})
    if standard in mm:
        return mm[standard]
    return None


def process_month(filename, ym, mapping):
    src = os.path.join(SRC_DIR, filename)
    name_map = mapping.get('name_map', {})
    cost_map = mapping.get('cost_map', {})
    cost_map_monthly = mapping.get('cost_map_monthly', {})

    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tf:
        tmp = tf.name
    try:
        shutil.copy(src, tmp)
        wb = openpyxl.load_workbook(tmp, read_only=True, data_only=True)
        if '원본' not in wb.sheetnames:
            wb.close()
            raise RuntimeError(f'{filename}: 원본 시트 없음')
        ws = wb['원본']

        rows = ws.iter_rows(values_only=True)
        headers = list(next(rows))

        col_order = find_col(headers, '주문번호')
        col_status = find_col(headers, '주문 상태')
        col_qty = find_col(headers, '수량')
        col_price = find_col(headers, '옵션+판매가')
        col_name = find_col(headers, '주문상품명(옵션포함)')
        col_ship = find_col(headers, '총 배송비(KRW)', prefix=True)
        col_coupon = find_col(headers, '쿠폰 할인금액(최초)')
        col_point = find_col(headers, '사용한 적립금액(최종)')
        col_refund = find_col(headers, '실제 환불금액')
        col_rank = find_col(headers, '회원등급 추가할인금액')
        col_app = find_col(headers, '앱 상품할인 금액(최종)')
        col_extra = find_col(headers, '상품별 추가할인금액')

        required = {
            '주문번호': col_order, '주문 상태': col_status, '수량': col_qty,
            '옵션+판매가': col_price, '주문상품명(옵션포함)': col_name,
            '총 배송비(KRW)': col_ship, '쿠폰 할인금액(최초)': col_coupon,
            '사용한 적립금액(최종)': col_point, '실제 환불금액': col_refund,
            '회원등급 추가할인금액': col_rank, '앱 상품할인 금액(최종)': col_app,
            '상품별 추가할인금액': col_extra,
        }
        missing = [k for k, v in required.items() if v < 0]
        if missing:
            raise RuntimeError(f'{filename}: 컬럼 누락 {missing}')

        gross_sum = 0.0
        coupon_sum = 0.0
        refund_sum = 0.0
        point_sum = 0.0
        rank_sum = 0.0
        app_sum = 0.0
        extra_sum = 0.0
        ship_sum = 0.0
        cost_sum = 0.0
        qty_sum = 0
        orders_included = set()
        orders_excluded = set()
        orders_first_seen = set()  # 주문번호 기준 쿠폰/적립금/배송비 1회만 적용
        unmapped = 0
        no_cost = 0
        no_cost_value = 0.0
        issues = []

        row_count = 0
        for row in rows:
            row_count += 1
            status = row[col_status]
            status_s = str(status).strip() if status is not None else ''
            order_no = row[col_order]
            order_key = str(order_no).strip() if order_no is not None else ''

            if status_s not in INCLUDE_STATUSES:
                if order_key:
                    orders_excluded.add(order_key)
                continue

            qty = to_num(row[col_qty])
            price = to_num(row[col_price])
            raw_name = row[col_name]
            raw_name_s = str(raw_name).strip() if raw_name is not None else ''

            line_gross = qty * price
            gross_sum += line_gross
            qty_sum += int(qty)

            # 행단위 차감 항목
            refund_sum += to_num(row[col_refund])
            rank_sum += to_num(row[col_rank])
            app_sum += to_num(row[col_app])
            extra_sum += to_num(row[col_extra])

            # 주문번호 첫 등장 시 쿠폰/적립금/배송비
            if order_key and order_key not in orders_first_seen:
                orders_first_seen.add(order_key)
                coupon_sum += to_num(row[col_coupon])
                point_sum += to_num(row[col_point])
                ship_sum += to_num(row[col_ship])

            if order_key:
                orders_included.add(order_key)

            # 원가
            standard = name_map.get(raw_name_s)
            if not standard:
                unmapped += 1
                no_cost_value += line_gross
                continue
            unit_cost = get_cost(standard, ym, cost_map, cost_map_monthly)
            if unit_cost is None:
                no_cost += 1
                no_cost_value += line_gross
                continue
            cost_sum += unit_cost * qty

        wb.close()

        total_discount = coupon_sum + refund_sum + point_sum + rank_sum + app_sum + extra_sum
        net_revenue = gross_sum - total_discount
        net_revenue_with_ship = net_revenue + ship_sum
        profit = net_revenue - cost_sum
        margin_rate = (profit / net_revenue) if net_revenue else 0.0

        return {
            'gross': round(gross_sum, 2),
            'coupon': round(coupon_sum, 2),
            'refund': round(refund_sum, 2),
            'point': round(point_sum, 2),
            'rank_discount': round(rank_sum, 2),
            'app_discount': round(app_sum, 2),
            'extra_discount': round(extra_sum, 2),
            'shipping': round(ship_sum, 2),
            'total_discount': round(total_discount, 2),
            'net_revenue': round(net_revenue, 2),
            'net_revenue_with_ship': round(net_revenue_with_ship, 2),
            'cost': round(cost_sum, 2),
            'profit': round(profit, 2),
            'margin_rate': round(margin_rate, 6),
            'qty': qty_sum,
            'order_count': len(orders_included),
            'excluded_order_count': len(orders_excluded - orders_included),
            'unmapped_count': unmapped,
            'no_cost_count': no_cost,
            'no_cost_value': round(no_cost_value, 2),
            '_row_count': row_count,
            '_issues': issues,
        }
    finally:
        try:
            os.remove(tmp)
        except OSError:
            pass


def main():
    print('[키노] 마스터 매핑 로드 중...')
    mapping = load_mapping()
    print(f'  name_map={len(mapping.get("name_map",{}))} cost_map={len(mapping.get("cost_map",{}))}')

    result = {
        'agent': 'kino',
        'method': 'aggregate-then-deduct (monthly settlement sheet style)',
        'months': {},
        'totals': {},
        'issues': [],
    }

    for fn, ym in MONTHS:
        print(f'[키노] {ym} 처리: {fn}')
        try:
            r = process_month(fn, ym, mapping)
        except Exception as e:
            print(f'  ERROR: {e}')
            result['issues'].append({'month': ym, 'error': str(e)})
            continue
        issues = r.pop('_issues', [])
        row_count = r.pop('_row_count', 0)
        if issues:
            result['issues'].extend([{'month': ym, **i} for i in issues])
        print(f'  rows={row_count} included_orders={r["order_count"]} gross={r["gross"]:,.0f} profit={r["profit"]:,.0f} margin={r["margin_rate"]*100:.2f}%')
        result['months'][ym] = r

    # totals 집계
    num_fields = ['gross', 'coupon', 'refund', 'point', 'rank_discount',
                  'app_discount', 'extra_discount', 'shipping', 'total_discount',
                  'net_revenue', 'net_revenue_with_ship', 'cost', 'profit',
                  'no_cost_value']
    int_fields = ['qty', 'order_count', 'excluded_order_count',
                  'unmapped_count', 'no_cost_count']
    totals = {k: 0.0 for k in num_fields}
    for k in int_fields:
        totals[k] = 0
    for ym, r in result['months'].items():
        for k in num_fields:
            totals[k] += r.get(k, 0.0)
        for k in int_fields:
            totals[k] += r.get(k, 0)
    for k in num_fields:
        totals[k] = round(totals[k], 2)
    totals['margin_rate'] = round(totals['profit'] / totals['net_revenue'], 6) if totals['net_revenue'] else 0.0
    result['totals'] = totals

    with open(RESULT_PATH, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    print(f'[키노] 결과 저장: {RESULT_PATH}')
    print(f'[키노] totals gross={totals["gross"]:,.0f} net={totals["net_revenue"]:,.0f} profit={totals["profit"]:,.0f} margin={totals["margin_rate"]*100:.2f}%')


if __name__ == '__main__':
    main()
