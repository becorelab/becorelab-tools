"""
카페24 매출 검증용 마스터 매핑 JSON 빌더
- 14개월 raw 품명 추출
- 12개 품명리스트 → raw→standard 매핑 통합
- 통합 원가 + 월별 fallback → standard→원가 매핑
"""
import sys, io, os, json, re, tempfile, shutil
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import openpyxl

BASE = r"N:\개인\Becorelab\03. 영업\20. 월별 매출정산\기타\자사몰 검증"
BASE_2026_01 = r"N:\개인\Becorelab\03. 영업\20. 월별 매출정산\2026.01"
BASE_2026_02 = r"N:\개인\Becorelab\03. 영업\20. 월별 매출정산\2026.02"
OUT_DIR = r"C:\Users\info\ClaudeAITeam\accounting\multi_agent_verify"

# ---------- 파일 매핑 ----------
RAW_FILES = {
    "2025-01": "25. 01 카페24 원본 데이터.xlsx",
    "2025-02": "25. 02 카페24 원본 데이터.xlsx",
    "2025-03": "25. 03 카페24 원본 데이터.xlsx",
    "2025-04": "25. 04 카페24 원본 데이터.xlsx",
    "2025-05": "25. 05 카페24 원본 데이터.xlsx",
    "2025-06": "25. 06 카페24 원본 데이터.xlsx",
    "2025-07": "25. 07 카페24 원본 데이터.xlsx",
    "2025-08": "25. 08 카페24 원본 데이터.xlsx",
    "2025-09": "25. 09 카페24 원본 데이터.xlsx",
    "2025-10": "25. 10 카페24 원본 데이터.xlsx",
    "2025-11": "25. 11 카페24 원본 데이터.xlsx",
    "2025-12": "25. 12 카페24 원본 데이터.xlsx",
    "2026-01": "26. 01 카페24 원본 데이터.xlsx",
    "2026-02": "26. 02 카페24 원본 데이터.xlsx",
}

# 25년 01월 품명리스트는 없음 (02월 매핑으로 fallback)
# 26년 01월/02월은 각 월 정산 폴더의 카페24 파일에 품명리스트 있음
NAMELIST_FILES = {
    "2025-02": os.path.join(BASE, "02월 카페24 - 복사본.xlsx"),
    "2025-03": os.path.join(BASE, "03월 자사몰.XLSX"),
    "2025-04": os.path.join(BASE, "04월 자사몰.xlsx"),
    "2025-05": os.path.join(BASE, "05월 카페24.xlsx"),
    "2025-06": os.path.join(BASE, "06월 카페24.xlsx"),
    "2025-07": os.path.join(BASE, "07월 카페24 2.xlsx"),
    "2025-08": os.path.join(BASE, "08월 카페24.xlsx"),
    "2025-09": os.path.join(BASE, "09월 카페24.xlsx"),
    "2025-10": os.path.join(BASE, "10월 카페24.xlsx"),
    "2025-11": os.path.join(BASE, "11월 카페24.xlsx"),
    "2025-12": os.path.join(BASE, "12월 카페24.xlsx"),
    "2026-01": os.path.join(BASE_2026_01, "01월 카페24.xlsx"),
    "2026-02": os.path.join(BASE_2026_02, "02월 카페24 (version 1).xlsx"),
}

SETTLEMENT_FILES = {
    "2025-01": os.path.join(BASE, "2025. 01 온라인 매출정산_0210_원가 수정.xlsx"),
    "2025-02": os.path.join(BASE, "2025. 02 온라인 매출정산.xlsx"),
    "2025-03": os.path.join(BASE, "2025. 03 온라인 매출정산_수정.xlsx"),
    "2025-04": os.path.join(BASE, "2025. 04 온라인 매출정산.xlsx"),
    "2025-05": os.path.join(BASE, "2025. 05 온라인 매출정산.xlsx"),
    "2025-06": os.path.join(BASE, "2025. 06 온라인 매출정산.xlsx"),
    "2025-07": os.path.join(BASE, "2025. 07 온라인 매출정산_광고비 적용 완료.xlsx"),
    "2025-08": os.path.join(BASE, "2025. 08 온라인 매출정산.xlsx"),
    "2025-09": os.path.join(BASE, "2025. 09 온라인 매출정산.xlsx"),
    "2025-10": os.path.join(BASE, "2025. 10 온라인 매출정산.xlsx"),
    "2025-11": os.path.join(BASE, "2025. 11 온라인 매출정산.xlsx"),
    "2025-12": os.path.join(BASE, "2025. 12 온라인 매출정산_자사몰 수정.xlsx"),
    "2026-01": os.path.join(BASE_2026_01, "2026. 01 온라인 매출정산.xlsx"),
    "2026-02": os.path.join(BASE_2026_02, "2026. 02 온라인 매출정산.xlsx"),
}

COST_MASTER_FILE = "상품명 및 원가 파일 전체 통합.xlsx"

RAW_NAME_COL_IDX = 51  # 0-based for 주문상품명(옵션포함)


def load_wb(src, read_only=False):
    fd, dst = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    shutil.copy(src, dst)
    wb = openpyxl.load_workbook(dst, data_only=True, read_only=read_only)
    return wb, dst


def norm(s):
    if s is None:
        return ""
    return re.sub(r'\s+', '', str(s)).strip()


# =============================================================
# Step 1: raw 품명 추출 (월별)
# =============================================================
def extract_raw_names():
    print("=" * 60)
    print("Step 1: 14개월 raw 품명 추출")
    print("=" * 60)
    raw_by_month = {}  # month -> set of raw names
    all_raw = set()
    for month, fname in RAW_FILES.items():
        src = os.path.join(BASE, fname)
        if not os.path.exists(src):
            print(f"  [!] {month}: 파일 없음 {fname}")
            raw_by_month[month] = set()
            continue
        wb, dst = load_wb(src, read_only=True)
        ws = wb["원본"]
        names = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) > RAW_NAME_COL_IDX:
                v = row[RAW_NAME_COL_IDX]
                if v is not None and str(v).strip():
                    names.add(str(v).strip())
        raw_by_month[month] = names
        all_raw.update(names)
        wb.close()
        os.remove(dst)
        print(f"  {month}: {len(names)}개 raw 품명 ({fname})")
    print(f"\n=> 전체 unique raw 품명: {len(all_raw)}개")
    return raw_by_month, all_raw


# =============================================================
# Step 2/3: 품명리스트 → raw→standard 매핑 구축
# =============================================================
def build_name_map():
    print("\n" + "=" * 60)
    print("Step 2/3: raw → standard 매핑 구축")
    print("=" * 60)
    # norm(raw) → {month: standard}
    per_month_maps = {}
    # 최종 매핑: norm(raw) → standard
    global_map = {}
    # norm(raw) → 원본 raw (첫 등장)
    raw_original = {}
    conflicts = []

    for month, src in NAMELIST_FILES.items():
        fname = os.path.basename(src)
        if not os.path.exists(src):
            print(f"  [!] {month}: 파일 없음 {fname}")
            per_month_maps[month] = {}
            continue
        wb, dst = load_wb(src)
        mp = {}
        # 품명리스트 시트와 품명리스트추가 시트 둘 다 수집
        for sheet_name in ["품명리스트", "품명리스트추가"]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for r in range(2, ws.max_row + 1):
                raw = ws.cell(r, 1).value
                std = ws.cell(r, 2).value
                if raw is None or std is None:
                    continue
                raw_s = str(raw).strip()
                std_s = str(std).strip()
                if not raw_s or not std_s:
                    continue
                n = norm(raw_s)
                mp[n] = std_s
                if n not in raw_original:
                    raw_original[n] = raw_s
        per_month_maps[month] = mp
        wb.close()
        os.remove(dst)
        print(f"  {month}: {len(mp)}개 매핑 ({fname})")

    # 통합: 가장 오래된 월 우선 (초기 매핑이 보통 더 단순/정확)
    # 월 순서: 2025-02 → 2026-02 (오래된 순)
    # 기존 상품은 오래된 매핑 유지, 신상품은 자동으로 최근 월에서 채움
    months_sorted = sorted(per_month_maps.keys(), reverse=False)
    for month in months_sorted:
        for raw_n, std in per_month_maps[month].items():
            if raw_n not in global_map:
                global_map[raw_n] = std
            else:
                if norm(global_map[raw_n]) != norm(std):
                    conflicts.append({
                        "raw": raw_original.get(raw_n, raw_n),
                        "existing_standard": global_map[raw_n],
                        "conflict_standard": std,
                        "conflict_month": month,
                    })
    print(f"\n=> 통합 매핑: {len(global_map)}개")
    print(f"=> 충돌: {len(conflicts)}건")
    return global_map, per_month_maps, raw_original, conflicts


# =============================================================
# Step 4: 원가 매핑 (통합 원가 우선)
# =============================================================
def build_cost_map():
    print("\n" + "=" * 60)
    print("Step 4: standard → 원가 매핑 구축")
    print("=" * 60)
    cost_map = {}  # norm(standard) → {"name":..., "cost":...}
    standard_display = {}  # norm → 원래 standard 이름

    # 통합 원가 파일
    src = os.path.join(BASE, COST_MASTER_FILE)
    wb, dst = load_wb(src)
    ws = wb["원가"]
    # Row 2~4: A=name, C=cost (판관비포함원가(VAT포함) 아닌 '원가특수')
    # 사양대로: A열=품명, C열=원가특수, E열=판관비포함원가(VAT포함)
    # 특수 영역 (Row 2~4): A=품명, C=원가특수 사용
    for r in range(2, 5):
        name = ws.cell(r, 1).value
        cost = ws.cell(r, 3).value
        if name and cost is not None:
            n = norm(name)
            cost_map[n] = float(cost)
            standard_display[n] = str(name).strip()
    # 본 데이터 (Row 7~): B=제품명, D 또는 E = 판관비포함원가(VAT포함)
    # 위 inspect에서 R6 header: [None, '제품명', '판관비포함원가(부가세별도)', '판관비포함원가(부가세포함)', '판관비포함원가(부가세포함)', ...]
    # E열(idx 5) = 판관비포함원가(VAT포함). 사양대로 E열 사용.
    for r in range(7, ws.max_row + 1):
        name = ws.cell(r, 2).value
        cost = ws.cell(r, 5).value  # E열
        if name is None or cost is None:
            continue
        try:
            cost_val = float(cost)
        except (ValueError, TypeError):
            continue
        n = norm(name)
        if n not in cost_map:
            cost_map[n] = cost_val
            standard_display[n] = str(name).strip()
    wb.close()
    os.remove(dst)
    print(f"  통합원가 파일에서: {len(cost_map)}개 standard")

    # 월별 정산시트의 원가 시트 (fallback용)
    cost_map_monthly = {}
    for month, src in SETTLEMENT_FILES.items():
        fname = os.path.basename(src)
        if not os.path.exists(src):
            cost_map_monthly[month] = {}
            continue
        try:
            wb, dst = load_wb(src)
        except Exception as e:
            print(f"  [!] {month} 로드 실패: {e}")
            cost_map_monthly[month] = {}
            continue
        if "원가" not in wb.sheetnames:
            cost_map_monthly[month] = {}
            wb.close()
            os.remove(dst)
            continue
        ws = wb["원가"]
        mcost = {}  # norm(standard) → cost
        # 같은 구조라고 가정 (특수 + 본데이터)
        # 안전하게: 헤더를 스캔해서 '제품명' 찾고 '판관비포함원가(부가세포함)' 열 찾기
        name_col = None
        cost_col = None
        header_row = None
        for r in range(1, min(10, ws.max_row + 1)):
            for c in range(1, min(15, ws.max_column + 1)):
                v = ws.cell(r, c).value
                if v and '제품명' in str(v):
                    name_col = c
                    header_row = r
                if v and '판관비포함' in str(v) and ('VAT포함' in str(v) or '부가세포함' in str(v)):
                    cost_col = c
        if name_col and cost_col and header_row:
            for r in range(header_row + 1, ws.max_row + 1):
                name = ws.cell(r, name_col).value
                cost = ws.cell(r, cost_col).value
                if name is None or cost is None:
                    continue
                try:
                    cost_val = float(cost)
                except (ValueError, TypeError):
                    continue
                n = norm(name)
                mcost[n] = cost_val
                if n not in standard_display:
                    standard_display[n] = str(name).strip()
        # 특수 영역 (A=품명, C=원가) 시도
        for r in range(2, 6):
            name = ws.cell(r, 1).value
            cost = ws.cell(r, 3).value
            if name and cost is not None:
                try:
                    cost_val = float(cost)
                    n = norm(name)
                    if n not in mcost:
                        mcost[n] = cost_val
                    if n not in standard_display:
                        standard_display[n] = str(name).strip()
                except (ValueError, TypeError):
                    pass
        cost_map_monthly[month] = mcost
        wb.close()
        os.remove(dst)
        print(f"  {month} 원가시트: {len(mcost)}개")

        # 통합 원가에 없는 건 fallback으로 보충
        for n, v in mcost.items():
            if n not in cost_map:
                cost_map[n] = v

    print(f"\n=> 최종 통합 cost_map: {len(cost_map)}개 standard")
    return cost_map, cost_map_monthly, standard_display


# =============================================================
# Step 5: JSON 생성
# =============================================================
def main():
    raw_by_month, all_raw = extract_raw_names()
    name_map_norm, per_month_maps, raw_original, conflicts = build_name_map()
    cost_map_norm, cost_map_monthly, standard_display = build_cost_map()

    # raw 품명별 매핑 시도 (원래 raw 텍스트를 키로)
    name_map_final = {}  # raw(원본) → standard
    unmapped_raw = []
    for raw in sorted(all_raw):
        n = norm(raw)
        if n in name_map_norm:
            name_map_final[raw] = name_map_norm[n]
        else:
            unmapped_raw.append(raw)

    # standard 수집
    all_standards = set(name_map_final.values())
    # 원가 있는/없는 standard
    standards_with_cost = []
    standards_without_cost = []
    cost_map_final = {}  # standard(원본) → cost
    for std in sorted(all_standards):
        n = norm(std)
        if n in cost_map_norm:
            cost_map_final[std] = cost_map_norm[n]
            standards_with_cost.append(std)
        else:
            standards_without_cost.append(std)

    # cost_map_monthly도 standard(원본) 단위로 rebuild
    cost_map_monthly_final = {}
    for month, mp in cost_map_monthly.items():
        out = {}
        for std in all_standards:
            n = norm(std)
            if n in mp:
                out[std] = mp[n]
        # 추가로, 월별 시트에만 있는 standard도 보존 (키는 display)
        for n, v in mp.items():
            disp = standard_display.get(n, n)
            if disp not in out:
                out[disp] = v
        cost_map_monthly_final[month] = out

    # 월별 raw 등장 정보
    raw_appearance = {}  # raw → [month,...]
    for month, raws in raw_by_month.items():
        for raw in raws:
            raw_appearance.setdefault(raw, []).append(month)

    stats = {
        "total_raw_names": len(all_raw),
        "mapped_raw_names": len(name_map_final),
        "unmapped_raw_names": len(unmapped_raw),
        "standards_total": len(all_standards),
        "standards_with_cost": len(standards_with_cost),
        "standards_without_cost": len(standards_without_cost),
        "namelist_files_count": len(NAMELIST_FILES),
        "raw_files_count": len(RAW_FILES),
    }

    issues = {
        "unmapped_raw_samples": unmapped_raw[:50],
        "unmapped_raw_total": len(unmapped_raw),
        "no_cost_standards": standards_without_cost,
        "conflicts": conflicts[:200],
        "conflicts_total": len(conflicts),
        "month_files_used": {month: os.path.basename(NAMELIST_FILES.get(month, "(없음-fallback)")) for month in RAW_FILES},
    }

    master = {
        "meta": {
            "raw_name_column": "주문상품명(옵션포함)",
            "raw_files": RAW_FILES,
            "namelist_files": {k: os.path.basename(v) for k, v in NAMELIST_FILES.items()},
            "namelist_file_paths": NAMELIST_FILES,
            "settlement_files": {k: os.path.basename(v) for k, v in SETTLEMENT_FILES.items()},
            "cost_master_file": COST_MASTER_FILE,
            "note": "name_map 키는 원본 raw 텍스트 (strip only). 정규화 시 re.sub(r'\\s+', '', s).",
        },
        "name_map": name_map_final,
        "cost_map": cost_map_final,
        "cost_map_monthly": cost_map_monthly_final,
        "raw_appearance": raw_appearance,
        "stats": stats,
        "issues": issues,
    }

    out_path = os.path.join(OUT_DIR, "master_mapping.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(master, f, ensure_ascii=False, indent=2)
    print(f"\n[OK] 저장: {out_path}")
    print(f"\n[STATS] {json.dumps(stats, ensure_ascii=False)}")
    print(f"[CONFLICTS] {len(conflicts)}")
    print(f"[UNMAPPED] {len(unmapped_raw)}")
    print(f"[NO COST] {len(standards_without_cost)}")
    return master


if __name__ == "__main__":
    main()
