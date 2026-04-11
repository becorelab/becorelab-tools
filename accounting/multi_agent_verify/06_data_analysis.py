"""
데이터 분석 에이전트 (두리) - 카페24 14개월 (2025.01 ~ 2026.02) 심층 분석
1) SKU별 매출/이익/수량 Top/Bottom
2) 카테고리별 매출/이익 분석
3) 회원 재구매 분석 (수령인 휴대전화 기준)
4) 가격대별 분포 + AOV 월별

로직: 모리 스크립트와 동일한 행처리/dedup/원가/차감
출력: 06_data_analysis.json, 06_data_analysis_report.md
"""
import sys, io, os, json, re, tempfile, shutil, traceback
from collections import defaultdict
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# -------- 경로 --------
MAPPING_FILE = 'C:/Users/info/ClaudeAITeam/accounting/multi_agent_verify/master_mapping.json'
SRC_DIR = 'N:/개인/Becorelab/03. 영업/20. 월별 매출정산/기타/자사몰 검증'
OUT_JSON = 'C:/Users/info/ClaudeAITeam/accounting/multi_agent_verify/06_data_analysis.json'
OUT_MD   = 'C:/Users/info/ClaudeAITeam/accounting/multi_agent_verify/06_data_analysis_report.md'

INCLUDE_STATUSES = {'배송 완료', '구매 확정', '배송중', '배송 준비중'}

# -------- 카테고리 키워드 (순서 중요: 먼저 걸리는 게 우선) --------
CATEGORY_RULES = [
    ('건조기시트', ['건조기시트', '건조기 시트', '건조기', 'dryer']),
    ('식기세척기', ['하트식세기', '식기세척기', '식세기', '식세', 'dishwasher']),
    ('캡슐세제',   ['캡슐세제', '캡슐 세제', '캡슐', '세탁세제', 'capsule']),
    ('얼룩제거제', ['얼룩제거제', '얼룩 제거제', '얼룩']),
    ('섬유탈취제', ['섬유탈취제', '섬유 탈취제', '탈취']),
    ('이염방지',   ['이염방지', '이염']),
]

def classify_category(standard_name: str) -> str:
    if not standard_name:
        return '기타'
    n = standard_name
    for cat, kws in CATEGORY_RULES:
        for kw in kws:
            if kw in n:
                return cat
    return '기타'

# -------- 유틸 --------
def sf(v):
    if v is None or v == '':
        return 0.0
    try:
        return float(str(v).replace(',', '').strip())
    except Exception:
        return 0.0

def si(v):
    return int(sf(v))

def find_col(headers, keyword, strict=False):
    for i, h in enumerate(headers):
        if h is None:
            continue
        s = str(h)
        if strict:
            if s.strip() == keyword:
                return i
        else:
            if keyword in s:
                return i
    return None

# -------- 매핑 로드 --------
print('[1/4] master_mapping.json 로드 중...')
with open(MAPPING_FILE, 'r', encoding='utf-8') as f:
    mapping = json.load(f)
name_map = mapping['name_map']
cost_map = mapping['cost_map']
cost_map_monthly = mapping.get('cost_map_monthly', {})
print(f'  name_map: {len(name_map)}개 / cost_map: {len(cost_map)}개')

# -------- 월별 파일 탐색 --------
print('[2/4] 원본 파일 탐색...')
month_files = []
for fn in sorted(os.listdir(SRC_DIR)):
    if '카페24 원본' in fn and fn.endswith('.xlsx'):
        m = re.match(r'(\d{2})\.\s*(\d{2})\s*카페24', fn)
        if m:
            label = f'20{m.group(1)}-{m.group(2)}'
            month_files.append((label, os.path.join(SRC_DIR, fn)))
print(f'  총 {len(month_files)}개 파일')

# -------- 집계 구조 --------
# SKU 단위 (standard 이름)
sku_agg = defaultdict(lambda: {'revenue': 0.0, 'profit': 0.0, 'cost': 0.0, 'qty': 0})
# 카테고리 월별
cat_monthly = defaultdict(lambda: defaultdict(lambda: {'revenue': 0.0, 'profit': 0.0, 'cost': 0.0, 'qty': 0}))
# 회원(휴대전화) 단위
# member[phone] = {'orders': set(), 'revenue': 0.0, 'first_ts': dt, 'last_ts': dt, 'order_months': set()}
from datetime import datetime
member_agg = {}
guest_order_count = 0
guest_revenue = 0.0
# 주문 단위 (주문번호 dedup). 주문 net 금액 = 행별 (gross - coupon/point/ship 차감) 합
order_agg = {}  # order_id -> {'net': 0.0, 'ts': dt, 'phone': str|None, 'month': 'YYYY-MM'}
# 월별 신규/재구매 매출
monthly_new_repeat = defaultdict(lambda: {'new_revenue': 0.0, 'repeat_revenue': 0.0,
                                           'new_orders': 0, 'repeat_orders': 0})

issues = []

def process_month(label, filepath):
    global guest_order_count, guest_revenue
    fd, dst = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    shutil.copy(filepath, dst)
    try:
        wb = openpyxl.load_workbook(dst, data_only=True, read_only=True)
        ws = wb['원본']
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = list(row)
            break

        IDX_ORDER  = find_col(headers, '주문번호')
        IDX_STATUS = find_col(headers, '주문 상태')
        IDX_QTY    = find_col(headers, '수량')
        IDX_OPTPR  = find_col(headers, '옵션+판매가')
        IDX_NAME   = find_col(headers, '주문상품명(옵션포함)')
        IDX_SHIP   = find_col(headers, '총 배송비(KRW)')
        IDX_COUPON = find_col(headers, '쿠폰 할인금액(최초)')
        IDX_POINT  = find_col(headers, '사용한 적립금액(최종)')
        IDX_REFUND = find_col(headers, '실제 환불금액')
        IDX_RANK   = find_col(headers, '회원등급 추가할인금액')
        IDX_APP    = find_col(headers, '앱 상품할인 금액(최종)')
        IDX_EXTRA  = find_col(headers, '상품별 추가할인금액')
        IDX_PHONE  = find_col(headers, '수령인 휴대전화')
        IDX_BUYER  = find_col(headers, '결제자')
        IDX_TS     = find_col(headers, '주문일시')
        # 품목별 주문번호를 별도로 잡아서 row-level dedup 용으로도 활용 가능
        # 품목별 주문번호가 order 단위 구분 key
        required = {'order': IDX_ORDER, 'status': IDX_STATUS, 'qty': IDX_QTY,
                    'opt_price': IDX_OPTPR, 'name': IDX_NAME}
        miss = [k for k, v in required.items() if v is None]
        if miss:
            issues.append(f'{label}: 컬럼 찾기 실패 {miss}')

        month_cost_fallback = cost_map_monthly.get(label, {})

        orders_seen = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if IDX_ORDER is None or row[IDX_ORDER] is None:
                continue
            order = str(row[IDX_ORDER]).strip()
            if not order:
                continue
            status = str(row[IDX_STATUS] or '').strip() if IDX_STATUS is not None else ''
            if status not in INCLUDE_STATUSES:
                continue

            qty = si(row[IDX_QTY])
            opt_price = sf(row[IDX_OPTPR])
            gross_row = qty * opt_price

            is_first = order not in orders_seen
            if is_first:
                orders_seen.add(order)

            coupon = sf(row[IDX_COUPON]) if (is_first and IDX_COUPON is not None) else 0.0
            point  = sf(row[IDX_POINT])  if (is_first and IDX_POINT  is not None) else 0.0
            ship   = sf(row[IDX_SHIP])   if (is_first and IDX_SHIP   is not None) else 0.0
            refund = sf(row[IDX_REFUND]) if IDX_REFUND is not None else 0.0
            rank_d = sf(row[IDX_RANK])   if IDX_RANK   is not None else 0.0
            app_d  = sf(row[IDX_APP])    if IDX_APP    is not None else 0.0
            extra_d= sf(row[IDX_EXTRA])  if IDX_EXTRA  is not None else 0.0

            raw_name = str(row[IDX_NAME] or '').strip()
            standard = name_map.get(raw_name, '')
            unit_cost = 0.0
            if standard:
                if standard in cost_map:
                    unit_cost = float(cost_map[standard])
                elif standard in month_cost_fallback:
                    unit_cost = float(month_cost_fallback[standard])
            cost_row = unit_cost * qty

            # 행 단위 순매출 (모리와 동일 방식: gross - coupon - refund - point - rank - app - extra)
            net_row = gross_row - coupon - refund - point - rank_d - app_d - extra_d
            profit_row = net_row - cost_row

            # SKU 집계 (standard 기준, 매핑 안되면 "미매핑" 한 곳에 모음)
            sku_key = standard if standard else f'[미매핑] {raw_name[:30]}' if raw_name else '[미매핑/공백]'
            sku_agg[sku_key]['revenue'] += net_row
            sku_agg[sku_key]['profit']  += profit_row
            sku_agg[sku_key]['cost']    += cost_row
            sku_agg[sku_key]['qty']     += qty

            # 카테고리 집계
            cat = classify_category(standard) if standard else '기타'
            c = cat_monthly[label][cat]
            c['revenue'] += net_row
            c['profit']  += profit_row
            c['cost']    += cost_row
            c['qty']     += qty

            # 주문 단위 집계
            if order not in order_agg:
                ts = None
                if IDX_TS is not None:
                    v = row[IDX_TS]
                    if isinstance(v, datetime):
                        ts = v
                    elif v:
                        try: ts = datetime.fromisoformat(str(v))
                        except Exception: ts = None
                phone = None
                if IDX_PHONE is not None:
                    p = row[IDX_PHONE]
                    if p:
                        phone = re.sub(r'[^0-9]', '', str(p))
                        if len(phone) < 9:
                            phone = None
                order_agg[order] = {
                    'net': 0.0, 'ts': ts, 'phone': phone, 'month': label,
                    'profit': 0.0, 'qty': 0,
                }
            order_agg[order]['net']    += net_row
            order_agg[order]['profit'] += profit_row
            order_agg[order]['qty']    += qty

        wb.close()
    finally:
        try: os.remove(dst)
        except Exception: pass

print('[3/4] 14개월 순회...')
for label, fp in month_files:
    print(f'  {label} ...', flush=True)
    try:
        process_month(label, fp)
    except Exception as e:
        traceback.print_exc()
        issues.append(f'{label}: 처리 에러 {e}')

# ================= 분석 1: SKU Top/Bottom =================
print('[4/4] 분석 집계...')
def margin_rate(rev, profit):
    return (profit / rev) if rev > 0 else 0.0

sku_rows = []
for name, d in sku_agg.items():
    rev = d['revenue']; pf = d['profit']
    sku_rows.append({
        'name': name,
        'revenue': round(rev, 0),
        'profit':  round(pf, 0),
        'cost':    round(d['cost'], 0),
        'qty':     d['qty'],
        'margin_rate': round(margin_rate(rev, pf), 4),
    })

by_revenue = sorted(sku_rows, key=lambda x: -x['revenue'])[:30]
by_profit  = sorted(sku_rows, key=lambda x: -x['profit'])[:30]
# 마진율 Top/Bottom은 수량 50개 이상만
qual = [r for r in sku_rows if r['qty'] >= 50 and r['revenue'] > 0]
by_margin_top = sorted(qual, key=lambda x: -x['margin_rate'])[:30]
by_margin_bot = sorted(qual, key=lambda x: x['margin_rate'])[:30]

# ================= 분석 2: 카테고리 =================
cat_totals = defaultdict(lambda: {'revenue': 0.0, 'profit': 0.0, 'qty': 0})
cat_monthly_out = {}
for label in sorted(cat_monthly.keys()):
    cat_monthly_out[label] = {}
    for cat, d in cat_monthly[label].items():
        cat_monthly_out[label][cat] = {
            'revenue': round(d['revenue'], 0),
            'profit':  round(d['profit'], 0),
            'qty':     d['qty'],
            'margin_rate': round(margin_rate(d['revenue'], d['profit']), 4),
        }
        cat_totals[cat]['revenue'] += d['revenue']
        cat_totals[cat]['profit']  += d['profit']
        cat_totals[cat]['qty']     += d['qty']

total_rev_all = sum(v['revenue'] for v in cat_totals.values())
cat_totals_out = {}
for cat, d in sorted(cat_totals.items(), key=lambda x: -x[1]['revenue']):
    cat_totals_out[cat] = {
        'revenue': round(d['revenue'], 0),
        'profit':  round(d['profit'], 0),
        'qty':     d['qty'],
        'margin_rate': round(margin_rate(d['revenue'], d['profit']), 4),
        'share':   round((d['revenue'] / total_rev_all) if total_rev_all > 0 else 0, 4),
    }

# ================= 분석 3: 회원 재구매 =================
# order_agg를 phone 기준으로 group
member = defaultdict(list)  # phone -> list of (ts, order_id, net, month)
for oid, o in order_agg.items():
    if o['phone']:
        member[o['phone']].append((o['ts'], oid, o['net'], o['month']))
    else:
        guest_order_count += 1
        guest_revenue += o['net']

# 정렬
for p in member:
    member[p].sort(key=lambda x: (x[0] or datetime.min))

total_unique = len(member)
new_only = 0
repeat_2plus = 0
repeat_5plus = 0
vip_revenue = 0.0
repeat_revenue_total = 0.0
interval_buckets = {'1주 이내':0, '1주-1개월':0, '1-3개월':0, '3-6개월':0, '6개월+':0}

for p, orders in member.items():
    n = len(orders)
    total_net = sum(o[2] for o in orders)
    if n == 1:
        new_only += 1
    else:
        repeat_2plus += 1
        # 재구매 매출 = 2번째 이후 주문 net 합
        repeat_revenue_total += sum(o[2] for o in orders[1:])
        # 재구매 간격 분포
        for i in range(1, n):
            if orders[i][0] and orders[i-1][0]:
                days = (orders[i][0] - orders[i-1][0]).days
                if days <= 7: interval_buckets['1주 이내'] += 1
                elif days <= 30: interval_buckets['1주-1개월'] += 1
                elif days <= 90: interval_buckets['1-3개월'] += 1
                elif days <= 180: interval_buckets['3-6개월'] += 1
                else: interval_buckets['6개월+'] += 1
    if n >= 5:
        repeat_5plus += 1
        vip_revenue += total_net

member_total_revenue = sum(sum(o[2] for o in orders) for orders in member.values())

# 월별 신규/재구매 매출
# "신규" 정의: 해당 회원의 첫 주문인 경우 = 신규, 그 이후는 재구매
for p, orders in member.items():
    for i, o in enumerate(orders):
        lab = o[3]
        if i == 0:
            monthly_new_repeat[lab]['new_revenue'] += o[2]
            monthly_new_repeat[lab]['new_orders']  += 1
        else:
            monthly_new_repeat[lab]['repeat_revenue'] += o[2]
            monthly_new_repeat[lab]['repeat_orders']  += 1

monthly_new_repeat_out = {}
for lab in sorted(monthly_new_repeat.keys()):
    d = monthly_new_repeat[lab]
    tot = d['new_revenue'] + d['repeat_revenue']
    monthly_new_repeat_out[lab] = {
        'new_revenue': round(d['new_revenue'], 0),
        'repeat_revenue': round(d['repeat_revenue'], 0),
        'new_orders': d['new_orders'],
        'repeat_orders': d['repeat_orders'],
        'repeat_share': round((d['repeat_revenue'] / tot) if tot > 0 else 0, 4),
    }

# ================= 분석 4: 가격대별 분포 + AOV =================
BANDS = [('<1만', 0, 10000), ('1-2만', 10000, 20000), ('2-3만', 20000, 30000),
         ('3-5만', 30000, 50000), ('5-10만', 50000, 100000), ('10만+', 100000, float('inf'))]
band_agg = {b[0]: {'orders': 0, 'revenue': 0.0, 'profit': 0.0} for b in BANDS}
monthly_aov = defaultdict(lambda: {'revenue': 0.0, 'orders': 0})
for oid, o in order_agg.items():
    net = o['net']
    for name, lo, hi in BANDS:
        if lo <= net < hi:
            b = band_agg[name]
            b['orders']  += 1
            b['revenue'] += net
            b['profit']  += o['profit']
            break
    monthly_aov[o['month']]['revenue'] += net
    monthly_aov[o['month']]['orders']  += 1

total_band_rev = sum(b['revenue'] for b in band_agg.values())
band_out = {}
for name, _, _ in BANDS:
    b = band_agg[name]
    band_out[name] = {
        'orders': b['orders'],
        'revenue': round(b['revenue'], 0),
        'profit':  round(b['profit'], 0),
        'margin_rate': round(margin_rate(b['revenue'], b['profit']), 4),
        'share':   round((b['revenue'] / total_band_rev) if total_band_rev > 0 else 0, 4),
        'avg_order': round((b['revenue'] / b['orders']) if b['orders'] > 0 else 0, 0),
    }

aov_out = {}
for lab in sorted(monthly_aov.keys()):
    d = monthly_aov[lab]
    aov_out[lab] = round((d['revenue'] / d['orders']) if d['orders'] > 0 else 0, 0)

# ================= JSON 출력 =================
output = {
    'agent': 'doori_data_analysis',
    'found': True,
    'period': '2025-01 ~ 2026-02 (14 months)',
    'member_id_column_used': '수령인 휴대전화 (phone digits)',
    'sku_top': {
        'by_revenue': by_revenue,
        'by_profit':  by_profit,
        'by_margin_rate_top': by_margin_top,
        'by_margin_rate_bottom': by_margin_bot,
        'qty_threshold_margin': 50,
    },
    'category': {
        'monthly': cat_monthly_out,
        'totals':  cat_totals_out,
    },
    'member': {
        'id_column_used': '수령인 휴대전화',
        'total_unique': total_unique,
        'new_only_count': new_only,
        'repeat_2plus_count': repeat_2plus,
        'repeat_5plus_count': repeat_5plus,
        'repeat_rate': round(repeat_2plus / total_unique, 4) if total_unique > 0 else 0,
        'vip_rate':   round(repeat_5plus / total_unique, 4) if total_unique > 0 else 0,
        'guest_orders': guest_order_count,
        'guest_revenue': round(guest_revenue, 0),
        'member_total_revenue': round(member_total_revenue, 0),
        'repeat_revenue_total': round(repeat_revenue_total, 0),
        'vip_revenue': round(vip_revenue, 0),
        'vip_revenue_share': round((vip_revenue / member_total_revenue) if member_total_revenue > 0 else 0, 4),
        'monthly': monthly_new_repeat_out,
        'repeat_interval_distribution': interval_buckets,
    },
    'price_band': {
        'bands': band_out,
        'aov_monthly': aov_out,
    },
    'issues': issues,
}

with open(OUT_JSON, 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)
print(f'저장: {OUT_JSON}')

# ================= Markdown Report =================
def fmt_krw(v):
    return f'{int(v):,}'

md = []
md.append('# 카페24 자사몰 심층 분석 리포트 (2025.01 ~ 2026.02, 14개월)')
md.append('')
md.append('> 로직: 모리 행단위 주간정산 (주문상태 필터, 주문별 dedup, 6개 차감, 원가)')
md.append(f'> 회원 식별: 수령인 휴대전화 (카페24 원본에 주문자ID/회원ID 컬럼 없음)')
md.append('')

# 1. SKU
md.append('## 1. SKU 분석')
md.append('')
md.append('### 1-1. 매출 Top 30')
md.append('')
md.append('| 순위 | SKU | 매출 | 이익 | 마진율 | 수량 |')
md.append('|---:|:---|---:|---:|---:|---:|')
for i, r in enumerate(by_revenue, 1):
    md.append(f'| {i} | {r["name"][:40]} | {fmt_krw(r["revenue"])} | {fmt_krw(r["profit"])} | {r["margin_rate"]*100:.1f}% | {r["qty"]} |')
md.append('')
md.append('### 1-2. 이익 Top 30')
md.append('')
md.append('| 순위 | SKU | 이익 | 매출 | 마진율 | 수량 |')
md.append('|---:|:---|---:|---:|---:|---:|')
for i, r in enumerate(by_profit, 1):
    md.append(f'| {i} | {r["name"][:40]} | {fmt_krw(r["profit"])} | {fmt_krw(r["revenue"])} | {r["margin_rate"]*100:.1f}% | {r["qty"]} |')
md.append('')
md.append('### 1-3. 마진율 Top 30 (수량 50개+)')
md.append('')
md.append('| 순위 | SKU | 마진율 | 매출 | 이익 | 수량 |')
md.append('|---:|:---|---:|---:|---:|---:|')
for i, r in enumerate(by_margin_top, 1):
    md.append(f'| {i} | {r["name"][:40]} | {r["margin_rate"]*100:.1f}% | {fmt_krw(r["revenue"])} | {fmt_krw(r["profit"])} | {r["qty"]} |')
md.append('')
md.append('### 1-4. 마진율 Bottom 30 (수량 50개+) — 수익성 낮은 SKU')
md.append('')
md.append('| 순위 | SKU | 마진율 | 매출 | 이익 | 수량 |')
md.append('|---:|:---|---:|---:|---:|---:|')
for i, r in enumerate(by_margin_bot, 1):
    md.append(f'| {i} | {r["name"][:40]} | {r["margin_rate"]*100:.1f}% | {fmt_krw(r["revenue"])} | {fmt_krw(r["profit"])} | {r["qty"]} |')
md.append('')
worst = by_margin_bot[:5]
md.append('**인사이트**: 마진율 워스트 5 → ' + ', '.join(f'{w["name"][:20]}({w["margin_rate"]*100:.1f}%)' for w in worst))
md.append('')

# 2. Category
md.append('## 2. 카테고리별 분석')
md.append('')
md.append('### 2-1. 14개월 합계')
md.append('')
md.append('| 카테고리 | 매출 | 이익 | 마진율 | 수량 | 매출 비중 |')
md.append('|:---|---:|---:|---:|---:|---:|')
for cat, d in cat_totals_out.items():
    md.append(f'| {cat} | {fmt_krw(d["revenue"])} | {fmt_krw(d["profit"])} | {d["margin_rate"]*100:.1f}% | {d["qty"]} | {d["share"]*100:.1f}% |')
md.append('')
md.append('### 2-2. 월별 매출 (카테고리)')
md.append('')
cats_order = list(cat_totals_out.keys())
header = '| 월 | ' + ' | '.join(cats_order) + ' |'
md.append(header)
md.append('|:---|' + '---:|' * len(cats_order))
for lab in sorted(cat_monthly_out.keys()):
    row = [lab]
    for cat in cats_order:
        d = cat_monthly_out[lab].get(cat, {})
        row.append(fmt_krw(d.get('revenue', 0)))
    md.append('| ' + ' | '.join(row) + ' |')
md.append('')
top_cat = next(iter(cat_totals_out))
top_cat_d = cat_totals_out[top_cat]
md.append(f'**인사이트**: 매출 1위 카테고리 = **{top_cat}** (매출 {fmt_krw(top_cat_d["revenue"])}원, 비중 {top_cat_d["share"]*100:.1f}%, 마진율 {top_cat_d["margin_rate"]*100:.1f}%)')
md.append('')

# 3. Member
md.append('## 3. 회원 재구매 분석')
md.append('')
md.append(f'- 회원 식별 컬럼: **수령인 휴대전화** (주문자ID/회원ID 컬럼 없음)')
md.append(f'- 총 unique 회원: **{total_unique:,}명** (전화번호 기반)')
md.append(f'- 1회 구매만: {new_only:,}명')
md.append(f'- 2회 이상 재구매: {repeat_2plus:,}명 (재구매율 **{(repeat_2plus/total_unique*100) if total_unique else 0:.1f}%**)')
md.append(f'- VIP (5회+): {repeat_5plus:,}명 ({(repeat_5plus/total_unique*100) if total_unique else 0:.1f}%)')
md.append(f'- 비회원/무전화 주문: {guest_order_count:,}건 (매출 {fmt_krw(guest_revenue)}원)')
md.append(f'- 재구매 매출 총액: {fmt_krw(repeat_revenue_total)}원 (회원매출 중 {(repeat_revenue_total/member_total_revenue*100) if member_total_revenue else 0:.1f}%)')
md.append(f'- VIP 매출 기여: {fmt_krw(vip_revenue)}원 ({(vip_revenue/member_total_revenue*100) if member_total_revenue else 0:.1f}%)')
md.append('')
md.append('### 재구매 간격 분포')
md.append('')
md.append('| 구간 | 건수 |')
md.append('|:---|---:|')
for k, v in interval_buckets.items():
    md.append(f'| {k} | {v:,} |')
md.append('')
md.append('### 월별 신규 vs 재구매 매출')
md.append('')
md.append('| 월 | 신규 매출 | 재구매 매출 | 재구매 비중 | 신규 주문 | 재구매 주문 |')
md.append('|:---|---:|---:|---:|---:|---:|')
for lab, d in monthly_new_repeat_out.items():
    md.append(f'| {lab} | {fmt_krw(d["new_revenue"])} | {fmt_krw(d["repeat_revenue"])} | {d["repeat_share"]*100:.1f}% | {d["new_orders"]} | {d["repeat_orders"]} |')
md.append('')
md.append(f'**인사이트**: 재구매 회원 비중 **{(repeat_2plus/total_unique*100) if total_unique else 0:.1f}%**, VIP(5회+) 단 {(repeat_5plus/total_unique*100) if total_unique else 0:.1f}%가 전체 회원매출의 **{(vip_revenue/member_total_revenue*100) if member_total_revenue else 0:.1f}%** 차지')
md.append('')

# 4. Price band
md.append('## 4. 가격대별 분포 + AOV')
md.append('')
md.append('### 4-1. 가격대 분포')
md.append('')
md.append('| 구간 | 주문수 | 매출 | 비중 | 평균 마진율 | 구간 AOV |')
md.append('|:---|---:|---:|---:|---:|---:|')
for name, _, _ in BANDS:
    b = band_out[name]
    md.append(f'| {name} | {b["orders"]:,} | {fmt_krw(b["revenue"])} | {b["share"]*100:.1f}% | {b["margin_rate"]*100:.1f}% | {fmt_krw(b["avg_order"])} |')
md.append('')
md.append('### 4-2. 월별 AOV (객단가)')
md.append('')
md.append('| 월 | AOV |')
md.append('|:---|---:|')
for lab, v in aov_out.items():
    md.append(f'| {lab} | {fmt_krw(v)} |')
md.append('')
aov_vals = list(aov_out.values())
if aov_vals:
    first_aov, last_aov = aov_vals[0], aov_vals[-1]
    delta = last_aov - first_aov
    md.append(f'**인사이트**: AOV {fmt_krw(first_aov)} → {fmt_krw(last_aov)}원 ({"+" if delta>=0 else ""}{fmt_krw(delta)}원, {(delta/first_aov*100) if first_aov else 0:+.1f}%)')
md.append('')

if issues:
    md.append('## ⚠️ 이슈')
    md.append('')
    for iss in issues:
        md.append(f'- {iss}')
    md.append('')

with open(OUT_MD, 'w', encoding='utf-8') as f:
    f.write('\n'.join(md))
print(f'저장: {OUT_MD}')
print()
print('=== 핵심 요약 ===')
print(f'카테고리 1위: {top_cat} 매출 {fmt_krw(top_cat_d["revenue"])}원, 비중 {top_cat_d["share"]*100:.1f}%')
print(f'재구매율: {(repeat_2plus/total_unique*100) if total_unique else 0:.1f}% (2회+ 회원 {repeat_2plus:,}/{total_unique:,}명)')
print(f'VIP 비중: {(vip_revenue/member_total_revenue*100) if member_total_revenue else 0:.1f}% ({repeat_5plus:,}명이 회원매출의')
print('마진율 Bottom 5:')
for w in by_margin_bot[:5]:
    print(f'  {w["name"][:40]}  {w["margin_rate"]*100:.1f}%  qty={w["qty"]}')
print('AOV 월별:', ', '.join(f'{k}:{fmt_krw(v)}' for k, v in aov_out.items()))
print('끝')
