"""
모리 에이전트 - 카페24 14개월 (2025.01 ~ 2026.02) 행단위 주간 정산 검증
방법: row-wise weekly settlement (per-order dedup for coupon/point/shipping)
출력: 02_mori_result.json
"""
import sys, io, os, json, re, tempfile, shutil, traceback
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# -------- 경로 --------
MAPPING_FILE = 'C:/Users/info/ClaudeAITeam/accounting/multi_agent_verify/master_mapping.json'
SRC_DIR = 'N:/개인/Becorelab/03. 영업/20. 월별 매출정산/기타/자사몰 검증'
OUT_JSON = 'C:/Users/info/ClaudeAITeam/accounting/multi_agent_verify/02_mori_result.json'

# -------- 주문 상태 필터 --------
INCLUDE_STATUSES = {'배송 완료', '구매 확정', '배송중', '배송 준비중'}

# -------- 유틸 --------
def sf(v):
    if v is None or v == '':
        return 0.0
    try:
        return float(str(v).replace(',', '').strip())
    except Exception:
        return 0.0

def si(v):
    return int(sf(v))

# -------- 매핑 로드 --------
print('[1/3] master_mapping.json 로드 중...')
with open(MAPPING_FILE, 'r', encoding='utf-8') as f:
    mapping = json.load(f)

name_map = mapping['name_map']
cost_map = mapping['cost_map']
cost_map_monthly = mapping.get('cost_map_monthly', {})
print(f'  name_map: {len(name_map)}개')
print(f'  cost_map: {len(cost_map)}개')
print(f'  cost_map_monthly: {len(cost_map_monthly)}개월')

# -------- 월별 파일 탐색 --------
print('[2/3] 원본 파일 탐색...')
month_files = []
for fn in sorted(os.listdir(SRC_DIR)):
    if '카페24 원본' in fn and fn.endswith('.xlsx'):
        m = re.match(r'(\d{2})\.\s*(\d{2})\s*카페24', fn)
        if m:
            label = f'20{m.group(1)}-{m.group(2)}'
            month_files.append((label, os.path.join(SRC_DIR, fn)))
print(f'  총 {len(month_files)}개 파일')

# -------- 컬럼 검색 --------
def find_col(headers, keyword, strict=False):
    """keyword가 포함된 첫 컬럼 인덱스. strict=True면 완전일치."""
    for i, h in enumerate(headers):
        if h is None:
            continue
        s = str(h)
        if strict:
            if s.strip() == keyword:
                return i
        else:
            if keyword in s:
                return i
    return None

# -------- 월 처리 --------
def process_month(label, filepath, issues):
    # N: 드라이브 권한 문제 회피
    fd, dst = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    shutil.copy(filepath, dst)
    try:
        wb = openpyxl.load_workbook(dst, data_only=True, read_only=True)
        ws = wb['원본']

        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = list(row)
            break

        IDX_ORDER  = find_col(headers, '주문번호')
        IDX_STATUS = find_col(headers, '주문 상태')
        IDX_QTY    = find_col(headers, '수량')
        IDX_OPTPR  = find_col(headers, '옵션+판매가')
        IDX_NAME   = find_col(headers, '주문상품명(옵션포함)')
        IDX_SHIP   = find_col(headers, '총 배송비(KRW)')
        IDX_COUPON = find_col(headers, '쿠폰 할인금액(최초)')
        IDX_POINT  = find_col(headers, '사용한 적립금액(최종)')
        IDX_REFUND = find_col(headers, '실제 환불금액')
        IDX_RANK   = find_col(headers, '회원등급 추가할인금액')
        IDX_APP    = find_col(headers, '앱 상품할인 금액(최종)')
        IDX_EXTRA  = find_col(headers, '상품별 추가할인금액')

        required = {
            'order': IDX_ORDER, 'status': IDX_STATUS, 'qty': IDX_QTY,
            'opt_price': IDX_OPTPR, 'name': IDX_NAME, 'ship': IDX_SHIP,
            'coupon': IDX_COUPON, 'point': IDX_POINT, 'refund': IDX_REFUND,
            'rank': IDX_RANK, 'app': IDX_APP, 'extra': IDX_EXTRA,
        }
        miss = [k for k, v in required.items() if v is None]
        if miss:
            issues.append(f'{label}: 컬럼 찾기 실패 {miss}')

        # 월별 원가 fallback
        month_cost_fallback = cost_map_monthly.get(label, {})

        sum_gross = 0.0
        sum_coupon = 0.0
        sum_refund = 0.0
        sum_point = 0.0
        sum_rank = 0.0
        sum_app = 0.0
        sum_extra = 0.0
        sum_ship = 0.0
        sum_cost = 0.0
        sum_qty = 0
        orders_seen = set()
        excluded_orders = set()
        unmapped_count = 0
        no_cost_count = 0
        no_cost_value = 0.0
        unmapped_samples = {}
        no_cost_samples = {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            if IDX_ORDER is None or row[IDX_ORDER] is None:
                continue
            order = str(row[IDX_ORDER]).strip()
            if not order:
                continue

            status = str(row[IDX_STATUS] or '').strip() if IDX_STATUS is not None else ''
            if status not in INCLUDE_STATUSES:
                excluded_orders.add(order)
                continue

            qty = si(row[IDX_QTY])
            opt_price = sf(row[IDX_OPTPR])
            gross_row = qty * opt_price

            is_first = order not in orders_seen
            if is_first:
                orders_seen.add(order)

            coupon = sf(row[IDX_COUPON]) if is_first else 0.0
            point  = sf(row[IDX_POINT])  if is_first else 0.0
            ship   = sf(row[IDX_SHIP])   if is_first else 0.0

            refund = sf(row[IDX_REFUND])
            rank_d = sf(row[IDX_RANK])
            app_d  = sf(row[IDX_APP])
            extra_d = sf(row[IDX_EXTRA])

            # 원가
            raw_name = str(row[IDX_NAME] or '').strip()
            standard = name_map.get(raw_name, '')
            unit_cost = 0.0
            if standard:
                if standard in cost_map:
                    unit_cost = float(cost_map[standard])
                elif standard in month_cost_fallback:
                    unit_cost = float(month_cost_fallback[standard])

            if not standard:
                unmapped_count += 1
                if raw_name and raw_name not in unmapped_samples:
                    unmapped_samples[raw_name] = gross_row
            elif unit_cost == 0:
                no_cost_count += 1
                no_cost_value += gross_row
                if standard not in no_cost_samples:
                    no_cost_samples[standard] = gross_row

            cost_row = unit_cost * qty

            sum_gross += gross_row
            sum_coupon += coupon
            sum_refund += refund
            sum_point += point
            sum_rank += rank_d
            sum_app += app_d
            sum_extra += extra_d
            sum_ship += ship
            sum_cost += cost_row
            sum_qty += qty

        wb.close()

        net_revenue = sum_gross - sum_coupon - sum_refund - sum_point - sum_rank - sum_app - sum_extra
        net_revenue_with_ship = net_revenue + sum_ship
        profit = net_revenue - sum_cost
        margin_rate = (profit / net_revenue) if net_revenue > 0 else 0.0

        result = {
            'gross': round(sum_gross, 2),
            'coupon': round(sum_coupon, 2),
            'refund': round(sum_refund, 2),
            'point': round(sum_point, 2),
            'rank_discount': round(sum_rank, 2),
            'app_discount': round(sum_app, 2),
            'extra_discount': round(sum_extra, 2),
            'shipping': round(sum_ship, 2),
            'net_revenue': round(net_revenue, 2),
            'net_revenue_with_ship': round(net_revenue_with_ship, 2),
            'cost': round(sum_cost, 2),
            'profit': round(profit, 2),
            'margin_rate': round(margin_rate, 6),
            'qty': sum_qty,
            'order_count': len(orders_seen),
            'excluded_order_count': len(excluded_orders - orders_seen),
            'unmapped_count': unmapped_count,
            'no_cost_count': no_cost_count,
            'no_cost_value': round(no_cost_value, 2),
        }

        # 이슈 리포트
        if unmapped_count > 0:
            top = sorted(unmapped_samples.items(), key=lambda x: -x[1])[:3]
            sample_str = '; '.join(f'{k[:40]}(₩{v:,.0f})' for k, v in top)
            issues.append(f'{label}: 미매핑 {unmapped_count}건 | 예: {sample_str}')
        if no_cost_count > 0:
            top = sorted(no_cost_samples.items(), key=lambda x: -x[1])[:3]
            sample_str = '; '.join(f'{k}(₩{v:,.0f})' for k, v in top)
            issues.append(f'{label}: 원가없음 {no_cost_count}건, ₩{no_cost_value:,.0f} | 예: {sample_str}')

        return result
    finally:
        try:
            os.remove(dst)
        except Exception:
            pass

# -------- 14개월 처리 --------
print('[3/3] 14개월 처리...')
issues = []
months_result = {}
for label, fp in month_files:
    print(f'  {label} ...', end=' ', flush=True)
    try:
        r = process_month(label, fp, issues)
        months_result[label] = r
        print(f'gross={r["gross"]:>13,.0f} net={r["net_revenue"]:>13,.0f} profit={r["profit"]:>12,.0f} margin={r["margin_rate"]*100:5.1f}%')
    except Exception as e:
        print(f'ERROR: {e}')
        traceback.print_exc()
        issues.append(f'{label}: 처리 에러 {e}')

# -------- totals --------
def sum_field(f):
    return sum(months_result[k][f] for k in months_result)

totals_gross = sum_field('gross')
totals_coupon = sum_field('coupon')
totals_refund = sum_field('refund')
totals_point = sum_field('point')
totals_rank = sum_field('rank_discount')
totals_app = sum_field('app_discount')
totals_extra = sum_field('extra_discount')
totals_ship = sum_field('shipping')
totals_net = sum_field('net_revenue')
totals_net_ship = sum_field('net_revenue_with_ship')
totals_cost = sum_field('cost')
totals_profit = sum_field('profit')
totals_qty = sum_field('qty')
totals_order = sum_field('order_count')
totals_excluded = sum_field('excluded_order_count')
totals_unmapped = sum_field('unmapped_count')
totals_nocost = sum_field('no_cost_count')
totals_nocost_val = sum_field('no_cost_value')
totals_margin_rate = (totals_profit / totals_net) if totals_net > 0 else 0.0

totals = {
    'gross': round(totals_gross, 2),
    'coupon': round(totals_coupon, 2),
    'refund': round(totals_refund, 2),
    'point': round(totals_point, 2),
    'rank_discount': round(totals_rank, 2),
    'app_discount': round(totals_app, 2),
    'extra_discount': round(totals_extra, 2),
    'shipping': round(totals_ship, 2),
    'net_revenue': round(totals_net, 2),
    'net_revenue_with_ship': round(totals_net_ship, 2),
    'cost': round(totals_cost, 2),
    'profit': round(totals_profit, 2),
    'margin_rate': round(totals_margin_rate, 6),
    'qty': totals_qty,
    'order_count': totals_order,
    'excluded_order_count': totals_excluded,
    'unmapped_count': totals_unmapped,
    'no_cost_count': totals_nocost,
    'no_cost_value': round(totals_nocost_val, 2),
}

output = {
    'agent': 'mori',
    'method': 'row-wise weekly settlement (per-order dedup for coupon/point/shipping)',
    'months': months_result,
    'totals': totals,
    'issues': issues,
}

with open(OUT_JSON, 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

print()
print('=' * 120)
print(f"{'월':<9} {'매출(gross)':>14} {'순매출':>14} {'원가':>13} {'이익':>13} {'마진율':>8} {'수량':>7} {'주문':>6}")
print('-' * 120)
for label in sorted(months_result.keys()):
    r = months_result[label]
    print(f"{label:<9} {r['gross']:>14,.0f} {r['net_revenue']:>14,.0f} {r['cost']:>13,.0f} {r['profit']:>13,.0f} {r['margin_rate']*100:>7.1f}% {r['qty']:>7} {r['order_count']:>6}")
print('-' * 120)
print(f"{'합계':<9} {totals['gross']:>14,.0f} {totals['net_revenue']:>14,.0f} {totals['cost']:>13,.0f} {totals['profit']:>13,.0f} {totals['margin_rate']*100:>7.1f}% {totals['qty']:>7} {totals['order_count']:>6}")
print()
print(f'저장: {OUT_JSON}')
print(f'이슈: {len(issues)}건')
for iss in issues:
    print(f'  - {iss}')
