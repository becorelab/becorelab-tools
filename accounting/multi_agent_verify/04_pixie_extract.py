"""
🦋 픽시 - 대표님 수동 정산시트에서 카페24 매출/이익 추출
- 14개월 (2025.01 ~ 2026.02)
- 3가지 시트별로 값 추출 (월별 온라인 매출정산 / 채널별 매출 이익 / 채널별 매출 이익 (광고비적용))
- 가짜 데이터 금지: 못 찾으면 found=false + null
"""
import sys
import io
import os
import json
import tempfile
import shutil
import traceback

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.utils import get_column_letter

BASE_VERIFY = r'N:/개인/Becorelab/03. 영업/20. 월별 매출정산/기타/자사몰 검증'
BASE_2026_01 = r'N:/개인/Becorelab/03. 영업/20. 월별 매출정산/2026.01'
BASE_2026_02 = r'N:/개인/Becorelab/03. 영업/20. 월별 매출정산/2026.02'

FILES = {
    '2025-01': (BASE_VERIFY, '2025. 01 온라인 매출정산_0210_원가 수정.xlsx', 1),
    '2025-02': (BASE_VERIFY, '2025. 02 온라인 매출정산.xlsx', 2),
    '2025-03': (BASE_VERIFY, '2025. 03 온라인 매출정산_수정.xlsx', 3),
    '2025-04': (BASE_VERIFY, '2025. 04 온라인 매출정산.xlsx', 4),
    '2025-05': (BASE_VERIFY, '2025. 05 온라인 매출정산.xlsx', 5),
    '2025-06': (BASE_VERIFY, '2025. 06 온라인 매출정산.xlsx', 6),
    '2025-07': (BASE_VERIFY, '2025. 07 온라인 매출정산_광고비 적용 완료.xlsx', 7),
    '2025-08': (BASE_VERIFY, '2025. 08 온라인 매출정산.xlsx', 8),
    '2025-09': (BASE_VERIFY, '2025. 09 온라인 매출정산.xlsx', 9),
    '2025-10': (BASE_VERIFY, '2025. 10 온라인 매출정산.xlsx', 10),
    '2025-11': (BASE_VERIFY, '2025. 11 온라인 매출정산.xlsx', 11),
    '2025-12': (BASE_VERIFY, '2025. 12 온라인 매출정산_자사몰 수정.xlsx', 12),
    '2026-01': (BASE_2026_01, '2026. 01 온라인 매출정산.xlsx', 1),
    '2026-02': (BASE_2026_02, '2026. 02 온라인 매출정산.xlsx', 2),
}

CAFE24_KEYWORDS = ['카페24', '자사몰', '네이버페이']


def copy_to_temp(src):
    """N 드라이브 파일을 임시 위치로 복사"""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tf:
        shutil.copy(src, tf.name)
        return tf.name


def num(v):
    """숫자 변환 (None/str/err 처리)"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip().replace(',', '')
        if not s or s.startswith('#'):
            return None
        try:
            return float(s)
        except ValueError:
            return None
    return None


def find_cafe24_row(ws, scan_col_range=(1, 3), row_max=200):
    """카페24 행 찾기 - B열(또는 A/C) 스캔"""
    for r in range(1, min(ws.max_row, row_max) + 1):
        for c in range(scan_col_range[0], scan_col_range[1] + 1):
            v = ws.cell(r, c).value
            if v and isinstance(v, str):
                for kw in CAFE24_KEYWORDS:
                    if kw in v:
                        return r
    return None


def find_header_rows(ws, max_scan=30):
    """
    채널별 시트에서 '월 헤더 행'과 '서브 헤더 행'을 자동 탐지.
    B열에 '채널'이 있는 행이 월 헤더 행, 그 다음 행이 서브 헤더.
    """
    for r in range(1, max_scan + 1):
        v = ws.cell(r, 2).value
        if v and isinstance(v, str) and v.strip() == '채널':
            return r, r + 1
    # fallback: '1월'이 나타나는 행을 찾기
    for r in range(1, max_scan + 1):
        for c in range(1, min(ws.max_column, 10) + 1):
            v = ws.cell(r, c).value
            if v and isinstance(v, str) and v.strip() == '1월':
                return r, r + 1
    return None, None


def find_month_columns(ws, month_idx, month_header_row, sub_header_row):
    """
    채널별 시트에서 month_idx 월의 (시작 컬럼, 끝 컬럼) 반환 (end exclusive)
    월 헤더는 병합셀이라 첫 셀에만 값이 있음
    """
    if month_header_row is None:
        return None, None
    month_cols = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(month_header_row, c).value
        if v and isinstance(v, str):
            s = v.strip()
            for n in range(1, 13):
                if s == f'{n}월' or s == f'{n:02d}월':
                    month_cols.append((n, c))
                    break
    month_cols.sort()
    for i, (n, c) in enumerate(month_cols):
        if n == month_idx:
            if i + 1 < len(month_cols):
                return c, month_cols[i + 1][1]
            for c2 in range(c + 1, ws.max_column + 2):
                v2 = ws.cell(month_header_row, c2).value
                if v2 and isinstance(v2, str) and '합계' in v2:
                    return c, c2
            return c, ws.max_column + 1
    return None, None


def extract_channel_sheet_pre_ad(ws, month_idx):
    """
    채널별 매출 이익 시트 (광고비 적용 전)
    - 행12 = 월 헤더, 행13 = 매출/이익/이익률
    - 헤더 동적 탐지
    """
    result = {
        'sheet_name': '채널별 매출 이익',
        'cafe24_revenue': None,
        'cafe24_profit': None,
        'cafe24_margin_rate': None,
        'found': False,
        'raw_row': None,
        'raw_columns': {},
        'note': None,
    }
    try:
        row = None
        for r in range(1, min(ws.max_row, 100) + 1):
            v = ws.cell(r, 2).value
            if v and isinstance(v, str) and '카페24' in v:
                row = r
                break
        if row is None:
            result['note'] = '카페24 행을 찾지 못함'
            return result

        mhr, shr = find_header_rows(ws)
        if mhr is None:
            result['note'] = '헤더 행을 찾지 못함'
            return result
        start_col, end_col = find_month_columns(ws, month_idx, mhr, shr)
        if start_col is None:
            result['note'] = f'{month_idx}월 헤더를 행{mhr}에서 찾지 못함'
            return result

        # 카페24 행을 다시 찾되, 월 헤더 행보다 아래
        row = None
        for r in range(shr + 1, min(ws.max_row, 100) + 1):
            v = ws.cell(r, 2).value
            if v and isinstance(v, str) and '카페24' in v:
                row = r
                break
        if row is None:
            result['note'] = '카페24 행을 찾지 못함'
            return result

        rev_c = prof_c = margin_c = None
        for c in range(start_col, end_col):
            v = ws.cell(shr, c).value
            if v and isinstance(v, str):
                s = v.strip()
                if s == '매출' and rev_c is None:
                    rev_c = c
                elif s == '이익' and prof_c is None:
                    prof_c = c
                elif s == '이익률' and margin_c is None:
                    margin_c = c

        if rev_c is None or prof_c is None:
            result['note'] = f'{month_idx}월 매출/이익 헤더를 찾지 못함'
            return result

        rev = num(ws.cell(row, rev_c).value)
        prof = num(ws.cell(row, prof_c).value)
        margin = num(ws.cell(row, margin_c).value) if margin_c else None

        result['cafe24_revenue'] = rev
        result['cafe24_profit'] = prof
        result['cafe24_margin_rate'] = margin
        result['found'] = rev is not None or prof is not None
        result['raw_row'] = row
        result['raw_columns'] = {
            '매출': get_column_letter(rev_c),
            '이익': get_column_letter(prof_c),
            '이익률': get_column_letter(margin_c) if margin_c else None,
        }
        result['header_check'] = {
            f'row{mhr}': str(ws.cell(mhr, start_col).value),
            f'row{shr}_revenue': str(ws.cell(shr, rev_c).value),
        }
    except Exception as e:
        result['note'] = f'error: {e}'
    return result


def extract_channel_sheet_post_ad(ws, month_idx):
    """
    채널별 매출 이익 (광고비적용) 시트
    - 행12 = 월 헤더, 행13 = 매출/이익/광고비/이익률
    - 헤더 동적 탐지
    """
    result = {
        'sheet_name': '채널별 매출 이익 (광고비적용)',
        'cafe24_revenue': None,
        'cafe24_profit': None,
        'cafe24_ad_cost': None,
        'cafe24_margin_rate': None,
        'found': False,
        'raw_row': None,
        'raw_columns': {},
        'note': None,
    }
    try:
        mhr, shr = find_header_rows(ws)
        if mhr is None:
            result['note'] = '헤더 행을 찾지 못함'
            return result
        start_col, end_col = find_month_columns(ws, month_idx, mhr, shr)
        if start_col is None:
            result['note'] = f'{month_idx}월 헤더를 행{mhr}에서 찾지 못함'
            return result

        row = None
        for r in range(shr + 1, min(ws.max_row, 100) + 1):
            v = ws.cell(r, 2).value
            if v and isinstance(v, str) and '카페24' in v:
                row = r
                break
        if row is None:
            result['note'] = '카페24 행을 찾지 못함'
            return result

        rev_c = prof_c = ad_c = margin_c = None
        for c in range(start_col, end_col):
            v = ws.cell(shr, c).value
            if v and isinstance(v, str):
                s = v.strip()
                if s == '매출' and rev_c is None:
                    rev_c = c
                elif s == '이익' and prof_c is None:
                    prof_c = c
                elif s == '광고비' and ad_c is None:
                    ad_c = c
                elif s == '이익률' and margin_c is None:
                    margin_c = c

        if rev_c is None or prof_c is None:
            result['note'] = f'{month_idx}월 매출/이익 헤더를 찾지 못함'
            return result

        rev = num(ws.cell(row, rev_c).value)
        prof = num(ws.cell(row, prof_c).value)
        ad = num(ws.cell(row, ad_c).value) if ad_c else None
        margin = num(ws.cell(row, margin_c).value) if margin_c else None

        result['cafe24_revenue'] = rev
        result['cafe24_profit'] = prof
        result['cafe24_ad_cost'] = ad
        result['cafe24_margin_rate'] = margin
        result['found'] = rev is not None or prof is not None
        result['raw_row'] = row
        result['raw_columns'] = {
            '매출': get_column_letter(rev_c),
            '이익': get_column_letter(prof_c),
            '광고비': get_column_letter(ad_c) if ad_c else None,
            '이익률': get_column_letter(margin_c) if margin_c else None,
        }
        result['header_check'] = {
            f'row{mhr}': str(ws.cell(mhr, start_col).value),
            f'row{shr}_revenue': str(ws.cell(shr, rev_c).value),
        }
    except Exception as e:
        result['note'] = f'error: {e}'
    return result


def extract_monthly_sheet(ws, month_idx):
    """
    월별 온라인 매출정산 시트
    - 행6 = 월 헤더 (병합셀) F6='1월', 다음 월은 열 패턴에 따라 달라짐
    - 행7 = 판매수량/매출액/배송비/매출+배송비/개별원가/이익/[이익률]/매출금액/이익합계
      - 신식(2025-03 이후): 월당 9열 (이익률 포함)
      - 구식(2025-01/02):    월당 8열 (이익률 없음)
    - 헤더 동적 파싱: 행6에서 n월 병합셀 시작 컬럼 찾기 + 행7에서 매출금액/이익합계 컬럼 찾기
    """
    result = {
        'sheet_name': '월별 온라인 매출정산',
        'cafe24_revenue': None,
        'cafe24_profit': None,
        'cafe24_margin_rate': None,
        'found': False,
        'raw_row': None,
        'raw_columns': {},
        'note': None,
    }
    try:
        # 1. 자사몰/카페24 시작 행 찾기
        start_row = None
        for r in range(1, min(ws.max_row, 200) + 1):
            a = ws.cell(r, 1).value
            b = ws.cell(r, 2).value
            if (a and isinstance(a, str) and '자사몰' in a) or \
               (b and isinstance(b, str) and '카페24' in b):
                start_row = r
                break
        if start_row is None:
            result['note'] = '자사몰/카페24 행을 찾지 못함'
            return result

        # 2. 월 헤더(행6)에서 각 월 시작 컬럼 수집 (병합셀의 첫 셀만 값 있음)
        month_cols = []  # [(month_num, col), ...]
        for c in range(1, ws.max_column + 1):
            v = ws.cell(6, c).value
            if v and isinstance(v, str):
                s = v.strip()
                # '1월', '01월', '2월', ...
                for n in range(1, 13):
                    if s == f'{n}월' or s == f'{n:02d}월':
                        month_cols.append((n, c))
                        break
        month_cols.sort()

        target_col = None
        next_col = None
        for i, (n, c) in enumerate(month_cols):
            if n == month_idx:
                target_col = c
                if i + 1 < len(month_cols):
                    next_col = month_cols[i + 1][1]
                else:
                    # 마지막 월이면 '합계' 또는 max_column 다음
                    for c2 in range(c + 1, ws.max_column + 2):
                        v2 = ws.cell(6, c2).value
                        if v2 and isinstance(v2, str) and '합계' in v2:
                            next_col = c2
                            break
                    if next_col is None:
                        next_col = ws.max_column + 1
                break

        if target_col is None:
            result['note'] = f'{month_idx}월 헤더를 행6에서 찾지 못함'
            return result

        # 3. 행7에서 target_col ~ next_col-1 범위의 헤더 탐색
        rev_sum_c = None
        prof_sum_c = None
        for c in range(target_col, next_col):
            v = ws.cell(7, c).value
            if v and isinstance(v, str):
                s = v.strip()
                if s == '매출금액' and rev_sum_c is None:
                    rev_sum_c = c
                elif s == '이익합계' and prof_sum_c is None:
                    prof_sum_c = c

        if rev_sum_c is None or prof_sum_c is None:
            result['note'] = f'{month_idx}월 매출금액/이익합계 헤더를 찾지 못함 (target_col={target_col}, next_col={next_col})'
            return result

        # 4. 자사몰 블록 스캔하여 값 읽기 (병합셀 첫 행에 값)
        rev = num(ws.cell(start_row, rev_sum_c).value)
        prof = num(ws.cell(start_row, prof_sum_c).value)

        if rev is None and prof is None:
            for r in range(start_row, start_row + 80):
                v_b = ws.cell(r, 2).value
                if v_b is None or (isinstance(v_b, str) and '카페24' not in v_b):
                    break
                rv = num(ws.cell(r, rev_sum_c).value)
                pv = num(ws.cell(r, prof_sum_c).value)
                if rv is not None and rev is None:
                    rev = rv
                if pv is not None and prof is None:
                    prof = pv
                if rev is not None and prof is not None:
                    break

        margin = None
        if rev and prof and rev != 0:
            margin = prof / rev

        result['cafe24_revenue'] = rev
        result['cafe24_profit'] = prof
        result['cafe24_margin_rate'] = margin
        result['found'] = rev is not None or prof is not None
        result['raw_row'] = start_row
        result['raw_columns'] = {
            '매출금액': get_column_letter(rev_sum_c),
            '이익합계': get_column_letter(prof_sum_c),
            'month_header_col': get_column_letter(target_col),
        }
        result['header_check'] = {
            'row6_month': str(ws.cell(6, target_col).value),
            'row7_rev_sum': str(ws.cell(7, rev_sum_c).value),
            'row7_prof_sum': str(ws.cell(7, prof_sum_c).value),
        }
    except Exception as e:
        result['note'] = f'error: {e}\n{traceback.format_exc()}'
    return result


def process_file(month_key, folder, filename, month_idx):
    src = os.path.join(folder, filename)
    src = src.replace('\\', '/')
    print(f'[{month_key}] {filename}')

    entry = {
        'source_file': filename,
        'month_idx': month_idx,
        'monthly_sheet': None,
        'channel_sheet_pre_ad': None,
        'channel_sheet_post_ad': None,
    }
    issues = []

    if not os.path.exists(src):
        msg = f'{month_key}: 파일 없음 - {src}'
        print('  ERR', msg)
        issues.append(msg)
        return entry, issues

    tmp = None
    try:
        tmp = copy_to_temp(src)
        wb = openpyxl.load_workbook(tmp, data_only=True)
        sheetnames = wb.sheetnames

        # 1. 월별 온라인 매출정산
        if '월별 온라인 매출정산' in sheetnames:
            ws = wb['월별 온라인 매출정산']
            entry['monthly_sheet'] = extract_monthly_sheet(ws, month_idx)
            if not entry['monthly_sheet']['found']:
                issues.append(f"{month_key}: 월별 온라인 매출정산 추출 실패 - {entry['monthly_sheet'].get('note')}")
        else:
            entry['monthly_sheet'] = {
                'sheet_name': '월별 온라인 매출정산',
                'found': False,
                'note': '시트 없음',
            }
            issues.append(f'{month_key}: 월별 온라인 매출정산 시트 없음')

        # 2. 채널별 매출 이익 (광고비 적용 전)
        if '채널별 매출 이익' in sheetnames:
            ws = wb['채널별 매출 이익']
            entry['channel_sheet_pre_ad'] = extract_channel_sheet_pre_ad(ws, month_idx)
            if not entry['channel_sheet_pre_ad']['found']:
                issues.append(f"{month_key}: 채널별 매출 이익 추출 실패 - {entry['channel_sheet_pre_ad'].get('note')}")
        else:
            entry['channel_sheet_pre_ad'] = {
                'sheet_name': '채널별 매출 이익',
                'found': False,
                'note': '시트 없음',
            }
            issues.append(f'{month_key}: 채널별 매출 이익 시트 없음')

        # 3. 채널별 매출 이익 (광고비적용)
        if '채널별 매출 이익 (광고비적용)' in sheetnames:
            ws = wb['채널별 매출 이익 (광고비적용)']
            entry['channel_sheet_post_ad'] = extract_channel_sheet_post_ad(ws, month_idx)
            if not entry['channel_sheet_post_ad']['found']:
                issues.append(f"{month_key}: 채널별 매출 이익 (광고비적용) 추출 실패 - {entry['channel_sheet_post_ad'].get('note')}")
        else:
            entry['channel_sheet_post_ad'] = {
                'sheet_name': '채널별 매출 이익 (광고비적용)',
                'found': False,
                'note': '시트 없음',
            }
            issues.append(f'{month_key}: 채널별 매출 이익 (광고비적용) 시트 없음')

    except Exception as e:
        msg = f'{month_key}: 파일 처리 에러 - {e}'
        print('  ERR', msg)
        issues.append(msg)
    finally:
        if tmp:
            try:
                os.unlink(tmp)
            except Exception:
                pass

    return entry, issues


def main():
    out = {
        'agent': 'pixie',
        'method': 'extract from manual settlement sheets',
        'months': {},
        'issues': [],
    }

    for month_key, (folder, filename, month_idx) in FILES.items():
        entry, issues = process_file(month_key, folder, filename, month_idx)
        out['months'][month_key] = entry
        out['issues'].extend(issues)

        # 요약 출력
        def fmt(v):
            if v is None:
                return 'None'
            return f'{v:>15,.0f}'

        m = entry.get('monthly_sheet') or {}
        p = entry.get('channel_sheet_pre_ad') or {}
        a = entry.get('channel_sheet_post_ad') or {}
        print(f'  월별시트     매출={fmt(m.get("cafe24_revenue"))} 이익={fmt(m.get("cafe24_profit"))}')
        print(f'  채널(pre)    매출={fmt(p.get("cafe24_revenue"))} 이익={fmt(p.get("cafe24_profit"))}')
        print(f'  채널(post)   매출={fmt(a.get("cafe24_revenue"))} 이익={fmt(a.get("cafe24_profit"))} 광고비={fmt(a.get("cafe24_ad_cost"))}')

    out_path = r'C:/Users/info/ClaudeAITeam/accounting/multi_agent_verify/04_pixie_result.json'
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=2, default=str)
    print(f'\nwrote: {out_path}')
    print(f'issues: {len(out["issues"])}건')
    for iss in out['issues']:
        print(f'  - {iss}')


if __name__ == '__main__':
    main()
