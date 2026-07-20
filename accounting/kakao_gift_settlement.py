# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import datetime

SRC='/Users/macmini_ky/Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브/Claude AI work space/mac window file transfer/카카오 선물하기_20260720061835.xlsx'
wb0=openpyxl.load_workbook(SRC, read_only=True, data_only=True)
ws0=wb0['Sheet0']; rows=list(ws0.iter_rows(values_only=True))
hdr=rows[0]; idx={h:i for i,h in enumerate(hdr)}; data=rows[1:]
def num(x):
    try: return float(x)
    except: return 0.0
CANCEL={'204 결제 취소 완료','303 결제 취소 완료','208 환불 완료'}
MAIN='[NEW 단독 런칭] 일비아 고체탈취제 2개 + 전용 홀더 1개'
COST=11938; FEE=0.10
wd=['월','화','수','목','금','토','일']
val=[r for r in data if r[idx['주문상태']] not in CANCEL]
canc=[r for r in data if r[idx['주문상태']] in CANCEL]

# 스타일
TITLE=Font(bold=True,size=14,color='FFFFFF')
HDR=Font(bold=True,color='FFFFFF')
HFILL=PatternFill('solid',fgColor='4472C4')
TFILL=PatternFill('solid',fgColor='2F5496')
SUMFILL=PatternFill('solid',fgColor='FFF2CC')
HITFILL=PatternFill('solid',fgColor='E2EFDA')
LEFT=Alignment(horizontal='left',vertical='center')
thin=Side(style='thin',color='CCCCCC')
BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
def style_hdr(ws,row,ncol):
    for c in range(1,ncol+1):
        cell=ws.cell(row=row,column=c); cell.font=HDR; cell.fill=HFILL; cell.alignment=LEFT; cell.border=BORD
def setw(ws,widths):
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w

wb=openpyxl.Workbook()

# ===== 시트1: 정산 요약 =====
ws=wb.active; ws.title='정산 요약'
ws.merge_cells('A1:D1')
ws['A1']='🎁 카카오 선물하기 포미위크 — 고체탈취제 런칭 정산 [최종]'
ws['A1'].font=TITLE; ws['A1'].fill=TFILL; ws['A1'].alignment=LEFT
ws.row_dimensions[1].height=24
r=3
info=[
 ('채널','카카오 선물하기 / 포미위크'),
 ('상품','일비아 고체탈취제(에어밤) 런칭 프로모션'),
 ('이벤트 기간','2026-07-13(월) ~ 07-19(일) · 07/20 00시 종료'),
 ('집계 기준','원본 958건 중 취소·환불 49건 제외 = 유효 909건'),
 ('원본 파일','카카오 선물하기_20260720061835.xlsx'),
 ('작성','정산 하치 · 2026-07-20'),
]
for k,v in info:
    ws.cell(r,1,k).font=Font(bold=True); ws.cell(r,1).alignment=LEFT
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=4)
    ws.cell(r,2,v).alignment=LEFT; r+=1
r+=1
# 손익 표
ws.cell(r,1,'💰 손익 요약 (VAT 포함 기준)').font=Font(bold=True,size=12); r+=1
hrow=r
for c,h in enumerate(['구분','금액','비고',''],1): ws.cell(r,c,h)
style_hdr(ws,r,4); r+=1
mset=[x for x in val if x[idx['상품명']]==MAIN]
mset_qty=sum(num(x[idx['수량']]) for x in mset)
mset_rev=sum(num(x[idx['정산기준금액']]) for x in mset)
other=[x for x in val if x[idx['상품명']]!=MAIN]
other_qty=sum(num(x[idx['수량']]) for x in other)
other_rev=sum(num(x[idx['정산기준금액']]) for x in other)
tot_rev=mset_rev+other_rev
scost=mset_qty*COST
snet=tot_rev*(1-FEE)
sprofit=mset_rev*(1-FEE)-scost   # 세트기준 순이익(곁다리 원가 미반영)
pl=[
 ('총매출 (정산기준금액)', tot_rev, '유효 909건 / 956수량'),
 ('  ├ 고체탈취제 세트', mset_rev, f'{int(mset_qty)}세트'),
 ('  └ 곁다리 상품', other_rev, f'{len(other)}건 / {int(other_qty)}개'),
 ('카카오 수수료 (10%)', -tot_rev*FEE, '정산기준금액 × 10%'),
 ('순정산 예상액', snet, '총매출 − 수수료'),
 ('세트 원가', -scost, f'세트원가 {COST:,}원 × {int(mset_qty)}세트'),
 ('★ 세트 순이익', sprofit, '순정산(세트분) − 세트원가 / 곁다리 원가 별도'),
]
for label,amt,note in pl:
    ws.cell(r,1,label).alignment=LEFT
    cc=ws.cell(r,2,int(amt)); cc.number_format='#,##0'; cc.alignment=LEFT
    ws.cell(r,3,note).alignment=LEFT
    if '★' in label:
        for c in range(1,5): ws.cell(r,c).fill=SUMFILL; ws.cell(r,c).font=Font(bold=True)
    r+=1
r+=1
ws.cell(r,1,'세트 순이익률 (매출 대비)').alignment=LEFT
ws.cell(r,2,f'{sprofit/mset_rev*100:.1f}%').alignment=LEFT; r+=1
ws.cell(r,1,'세트 개당 순이익').alignment=LEFT
ws.cell(r,2,int(sprofit/mset_qty)).number_format='#,##0'; ws.cell(r,2).alignment=LEFT; r+=1
ws.cell(r,1,'취소·환불').alignment=LEFT
ws.cell(r,2,f'{len(canc)}건 / {int(sum(num(x[idx["상품금액"]]) for x in canc)):,}원 (취소율 {len(canc)/len(data)*100:.1f}%)').alignment=LEFT; r+=2

# 정산 일정 (실제 계약 — 월 2회)
ws.cell(r,1,'📅 정산 일정 (실제 계약 기준 · 월 2회)').font=Font(bold=True,size=12); r+=1
for c,h in enumerate(['회차','정산대상 기간','확정일','세금계산서','지급일'],1):
    cell=ws.cell(r,c,h); cell.font=HDR; cell.fill=HFILL; cell.alignment=LEFT; cell.border=BORD
r+=1
for row in [('1회','당월 1~15일','당월 20일','익월 10일까지','당월 25일까지'),
            ('2회','당월 16~말일','익월 5일','익월 10일까지','익월 10일까지')]:
    for c,vv in enumerate(row,1):
        cell=ws.cell(r,c,vv); cell.alignment=LEFT; cell.border=BORD
    r+=1
r+=1
for k,v in [
 ('정산 기준','구매확정 처리된 상품의 판매금액 − 수수료·차감금액·보류금액. 휴일이면 익영업일에 확정/지급'),
 ('수수료','10% (대표님 확정)'),
 ('세금계산서','구매결정/구매결정취소/환불분 발행 (for Biz: 을 및 을의 고객사분)'),
 ('⚠️ 주의','판매일 아닌 "구매확정일" 기준으로 회차 결정. 선물하기는 배송완료 후 자동 구매확정 → 시점차 발생')]:
    ws.cell(r,1,k).font=Font(bold=True); ws.cell(r,1).alignment=LEFT
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=5)
    ws.cell(r,2,v).alignment=Alignment(horizontal='left',vertical='center',wrap_text=True); r+=1
r+=1
ws.cell(r,1,'🗓️ 본 이벤트 구매확정 현황 (2026-07-20 기준)').font=Font(bold=True,size=12); r+=1
stat=defaultdict(lambda:[0,0.0])
for x in data:
    stat[x[idx['주문상태']]][0]+=1; stat[x[idx['주문상태']]][1]+=num(x[idx['정산기준금액']])
conf=stat['601 구매 결정']; done=stat['305 배송 완료']
ship=stat['202 배송 요청'][0]+stat['201 결제 완료'][0]+stat['301 배송 준비 중'][0]+stat['304 배송 중'][0]
ship_amt=stat['202 배송 요청'][1]+stat['201 결제 완료'][1]+stat['301 배송 준비 중'][1]+stat['304 배송 중'][1]
for c,h in enumerate(['상태','건수','금액','정산 회차 귀속'],1):
    cell=ws.cell(r,c,h); cell.font=HDR; cell.fill=HFILL; cell.alignment=LEFT; cell.border=BORD
r+=1
for label,cnt,amt,belong in [
 ('구매확정 완료 (601)',conf[0],int(conf[1]),'확정일 속한 회차 (대부분 7월 2회차: 8/5확정·8/10지급)'),
 ('배송완료·확정대기 (305)',done[0],int(done[1]),'자동 구매확정 후 → 7월 2회차 또는 8월 회차'),
 ('배송중·요청·결제완료',ship,int(ship_amt),'배송완료→구매확정 후 → 8월 회차 이월 가능')]:
    ws.cell(r,1,label).alignment=LEFT; ws.cell(r,1).border=BORD
    ws.cell(r,2,cnt).number_format='#,##0'; ws.cell(r,2).alignment=LEFT; ws.cell(r,2).border=BORD
    ws.cell(r,3,amt).number_format='#,##0'; ws.cell(r,3).alignment=LEFT; ws.cell(r,3).border=BORD
    ws.cell(r,4,belong).alignment=Alignment(horizontal='left',wrap_text=True); ws.cell(r,4).border=BORD
    r+=1
r+=1
ws.cell(r,1,'※ 오늘 기준 실제 구매확정은 46건뿐. 대금 대부분은 배송완료→자동구매확정 이후 정산. 정확한 회차는 구매확정일로 최종 확정.').alignment=Alignment(horizontal='left',wrap_text=True)
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5)
setw(ws,[24,16,14,42,14])

# ===== 시트2: 일별 판매량 =====
ws=wb.create_sheet('일별 판매량')
ws.cell(1,1,'📊 일별 판매량').font=Font(bold=True,size=13)
head=['날짜','요일','건수','총수량','세트수량','매출','수수료10%','순정산']
for c,h in enumerate(head,1): ws.cell(3,c,h)
style_hdr(ws,3,len(head))
day=defaultdict(lambda:[0,0.0,0.0,0.0])
for x in val:
    d=str(x[idx['주문일']])[:10]
    day[d][0]+=1; day[d][1]+=num(x[idx['수량']]); day[d][2]+=num(x[idx['정산기준금액']])
    if x[idx['상품명']]==MAIN: day[d][3]+=num(x[idx['수량']])
rr=4; tb=[0,0,0,0]
for k in sorted(day):
    y,m,dd=map(int,k.split('-')); w=wd[datetime.date(y,m,dd).weekday()]
    v=day[k]
    vals=[k,w,v[0],int(v[1]),int(v[3]),int(v[2]),int(v[2]*FEE),int(v[2]*(1-FEE))]
    for c,vv in enumerate(vals,1):
        cell=ws.cell(rr,c,vv); cell.alignment=LEFT; cell.border=BORD
        if c>=3 and c!=2: cell.number_format='#,##0'
    if v[0]==max(d2[0] for d2 in day.values()):
        for c in range(1,len(head)+1): ws.cell(rr,c).fill=HITFILL
    tb[0]+=v[0]; tb[1]+=v[1]; tb[2]+=v[3]; tb[3]+=v[2]; rr+=1
tv=['합계','',tb[0],int(tb[1]),int(tb[2]),int(tb[3]),int(tb[3]*FEE),int(tb[3]*(1-FEE))]
for c,vv in enumerate(tv,1):
    cell=ws.cell(rr,c,vv); cell.font=Font(bold=True); cell.fill=SUMFILL; cell.alignment=LEFT
    if c>=3 and c!=2: cell.number_format='#,##0'
setw(ws,[13,6,8,9,10,14,12,14])

# ===== 시트3: 상품별 =====
ws=wb.create_sheet('상품별')
ws.cell(1,1,'🎯 상품별 판매').font=Font(bold=True,size=13)
head=['상품명','건수','수량','매출','비중']
for c,h in enumerate(head,1): ws.cell(3,c,h)
style_hdr(ws,3,len(head))
prod=defaultdict(lambda:[0,0.0,0.0])
for x in val:
    nm=x[idx['상품명']]; prod[nm][0]+=1; prod[nm][1]+=num(x[idx['수량']]); prod[nm][2]+=num(x[idx['정산기준금액']])
rr=4
for nm,v in sorted(prod.items(),key=lambda z:-z[1][2]):
    vals=[nm,v[0],int(v[1]),int(v[2]),f'{v[2]/tot_rev*100:.1f}%']
    for c,vv in enumerate(vals,1):
        cell=ws.cell(rr,c,vv); cell.alignment=LEFT; cell.border=BORD
        if c in (2,3,4): cell.number_format='#,##0'
    rr+=1
setw(ws,[46,8,8,14,8])

# ===== 시트4: 시간대별 =====
ws=wb.create_sheet('시간대별')
ws.cell(1,1,'⏰ 시간대별 주문 (전체 유효)').font=Font(bold=True,size=13)
for c,h in enumerate(['시','건수','비중'],1): ws.cell(3,c,h)
style_hdr(ws,3,3)
hourd=defaultdict(int)
for x in val:
    t=str(x[idx['주문일']])
    if len(t)>=13: hourd[int(t[11:13])]+=1
rr=4; th=sum(hourd.values())
for h in sorted(hourd):
    for c,vv in enumerate([f'{h:02}시',hourd[h],f'{hourd[h]/th*100:.1f}%'],1):
        cell=ws.cell(rr,c,vv); cell.alignment=LEFT; cell.border=BORD
        if c==2: cell.number_format='#,##0'
    if hourd[h]==max(hourd.values()):
        for c in range(1,4): ws.cell(rr,c).fill=HITFILL
    rr+=1
setw(ws,[10,10,10])

# ===== 시트: 행사 완전손익 (현보님 마케팅 시트 반영) — 정산요약 다음(index 1) =====
ws=wb.create_sheet('행사 완전손익', 1)
ws.merge_cells('A1:C1')
ws['A1']='📊 행사 완전손익 (마케팅 비용 전체 반영 · 현보님 시트 기준)'
ws['A1'].font=TITLE; ws['A1'].fill=TFILL; ws['A1'].alignment=LEFT
ws.row_dimensions[1].height=24
r=3
ws.cell(r,1,'구분').font=HDR; ws.cell(r,2,'금액').font=HDR; ws.cell(r,3,'비고').font=HDR
for c in range(1,4): ws.cell(r,c).fill=HFILL; ws.cell(r,c).alignment=LEFT; ws.cell(r,c).border=BORD
r+=1
pnl=[
 ('매출 (가구매 300건 포함)', 29890300, '937세트 = 실판매 637 + 가구매 300', False),
 ('카카오 수수료 (전체 10%)', -2989030, '937건 정산기준금액 × 10%', False),
 ('제품원가 + 배송비', -13336321, '건당 14,233원 × 937 (배송비 원가 포함)', False),
 ('= 판매 순매출', 13564949, '매출 − 수수료 − 원가', True),
 ('광고선전비', -4834585, '실광고 + 순위부스팅(가구매·위시·랭킹)', False),
 ('사은품비', -1506800, '본품증정 + 상품권 + 호텔숙박권', False),
 ('★ 순이익', 7223564, '순이익률 24%', True),
]
for label,amt,note,hi in pnl:
    ws.cell(r,1,label).alignment=LEFT; ws.cell(r,1).border=BORD
    cc=ws.cell(r,2,amt); cc.number_format='#,##0'; cc.alignment=LEFT; cc.border=BORD
    ws.cell(r,3,note).alignment=LEFT; ws.cell(r,3).border=BORD
    if hi:
        for c in range(1,4):
            ws.cell(r,c).fill=SUMFILL; ws.cell(r,c).font=Font(bold=True)
    r+=1
r+=2
# 광고선전비 내역
ws.cell(r,1,'📢 광고선전비 내역 (₩4,834,585 · 단가 VAT별도 → 시트 VAT포함)').font=Font(bold=True,size=12); r+=1
for c,h in enumerate(['항목','계산(공급가)','VAT포함','성격'],1):
    cell=ws.cell(r,c,h); cell.font=HDR; cell.fill=HFILL; cell.alignment=LEFT; cell.border=BORD
r+=1
adrows=[
 ('가구매 비용(대행)','3,500 × 300건',1155000,'순위 부스팅'),
 ('좋아요(위시) 작업','300 × 850건',280500,'순위 부스팅'),
 ('가구매 수수료(카카오10%)','3,190 × 300건',957000,'순위 부스팅'),
 ('카카오 일비아 회원 푸쉬','—',651237,'광고'),
 ('실시간 랭킹','350,000 × 3건',1155000,'광고'),
 ('카카오 비즈보드 광고','—',635848,'광고'),
]
for label,calc,amt,kind in adrows:
    ws.cell(r,1,label).alignment=LEFT; ws.cell(r,1).border=BORD
    ws.cell(r,2,calc).alignment=LEFT; ws.cell(r,2).border=BORD
    cc=ws.cell(r,3,amt); cc.number_format='#,##0'; cc.alignment=LEFT; cc.border=BORD
    ws.cell(r,4,kind).alignment=LEFT; ws.cell(r,4).border=BORD
    r+=1
ws.cell(r,1,'합계').font=Font(bold=True); ws.cell(r,1).fill=SUMFILL; ws.cell(r,1).border=BORD
ws.cell(r,2).fill=SUMFILL; ws.cell(r,3,4834585).number_format='#,##0'; ws.cell(r,3).font=Font(bold=True); ws.cell(r,3).fill=SUMFILL; ws.cell(r,3).alignment=LEFT; ws.cell(r,3).border=BORD
ws.cell(r,4).fill=SUMFILL
r+=2
ws.cell(r,1,'   ├ 진짜 광고 (회원푸쉬+랭킹+비즈보드)').alignment=LEFT; ws.cell(r,2,2442085).number_format='#,##0'; ws.cell(r,2).alignment=LEFT; ws.cell(r,3,'50.5%').alignment=LEFT; r+=1
ws.cell(r,1,'   └ 순위 부스팅 (가구매+위시+가구매수수료)').alignment=LEFT; ws.cell(r,2,2392500).number_format='#,##0'; ws.cell(r,2).alignment=LEFT; ws.cell(r,3,'49.5%').alignment=LEFT; r+=2
# 사은품비 내역
ws.cell(r,1,'🎁 사은품비 내역 (₩1,506,800)').font=Font(bold=True,size=12); r+=1
for c,h in enumerate(['항목','단가','수량','금액'],1):
    cell=ws.cell(r,c,h); cell.font=HDR; cell.fill=HFILL; cell.alignment=LEFT; cell.border=BORD
r+=1
giftrows=[('본품 증정(선착순)','3,034','200개',606800),('상품권 1만원(추첨)','10,000','30개',300000),('5성급 호텔 숙박권(최다구매)','600,000','1개',600000)]
for label,unit,qty,amt in giftrows:
    ws.cell(r,1,label).alignment=LEFT; ws.cell(r,1).border=BORD
    ws.cell(r,2,unit).alignment=LEFT; ws.cell(r,2).border=BORD
    ws.cell(r,3,qty).alignment=LEFT; ws.cell(r,3).border=BORD
    cc=ws.cell(r,4,amt); cc.number_format='#,##0'; cc.alignment=LEFT; cc.border=BORD
    r+=1
ws.cell(r,1,'합계').font=Font(bold=True); ws.cell(r,1).fill=SUMFILL; ws.cell(r,1).border=BORD
ws.cell(r,2).fill=SUMFILL; ws.cell(r,3).fill=SUMFILL; ws.cell(r,4,1506800).number_format='#,##0'; ws.cell(r,4).font=Font(bold=True); ws.cell(r,4).fill=SUMFILL; ws.cell(r,4).alignment=LEFT; ws.cell(r,4).border=BORD
r+=2
ws.cell(r,1,'※ 제 제품마진 관점(15,715,364)과 다른 이유 = 위 배송비·광고선전비·사은품비를 모두 반영한 「행사 완전손익」이기 때문. 배송비는 제품원가 정밀화 시 미세조정 예정.').alignment=Alignment(horizontal='left',wrap_text=True)
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
setw(ws,[34,18,14,22])

# ===== 시트: 주문내역(유효) =====
ws=wb.create_sheet('주문내역')
head=['주문일','주문번호','상품명','수량','정산기준금액','주문상태']
for c,h in enumerate(head,1): ws.cell(1,c,h)
style_hdr(ws,1,len(head))
rr=2
for x in sorted(val,key=lambda z:str(z[idx['주문일']])):
    vals=[str(x[idx['주문일']]),str(x[idx['주문번호']]),x[idx['상품명']],int(num(x[idx['수량']])),int(num(x[idx['정산기준금액']])),x[idx['주문상태']]]
    for c,vv in enumerate(vals,1):
        cell=ws.cell(rr,c,vv); cell.alignment=LEFT
        if c in (4,5): cell.number_format='#,##0'
    rr+=1
setw(ws,[20,14,44,7,14,18])
ws.freeze_panes='A2'

OUT='/Users/macmini_ky/Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브/Claude AI work space/mac window file transfer/카카오 선물하기_고체탈취제 런칭 정산_260720.xlsx'
wb.save(OUT)
print('저장 완료:', OUT)
print(f'세트순이익 {int(sprofit):,} / 순이익률 {sprofit/mset_rev*100:.1f}% / 개당 {int(sprofit/mset_qty):,}')
