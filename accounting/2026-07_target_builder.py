# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── 공통 스타일 ──
HEAD = PatternFill("solid", fgColor="1F4E78")
SUB  = PatternFill("solid", fgColor="D9E1F2")
TOT  = PatternFill("solid", fgColor="FFE699")
RISK = PatternFill("solid", fgColor="FFC7CE")   # 품절위험
OKF  = PatternFill("solid", fgColor="E2EFDA")    # 입고확정
WHITE = Font(color="FFFFFF", bold=True, size=11)
BOLD = Font(bold=True)
RISKF = Font(color="9C0006", bold=True)
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
L = Alignment(horizontal="left", vertical="center", wrap_text=True)
C = Alignment(horizontal="center", vertical="center")
R = Alignment(horizontal="right", vertical="center")

def style_header(ws, row, ncol):
    for c in range(1, ncol+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD; cell.font = WHITE; cell.alignment = C; cell.border = BORDER

def box(ws, r1, r2, c1, c2):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(row=r, column=c).border = BORDER

# ════════════════════════════════════════════
# 시트1: 채널 요약
# ════════════════════════════════════════════
ws = wb.active
ws.title = "채널 요약"
ws["A1"] = "2026년 7월 매출 목표안 — 채널별 › 품목별"
ws["A1"].font = Font(bold=True, size=15, color="1F4E78")
ws.merge_cells("A1:E1")
ws["A2"] = "기준: 목표 2억원 / 단위: 만원 (VAT포함) / 6월=ERP실적+라이플로우(수동) / 작성: 세일즈 하치"
ws["A2"].font = Font(italic=True, size=9, color="808080")
ws.merge_cells("A2:E2")

hdr = ["채널", "6월 실적", "7월 목표", "비중", "비고"]
hr = 4
for i, h in enumerate(hdr, 1):
    ws.cell(row=hr, column=i, value=h)
style_header(ws, hr, 5)

# (채널, 6월실적, 7월목표(없으면 None=수식참조), 비고)
ch = [
    ("쿠팡 로켓(1P)",      4681, "=,로켓 품목,",   "1P 직매입 · 품목 상세 시트 참조"),
    ("쿠팡 그로스(3P)",    2800, "=,그로스 품목,", "3P · 바이올렛 7/16 입고 · 품목 상세 시트 참조"),
    ("자사몰(카페24)",     2471, 3500,             "공구·이벤트 의존(변동성)"),
    ("카카오선물하기",      0,    3000,             "🆕 7월 런칭 이벤트"),
    ("스마트스토어",        4419, 1500,             "6월은 공구 일회성 → 7월 베이스라인"),
    ("29CM",               0,    1500,             "🆕 입점 행사"),
    ("라이플로우",          2000, 1000,             "수동발주(ERP 외) · 보수반영"),
    ("기타채널(11번가 등)", 256,  1000,             "자연 유입"),
]
r = hr + 1
data_start = r
for name, m6, m7, memo in ch:
    ws.cell(row=r, column=1, value=name).alignment = L
    ws.cell(row=r, column=2, value=m6).alignment = R
    if isinstance(m7, str) and m7.startswith("="):
        sheet = m7.split(",")[1]
        ws.cell(row=r, column=3, value=f"='{sheet}'!D100").alignment = R
    else:
        ws.cell(row=r, column=3, value=m7).alignment = R
    ws.cell(row=r, column=5, value=memo).alignment = L
    r += 1
data_end = r - 1
# 비중 수식
for rr in range(data_start, data_end+1):
    ws.cell(row=rr, column=4, value=f"=C{rr}/$C${data_end+1}").alignment = C
    ws.cell(row=rr, column=4).number_format = "0.0%"
# 합계행
ws.cell(row=r, column=1, value="합계").font = BOLD
ws.cell(row=r, column=2, value=f"=SUM(B{data_start}:B{data_end})")
ws.cell(row=r, column=3, value=f"=SUM(C{data_start}:C{data_end})")
ws.cell(row=r, column=4, value=f"=SUM(D{data_start}:D{data_end})")
ws.cell(row=r, column=4).number_format = "0.0%"
for c in range(1,6):
    ws.cell(row=r, column=c).fill = TOT; ws.cell(row=r, column=c).font = BOLD
totrow = r
# 목표 대비
r += 1
ws.cell(row=r, column=1, value="목표 2억 대비")
ws.cell(row=r, column=3, value=f"=C{totrow}-20000")
ws.cell(row=r, column=3).number_format = "+#,##0;-#,##0"
ws.cell(row=r, column=5, value="양수=초과달성 / 음수=미달")
ws.cell(row=r, column=5).alignment = L
for c in [1,2,3,4]:
    ws.cell(row=hr, column=c)
box(ws, hr, totrow, 1, 5)
for rr in range(data_start, totrow):
    ws.cell(row=rr, column=2).number_format = "#,##0"
    ws.cell(row=rr, column=3).number_format = "#,##0"
ws.cell(row=totrow, column=2).number_format="#,##0"; ws.cell(row=totrow,column=3).number_format="#,##0"
ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 9
ws.column_dimensions["E"].width = 40

# ════════════════════════════════════════════
# 품목 상세 시트 생성 함수
# ════════════════════════════════════════════
def make_item_sheet(title, rows, target_total):
    ws = wb.create_sheet(title)
    ws["A1"] = f"{title} — 7월 품목별 목표 (단위: 만원)"
    ws["A1"].font = Font(bold=True, size=13, color="1F4E78")
    ws.merge_cells("A1:H1")
    hdr = ["품목", "5월 실적", "7월 목표", "현재고", "입고예정", "입고예정일", "발주필요", "비고"]
    hr = 3
    for i, h in enumerate(hdr, 1):
        ws.cell(row=hr, column=i, value=h)
    style_header(ws, hr, 8)
    r = hr + 1
    start = r
    for it in rows:
        name, m5, m7, stock, inq, inday, need, memo, risk = it
        ws.cell(row=r, column=1, value=name).alignment = L
        ws.cell(row=r, column=2, value=m5).alignment = R
        ws.cell(row=r, column=3, value=m7).alignment = R
        ws.cell(row=r, column=4, value=stock).alignment = R
        ws.cell(row=r, column=5, value=inq).alignment = R
        ws.cell(row=r, column=6, value=inday).alignment = C
        ws.cell(row=r, column=7, value=need).alignment = C
        ws.cell(row=r, column=8, value=memo).alignment = L
        ws.cell(row=r, column=2).number_format = "#,##0"
        ws.cell(row=r, column=3).number_format = "#,##0"
        ws.cell(row=r, column=4).number_format = "#,##0"
        ws.cell(row=r, column=5).number_format = "#,##0"
        if risk == "risk":
            ws.cell(row=r, column=4).fill = RISK; ws.cell(row=r, column=4).font = RISKF
        if inday and "07-16" in str(inday):
            ws.cell(row=r, column=6).fill = OKF; ws.cell(row=r, column=6).font = BOLD
        r += 1
    end = r - 1
    # 소계 (D100에 고정 배치 → 요약시트가 참조)
    ws.cell(row=r, column=1, value="합 계").font = BOLD
    ws.cell(row=r, column=2, value=f"=SUM(B{start}:B{end})").number_format="#,##0"
    ws.cell(row=r, column=3, value=f"=SUM(C{start}:C{end})").number_format="#,##0"
    for c in range(1, 9):
        ws.cell(row=r, column=c).fill = SUB; ws.cell(row=r, column=c).font = BOLD
    subrow = r
    box(ws, hr, subrow, 1, 8)
    # 요약시트 참조용 D100에 합계 복제
    ws.cell(row=100, column=4, value=f"=C{subrow}")
    # 목표 대비 표시
    r += 1
    ws.cell(row=r, column=1, value=f"채널 목표(가안)")
    ws.cell(row=r, column=3, value=target_total).number_format="#,##0"
    widths = [30, 10, 10, 9, 9, 16, 9, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws

# 로켓 품목 (name,5월,7월,재고,입고예정,입고일,발주필요,비고,risk)
rocket = [
 ("건조기시트 코튼블루(단품+2입)", 3518, 3400, 625, 5400, "ERP 미갱신·확인요", "", "주력 · 입고로 커버(월수요 빠듯)", "risk"),
 ("캡슐표백제",                    675, 1300, 4414, 0, "", "", "그로스서 이관 · 재고 넉넉 → 확대", ""),
 ("하트식세기(+선물세트)",          530, 700, 1047, 5000, "ERP 미갱신·확인요", "", "입고 충분", ""),
 ("섬유탈취제 100ml",               362, 500, 242, 0, "", "O", "여름시즌 · 발주 필요", "risk"),
 ("얼룩제거제 350ml",               116, 400, 2504, 0, "", "", "재고 충분 → 확대", ""),
 ("캡슐세제",                       130, 200, 745, 0, "", "", "", ""),
 ("식기세척기 2개",                  97, 150, 57, 0, "", "O", "일반 품절 → 하트 대체", "risk"),
 ("건조기시트 바이올렛머스크",        244, 150, 22, "", "2026-07-16", "", "7/16 입고 후 판매", "risk"),
 ("섬유탈취제 400ml",               147, 100, 28, 0, "", "O", "품절 임박 · 발주 필요", "risk"),
 ("기타/완충",                        0, 100, "", "", "", "", "", ""),
]
# 그로스 품목
gross = [
 ("건조기시트 베이비크림",          347, 800, 917, 0, "", "O", "로켓서 이관 · 주력화 · 발주 필요", "risk"),
 ("건조기시트 바이올렛머스크",      1026, 650, 22, "", "2026-07-16", "", "그로스 1등 · 7/16 입고 후 반월", "risk"),
 ("얼룩제거제 350ml",               325, 500, 2504, 0, "", "", "재고 충분", ""),
 ("식기세척기세제(하트)",            332, 500, 1047, 5000, "ERP 미갱신·확인요", "", "세이랩", ""),
 ("입테이프(오모모)",               359, 400, "", "", "확인요", "", "재고 확인 필요", ""),
 ("수세미",                         153, 200, 933, 0, "", "", "", ""),
]

make_item_sheet("로켓 품목", rocket, 7000)
make_item_sheet("그로스 품목", gross, 3050)

out = "/Users/macmini_ky/Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브/Claude AI work space/02. 매출 정산/2026-07 매출 목표안.xlsx"
wb.save(out)
print("저장 완료:", out)
