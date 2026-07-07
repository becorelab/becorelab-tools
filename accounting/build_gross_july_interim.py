# -*- coding: utf-8 -*-
"""로켓그로스 2026년 7월 '중간' 정산 — 매출인식 7/1~7/5 (첫 주기, 주기종료 7/5)
채움컴퍼니(A00940134) + 세이랩컴퍼니(A01707416) 통합.
build_gross_june_full.py / build_gross_june_seirab.py(6월 완성판)와 동일 방법론.
⚠️ 중간 집계(월말 아님). 광고비: 채움 = JSON 일별수집(소급 차이 가능), 세이랩 = 미확보(0 처리).
작성 2026-07-07."""
import openpyxl, glob, os, json, re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = "/Users/macmini_ky/Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브/Claude AI work space/02. 매출 정산/26.07"
OUT  = os.path.join(BASE, "26.07 로켓그로스 중간정산_7.1-7.5.xlsx")
AD   = "/Users/macmini_ky/ClaudeAITeam/marketing/coupang_data"
P1, P2 = "2026-07-01", "2026-07-05"

# ---- 원가 (옵션ID -> 세트원가, VAT별도) ----
# 6월 완성판 COST 그대로 + 7월 신규 옵션ID (제품 기본단가 × 개수로 산출, 근거 메모 시트 명시)
COST_CHAEUM = {
    '95304424363': 3932, '95363500368': 7864, '95363500367': 11796,   # 바이올렛 1/2/3
    '95602285926': 3932,                                              # 바이올렛 1 (신규 옵션ID, 6월 등록)
    '95377476928': 3932, '95377476939': 7864, '95377476938': 11796,   # 베이비크림 1/2/3
    '95304888338': 3199, '95304888360': 6398, '95304888356': 9597,    # 얼룩 1/2/3
    '95304888355': 19194,                                             # 얼룩 6
    '95060453340': 2200, '95060453345': 4400, '95060453346': 6600,    # 입테 1/2/3
    '95060453344': 8800, '95060453347': 11000, '95060453343': 13200,  # 입테 4/5/6
    # -- 7월 신규: 섬유탈취제 코튼블루 100ml (기본단가 1,557 × 개수) --
    '95585733312': 1557,   # 100ml 1개
    '95585733317': 3114,   # 100ml 2개 (1,557×2)
}
COST_SEIRAB = {
    '95510001821': 3304,  '95556504824': 6608,  '95556504822': 9912,  '95556504823': 13216,  # 식세기 1/2/3/4
    '95463771120': 1860,  '95607833748': 1860,                                               # 수세미 1 (ID 2종)
    '95463791496': 3720,  '95463791493': 5580,  '95463791494': 7440,                          # 수세미 2/3/4
    '95463791492': 9300,  '95463791495': 11160,                                              # 수세미 5/6
    # -- 7월 신규: 식세기 5개입 (기본단가 3,304 × 5) --
    '95556504821': 16520,
}

def bucket_chaeum(name):
    n = name or ''
    if '바이올렛' in n: return '바이올렛'
    if '베이비' in n: return '베이비크림'
    if '얼룩' in n or '목때' in n: return '얼룩제거제'
    if '입벌림' in n or '오모모' in n: return '입테이프'
    if '섬유탈취제' in n or '탈취제' in n: return '섬유탈취제'
    return '기타'

def bucket_seirab(name):
    n = name or ''
    if '식기세척기' in n or '식세기' in n: return '식기세척기세제'
    if '수세미' in n: return '한장수세미'
    return '기타'

DATE_RE = re.compile(r'^\d{4}-\d{2}-\d{2}$')

# ---- CATEGORY_TR: 매출/정산대상액 (매출인식일 7/1~7/5) ----
def load_sales(src):
    prod = defaultdict(lambda: {'옵션명':'', '수량':0, '판매액':0, '정산대상액':0})
    day  = defaultdict(lambda: {'판매액':0, '정산대상액':0, '수량':0})
    for f in glob.glob(os.path.join(src, '*CATEGORY_TR*')):
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True); ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None: continue
            rec = str(row[4])[:10]
            if not (P1 <= rec <= P2): continue
            oid = str(row[11]).split('.')[0]
            try: qty=float(row[16] or 0); pa=float(row[17] or 0); st=float(row[23] or 0)
            except: continue
            p = prod[oid]; p['옵션명'] = (str(row[13]) + ' / ' + str(row[14]))[:60]
            p['수량']+=qty; p['판매액']+=pa; p['정산대상액']+=st
            day[rec]['판매액']+=pa; day[rec]['정산대상액']+=st; day[rec]['수량']+=qty
        wb.close()
    return prod, day

# ---- 물류비: 요약행(min_row=3) 최종비용(VAT포함), 주기 = 7/5 ----
def logi_total(src, pat, sheet):
    seen=set(); total=0.0
    for f in glob.glob(os.path.join(src, pat)):
        wb=openpyxl.load_workbook(f,read_only=True,data_only=True)
        if sheet not in wb.sheetnames: wb.close(); continue
        for row in wb[sheet].iter_rows(min_row=3, values_only=True):
            if not row or row[0] is None: continue
            end=str(row[0])[:10]
            if not DATE_RE.match(end) or end in seen: continue
            if end != '2026-07-05': continue
            seen.add(end)
            try: total+=float(row[3] or 0)
            except: pass
        wb.close()
    return total

# ---- 광고비: 채움 JSON 일별 (7/1~7/5) ----
def load_ad_chaeum():
    ad_daily=defaultdict(float); ad_prod=defaultdict(float); missing=[]
    for d in range(1,6):
        fn=os.path.join(AD, f"A00940134_pa_daily_keyword_202607{d:02d}_202607{d:02d}.json")
        if not os.path.exists(fn):
            missing.append(f"2026-07-{d:02d}"); continue
        for rec in json.load(open(fn)):
            try: v=float(str(rec.get('광고비',0)).replace(',','') or 0)
            except: continue
            ad_daily[f"2026-07-{d:02d}"]+=v
            ad_prod[bucket_chaeum(str(rec.get('광고집행 상품명','') or ''))]+=v
    return ad_daily, ad_prod, missing

# ================= 데이터 로드 =================
accounts = {}
for key, folder, cost, bucket in [
        ('채움', '채움컴퍼니', COST_CHAEUM, bucket_chaeum),
        ('세이랩', '세이랩컴퍼니', COST_SEIRAB, bucket_seirab)]:
    src = os.path.join(BASE, "그로스", folder)  # 월별/그로스/{계정} 구조 (26.06부터, 대표님 재편)
    if not os.path.isdir(src): src = os.path.join(BASE, folder)  # 구 구조 폴백
    prod, day = load_sales(src)
    unknown = [o for o in prod if o not in cost]
    io   = logi_total(src, '*WAREHOUSING_SHIPPING*', '입출고비')
    ship = logi_total(src, '*WAREHOUSING_SHIPPING*', '배송비')
    stor = logi_total(src, '*STORAGE_FEE*', '보관비')
    accounts[key] = dict(prod=prod, day=day, unknown=unknown, io=io, ship=ship, stor=stor,
                         cost=cost, bucket=bucket)

ad_daily, ad_prod, ad_missing = load_ad_chaeum()
accounts['채움']['ad'] = sum(ad_daily.values())
accounts['채움']['ad_prod'] = ad_prod
accounts['세이랩']['ad'] = 0.0   # 광고비 미확보 (리포트 필요) — 0 처리
accounts['세이랩']['ad_prod'] = {}

# 계정별 합계 계산
for key, A in accounts.items():
    prod, cost = A['prod'], A['cost']
    A['판매액'] = sum(p['판매액'] for p in prod.values())
    A['정산']   = sum(p['정산대상액'] for p in prod.values())
    A['수량']   = sum(p['수량'] for p in prod.values())
    A['원가']   = sum(cost[o]*p['수량'] for o,p in prod.items() if o in cost)
    A['원가미상수량'] = sum(p['수량'] for o,p in prod.items() if o not in cost)
    A['물류']   = A['io']+A['ship']+A['stor']
    A['순이익'] = A['정산']-A['원가']-A['물류']-A['ad']

# ================= 엑셀 작성 (6월 스타일 함수 재사용) =================
thin=Side(style='thin',color='CCCCCC'); border=Border(thin,thin,thin,thin)
hf=PatternFill('solid',fgColor='4472C4'); hfont=Font(color='FFFFFF',bold=True,size=10)
tot=PatternFill('solid',fgColor='FFF2CC')
def hdr(ws,r,n):
    for c in range(1,n+1):
        cell=ws.cell(row=r,column=c); cell.fill=hf; cell.font=hfont
        cell.alignment=Alignment(horizontal='center',vertical='center'); cell.border=border
def fmt(ws,r1,r2,c1,c2,f='#,##0'):
    for row in ws.iter_rows(min_row=r1,max_row=r2,min_col=c1,max_col=c2):
        for cell in row:
            if isinstance(cell.value,(int,float)): cell.number_format=f
def pct(ws,r1,r2,c):
    for row in ws.iter_rows(min_row=r1,max_row=r2,min_col=c,max_col=c):
        for cell in row:
            if isinstance(cell.value,(int,float)): cell.number_format='0.0%'
def totrow(ws,r,n):
    for c in range(1,n+1):
        cell=ws.cell(row=r,column=c); cell.fill=tot; cell.font=Font(bold=True)

wb=Workbook()

def sheet_products(ws, key, title):
    A=accounts[key]; prod, cost, bucket = A['prod'], A['cost'], A['bucket']
    ws['A1']=title; ws['A1'].font=Font(bold=True,size=13)
    H=['옵션ID','상품/옵션','수량','판매액','정산대상액','원가(단가)','원가합계','매출이익(정산-원가)','이익률']
    ws.append([]); ws.append(H); hdr(ws,3,len(H))
    r=4; T=defaultdict(float)
    for oid in sorted(prod,key=lambda x:-prod[x]['판매액']):
        p=prod[oid]
        if oid in cost:
            uc=cost[oid]; cc=uc*p['수량']; prof=p['정산대상액']-cc
            ws.append([oid,p['옵션명'],p['수량'],p['판매액'],p['정산대상액'],uc,cc,prof,prof/p['판매액'] if p['판매액'] else 0])
            T['원가합계']+=cc; T['이익']+=prof
        else:
            ws.append([oid,p['옵션명'],p['수량'],p['판매액'],p['정산대상액'],'원가미상','원가미상','',''])
        T['수량']+=p['수량']; T['판매액']+=p['판매액']; T['정산대상액']+=p['정산대상액']
        r+=1
    ws.append(['','합계',T['수량'],T['판매액'],T['정산대상액'],'',T['원가합계'],T['이익'],T['이익']/T['판매액'] if T['판매액'] else 0])
    totrow(ws,r,9); fmt(ws,4,r,3,8); pct(ws,4,r,9)
    # 일별 미니 요약
    r2=r+2
    ws.cell(row=r2,column=1,value='[일별] 매출인식일 기준').font=Font(bold=True)
    ws.cell(row=r2+1,column=1,value='날짜'); ws.cell(row=r2+1,column=2,value='수량')
    ws.cell(row=r2+1,column=3,value='판매액'); ws.cell(row=r2+1,column=4,value='정산대상액')
    hdr(ws,r2+1,4)
    rr=r2+2
    for d in sorted(A['day']):
        c=A['day'][d]; ws.append([]) # placeholder to keep row cursor simple
        ws.cell(row=rr,column=1,value=d); ws.cell(row=rr,column=2,value=c['수량'])
        ws.cell(row=rr,column=3,value=c['판매액']); ws.cell(row=rr,column=4,value=c['정산대상액'])
        rr+=1
    fmt(ws,r2+2,rr-1,2,4)
    # 제품군 순마진
    rg=rr+1
    ws.cell(row=rg,column=1,value='[제품군 순마진] 정산대상액 − 원가 − 물류(제품 미배분→합계행만) − 광고비').font=Font(bold=True)
    HG=['제품군','수량','판매액','정산대상액','원가','광고비','매출이익(정산-원가-광고)']
    for i,h in enumerate(HG,1): ws.cell(row=rg+1,column=i,value=h)
    hdr(ws,rg+1,len(HG))
    pg=defaultdict(lambda: defaultdict(float))
    for oid,p in prod.items():
        b=bucket(p['옵션명']); pg[b]['수량']+=p['수량']; pg[b]['판매액']+=p['판매액']; pg[b]['정산']+=p['정산대상액']
        pg[b]['원가']+= cost.get(oid,0)*p['수량']
    for b,v in A['ad_prod'].items():
        if v: pg[b]['광고']+=v
    rr2=rg+2
    for b in sorted(pg,key=lambda x:-pg[x]['판매액']):
        d=pg[b]; net=d['정산']-d['원가']-d['광고']
        for i,v in enumerate([b,d['수량'],d['판매액'],d['정산'],d['원가'],d['광고'],net],1):
            ws.cell(row=rr2,column=i,value=v)
        rr2+=1
    fmt(ws,rg+2,rr2-1,2,7)
    ws.cell(row=rr2+1,column=1,value='※ 물류비는 주기 요약값(계정 합계)만 확보 — 제품군 배분 없음. 순이익은 요약 시트 참조.')
    for col,w in zip('ABCDEFGHI',[13,46,7,12,12,11,11,15,8]): ws.column_dimensions[col].width=w
    A['pg']=pg
    return T

ws=wb.active; ws.title='채움 제품별'
T_ch = sheet_products(ws,'채움','로켓그로스(채움컴퍼니 A00940134) 7월 중간 — 매출인식 7/1~7/5 ⚠️중간집계(월말 아님)')
ws2=wb.create_sheet('세이랩 제품별')
T_se = sheet_products(ws2,'세이랩','로켓그로스(세이랩컴퍼니 A01707416) 7월 중간 — 매출인식 7/1~7/5 ⚠️중간집계(월말 아님)')

# ---- 요약 (두 계정 비교) ----
ws3=wb.create_sheet('요약')
ws3['A1']='로켓그로스 7월 중간정산 요약 — 매출인식 7/1~7/5 (첫 주기, 주기종료 7/5) ⚠️중간집계'; ws3['A1'].font=Font(bold=True,size=13)
H3=['항목','채움컴퍼니','세이랩컴퍼니','합계']
ws3.append([]); ws3.append(H3); hdr(ws3,3,len(H3))
C, S = accounts['채움'], accounts['세이랩']
rows=[
    ('수량', C['수량'], S['수량']),
    ('판매액', C['판매액'], S['판매액']),
    ('공제(수수료·쿠폰)', C['판매액']-C['정산'], S['판매액']-S['정산']),
    ('정산대상액', C['정산'], S['정산']),
    ('입출고비(VAT포함)', C['io'], S['io']),
    ('배송비(VAT포함)', C['ship'], S['ship']),
    ('보관비(VAT포함)', C['stor'], S['stor']),
    ('물류계', C['물류'], S['물류']),
    ('광고비', C['ad'], S['ad']),
    ('원가(VAT별도)', C['원가'], S['원가']),
    ('순이익(정산-원가-물류-광고)', C['순이익'], S['순이익']),
]
r3=4
for name,a,b in rows:
    ws3.append([name,a,b,a+b]); r3+=1
ws3.append(['순마진율(판매액대비)', C['순이익']/C['판매액'] if C['판매액'] else 0,
            S['순이익']/S['판매액'] if S['판매액'] else 0,
            (C['순이익']+S['순이익'])/(C['판매액']+S['판매액']) if (C['판매액']+S['판매액']) else 0])
pct(ws3,r3,r3,2); pct(ws3,r3,r3,3); pct(ws3,r3,r3,4)
totrow(ws3,r3-1,4)  # 순이익 행 강조
fmt(ws3,4,r3-1,2,4)
r3+=1
# 6월 일평균 비교
ws3.cell(row=r3+1,column=1,value='[6월 일평균 대비]').font=Font(bold=True)
ws3.cell(row=r3+2,column=1,value='7월 1~5일 일평균 판매액')
ws3.cell(row=r3+2,column=2,value=C['판매액']/5); ws3.cell(row=r3+2,column=3,value=S['판매액']/5)
ws3.cell(row=r3+3,column=1,value='6월 일평균 판매액')
ws3.cell(row=r3+3,column=2,value=937609); ws3.cell(row=r3+3,column=3,value=301014)
ws3.cell(row=r3+4,column=1,value='증감률')
ws3.cell(row=r3+4,column=2,value=C['판매액']/5/937609-1); ws3.cell(row=r3+4,column=3,value=S['판매액']/5/301014-1)
fmt(ws3,r3+2,r3+3,2,3); pct(ws3,r3+4,r3+4,2); pct(ws3,r3+4,r3+4,3)
# 채움 광고 일별
ws3.cell(row=r3+6,column=1,value='[채움 광고비 일별 (JSON 수집)]').font=Font(bold=True)
rr3=r3+7
for d in sorted(ad_daily):
    ws3.cell(row=rr3,column=1,value=d); ws3.cell(row=rr3,column=2,value=ad_daily[d]); rr3+=1
fmt(ws3,r3+7,rr3-1,2,2)
for col,w in zip('ABCD',[28,15,15,15]): ws3.column_dimensions[col].width=w

# ---- 메모 ----
ws4=wb.create_sheet('메모')
notes=[
 '[로켓그로스 7월 중간정산 메모 / 작성 2026-07-07]','',
 '⚠️ 본 표는 7월 "중간 집계"입니다 (매출인식 7/1~7/5, 첫 정산주기만). 월말 정산이 아닙니다.',
 '',
 '[범위] 매출인식일 2026-07-01 ~ 07-05 / 정산주기(종료일) 2026-07-05 1개.',
 '[기준] 6월 완성판(build_gross_june_full.py / build_gross_june_seirab.py)과 동일 방법론.',
 ' - 정산대상액 = CATEGORY_TR col23 (판매수수료·할인쿠폰 차감 후). 물류비 = 주기 요약행 최종비용(VAT포함).',
 ' - 원가 = VAT별도 세트원가 (6/26 대표님 원가표).',
 '',
 '[광고비 한계 — 반드시 유의]',
 ' - 채움: JSON 일별수집(7/1~7/5) 합산. 쿠팡 광고비는 소급 반영으로 JSON이 공식 XLSX보다 적게 잡힐 수 있음',
 '   (6월 사례: JSON이 XLSX 대비 -114,032). 월말에 광고센터 공식 XLSX로 확정 필요.',
 ' - 세이랩: 광고비 미확보 (로컬 수집 없음) → 0 처리. 리포트 필요. 세이랩 순이익은 광고비만큼 과대계상 상태.',
 '',
 '[7월 신규 옵션ID 원가 매핑 근거]',
 ' - 채움 95585733312 = 1,557 (섬유탈취제 100ml 기본단가 1,557 × 1개)',
 ' - 채움 95585733317 = 3,114 (섬유탈취제 100ml 1,557 × 2개)',
 ' - 세이랩 95556504821 = 16,520 (식세기 기본단가 3,304 × 5개)',
 ' - 그 외 옵션ID는 6월 완성판 COST 딕셔너리 그대로.',
 '',
 '[물류비] 세이랩 = 0원 (신규 계정 비용제로 면제 구간 지속, 6월과 동일). 채움 보관비 0원 = 이번 주기 미부과.',
 '[수금] 주정산·셀러월렛(빠른정산) 수금 일정은 중간집계 단계라 미포함 — 월말 정산에서 반영.',
]
for i,n in enumerate(notes,1): ws4.cell(row=i,column=1,value=n)
ws4.cell(row=1,column=1).font=Font(bold=True,size=12); ws4.column_dimensions['A'].width=110

wb.save(OUT)

# ================= 콘솔 요약 + 검증 =================
print('저장:', OUT); print()
for key in ['채움','세이랩']:
    A=accounts[key]
    pa=A['판매액']
    print(f"=== {key} 7월 중간 (7/1~7/5) ===")
    print(f"수량          {int(A['수량']):>10,}")
    print(f"판매액        {int(pa):>10,}")
    print(f"공제          {int(pa-A['정산']):>10,}")
    print(f"정산대상액    {int(A['정산']):>10,}")
    print(f"물류비        {int(A['물류']):>10,}  (입출고 {int(A['io']):,} + 배송 {int(A['ship']):,} + 보관 {int(A['stor']):,})")
    print(f"광고비        {int(A['ad']):>10,}" + ("  (JSON 일별, 소급차 가능)" if key=='채움' else "  (미확보 0 처리)"))
    print(f"원가          {int(A['원가']):>10,}")
    print(f"순이익        {int(A['순이익']):>10,}  ({A['순이익']/pa*100 if pa else 0:.1f}%)")
    if A['unknown']: print(f"⚠️ 원가미상 옵션ID: {A['unknown']} (수량 {A['원가미상수량']:.0f})")
    print()
print('=== 광고비 (채움, 일별) ===')
for d in sorted(ad_daily): print(f"  {d}: {ad_daily[d]:>10,.0f}")
if ad_missing: print('  ⚠️ 누락 날짜:', ad_missing)
print('  제품군 배분:', {k: round(v) for k,v in accounts['채움']['ad_prod'].items()})
print()
print('=== 6월 일평균 비교 ===')
print(f"채움: 7월 일평균 {accounts['채움']['판매액']/5:,.0f} vs 6월 937,609 ({accounts['채움']['판매액']/5/937609*100-100:+.1f}%)")
print(f"세이랩: 7월 일평균 {accounts['세이랩']['판매액']/5:,.0f} vs 6월 301,014 ({accounts['세이랩']['판매액']/5/301014*100-100:+.1f}%)")
