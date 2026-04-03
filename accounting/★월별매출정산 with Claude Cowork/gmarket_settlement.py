"""
지마켓 매출 정산 스크립트
매출기준_상품판매_상세내역 + 배송비 → 품명 매핑 → 피벗 집계

사용법: python gmarket_settlement.py <매출기준파일> [배송비파일] [구매확정파일]
  - .xls / .xlsx 모두 지원
"""
import sys, io, os, json
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(SCRIPT_DIR)
NAME_MAP_PATH = os.path.join(PARENT_DIR, 'gmarket_name_map.json')

with open(NAME_MAP_PATH, 'r', encoding='utf-8') as f:
    NAME_MAP = json.load(f)
print(f"[1/5] 품명리스트 로드: {len(NAME_MAP)}개 매핑")

if len(sys.argv) < 2:
    print("사용법: python gmarket_settlement.py <매출기준파일> [배송비파일] [구매확정파일]")
    sys.exit(1)

sales_path = sys.argv[1]
ship_path = sys.argv[2] if len(sys.argv) > 2 else None
confirm_path = sys.argv[3] if len(sys.argv) > 3 else None

def safe_float(val):
    if val is None or val == '' or val == 'None': return 0.0
    return float(str(val).replace(',', ''))

def safe_int(val):
    if val is None or val == '' or val == 'None': return 0
    return int(float(str(val).replace(',', '')))

def read_file(path):
    """xls/xlsx 모두 읽어서 (headers, rows) 반환"""
    ext = os.path.splitext(path)[1].lower()
    if ext == '.xls':
        import xlrd
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_index(0)
        headers = [ws.cell_value(0, c) for c in range(ws.ncols)]
        rows = []
        for r in range(1, ws.nrows):
            rows.append([ws.cell_value(r, c) for c in range(ws.ncols)])
        return headers, rows
    else:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = [cell.value for cell in ws[1]]
        rows = [list(r) for r in ws.iter_rows(min_row=2, values_only=True)]
        return headers, rows

def col_idx(headers, keyword):
    for i, h in enumerate(headers):
        if h and keyword in str(h):
            return i
    return None

# ─── 구매확정에서 옵션 로드 ───
print(f"[2/5] 옵션 정보 로드...")
option_map = {}
if confirm_path and os.path.exists(confirm_path):
    h_c, rows_c = read_file(confirm_path)
    idx_order = col_idx(h_c, '주문번호')
    idx_name = col_idx(h_c, '상품명')
    idx_opt = col_idx(h_c, '옵션')
    for row in rows_c:
        ono = str(row[idx_order] or '').strip()
        name = str(row[idx_name] or '').strip()
        opt = str(row[idx_opt] or '').strip() if idx_opt else ''
        if ono and name:
            option_map[ono] = f"{name}{opt}" if opt else name
    print(f"  구매확정: {len(option_map)}건")
else:
    print(f"  구매확정 없음, 스킵")

# ─── 매출기준 읽기 ───
print(f"[3/5] 매출기준 읽는 중: {os.path.basename(sales_path)}")
h_sales, rows_sales = read_file(sales_path)
IDX_ORDER = col_idx(h_sales, '주문번호')
IDX_NAME = col_idx(h_sales, '상품명')
IDX_QTY = col_idx(h_sales, '주문수량')
IDX_SETTLE = col_idx(h_sales, '판매자 최종정산금')
IDX_OPTION = col_idx(h_sales, '옵션상품')
print(f"  {len(rows_sales)}행, 컬럼: 상품명={IDX_NAME}, 수량={IDX_QTY}, 정산금={IDX_SETTLE}")

# ─── 배송비 읽기 ───
ship_by_order = {}
total_ship_settle = 0
if ship_path and os.path.exists(ship_path):
    print(f"[4/5] 배송비 읽는 중: {os.path.basename(ship_path)}")
    h_ship, rows_ship = read_file(ship_path)
    IDX_SHIP_ORDER = col_idx(h_ship, '매출주문번호') or col_idx(h_ship, '대표주문번호')
    IDX_SHIP_SETTLE = col_idx(h_ship, '배송비정산금액')
    for row in rows_ship:
        ono = str(row[IDX_SHIP_ORDER] or '').strip()
        settle = safe_float(row[IDX_SHIP_SETTLE])
        if ono:
            ship_by_order[ono] = ship_by_order.get(ono, 0) + settle
            total_ship_settle += settle
    print(f"  {len(rows_ship)}행, 총 배송비정산: {total_ship_settle:,.0f}")
else:
    print(f"[4/5] 배송비 파일 없음, 스킵")

# ─── 피벗 집계 ───
print(f"[5/5] 데이터 처리...")
pivot = defaultdict(lambda: {'qty': 0, 'settle': 0, 'ship': 0})
unmapped = []

for row in rows_sales:
    product_name = str(row[IDX_NAME] or '').strip()
    if not product_name:
        continue

    order_no = str(row[IDX_ORDER] or '').strip()
    qty = safe_int(row[IDX_QTY])
    settle = safe_float(row[IDX_SETTLE])

    # 매핑
    final = NAME_MAP.get(product_name, None)
    if not final and order_no in option_map:
        final = NAME_MAP.get(option_map[order_no], None)
    if not final:
        for mk, mv in NAME_MAP.items():
            if product_name in mk:
                final = mv
                break

    if not final:
        unmapped.append(product_name)
        continue

    # 배송비 (주문번호로 매칭)
    ship = ship_by_order.pop(order_no, 0)

    pivot[final]['qty'] += qty
    pivot[final]['settle'] += settle
    pivot[final]['ship'] += ship

# ─── 결과 출력 ───
print(f"\n{'='*75}")
print(f"  지마켓 매출 피벗 결과")
print(f"{'='*75}")
print(f"  {len(rows_sales)}행 → {len(pivot)}개 상품 카테고리\n")

sorted_items = sorted(pivot.items(), key=lambda x: x[0])

print(f"{'#':>3}  {'상품명':<35} {'수량':>8} {'판매자 최종정산금':>16} {'배송비':>12}")
print(f"{'─'*3}  {'─'*35} {'─'*8} {'─'*16} {'─'*12}")

total_qty = total_settle = total_ship = 0
for i, (name, data) in enumerate(sorted_items, 1):
    q, s, sh = data['qty'], data['settle'], data['ship']
    total_qty += q
    total_settle += s
    total_ship += sh
    print(f"{i:>3}  {name:<35} {q:>8,} {s:>16,.0f} {sh:>12,.0f}")

print(f"{'─'*3}  {'─'*35} {'─'*8} {'─'*16} {'─'*12}")
print(f"{'':>3}  {'총합계':<35} {total_qty:>8,} {total_settle:>16,.0f} {total_ship:>12,.0f}")
print(f"\n  정산금 + 배송비 = {total_settle + total_ship:,.0f}")

if unmapped:
    unique_unmapped = sorted(set(unmapped))
    print(f"\n{'='*75}")
    print(f"  ⚠️ 미매핑: {len(unique_unmapped)}개 ({len(unmapped)}행)")
    print(f"{'='*75}")
    for name in unique_unmapped:
        cnt = unmapped.count(name)
        print(f"  - {name[:65]} ({cnt}건)")
else:
    print(f"\n  ✅ 미매핑 상품 없음 — 전체 매핑 성공!")

print()
