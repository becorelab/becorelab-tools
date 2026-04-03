"""
스마트스토어 매출 정산 스크립트
결제정산-일별상세 + 구매확정내역/주문조회 → 상품명 매핑 → 피벗 집계

사용법: python smartstore_settlement.py <결제정산파일.xlsx> [구매확정파일.xlsx] [주문조회파일.xlsx]
"""
import sys
import io
import os
import json
import openpyxl
from collections import defaultdict, Counter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(SCRIPT_DIR)
NAME_MAP_PATH = os.path.join(PARENT_DIR, 'smartstore_name_map.json')

with open(NAME_MAP_PATH, 'r', encoding='utf-8') as f:
    NAME_MAP = json.load(f)
print(f"[1/5] 품명리스트 로드: {len(NAME_MAP)}개 매핑")

if len(sys.argv) < 2:
    print("사용법: python smartstore_settlement.py <결제정산.xlsx> [구매확정.xlsx] [주문조회.xlsx]")
    sys.exit(1)

settle_path = sys.argv[1]
confirm_path = sys.argv[2] if len(sys.argv) > 2 else None
order_path = sys.argv[3] if len(sys.argv) > 3 else None

def safe_float(val):
    if val is None or val == '' or val == 'None': return 0.0
    return float(str(val).replace(',', ''))

def col_idx(headers, keyword):
    for i, h in enumerate(headers):
        if h and keyword in str(h):
            return i
    return None

# ─── 구매확정/주문조회에서 옵션 정보 로드 ───
print(f"[2/5] 옵션 정보 로드...")
option_map = {}  # 상품주문번호 → (상품명+옵션)

def load_options(path, label):
    if not path or not os.path.exists(path):
        print(f"  {label}: 파일 없음, 스킵")
        return
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in ws[1]]
    idx_item_no = col_idx(headers, '상품주문번호')
    idx_name = col_idx(headers, '상품명')
    idx_option = col_idx(headers, '옵션정보')
    cnt = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        item_no = str(row[idx_item_no] or '').strip()
        name = str(row[idx_name] or '').strip()
        option = str(row[idx_option] or '').strip()
        if item_no and name:
            full = f"{name}{option}" if option else name
            option_map[item_no] = full
            cnt += 1
    print(f"  {label}: {cnt}건 로드")

load_options(confirm_path, "구매확정")
load_options(order_path, "주문조회")

# ─── 결제정산 읽기 ───
print(f"[3/5] 결제정산 읽는 중: {os.path.basename(settle_path)}")
wb = openpyxl.load_workbook(settle_path, data_only=True)
ws = wb[wb.sheetnames[0]]
headers = [cell.value for cell in ws[1]]

IDX_ORDER_NO = col_idx(headers, '주문번호')
IDX_ITEM_NO = col_idx(headers, '상품주문번호')
IDX_TYPE = col_idx(headers, '구분')
IDX_NAME = col_idx(headers, '상품명')
IDX_SETTLE_AMT = col_idx(headers, '정산기준금액')
IDX_NPAY = col_idx(headers, 'Npay 수수료')
IDX_SALES_FEE = col_idx(headers, '매출연동 수수료')
IDX_INSTALLMENT = col_idx(headers, '무이자할부 수수료')

# ─── 데이터 처리 ───
print(f"[4/5] 데이터 처리...")
pivot = defaultdict(lambda: {'count': 0, 'net_amount': 0})
unmapped = []
total_rows = 0
shipping_total = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    row_type = str(row[IDX_TYPE] or '').strip()
    item_no = str(row[IDX_ITEM_NO] or '').strip()
    product_name = str(row[IDX_NAME] or '').strip()

    # 정산금액 계산
    settle_amt = safe_float(row[IDX_SETTLE_AMT])
    npay_fee = safe_float(row[IDX_NPAY])
    sales_fee = safe_float(row[IDX_SALES_FEE])
    installment_fee = safe_float(row[IDX_INSTALLMENT]) if IDX_INSTALLMENT else 0
    net = settle_amt + npay_fee + sales_fee + installment_fee  # 수수료는 음수

    if row_type == '배송비':
        shipping_total += net
        continue

    if row_type != '상품주문':
        continue

    total_rows += 1

    # 상품명 매핑: 1) 결제정산 상품명 직접 매칭
    final_name = NAME_MAP.get(product_name, None)

    # 2) 실패 시 → 구매확정/주문조회에서 옵션 포함 이름으로 매칭
    if not final_name and item_no in option_map:
        full_name = option_map[item_no]
        final_name = NAME_MAP.get(full_name, None)

    # 3) 부분 매칭 시도 (상품명이 매핑 키에 포함되는 경우)
    if not final_name:
        for map_key, map_val in NAME_MAP.items():
            if product_name and product_name in map_key:
                final_name = map_val
                break

    if not final_name:
        unmapped.append(product_name)
        continue

    pivot[final_name]['count'] += 1
    pivot[final_name]['net_amount'] += net

# ─── 결과 출력 ───
print(f"[5/5] 결과 출력\n")
print(f"{'='*70}")
print(f"  스마트스토어 매출 피벗 결과")
print(f"{'='*70}")
print(f"  상품주문 {total_rows}행 → {len(pivot)}개 상품 카테고리\n")

sorted_items = sorted(pivot.items(), key=lambda x: x[0])

print(f"{'#':>3}  {'상품명':<40} {'수량':>8} {'최종정산금':>15}")
print(f"{'─'*3}  {'─'*40} {'─'*8} {'─'*15}")

total_count = 0
total_net = 0

for i, (name, data) in enumerate(sorted_items, 1):
    c = data['count']
    n = data['net_amount']
    total_count += c
    total_net += n
    print(f"{i:>3}  {name:<40} {c:>8,} {n:>15,.0f}")

print(f"{'─'*3}  {'─'*40} {'─'*8} {'─'*15}")
print(f"{'':>3}  {'총합계':<40} {total_count:>8,} {total_net:>15,.0f}")
print(f"\n  배송비 정산금 합계: {shipping_total:,.0f}")
print(f"  상품 + 배송비 합계: {total_net + shipping_total:,.0f}")

if unmapped:
    unique_unmapped = sorted(set(unmapped))
    print(f"\n{'='*70}")
    print(f"  ⚠️ 미매핑 상품: {len(unique_unmapped)}개 ({len(unmapped)}행)")
    print(f"{'='*70}")
    for name in unique_unmapped:
        cnt = unmapped.count(name)
        print(f"  - {name[:70]} ({cnt}건)")
else:
    print(f"\n  ✅ 미매핑 상품 없음 — 전체 매핑 성공!")

print()
