"""
11번가 매출 정산 스크립트
정산_확정건 → 품명 매핑 → 피벗 집계

사용법: python 11st_settlement.py <정산파일.xls 또는 .xlsx>
  - 헤더가 Row 6에 위치 (11번가 고유 형식)
"""
import sys, io, os, json
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(SCRIPT_DIR)
NAME_MAP_PATH = os.path.join(PARENT_DIR, '11st_name_map.json')

with open(NAME_MAP_PATH, 'r', encoding='utf-8') as f:
    NAME_MAP = json.load(f)
print(f"[1/3] 품명리스트 로드: {len(NAME_MAP)}개 매핑")

if len(sys.argv) < 2:
    print("사용법: python 11st_settlement.py <정산파일>")
    sys.exit(1)

data_path = sys.argv[1]

def safe_float(val):
    if val is None or val == '' or val == 'None': return 0.0
    return float(str(val).replace(',', ''))

def safe_int(val):
    if val is None or val == '' or val == 'None': return 0
    return int(float(str(val).replace(',', '')))

def read_11st(path):
    """11번가 파일 읽기 (헤더 Row 6, 데이터 Row 7~)"""
    ext = os.path.splitext(path)[1].lower()
    if ext == '.xls':
        import xlrd
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_index(0)
        # 헤더 찾기 (NO가 있는 행)
        header_row = 5  # 0-based, Row 6
        for r in range(10):
            if ws.cell_value(r, 0) == 'NO':
                header_row = r
                break
        headers = [ws.cell_value(header_row, c) for c in range(ws.ncols)]
        rows = []
        for r in range(header_row + 1, ws.nrows):
            rows.append([ws.cell_value(r, c) for c in range(ws.ncols)])
        return headers, rows
    else:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        # 헤더 찾기
        header_row = 6
        for r in range(1, 11):
            if ws.cell(row=r, column=1).value == 'NO':
                header_row = r
                break
        headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
        rows = []
        for r in range(header_row + 1, ws.max_row + 1):
            rows.append([ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)])
        return headers, rows

print(f"[2/3] 데이터 읽는 중: {os.path.basename(data_path)}")
headers, rows = read_11st(data_path)

def col_idx(keyword):
    for i, h in enumerate(headers):
        if h and keyword in str(h):
            return i
    return None

IDX_STATUS = col_idx('주문상태')
IDX_NAME = col_idx('상품명')
IDX_OPTION = col_idx('옵션명')
IDX_QTY = col_idx('수량')
IDX_SETTLE = col_idx('정산금액')
IDX_SHIP = col_idx('선결제배송비')
IDX_AD = col_idx('후불광고비')

print(f"  {len(rows)}행, 상품명={IDX_NAME}, 수량={IDX_QTY}, 정산={IDX_SETTLE}, 배송비={IDX_SHIP}, 광고비={IDX_AD}")

# ─── 피벗 집계 ───
print(f"[3/3] 데이터 처리...")
pivot = defaultdict(lambda: {'qty': 0, 'settle': 0, 'ship': 0, 'ad': 0})
unmapped = []

for row in rows:
    name = str(row[IDX_NAME] or '').strip()
    if not name:
        continue

    option = str(row[IDX_OPTION] or '').strip() if IDX_OPTION else ''
    qty = safe_int(row[IDX_QTY])
    settle = safe_float(row[IDX_SETTLE])
    ship = safe_float(row[IDX_SHIP]) if IDX_SHIP else 0
    ad = safe_float(row[IDX_AD]) if IDX_AD else 0

    # 매핑: 상품명으로 직접 매칭
    final = NAME_MAP.get(name, None)
    # 상품명+옵션으로 매칭
    if not final:
        full = f"{name}{option}" if option else name
        final = NAME_MAP.get(full, None)
    # 부분 매칭
    if not final:
        for mk, mv in NAME_MAP.items():
            if name in mk or mk in name:
                final = mv
                break

    if not final:
        unmapped.append(f"{name} | {option}")
        continue

    # 옵션에서 수량 보정 (예: "2개" → 실제 구성 수량)
    # 11번가는 옵션명이 "1개", "2개" 등으로 구성 수량을 나타냄
    # 피벗에서는 이미 품명리스트에서 매핑 처리됨

    pivot[final]['qty'] += qty
    pivot[final]['settle'] += settle
    pivot[final]['ship'] += ship
    pivot[final]['ad'] += ad

# ─── 결과 출력 ───
sorted_items = sorted(pivot.items(), key=lambda x: x[0])
total_qty = total_settle = total_ship = total_ad = 0

print(f"\n{'='*80}")
print(f"  11번가 매출 피벗 결과")
print(f"{'='*80}")
print(f"  {len(rows)}행 → {len(pivot)}개 상품 카테고리\n")

print(f"{'#':>3}  {'상품명':<25} {'수량':>6} {'정산금액':>14} {'배송비':>10} {'후불광고비':>12}")
print(f"{'─'*3}  {'─'*25} {'─'*6} {'─'*14} {'─'*10} {'─'*12}")

for i, (name, data) in enumerate(sorted_items, 1):
    q, s, sh, ad = data['qty'], data['settle'], data['ship'], data['ad']
    total_qty += q; total_settle += s; total_ship += sh; total_ad += ad
    sh_str = f"{sh:>10,.0f}" if sh > 0 else f"{'':>10}"
    ad_str = f"{ad:>12,.0f}" if ad > 0 else f"{'':>12}"
    print(f"{i:>3}  {name:<25} {q:>6,} {s:>14,.0f} {sh_str} {ad_str}")

print(f"{'─'*3}  {'─'*25} {'─'*6} {'─'*14} {'─'*10} {'─'*12}")
print(f"{'':>3}  {'총합계':<25} {total_qty:>6,} {total_settle:>14,.0f} {total_ship:>10,.0f} {total_ad:>12,.0f}")
print(f"\n  정산금 + 배송비 = {total_settle + total_ship:,.0f}")

if unmapped:
    unique_unmapped = sorted(set(unmapped))
    print(f"\n{'='*80}")
    print(f"  ⚠️ 미매핑: {len(unique_unmapped)}개 ({len(unmapped)}행)")
    print(f"{'='*80}")
    for name in unique_unmapped:
        cnt = unmapped.count(name)
        print(f"  - {name[:70]} ({cnt}건)")
else:
    print(f"\n  ✅ 미매핑 상품 없음 — 전체 매핑 성공!")

print()
