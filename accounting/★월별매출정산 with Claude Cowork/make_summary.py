"""
3월 정산 종합 매핑 파일 생성
각 채널별 3월 피벗 결과를 매출정산 시트 행번호와 매칭하여
대표님이 복붙하기 편한 엑셀 생성
"""
import sys, io, os, json, openpyxl
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# 매출정산 시트에서 E열 상품명 + 행번호 추출
wb_main = openpyxl.load_workbook(os.path.join(os.path.dirname(SCRIPT_DIR), '2026. 02 온라인 매출정산.xlsx'), data_only=True)
ws_main = wb_main['월별 온라인 매출정산']

# 채널별 행 범위
CHANNEL_RANGES = {
    '카페24': (10, 170),
    '스마트스토어': (171, 249),
    '카카오선물하기': (278, 292),
    '로켓배송': (293, 312),
    '옥션': (313, 395),
    '지마켓': (396, 514),
    '11번가': (515, 536),
    '신세계': (596, 651),
    '오늘의집': (710, 732),
    '지엠홀딩스': (998, 1024),
    '에드가': (1028, 1040),
    '두버': (1049, 1066),
}

# E열 상품명 → 행번호 매핑
channel_products = {}
for ch, (start, end) in CHANNEL_RANGES.items():
    products = {}
    for r in range(start, end + 1):
        name = ws_main.cell(row=r, column=5).value  # E열
        if name:
            name = str(name).strip()
            products[name] = r
    channel_products[ch] = products

# 3월 정산 결과 로드 (각 채널 엑셀에서)
BASE = os.path.join(SCRIPT_DIR, '3월 매출 정산')
results = {}

def load_pivot(path, qty_col=1, settle_col=2, ship_col=None):
    """피벗 시트에서 상품명→(수량,정산금,배송비) 로드"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['피벗']
    data = {}
    skip_names = {'총합계', '(비어 있음)', '행 레이블', None, ''}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        if name and str(name).strip() not in skip_names:
            name = str(name).strip()
            # 숫자가 아닌 경우만 (헤더/타이틀 스킵)
            try:
                qty_val = row[qty_col]
                if qty_val is None or isinstance(qty_val, str):
                    continue
                qty = int(qty_val)
                settle = float(row[settle_col] or 0)
                ship = float(row[ship_col]) if ship_col and len(row) > ship_col and row[ship_col] else 0
                data[name] = {'qty': qty, 'settle': round(settle), 'ship': round(ship)}
            except (ValueError, TypeError):
                continue
    return data

# 각 채널별 결과 로드
channel_files = {
    '카페24': ('03월 카페24/카페24_정산결과_becorelab2_20260403_356_5195.xlsx', 1, 2, 3),
    '스마트스토어': ('03월 스마트스토어/03월 스마트스토어.xlsx', 1, 2, None),
    '카카오선물하기': ('03월 카카오 선물하기/03월 카카오선물하기.xlsx', 1, 2, None),
    '로켓배송': ('03월 로켓배송/03월 로켓배송.xlsx', 1, 2, None),
    '옥션': ('03월 옥션/03월 옥션.xlsx', 1, 2, 3),
    '지마켓': ('03월 지마켓/03월 지마켓.xlsx', 1, 2, 3),
    '11번가': ('03월 11번가/03월 11번가.xlsx', 1, 2, 3),
    '신세계': ('03월 신세계/03월 신세계.xlsx', 1, 2, 3),
    '오늘의집': ('03월 오늘의집/03월 오늘의집.xlsx', 1, 2, None),
    '지엠홀딩스': ('03월 지엠홀딩스/03월 지엠홀딩스.xlsx', 1, 2, 3),
    '에드가': ('03월 에드가/03월 에드가.xlsx', 1, 2, 3),
    '두버': ('03월 두버/03월 두버.xlsx', 1, 2, 3),
}

for ch, (fpath, qc, sc, shc) in channel_files.items():
    full = os.path.join(BASE, fpath)
    if os.path.exists(full):
        try:
            results[ch] = load_pivot(full, qc, sc, shc)
            print(f"  ✅ {ch}: {len(results[ch])}개 상품")
        except Exception as e:
            print(f"  ❌ {ch}: {e}")
    else:
        print(f"  ⚠️ {ch}: 파일 없음 ({fpath})")

# === 종합 엑셀 생성 ===
print(f"\n엑셀 생성 중...")

HF = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
HFont = Font(bold=True, size=11, color='FFFFFF')
CH_FILL = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
CH_FONT = Font(bold=True, size=12, color='FFFFFF')
MATCH_FILL = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
NEW_FILL = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
WARN_FILL = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
NUM_FONT = Font(size=10, name='Consolas')
BOLD_NUM = Font(bold=True, size=11, name='Consolas', color='1565C0')
thin_border = Border(bottom=Side(style='thin', color='E0E0E0'))

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '3월 정산 종합'

# 헤더
headers = ['채널', '정산시트\n행번호', '정산시트\n상품명(E열)', '3월 피벗\n상품명', '판매수량\n(X열)', '매출액\n(Y열)', '배송비\n(Z열)', '매칭상태', '비고']
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = HFont
    cell.fill = HF
    cell.alignment = Alignment(horizontal='center', wrap_text=True)

r = 2
total_matched = 0
total_new = 0

for ch in CHANNEL_RANGES:
    if ch not in results:
        continue

    pivot_data = results[ch]
    e_products = channel_products[ch]

    # 채널 헤더
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    cell = ws.cell(row=r, column=1, value=f"📦 {ch}")
    cell.font = CH_FONT
    cell.fill = CH_FILL
    r += 1

    matched_names = set()

    # 피벗 데이터 매칭
    for pivot_name, data in sorted(pivot_data.items()):
        if data['qty'] == 0 and data['settle'] == 0:
            continue

        # E열에서 매칭
        row_num = None
        e_name = None

        # 정확 매칭만 (부분 매칭 금지 — 오매칭 방지)
        if pivot_name in e_products:
            row_num = e_products[pivot_name]
            e_name = pivot_name
        else:
            # 공백/특수문자 제거 후 정확 비교만
            for ename, erow in e_products.items():
                p_clean = pivot_name.replace(' ', '').replace('_', '')
                e_clean = ename.replace(' ', '').replace('_', '')
                if p_clean == e_clean:
                    row_num = erow
                    e_name = ename
                    break

        if row_num:
            matched_names.add(e_name)
            status = '✅ 매칭'
            fill = MATCH_FILL
            total_matched += 1
        else:
            status = '⚠️ 행 없음'
            fill = NEW_FILL
            total_new += 1

        ws.cell(row=r, column=1, value=ch)
        ws.cell(row=r, column=2, value=row_num).font = BOLD_NUM if row_num else Font(color='E53935')
        ws.cell(row=r, column=3, value=e_name or '')
        ws.cell(row=r, column=4, value=pivot_name)
        ws.cell(row=r, column=5, value=data['qty']).font = NUM_FONT
        ws.cell(row=r, column=5).number_format = '#,##0'
        ws.cell(row=r, column=6, value=data['settle']).font = NUM_FONT
        ws.cell(row=r, column=6).number_format = '#,##0'
        ws.cell(row=r, column=7, value=data['ship'] if data['ship'] > 0 else None)
        if data['ship']:
            ws.cell(row=r, column=7).font = NUM_FONT
            ws.cell(row=r, column=7).number_format = '#,##0'
        ws.cell(row=r, column=8, value=status)

        for c2 in range(1, 10):
            ws.cell(row=r, column=c2).fill = fill
            ws.cell(row=r, column=c2).border = thin_border
        r += 1

    r += 1  # 채널 간 빈 행

# 컬럼 너비
ws.column_dimensions['A'].width = 16
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 40
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 14
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 12
ws.column_dimensions['I'].width = 25
ws.auto_filter.ref = f"A1:I{r}"
ws.freeze_panes = 'A2'

# 요약 시트
ws_sum = wb.create_sheet('채널별 요약')
ws_sum.merge_cells('A1:D1')
ws_sum['A1'] = '3월 매출 정산 채널별 요약'
ws_sum['A1'].font = Font(bold=True, size=14, color='1B2A4A')
for c, h in enumerate(['채널', '상품수', '매출/정산금', '배송비'], 1):
    cell = ws_sum.cell(row=3, column=c, value=h)
    cell.font = HFont; cell.fill = HF
sr = 4
grand_settle = grand_ship = 0
for ch in CHANNEL_RANGES:
    if ch not in results: continue
    data = results[ch]
    ch_settle = sum(d['settle'] for d in data.values())
    ch_ship = sum(d['ship'] for d in data.values())
    grand_settle += ch_settle
    grand_ship += ch_ship
    ws_sum.cell(row=sr, column=1, value=ch)
    ws_sum.cell(row=sr, column=2, value=len(data))
    ws_sum.cell(row=sr, column=3, value=ch_settle).number_format = '#,##0'
    ws_sum.cell(row=sr, column=4, value=ch_ship if ch_ship > 0 else None)
    if ch_ship: ws_sum.cell(row=sr, column=4).number_format = '#,##0'
    sr += 1
# 합계
for c2 in range(1, 5):
    ws_sum.cell(row=sr, column=c2).fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
ws_sum.cell(row=sr, column=1, value='합계').font = Font(bold=True)
ws_sum.cell(row=sr, column=3, value=grand_settle).font = BOLD_NUM
ws_sum.cell(row=sr, column=3).number_format = '#,##0'
ws_sum.column_dimensions['A'].width = 18; ws_sum.column_dimensions['B'].width = 10
ws_sum.column_dimensions['C'].width = 18; ws_sum.column_dimensions['D'].width = 14

outpath = os.path.join(BASE, '★3월 정산 종합 매핑_v2.xlsx')
wb.save(outpath)
print(f"\n저장: {outpath}")
print(f"매칭: {total_matched}개, 행 없음: {total_new}개")
