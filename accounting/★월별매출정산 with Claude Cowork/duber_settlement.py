"""
두버 매출 정산 스크립트
바이피엘 출고요청 파일 → 두버 필터 → 공급가 적용 → 피벗
"""
import sys, io, os, json, re
from collections import defaultdict
import openpyxl, xlrd
from openpyxl.styles import Font, PatternFill, Border, Side

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(SCRIPT_DIR)

def sf(v):
    if v is None or v == '': return 0.0
    s = str(v).replace(',', '').replace('원', '').strip()
    try: return float(s)
    except: return 0.0

def si(v):
    if v is None or v == '': return 0
    return int(float(str(v).replace(',', '')))

# 1. 품명리스트 + 공급가
wb_ref = openpyxl.load_workbook('N:/개인/Becorelab/03. 영업/20. 월별 매출정산/2026.02/02월 두버.xlsx', data_only=True)
duber_map = {}
for row in wb_ref['품명리스트'].iter_rows(min_row=2, values_only=True):
    orig = str(row[1] or '').strip()
    mapped = str(row[2] or '').strip()
    if orig and mapped:
        duber_map[orig] = mapped
print(f"[1/4] 품명리스트: {len(duber_map)}개")

with open(os.path.join(PARENT_DIR, 'duber_name_map.json'), 'w', encoding='utf-8') as f:
    json.dump(duber_map, f, ensure_ascii=False, indent=2)

# 공급가 (품명리스트 최종상품명 → 공급가)
cost_manual = {
    '건조기시트 1개': 6200, '식세기세제': 6150, '하트 식세기세제': 6950,
    '하트식세기세제 선물세트': 15800, '하트 식세기 선물세트': 15800, '올인원 수세미': 2500, '캡슐세제': 9450,
    '섬유탈취제 400': 5940, '섬유탈취제 100': 3950, '얼룩제거제 350': 5950,
    '얼룩제거제 100': 3450, '세탁세제': 6150, '하트 세탁세제': 7450,
    '이염방지시트': 4450, '다목적세정제': 5400, '하트 세탁세제 선물세트': 16700,
}
print(f"[2/4] 공급가: {len(cost_manual)}개")

# 2. 발송 파일에서 두버 추출 (3월만)
BASE = os.path.join(SCRIPT_DIR, '3월 수동 발주', '3월 발송 내역')
all_duber = []
for fname in sorted(os.listdir(BASE)):
    date_match = re.search(r'_(\d{4})', fname)
    if date_match and int(date_match.group(1)[:2]) != 3:
        continue
    fpath = os.path.join(BASE, fname)
    ext = os.path.splitext(fname)[1].lower()
    try:
        if ext == '.xlsx':
            wb = openpyxl.load_workbook(fpath, data_only=True)
            ws = wb[wb.sheetnames[0]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if '두버' in str(row[0] or ''):
                    all_duber.append({'file': fname, 'order': str(row[1] or '').strip(),
                                      'name': str(row[7] or '').strip(),
                                      'option': str(row[8] or '').strip(), 'qty': si(row[9])})
        elif ext == '.xls':
            wb = xlrd.open_workbook(fpath)
            ws = wb.sheet_by_index(0)
            for r in range(1, ws.nrows):
                if '두버' in str(ws.cell_value(r, 0)):
                    all_duber.append({'file': fname, 'order': str(ws.cell_value(r, 1)).strip(),
                                      'name': str(ws.cell_value(r, 7)).strip(),
                                      'option': str(ws.cell_value(r, 8)).strip(), 'qty': si(ws.cell_value(r, 9))})
    except:
        pass

print(f"[3/4] 두버 3월 발송: {len(all_duber)}건")
for item in all_duber:
    print(f"  {item['file'][:30]} | {item['name'][:30]} | {item['option'][:15]} | qty={item['qty']}")

# 3. 매핑 + 공급가 적용
SHIP_COST = 3000
pivot = defaultdict(lambda: {'qty': 0, 'supply': 0, 'ship': 0})
unmapped = []
proc_rows = []
shipped_orders = set()  # 주문번호 기준 묶음배송

for item in all_duber:
    name, option, qty = item['name'], item['option'], item['qty']
    full = f"{name}{option}" if option else name
    final = duber_map.get(full) or duber_map.get(name)
    if not final:
        for mk, mv in duber_map.items():
            if name and (name in mk or mk in name):
                final = mv
                break
    if not final:
        unmapped.append(f"{name} | {option}")
        proc_rows.append({'file': item['file'], 'name': name, 'option': option,
                          'final': '', 'qty': qty, 'supply': 0, 'ship': 0})
        continue

    supply = cost_manual.get(final, 0)
    # 배송비: 주문번호 기준 1회만
    order_no = item.get('order', '')
    ship = 0
    if order_no and order_no not in shipped_orders:
        ship = SHIP_COST
        shipped_orders.add(order_no)
    elif not order_no:
        ship = SHIP_COST  # 주문번호 없으면 건당

    pivot[final]['qty'] += qty
    pivot[final]['supply'] += supply * qty
    pivot[final]['ship'] += ship
    proc_rows.append({'file': item['file'], 'name': name, 'option': option,
                      'final': final, 'qty': qty, 'supply': supply * qty, 'ship': ship})

# 4. 결과
sorted_items = sorted(pivot.items())
tq = ts = tsh = 0
print(f"\n[4/4] 결과\n")
print(f"{'#':>3}  {'상품명':<25} {'수량':>6} {'공급가합계':>14} {'배송비':>10}")
print(f"{'---':>3}  {'---':<25} {'---':>6} {'---':>14} {'---':>10}")
for i, (name, data) in enumerate(sorted_items, 1):
    tq += data['qty']; ts += data['supply']; tsh += data['ship']
    print(f"{i:>3}  {name:<25} {data['qty']:>6,} {data['supply']:>14,.0f} {data['ship']:>10,.0f}")
print(f"{'':>3}  {'총합계':<25} {tq:>6,} {ts:>14,.0f} {tsh:>10,.0f}")
print(f"\n  공급가+배송비 = {ts + tsh:,.0f}")

if unmapped:
    unique = sorted(set(unmapped))
    print(f"\n미매핑 {len(unique)}개:")
    for u in unique:
        print(f"  - {u[:60]} ({unmapped.count(u)}건)")
else:
    print(f"\n미매핑 없음")

# 엑셀
HF = PatternFill(start_color='2D4A7A', end_color='2D4A7A', fill_type='solid')
HFont = Font(bold=True, size=11, color='FFFFFF')
SF = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
TF = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
BN = Font(bold=True, size=11, name='Consolas')
tb = Border(top=Side(style='medium', color='1B2A4A'), bottom=Side(style='medium', color='1B2A4A'))

wbo = openpyxl.Workbook()
wsp = wbo.active
wsp.title = '피벗'
wsp.merge_cells('A1:D1')
wsp['A1'] = '두버 매출 피벗 (3월)'
wsp['A1'].font = Font(bold=True, size=14, color='1B2A4A')
for c, hd in enumerate(['행 레이블', '수량', '공급가합계', '배송비'], 1):
    cell = wsp.cell(row=3, column=c, value=hd)
    cell.font = HFont; cell.fill = HF
r = 4
for name, data in sorted_items:
    wsp.cell(row=r, column=1, value=name)
    wsp.cell(row=r, column=2, value=data['qty']).number_format = '#,##0'
    wsp.cell(row=r, column=3, value=round(data['supply'])).number_format = '#,##0'
    wsp.cell(row=r, column=4, value=round(data['ship'])).number_format = '#,##0'
    r += 1
for c2 in range(1, 5):
    wsp.cell(row=r, column=c2).fill = SF
    wsp.cell(row=r, column=c2).border = tb
wsp.cell(row=r, column=1, value='총합계').font = Font(bold=True)
wsp.cell(row=r, column=2, value=tq).font = BN; wsp.cell(row=r, column=2).number_format = '#,##0'
wsp.cell(row=r, column=3, value=round(ts)).font = BN; wsp.cell(row=r, column=3).number_format = '#,##0'
wsp.cell(row=r, column=4, value=round(tsh)).font = BN; wsp.cell(row=r, column=4).number_format = '#,##0'
r += 2
for c2 in range(3, 5):
    wsp.cell(row=r, column=c2).fill = TF; wsp.cell(row=r, column=c2).border = tb
wsp.cell(row=r, column=3, value='공급가+배송비').font = Font(bold=True, size=10, color='1565C0')
wsp.cell(row=r, column=4, value=round(ts + tsh)).font = Font(bold=True, size=12, color='1565C0', name='Consolas')
wsp.cell(row=r, column=4).number_format = '#,##0'
wsp.column_dimensions['A'].width = 30; wsp.column_dimensions['B'].width = 10
wsp.column_dimensions['C'].width = 16; wsp.column_dimensions['D'].width = 14

wp = wbo.create_sheet('가공')
for c, hd in enumerate(['파일', '상품명', '옵션', '최종상품명', '수량', '공급가', '배송비'], 1):
    cell = wp.cell(row=1, column=c, value=hd); cell.font = HFont; cell.fill = HF
for i, pr in enumerate(proc_rows, 2):
    wp.cell(row=i, column=1, value=pr['file']); wp.cell(row=i, column=2, value=pr['name'])
    wp.cell(row=i, column=3, value=pr['option']); wp.cell(row=i, column=4, value=pr['final'])
    wp.cell(row=i, column=5, value=pr['qty']).number_format = '#,##0'
    wp.cell(row=i, column=6, value=round(pr['supply'])).number_format = '#,##0'
    wp.cell(row=i, column=7, value=round(pr['ship']) if pr['ship'] > 0 else None)
wp.column_dimensions['A'].width = 35; wp.column_dimensions['B'].width = 40
wp.column_dimensions['C'].width = 25; wp.column_dimensions['D'].width = 20
wp.column_dimensions['E'].width = 8; wp.column_dimensions['F'].width = 12; wp.column_dimensions['G'].width = 10

outdir = os.path.join(SCRIPT_DIR, '3월 매출 정산', '03월 두버')
os.makedirs(outdir, exist_ok=True)
outpath = os.path.join(outdir, '03월 두버.xlsx')
wbo.save(outpath)
print(f"\n엑셀 저장: {outpath}")
