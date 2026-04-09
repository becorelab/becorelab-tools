"""
카페24 매출 정산 스크립트
원본 데이터 → 상품명 매핑 → 피벗 집계 → 결과 출력

사용법: python cafe24_settlement.py <원본파일.xlsx 또는 .csv>
"""
import sys
import io
import os
import json
import csv
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ─── 설정 ───
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(SCRIPT_DIR)
NAME_MAP_PATH = os.path.join(PARENT_DIR, 'cafe24_name_map_temp.json')

# ─── 품명리스트 로드 ───
with open(NAME_MAP_PATH, 'r', encoding='utf-8') as f:
    NAME_MAP = json.load(f)
print(f"[1/4] 품명리스트 로드: {len(NAME_MAP)}개 매핑")

# ─── 원본 파일 읽기 ───
if len(sys.argv) < 2:
    print("사용법: python cafe24_settlement.py <원본파일.xlsx 또는 .csv>")
    sys.exit(1)

raw_path = sys.argv[1]
print(f"[2/4] 원본 파일 읽는 중: {os.path.basename(raw_path)}")

def col_idx(headers, keyword):
    """컬럼명에 keyword가 포함된 인덱스 찾기"""
    for i, h in enumerate(headers):
        if h and keyword in str(h):
            return i
    return None

def read_file(path):
    """xlsx 또는 csv 파일을 읽어서 (headers, rows) 반환"""
    ext = os.path.splitext(path)[1].lower()

    if ext == '.csv':
        # CSV: 인코딩 자동 감지 (utf-8-sig → euc-kr → cp949)
        for enc in ['utf-8-sig', 'utf-8', 'euc-kr', 'cp949']:
            try:
                with open(path, 'r', encoding=enc) as f:
                    reader = csv.reader(f)
                    headers = next(reader)
                    rows = list(reader)
                print(f"  CSV 인코딩: {enc}, {len(rows)}행")
                return headers, rows
            except (UnicodeDecodeError, UnicodeError):
                continue
        raise ValueError("CSV 인코딩을 감지할 수 없습니다")

    else:  # xlsx
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(list(row))
        print(f"  Excel: {len(rows)}행")
        return headers, rows

headers, rows = read_file(raw_path)

# 필요한 컬럼 인덱스
IDX_ORDER_NO = col_idx(headers, '주문번호')
IDX_PRODUCT_NAME = col_idx(headers, '주문상품명(옵션포함)') or col_idx(headers, '주문상품명')
IDX_OPTION_PRICE = col_idx(headers, '옵션+판매가')
IDX_QTY = col_idx(headers, '수량')
IDX_SHIPPING = col_idx(headers, '총 배송비(KRW)')

print(f"  컬럼: 주문번호={IDX_ORDER_NO}, 상품명={IDX_PRODUCT_NAME}, "
      f"옵션+판매가={IDX_OPTION_PRICE}, 수량={IDX_QTY}, 배송비={IDX_SHIPPING}")

# ─── 데이터 처리 ───
print(f"[3/4] 데이터 처리 중...")

order_shipping_used = set()
pivot = defaultdict(lambda: {'qty': 0, 'amount': 0, 'shipping': 0})
unmapped = []
total_rows = 0

def safe_float(val):
    """문자열/None → float 변환"""
    if val is None or val == '' or val == 'None':
        return 0.0
    return float(str(val).replace(',', ''))

def safe_int(val):
    """문자열/None → int 변환"""
    if val is None or val == '' or val == 'None':
        return 0
    return int(float(str(val).replace(',', '')))

# 차감 항목 컬럼 인덱스
IDX_REFUND = col_idx(headers, '실제 환불금액')
IDX_COUPON_ORDER = col_idx(headers, '주문서 쿠폰 할인금액')
IDX_APP_DISCOUNT = col_idx(headers, '앱 상품할인 금액(최종)')

# 차감 합계
sum_refund = 0.0
sum_coupon = 0.0
sum_app_discount = 0.0
coupon_counted_orders = set()  # 쿠폰은 주문당 1번만
refund_counted_orders = set()  # 환불도 주문당 1번만

for row in rows:
    total_rows += 1

    # 상품명 매핑
    product_name_raw = str(row[IDX_PRODUCT_NAME] or '').strip()
    final_name = NAME_MAP.get(product_name_raw, None)

    if not final_name:
        unmapped.append(product_name_raw)
        continue

    # 수량
    qty = safe_int(row[IDX_QTY])

    # 결제총액 = 옵션+판매가 × 수량
    option_price = safe_float(row[IDX_OPTION_PRICE])
    amount = option_price * qty

    # 배송비 (주문 단위, 첫 품목에만)
    order_no = str(row[IDX_ORDER_NO] or '').strip()
    shipping = 0.0
    if order_no and order_no not in order_shipping_used:
        shipping = safe_float(row[IDX_SHIPPING])
        order_shipping_used.add(order_no)

    pivot[final_name]['qty'] += qty
    pivot[final_name]['amount'] += amount
    pivot[final_name]['shipping'] += shipping

    # 차감 항목 누적 (주문번호 기준 중복 제거)
    # 환불: 주문 단위 (같은 주문의 여러 상품에 같은 금액 반복 → 1번만)
    if IDX_REFUND is not None and order_no not in refund_counted_orders:
        refund_val = safe_float(row[IDX_REFUND])
        if refund_val > 0:
            sum_refund += refund_val
            refund_counted_orders.add(order_no)
    # 쿠폰: 주문 단위 (첫 품목에만 카운트)
    if IDX_COUPON_ORDER is not None and order_no not in coupon_counted_orders:
        sum_coupon += safe_float(row[IDX_COUPON_ORDER])
        coupon_counted_orders.add(order_no)
    # 앱 할인: 품목 단위
    if IDX_APP_DISCOUNT is not None:
        sum_app_discount += safe_float(row[IDX_APP_DISCOUNT])

# ─── 결과 출력 ───
print(f"[4/4] 결과 출력\n")
print(f"{'='*80}")
print(f"  카페24 매출 피벗 결과")
print(f"{'='*80}")
print(f"  총 {total_rows}행 처리 → {len(pivot)}개 상품 카테고리\n")

sorted_items = sorted(pivot.items(), key=lambda x: x[0])

print(f"{'#':>3}  {'상품명':<45} {'수량':>8} {'결제총액':>15} {'배송비':>12}")
print(f"{'─'*3}  {'─'*45} {'─'*8} {'─'*15} {'─'*12}")

total_qty = 0
total_amount = 0
total_shipping = 0

for i, (name, data) in enumerate(sorted_items, 1):
    q = data['qty']
    a = data['amount']
    s = data['shipping']
    total_qty += q
    total_amount += a
    total_shipping += s
    print(f"{i:>3}  {name:<45} {q:>8,} {a:>15,.0f} {s:>12,.0f}")

print(f"{'─'*3}  {'─'*45} {'─'*8} {'─'*15} {'─'*12}")
print(f"{'':>3}  {'총합계':<45} {total_qty:>8,} {total_amount:>15,.0f} {total_shipping:>12,.0f}")

# 하단 정산 요약
gross = total_amount + total_shipping
net_revenue = gross - sum_refund - sum_coupon - sum_app_discount
net_profit = net_revenue  # 환불/쿠폰/앱할인은 매출 차감 = 이익 차감 (원가 없으므로)
print(f"\n{'─'*80}")
print(f"  {'결제총액 + 배송비':<40} {gross:>15,.0f}")
print(f"  {'환불 (주문 중복제거)':<40} {sum_refund:>15,.0f}")
print(f"  {'쿠폰 적용':<40} {sum_coupon:>15,.0f}")
print(f"  {'앱 상품 할인':<40} {sum_app_discount:>15,.0f}")
print(f"  {'최종 매출':<40} {net_revenue:>15,.0f}")
print(f"")
print(f"  ※ 환불/쿠폰/앱할인은 매출과 이익 모두에서 차감해야 합니다.")
print(f"  ※ 정산시트 기입 시: 매출액(Y열)에 차감 전 금액, 하단 환불/쿠폰/앱할인 행에")
print(f"    매출(P열)과 이익(T열) 모두 마이너스 값을 넣어주세요.")

# 미매핑 상품
if unmapped:
    unique_unmapped = sorted(set(unmapped))
    print(f"\n{'='*80}")
    print(f"  ⚠️ 미매핑 상품: {len(unique_unmapped)}개 ({len(unmapped)}행)")
    print(f"{'='*80}")
    for name in unique_unmapped:
        cnt = unmapped.count(name)
        print(f"  - {name[:80]} ({cnt}건)")
else:
    print(f"\n  ✅ 미매핑 상품 없음 — 전체 매핑 성공!")

print()
