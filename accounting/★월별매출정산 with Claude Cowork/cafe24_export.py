"""
카페24 정산 결과 → 엑셀 출력
원본/피벗/정산요약 시트 구성

사용법: python cafe24_export.py <원본파일.xlsx 또는 .csv>
"""
import sys
import io
import os
import json
import csv
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(SCRIPT_DIR)
NAME_MAP_PATH = os.path.join(PARENT_DIR, 'cafe24_name_map_temp.json')

with open(NAME_MAP_PATH, 'r', encoding='utf-8') as f:
    NAME_MAP = json.load(f)

if len(sys.argv) < 2:
    print("사용법: python cafe24_export.py <원본파일>")
    sys.exit(1)

raw_path = sys.argv[1]
basename = os.path.splitext(os.path.basename(raw_path))[0]

# ─── 파일 읽기 ───
def col_idx(headers, keyword):
    for i, h in enumerate(headers):
        if h and keyword in str(h):
            return i
    return None

def safe_float(val):
    if val is None or val == '' or val == 'None': return 0.0
    return float(str(val).replace(',', ''))

def safe_int(val):
    if val is None or val == '' or val == 'None': return 0
    return int(float(str(val).replace(',', '')))

ext = os.path.splitext(raw_path)[1].lower()
if ext == '.csv':
    for enc in ['utf-8-sig', 'utf-8', 'euc-kr', 'cp949']:
        try:
            with open(raw_path, 'r', encoding=enc) as f:
                reader = csv.reader(f)
                headers = next(reader)
                rows = list(reader)
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
else:
    import openpyxl as oxl
    wb_in = oxl.load_workbook(raw_path, data_only=True)
    ws_in = wb_in[wb_in.sheetnames[0]]
    headers = [cell.value for cell in ws_in[1]]
    rows = [list(r) for r in ws_in.iter_rows(min_row=2, values_only=True)]

IDX_ORDER_NO = col_idx(headers, '주문번호')
IDX_PRODUCT_NAME = col_idx(headers, '주문상품명(옵션포함)') or col_idx(headers, '주문상품명')
IDX_OPTION_PRICE = col_idx(headers, '옵션+판매가')
IDX_QTY = col_idx(headers, '수량')
IDX_SHIPPING = col_idx(headers, '총 배송비(KRW)')
IDX_COUPON_ORDER = col_idx(headers, '주문서 쿠폰 할인금액')
IDX_APP_DISCOUNT = col_idx(headers, '앱 상품할인 금액(최종)')
IDX_REFUND = col_idx(headers, '실제 환불금액')
IDX_STATUS = col_idx(headers, '주문 상태')

print(f"[1/3] 데이터 처리 중... ({len(rows)}행)")

# ─── 처리 ───
order_shipping_used = set()
coupon_counted_orders = set()
pivot = defaultdict(lambda: {'qty': 0, 'amount': 0, 'shipping': 0})
unmapped = []
processed_rows = []  # 가공 데이터
sum_refund = 0.0
sum_coupon = 0.0
sum_app_discount = 0.0

for row in rows:
    product_name_raw = str(row[IDX_PRODUCT_NAME] or '').strip()
    final_name = NAME_MAP.get(product_name_raw, '')
    qty = safe_int(row[IDX_QTY])
    option_price = safe_float(row[IDX_OPTION_PRICE])
    amount = option_price * qty
    order_no = str(row[IDX_ORDER_NO] or '').strip()

    shipping = 0.0
    if order_no and order_no not in order_shipping_used:
        shipping = safe_float(row[IDX_SHIPPING])
        order_shipping_used.add(order_no)

    if final_name:
        pivot[final_name]['qty'] += qty
        pivot[final_name]['amount'] += amount
        pivot[final_name]['shipping'] += shipping
    else:
        unmapped.append(product_name_raw)

    # 차감 항목
    if IDX_REFUND is not None:
        sum_refund += safe_float(row[IDX_REFUND])
    if IDX_COUPON_ORDER is not None and order_no not in coupon_counted_orders:
        sum_coupon += safe_float(row[IDX_COUPON_ORDER])
        coupon_counted_orders.add(order_no)
    if IDX_APP_DISCOUNT is not None:
        sum_app_discount += safe_float(row[IDX_APP_DISCOUNT])

    processed_rows.append({
        'order_no': order_no,
        'status': str(row[IDX_STATUS] or '') if IDX_STATUS else '',
        'product_raw': product_name_raw,
        'final_name': final_name,
        'qty': qty,
        'option_price': option_price,
        'amount': amount,
        'shipping': shipping,
    })

# ─── 스타일 정의 ───
NAVY = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
DARK_BG = PatternFill(start_color='1E1E2E', end_color='1E1E2E', fill_type='solid')
HEADER_FILL = PatternFill(start_color='2D4A7A', end_color='2D4A7A', fill_type='solid')
LIGHT_ROW = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
SUMMARY_FILL = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
DEDUCT_FILL = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
TOTAL_FILL = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
WARN_FILL = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')

HEADER_FONT = Font(bold=True, size=11, color='FFFFFF')
TITLE_FONT = Font(bold=True, size=14, color='1B2A4A')
SUBTITLE_FONT = Font(bold=True, size=11, color='546E7A')
NUM_FONT = Font(size=10, name='Consolas')
BOLD_NUM = Font(bold=True, size=11, name='Consolas')

thin_border = Border(
    bottom=Side(style='thin', color='E0E0E0')
)
thick_border = Border(
    top=Side(style='medium', color='1B2A4A'),
    bottom=Side(style='medium', color='1B2A4A')
)

NUM_FMT = '#,##0'
PCT_FMT = '0.0%'

# ─── 엑셀 생성 ───
print(f"[2/3] 엑셀 생성 중...")
wb = Workbook()

# === Sheet 1: 피벗 ===
ws_pivot = wb.active
ws_pivot.title = '피벗'
ws_pivot.sheet_properties.tabColor = '4CAF50'

# 타이틀
ws_pivot.merge_cells('A1:D1')
ws_pivot['A1'] = f'📊 카페24 매출 피벗'
ws_pivot['A1'].font = TITLE_FONT
ws_pivot.merge_cells('A2:D2')
ws_pivot['A2'] = f'파일: {basename}'
ws_pivot['A2'].font = SUBTITLE_FONT

# 헤더
pivot_headers = ['행 레이블', '합계 : 수량', '합계 : 결제총액', '합계 : 총 배송비(KRW)']
for c, h in enumerate(pivot_headers, 1):
    cell = ws_pivot.cell(row=4, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center')

# 데이터
sorted_items = sorted(pivot.items(), key=lambda x: x[0])
total_qty = total_amount = total_shipping = 0
r = 5
for name, data in sorted_items:
    q, a, s = data['qty'], data['amount'], data['shipping']
    total_qty += q
    total_amount += a
    total_shipping += s
    ws_pivot.cell(row=r, column=1, value=name).font = Font(size=10)
    ws_pivot.cell(row=r, column=2, value=q).font = NUM_FONT
    ws_pivot.cell(row=r, column=2).number_format = NUM_FMT
    ws_pivot.cell(row=r, column=2).alignment = Alignment(horizontal='right')
    ws_pivot.cell(row=r, column=3, value=a).font = NUM_FONT
    ws_pivot.cell(row=r, column=3).number_format = NUM_FMT
    ws_pivot.cell(row=r, column=3).alignment = Alignment(horizontal='right')
    ws_pivot.cell(row=r, column=4, value=s if s > 0 else None).font = NUM_FONT
    ws_pivot.cell(row=r, column=4).number_format = NUM_FMT
    ws_pivot.cell(row=r, column=4).alignment = Alignment(horizontal='right')
    if (r - 5) % 2 == 1:
        for c in range(1, 5):
            ws_pivot.cell(row=r, column=c).fill = LIGHT_ROW
    for c in range(1, 5):
        ws_pivot.cell(row=r, column=c).border = thin_border
    r += 1

# 총합계
for c in range(1, 5):
    ws_pivot.cell(row=r, column=c).fill = SUMMARY_FILL
    ws_pivot.cell(row=r, column=c).border = thick_border
ws_pivot.cell(row=r, column=1, value='총합계').font = Font(bold=True, size=11)
ws_pivot.cell(row=r, column=2, value=total_qty).font = BOLD_NUM
ws_pivot.cell(row=r, column=2).number_format = NUM_FMT
ws_pivot.cell(row=r, column=3, value=total_amount).font = BOLD_NUM
ws_pivot.cell(row=r, column=3).number_format = NUM_FMT
ws_pivot.cell(row=r, column=4, value=total_shipping).font = BOLD_NUM
ws_pivot.cell(row=r, column=4).number_format = NUM_FMT
r += 1

# 하단 정산 요약
gross = total_amount + total_shipping
summary_data = [
    ('결제총액 + 배송비', gross, TOTAL_FILL),
    ('환불', sum_refund, WARN_FILL),
    ('쿠폰 적용', sum_coupon, DEDUCT_FILL),
    ('앱 상품 할인', sum_app_discount, DEDUCT_FILL),
]
r += 1
for label, val, fill in summary_data:
    ws_pivot.cell(row=r, column=3, value=label).font = Font(bold=True, size=10)
    ws_pivot.cell(row=r, column=4, value=val).font = BOLD_NUM
    ws_pivot.cell(row=r, column=4).number_format = NUM_FMT
    for c in range(3, 5):
        ws_pivot.cell(row=r, column=c).fill = fill
    r += 1

# 최종 정산
net = gross - sum_refund - sum_coupon - sum_app_discount
ws_pivot.cell(row=r, column=3, value='(결제총액+배송비-환불-쿠폰-앱 할인)').font = Font(bold=True, size=10, color='1565C0')
ws_pivot.cell(row=r, column=4, value=net).font = Font(bold=True, size=12, color='1565C0', name='Consolas')
ws_pivot.cell(row=r, column=4).number_format = NUM_FMT
for c in range(3, 5):
    ws_pivot.cell(row=r, column=c).fill = TOTAL_FILL
    ws_pivot.cell(row=r, column=c).border = thick_border
r += 2

# 환불 주의사항
ws_pivot.cell(row=r, column=1, value='⚠️ 환불 금액은 수동 확인이 필요합니다 (반품 환불만 적용하려면 직접 수정해주세요)').font = Font(size=9, color='E53935', italic=True)

# 컬럼 너비
ws_pivot.column_dimensions['A'].width = 50
ws_pivot.column_dimensions['B'].width = 15
ws_pivot.column_dimensions['C'].width = 20
ws_pivot.column_dimensions['D'].width = 20

# === Sheet 2: 가공 ===
ws_proc = wb.create_sheet('가공')
ws_proc.sheet_properties.tabColor = '2196F3'

proc_headers = ['주문번호', '주문상태', '원본 상품명', '최종상품명', '수량', '옵션+판매가', '결제총액', '배송비']
for c, h in enumerate(proc_headers, 1):
    cell = ws_proc.cell(row=1, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center')

for i, pr in enumerate(processed_rows, 2):
    ws_proc.cell(row=i, column=1, value=pr['order_no'])
    ws_proc.cell(row=i, column=2, value=pr['status'])
    ws_proc.cell(row=i, column=3, value=pr['product_raw'])
    cell_fn = ws_proc.cell(row=i, column=4, value=pr['final_name'])
    if not pr['final_name']:
        cell_fn.fill = WARN_FILL
    ws_proc.cell(row=i, column=5, value=pr['qty']).number_format = NUM_FMT
    ws_proc.cell(row=i, column=6, value=pr['option_price']).number_format = NUM_FMT
    ws_proc.cell(row=i, column=7, value=pr['amount']).number_format = NUM_FMT
    ws_proc.cell(row=i, column=8, value=pr['shipping'] if pr['shipping'] > 0 else None)
    if pr['shipping'] > 0:
        ws_proc.cell(row=i, column=8).number_format = NUM_FMT

ws_proc.column_dimensions['A'].width = 22
ws_proc.column_dimensions['B'].width = 18
ws_proc.column_dimensions['C'].width = 60
ws_proc.column_dimensions['D'].width = 35
ws_proc.column_dimensions['E'].width = 10
ws_proc.column_dimensions['F'].width = 15
ws_proc.column_dimensions['G'].width = 15
ws_proc.column_dimensions['H'].width = 12
ws_proc.auto_filter.ref = f"A1:H{len(processed_rows)+1}"

# === Sheet 3: 미매핑 (있을 경우) ===
if unmapped:
    ws_unmap = wb.create_sheet('미매핑')
    ws_unmap.sheet_properties.tabColor = 'F44336'
    unmap_headers = ['#', '원본 상품명', '건수', '매핑 입력']
    for c, h in enumerate(unmap_headers, 1):
        cell = ws_unmap.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = PatternFill(start_color='C62828', end_color='C62828', fill_type='solid')
    from collections import Counter
    unmap_counts = Counter(unmapped)
    for i, (name, cnt) in enumerate(sorted(unmap_counts.items()), 1):
        ws_unmap.cell(row=i+1, column=1, value=i)
        ws_unmap.cell(row=i+1, column=2, value=name)
        ws_unmap.cell(row=i+1, column=3, value=cnt)
    ws_unmap.column_dimensions['B'].width = 80
    ws_unmap.column_dimensions['D'].width = 30

# 저장
output_dir = os.path.dirname(raw_path)
output_name = f"카페24_정산결과_{basename}.xlsx"
output_path = os.path.join(output_dir, output_name)
wb.save(output_path)

print(f"[3/3] 저장 완료!")
print(f"  📁 {output_path}")
print(f"  📊 피벗: {len(pivot)}개 상품")
print(f"  📋 가공: {len(processed_rows)}행")
if unmapped:
    print(f"  ⚠️ 미매핑: {len(set(unmapped))}개")
else:
    print(f"  ✅ 전체 매핑 성공")
