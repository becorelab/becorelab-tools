"""
옥션 매출 정산 스크립트 (ESM 플랫폼, 지마켓과 동일 형식)
매출기준_상품판매_상세내역 + 배송비 → 품명 매핑 → 집계 → 원가/이익 → xlsx

사용법: python auction_settlement.py <매출기준파일> <배송비파일> <출력xlsx>
"""
import sys, io, os, json
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(SCRIPT_DIR)
NAME_MAP_PATH = os.path.join(PARENT_DIR, 'auction_name_map.json')
COST_MASTER_PATH = '/tmp/cost_master.json'

with open(NAME_MAP_PATH, 'r', encoding='utf-8') as f:
    NAME_MAP = json.load(f)
NAME_MAP.pop('상품명', None)  # 헤더 제거
with open(COST_MASTER_PATH, 'r', encoding='utf-8') as f:
    COST = json.load(f)
print(f"[1/6] 매핑 {len(NAME_MAP)}개, 원가 {len(COST)}개 로드")

sales_path = sys.argv[1]
ship_path = sys.argv[2]
out_path = sys.argv[3]

def safe_float(val):
    if val is None or val == '' or val == 'None': return 0.0
    return float(str(val).replace(',', ''))

def safe_int(val):
    if val is None or val == '' or val == 'None': return 0
    return int(float(str(val).replace(',', '')))

def read_xls(path):
    import xlrd
    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_index(0)
    headers = [ws.cell_value(0, c) for c in range(ws.ncols)]
    rows = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(1, ws.nrows)]
    return headers, rows

def col_idx(headers, keyword):
    for i, h in enumerate(headers):
        if h and keyword in str(h):
            return i
    return None

# ─── 배송비: 묶음배송번호별 정산액 합산 ───
print(f"[2/6] 배송비 읽는 중")
h_ship, rows_ship = read_xls(ship_path)
IDX_SH_BUNDLE = col_idx(h_ship, '묶음배송번호')
IDX_SH_SETTLE = col_idx(h_ship, '배송비 정산액')
ship_by_bundle = defaultdict(float)
total_ship = 0.0
for row in rows_ship:
    bundle = str(int(safe_float(row[IDX_SH_BUNDLE]))) if row[IDX_SH_BUNDLE] else ''
    settle = safe_float(row[IDX_SH_SETTLE])
    if bundle:
        ship_by_bundle[bundle] += settle
    total_ship += settle
print(f"  배송비 {len(rows_ship)}행, 총 배송비정산액 {total_ship:,.0f}")

# ─── 매출기준 ───
print(f"[3/6] 매출기준 읽는 중")
h_sales, rows_sales = read_xls(sales_path)
IDX_NAME = col_idx(h_sales, '상품명')
IDX_QTY = col_idx(h_sales, '주문수량')
IDX_SETTLE = col_idx(h_sales, '판매자 최종정산금')
IDX_BUNDLE = col_idx(h_sales, '묶음배송번호')
print(f"  상품판매 {len(rows_sales)}행 (상품명={IDX_NAME} 수량={IDX_QTY} 정산금={IDX_SETTLE} 묶음={IDX_BUNDLE})")

def normalize(name):
    n = name.strip()
    if n.startswith('(스타배송)'):
        n = n[len('(스타배송)'):].strip()
    return n

def map_name(name):
    n = normalize(name)
    if n in NAME_MAP:
        return NAME_MAP[n]
    # 부분 일치: 원본 상품명이 매핑 키를 포함하거나 키가 원본을 포함
    for mk, mv in NAME_MAP.items():
        if n == mk:
            return mv
    for mk, mv in NAME_MAP.items():
        if mk in n or n in mk:
            return mv
    return None

# ─── 집계 (정산금액 = 최종정산금 + 묶음배송비) ───
print(f"[4/6] 집계")
pivot = defaultdict(lambda: {'qty': 0, 'settle': 0})
unmapped = defaultdict(int)
seen_bundles = set()
total_settle_raw = 0.0

for row in rows_sales:
    name = str(row[IDX_NAME] or '').strip()
    if not name:
        continue
    qty = safe_int(row[IDX_QTY])
    settle = safe_float(row[IDX_SETTLE])
    total_settle_raw += settle
    bundle = str(int(safe_float(row[IDX_BUNDLE]))) if row[IDX_BUNDLE] else ''

    # 배송비: 묶음배송번호당 1회만 더함
    ship = 0.0
    if bundle and bundle not in seen_bundles:
        ship = ship_by_bundle.get(bundle, 0.0)
        seen_bundles.add(bundle)

    final = map_name(name)
    if not final:
        unmapped[name] += 1
        continue

    pivot[final]['qty'] += qty
    pivot[final]['settle'] += settle + ship

# 매핑 안 된 행의 배송비는 위에서 더해지지 않았을 수 있으나, 미매핑 0건 가정
mapped_ship = sum(ship_by_bundle.get(b, 0.0) for b in seen_bundles)

# ─── 원가/이익 ───
print(f"[5/6] 원가 계산")
rows_out = []
cost_missing = set()
for name, d in pivot.items():
    qty, settle = d['qty'], d['settle']
    unit_cost = COST.get(name)
    if unit_cost is None:
        cost_total = None
        profit = None
        cost_missing.add(name)
    else:
        cost_total = qty * unit_cost
        profit = settle - cost_total
    rows_out.append({'품명': name, '수량': qty, '정산금액': round(settle),
                     '원가': cost_total, '이익': profit})

rows_out.sort(key=lambda x: x['정산금액'], reverse=True)

# ─── xlsx 출력 ───
print(f"[6/6] xlsx 작성")
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
wb = openpyxl.Workbook()
ws = wb.active
ws.title = '옥션 5월'

thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_fill = PatternFill('solid', fgColor='2F5496')
hdr_font = Font(bold=True, color='FFFFFF')
sum_fill = PatternFill('solid', fgColor='D9E1F2')
sum_font = Font(bold=True)

cols = ['품명', '수량', '정산금액', '원가', '이익']
ws.append(cols)
for c in range(1, 6):
    cell = ws.cell(1, c)
    cell.fill = hdr_fill; cell.font = hdr_font
    cell.alignment = Alignment(horizontal='center'); cell.border = border

tot_qty = tot_settle = 0
tot_cost = tot_profit = 0
for r in rows_out:
    cost = r['원가'] if r['원가'] is not None else '원가미상'
    profit = r['이익'] if r['이익'] is not None else '원가미상'
    ws.append([r['품명'], r['수량'], r['정산금액'], cost, profit])
    tot_qty += r['수량']; tot_settle += r['정산금액']
    if r['원가'] is not None:
        tot_cost += r['원가']; tot_profit += r['이익']

# 합계행
ws.append(['합계', tot_qty, tot_settle, tot_cost, tot_profit])

# 서식
nrows = ws.max_row
for r in range(2, nrows + 1):
    for c in range(1, 6):
        cell = ws.cell(r, c)
        cell.border = border
        if c >= 2 and isinstance(cell.value, (int, float)):
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='right')
    if r == nrows:
        for c in range(1, 6):
            ws.cell(r, c).fill = sum_fill
            ws.cell(r, c).font = sum_font

ws.column_dimensions['A'].width = 28
for col in ['B', 'C', 'D', 'E']:
    ws.column_dimensions[col].width = 14

os.makedirs(os.path.dirname(out_path), exist_ok=True)
wb.save(out_path)

# ─── 콘솔 요약 + 검증 ───
print(f"\n{'='*70}")
print(f"  옥션 5월 정산 결과")
print(f"{'='*70}")
print(f"  {'품명':<26} {'수량':>5} {'정산금액':>11} {'원가':>11} {'이익':>11}")
for r in rows_out:
    cost = f"{r['원가']:,}" if r['원가'] is not None else '원가미상'
    profit = f"{r['이익']:,}" if r['이익'] is not None else '원가미상'
    print(f"  {r['품명']:<26} {r['수량']:>5} {r['정산금액']:>11,} {cost:>11} {profit:>11}")
print(f"  {'-'*68}")
print(f"  {'합계':<26} {tot_qty:>5} {tot_settle:>11,} {tot_cost:>11,} {tot_profit:>11,}")

print(f"\n[검증]")
print(f"  원본 최종정산금 합계: {total_settle_raw:,.0f}")
print(f"  원본 배송비정산액 합계: {total_ship:,.0f}")
print(f"  매핑된 배송비: {mapped_ship:,.0f}")
print(f"  원본 정산금+배송비 = {total_settle_raw + total_ship:,.0f}")
print(f"  출력 정산금액 합계 = {tot_settle:,}")
diff = (total_settle_raw + total_ship) - tot_settle
print(f"  차이 = {diff:,.0f}")

if unmapped:
    print(f"\n  ⚠️ 미매핑 {len(unmapped)}종:")
    for n, c in sorted(unmapped.items()):
        print(f"    - {n} ({c}건)")
else:
    print(f"\n  ✅ 미매핑 없음")
if cost_missing:
    print(f"\n  ⚠️ 원가미상 {len(cost_missing)}종: {', '.join(sorted(cost_missing))}")
else:
    print(f"\n  ✅ 원가미상 없음")
print(f"\n  저장: {out_path}")
