"""
쿠팡 로켓배송 정산 결과 → 엑셀 출력 (피벗 + 가공 시트)

사용법: python rocket_export.py <매출데이터.xlsx> [광고보고서.xlsx]
"""
import sys, io, os, json, openpyxl
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(SCRIPT_DIR)
NAME_MAP_PATH = os.path.join(PARENT_DIR, 'rocket_name_map.json')

with open(NAME_MAP_PATH, 'r', encoding='utf-8') as f:
    NAME_MAP = json.load(f)

def safe_float(val):
    if val is None or val == '' or val == 'None': return 0.0
    return float(str(val).replace(',', ''))

def safe_int(val):
    if val is None or val == '' or val == 'None': return 0
    return int(float(str(val).replace(',', '')))

def col_idx(headers, keyword):
    for i, h in enumerate(headers):
        if h and keyword in str(h):
            return i
    return None

data_path = sys.argv[1]
ad_path = sys.argv[2] if len(sys.argv) > 2 else None

# 매출 데이터
wb_in = openpyxl.load_workbook(data_path, data_only=True)
ws_in = wb_in['data'] if 'data' in wb_in.sheetnames else wb_in[wb_in.sheetnames[0]]
headers = [cell.value for cell in ws_in[1]]
IDX_TYPE = col_idx(headers, '구분')
IDX_SKU = col_idx(headers, 'SKU명')
IDX_QTY = col_idx(headers, '수량')
IDX_TOTAL_PRICE = col_idx(headers, '총단가')
IDX_SUPPLY = col_idx(headers, '총공급가액')
IDX_TAX = col_idx(headers, '총세액')

pivot = defaultdict(lambda: {'qty': 0, 'total_price': 0})
proc_rows = []

for row in ws_in.iter_rows(min_row=2, values_only=True):
    if row[IDX_TYPE] is None:
        continue
    sku = str(row[IDX_SKU] or '').strip()
    if not sku:
        continue
    sku_clean = sku.replace('\xa0', ' ').strip()
    final = NAME_MAP.get(sku, None) or NAME_MAP.get(sku_clean, None)
    qty = safe_int(row[IDX_QTY])
    tp = safe_float(row[IDX_TOTAL_PRICE])

    if final:
        pivot[final]['qty'] += qty
        pivot[final]['total_price'] += tp

    proc_rows.append({
        'type': str(row[IDX_TYPE] or ''),
        'sku': sku,
        'final': final or '',
        'qty': qty,
        'total_price': tp,
    })

# 광고비
ad_by_product = defaultdict(lambda: {'smart': 0, 'campaign': 0})
total_ad = 0

if ad_path and os.path.exists(ad_path):
    wb_ad = openpyxl.load_workbook(ad_path, data_only=True)
    ws_ad = wb_ad[wb_ad.sheetnames[0]]
    ad_h = [cell.value for cell in ws_ad[1]]
    IDX_CAMP = col_idx(ad_h, '캠페인명')
    IDX_AD_PROD = col_idx(ad_h, '광고집행 상품명')
    IDX_AD_COST = col_idx(ad_h, '광고비')

    for row in ws_ad.iter_rows(min_row=2, values_only=True):
        camp = str(row[IDX_CAMP] or '').strip()
        ad_prod = str(row[IDX_AD_PROD] or '').strip()
        cost = safe_float(row[IDX_AD_COST])
        if cost <= 0:
            continue
        total_ad += cost
        mapped = NAME_MAP.get(ad_prod, None)
        if not mapped:
            for mk, mv in NAME_MAP.items():
                if ad_prod and ad_prod in mk:
                    mapped = mv
                    break
        if mapped:
            is_smart = '스마트' in camp or 'AI' in camp
            if is_smart:
                ad_by_product[mapped]['smart'] += cost
            else:
                ad_by_product[mapped]['campaign'] += cost

# === 엑셀 생성 ===
HEADER_FILL = PatternFill(start_color='2D4A7A', end_color='2D4A7A', fill_type='solid')
HEADER_FONT = Font(bold=True, size=11, color='FFFFFF')
LIGHT_ROW = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
SUMMARY_FILL = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
TOTAL_FILL = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
AD_FILL = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
TITLE_FONT = Font(bold=True, size=14, color='1B2A4A')
NUM_FONT = Font(size=10, name='Consolas')
BOLD_NUM = Font(bold=True, size=11, name='Consolas')
thin_border = Border(bottom=Side(style='thin', color='E0E0E0'))
thick_border = Border(top=Side(style='medium', color='1B2A4A'), bottom=Side(style='medium', color='1B2A4A'))

wb = openpyxl.Workbook()

# === 피벗 시트 ===
ws_pv = wb.active
ws_pv.title = '피벗'
ws_pv.sheet_properties.tabColor = '4CAF50'
ws_pv.merge_cells('A1:F1')
ws_pv['A1'] = '쿠팡 로켓배송 매출 피벗'
ws_pv['A1'].font = TITLE_FONT

has_ads = total_ad > 0
if has_ads:
    pv_h = ['행 레이블', '수량', '총단가', '광고비(스마트)', '광고비(독립캠페인)']
else:
    pv_h = ['행 레이블', '수량', '총단가']

for c, h in enumerate(pv_h, 1):
    cell = ws_pv.cell(row=3, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center')

sorted_items = sorted(pivot.items(), key=lambda x: x[0])
total_qty = total_price = total_smart = total_campaign = 0
r = 4
for name, data in sorted_items:
    total_qty += data['qty']
    total_price += data['total_price']
    smart = ad_by_product[name]['smart']
    campaign = ad_by_product[name]['campaign']
    total_smart += smart
    total_campaign += campaign

    ws_pv.cell(row=r, column=1, value=name).font = Font(size=10)
    ws_pv.cell(row=r, column=2, value=data['qty']).font = NUM_FONT
    ws_pv.cell(row=r, column=2).number_format = '#,##0'
    ws_pv.cell(row=r, column=3, value=round(data['total_price'])).font = NUM_FONT
    ws_pv.cell(row=r, column=3).number_format = '#,##0'
    if has_ads:
        ws_pv.cell(row=r, column=4, value=round(smart) if smart > 0 else None).font = NUM_FONT
        ws_pv.cell(row=r, column=4).number_format = '#,##0'
        ws_pv.cell(row=r, column=5, value=round(campaign) if campaign > 0 else None).font = NUM_FONT
        ws_pv.cell(row=r, column=5).number_format = '#,##0'

    cols = 6 if has_ads else 4
    if (r - 4) % 2 == 1:
        for c in range(1, cols):
            ws_pv.cell(row=r, column=c).fill = LIGHT_ROW
    for c in range(1, cols):
        ws_pv.cell(row=r, column=c).border = thin_border
    r += 1

# 총합계
cols = 6 if has_ads else 4
for c in range(1, cols):
    ws_pv.cell(row=r, column=c).fill = SUMMARY_FILL
    ws_pv.cell(row=r, column=c).border = thick_border
ws_pv.cell(row=r, column=1, value='총합계').font = Font(bold=True, size=11)
ws_pv.cell(row=r, column=2, value=total_qty).font = BOLD_NUM
ws_pv.cell(row=r, column=2).number_format = '#,##0'
ws_pv.cell(row=r, column=3, value=round(total_price)).font = BOLD_NUM
ws_pv.cell(row=r, column=3).number_format = '#,##0'
if has_ads:
    ws_pv.cell(row=r, column=4, value=round(total_smart)).font = BOLD_NUM
    ws_pv.cell(row=r, column=4).number_format = '#,##0'
    ws_pv.cell(row=r, column=5, value=round(total_campaign)).font = BOLD_NUM
    ws_pv.cell(row=r, column=5).number_format = '#,##0'
    r += 2
    for c in range(3, 6):
        ws_pv.cell(row=r, column=c).fill = AD_FILL
    ws_pv.cell(row=r, column=3, value='총 광고비').font = Font(bold=True, size=10)
    ws_pv.cell(row=r, column=4, value=round(total_smart + total_campaign)).font = BOLD_NUM
    ws_pv.cell(row=r, column=4).number_format = '#,##0'

ws_pv.column_dimensions['A'].width = 35
ws_pv.column_dimensions['B'].width = 10
ws_pv.column_dimensions['C'].width = 16
ws_pv.column_dimensions['D'].width = 16
ws_pv.column_dimensions['E'].width = 18

# === 가공 시트 ===
ws_proc = wb.create_sheet('가공')
ws_proc.sheet_properties.tabColor = '2196F3'
proc_h = ['구분', 'SKU명', '최종상품명', '수량', '총단가']
for c, h in enumerate(proc_h, 1):
    cell = ws_proc.cell(row=1, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
for i, pr in enumerate(proc_rows, 2):
    ws_proc.cell(row=i, column=1, value=pr['type'])
    ws_proc.cell(row=i, column=2, value=pr['sku'])
    ws_proc.cell(row=i, column=3, value=pr['final'])
    ws_proc.cell(row=i, column=4, value=pr['qty']).number_format = '#,##0'
    ws_proc.cell(row=i, column=5, value=round(pr['total_price'])).number_format = '#,##0'

ws_proc.column_dimensions['A'].width = 10
ws_proc.column_dimensions['B'].width = 60
ws_proc.column_dimensions['C'].width = 30
ws_proc.column_dimensions['D'].width = 10
ws_proc.column_dimensions['E'].width = 15
ws_proc.auto_filter.ref = f"A1:E{len(proc_rows)+1}"

outpath = os.path.join(os.path.dirname(data_path), '03월 로켓배송.xlsx')
wb.save(outpath)
print(f"저장 완료: {outpath}")
print(f"피벗: {len(pivot)}개, 매출: {round(total_price):,}원, 광고비: {round(total_smart+total_campaign):,}원")
