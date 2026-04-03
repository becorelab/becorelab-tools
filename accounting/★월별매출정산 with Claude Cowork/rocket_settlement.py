"""
쿠팡 로켓배송 매출 정산 스크립트
Coupang_Stocked_Data_List → 품명 매핑 → 피벗 집계 + 광고비 반영

사용법: python rocket_settlement.py <매출데이터.xlsx> [광고보고서.xlsx]
"""
import sys
import io
import os
import json
import openpyxl
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(SCRIPT_DIR)
NAME_MAP_PATH = os.path.join(PARENT_DIR, 'rocket_name_map.json')

# 품명리스트 로드
with open(NAME_MAP_PATH, 'r', encoding='utf-8') as f:
    NAME_MAP = json.load(f)
print(f"[1/5] 품명리스트 로드: {len(NAME_MAP)}개 매핑")

if len(sys.argv) < 2:
    print("사용법: python rocket_settlement.py <매출데이터.xlsx> [광고보고서.xlsx]")
    sys.exit(1)

data_path = sys.argv[1]
ad_path = sys.argv[2] if len(sys.argv) > 2 else None

def safe_float(val):
    if val is None or val == '' or val == 'None': return 0.0
    return float(str(val).replace(',', ''))

def safe_int(val):
    if val is None or val == '' or val == 'None': return 0
    return int(float(str(val).replace(',', '')))

def col_idx(headers, keyword):
    for i, h in enumerate(headers):
        if h and keyword in str(h):
            return i
    return None

# ─── 매출 데이터 읽기 ───
print(f"[2/5] 매출 데이터 읽는 중: {os.path.basename(data_path)}")
wb = openpyxl.load_workbook(data_path, data_only=True)
ws = wb['data'] if 'data' in wb.sheetnames else wb[wb.sheetnames[0]]
headers = [cell.value for cell in ws[1]]

IDX_TYPE = col_idx(headers, '구분')
IDX_SKU = col_idx(headers, 'SKU명')
IDX_QTY = col_idx(headers, '수량')
IDX_TOTAL_PRICE = col_idx(headers, '총단가')

print(f"  컬럼: SKU명={IDX_SKU}, 수량={IDX_QTY}, 총단가={IDX_TOTAL_PRICE}")

# ─── 피벗 집계 ───
print(f"[3/5] 데이터 처리...")
pivot = defaultdict(lambda: {'qty': 0, 'total_price': 0})
unmapped = []
total_rows = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    if row[IDX_TYPE] is None:
        continue
    total_rows += 1

    sku_name = str(row[IDX_SKU] or '').strip()
    if not sku_name:
        continue
    # non-breaking space 처리
    sku_name_clean = sku_name.replace('\xa0', ' ').strip()

    final_name = NAME_MAP.get(sku_name, None) or NAME_MAP.get(sku_name_clean, None)

    if not final_name:
        unmapped.append(sku_name)
        continue

    qty = safe_int(row[IDX_QTY])
    total_price = safe_float(row[IDX_TOTAL_PRICE])

    pivot[final_name]['qty'] += qty
    pivot[final_name]['total_price'] += total_price

# ─── 광고비 처리 ───
ad_by_product = defaultdict(lambda: {'smart': 0, 'campaign': 0})
total_ad = 0

if ad_path and os.path.exists(ad_path):
    print(f"[4/5] 광고비 처리: {os.path.basename(ad_path)}")
    wb_ad = openpyxl.load_workbook(ad_path, data_only=True)
    ws_ad = wb_ad[wb_ad.sheetnames[0]]
    ad_headers = [cell.value for cell in ws_ad[1]]

    IDX_CAMP_NAME = col_idx(ad_headers, '캠페인명')
    IDX_AD_PRODUCT = col_idx(ad_headers, '광고집행 상품명')
    IDX_AD_COST = col_idx(ad_headers, '광고비')

    # 캠페인별 상품별 광고비 집계
    camp_product_cost = defaultdict(lambda: defaultdict(float))

    for row in ws_ad.iter_rows(min_row=2, values_only=True):
        camp = str(row[IDX_CAMP_NAME] or '').strip()
        ad_product = str(row[IDX_AD_PRODUCT] or '').strip()
        cost = safe_float(row[IDX_AD_COST])
        if cost > 0:
            camp_product_cost[camp][ad_product] += cost

    # 캠페인 요약 출력
    print(f"  캠페인별 광고비:")
    for camp, products in camp_product_cost.items():
        camp_total = sum(products.values())
        total_ad += camp_total
        print(f"    {camp[:50]}: {camp_total:,.0f}원 ({len(products)}개 상품)")

    # 광고비를 상품에 매핑
    # 광고집행 상품명 → 품명리스트로 매핑 후 배분
    for camp, products in camp_product_cost.items():
        is_smart = '스마트' in camp or 'AI' in camp
        for ad_prod, cost in products.items():
            # 광고 상품명으로 품명 매핑
            mapped = NAME_MAP.get(ad_prod, None)
            if not mapped:
                # 부분 매칭
                for mk, mv in NAME_MAP.items():
                    if ad_prod and ad_prod in mk:
                        mapped = mv
                        break
            if mapped:
                if is_smart:
                    ad_by_product[mapped]['smart'] += cost
                else:
                    ad_by_product[mapped]['campaign'] += cost

    print(f"  총 광고비: {total_ad:,.0f}원")
else:
    print(f"[4/5] 광고보고서 없음, 스킵")

# ─── 원가 계산 (2월 피벗 기준 이익률 참조) ───
# 로켓배송은 총단가에서 원가를 빼서 이익 계산
# 원가 = 총단가 - 이익 인데, 이익률이 있으므로: 이익 = 총단가 × 이익률 / (1 + 이익률) 아님
# 실제로는: 이익 = 총단가 - (수량 × 개별원가) 방식
# 여기서는 원가 정보가 없으므로 총단가만 표시

# ─── 결과 출력 ───
print(f"[5/5] 결과 출력\n")
print(f"{'='*90}")
print(f"  쿠팡 로켓배송 매출 피벗 결과")
print(f"{'='*90}")
print(f"  총 {total_rows}행 처리 → {len(pivot)}개 상품 카테고리\n")

sorted_items = sorted(pivot.items(), key=lambda x: x[0])

if total_ad > 0:
    print(f"{'#':>3}  {'상품명':<30} {'수량':>8} {'총단가':>14} {'광고비(스마트)':>14} {'광고비(독립)':>14}")
    print(f"{'─'*3}  {'─'*30} {'─'*8} {'─'*14} {'─'*14} {'─'*14}")
else:
    print(f"{'#':>3}  {'상품명':<30} {'수량':>8} {'총단가':>14}")
    print(f"{'─'*3}  {'─'*30} {'─'*8} {'─'*14}")

total_qty = 0
total_price = 0
total_smart = 0
total_campaign = 0

for i, (name, data) in enumerate(sorted_items, 1):
    q = data['qty']
    p = data['total_price']
    smart = ad_by_product[name]['smart']
    campaign = ad_by_product[name]['campaign']
    total_qty += q
    total_price += p
    total_smart += smart
    total_campaign += campaign

    if total_ad > 0:
        s_str = f"{smart:>14,.0f}" if smart > 0 else f"{'':>14}"
        c_str = f"{campaign:>14,.0f}" if campaign > 0 else f"{'':>14}"
        print(f"{i:>3}  {name:<30} {q:>8,} {p:>14,.0f} {s_str} {c_str}")
    else:
        print(f"{i:>3}  {name:<30} {q:>8,} {p:>14,.0f}")

if total_ad > 0:
    print(f"{'─'*3}  {'─'*30} {'─'*8} {'─'*14} {'─'*14} {'─'*14}")
    print(f"{'':>3}  {'총합계':<30} {total_qty:>8,} {total_price:>14,.0f} {total_smart:>14,.0f} {total_campaign:>14,.0f}")
    print(f"\n  총 광고비: {total_smart + total_campaign:,.0f}원 (스마트: {total_smart:,.0f} + 독립: {total_campaign:,.0f})")
else:
    print(f"{'─'*3}  {'─'*30} {'─'*8} {'─'*14}")
    print(f"{'':>3}  {'총합계':<30} {total_qty:>8,} {total_price:>14,.0f}")

if unmapped:
    unique_unmapped = sorted(set(unmapped))
    print(f"\n{'='*90}")
    print(f"  ⚠️ 미매핑 상품: {len(unique_unmapped)}개 ({len(unmapped)}행)")
    print(f"{'='*90}")
    for name in unique_unmapped:
        cnt = unmapped.count(name)
        print(f"  - {name[:70]} ({cnt}건)")
else:
    print(f"\n  ✅ 미매핑 상품 없음 — 전체 매핑 성공!")

print()
