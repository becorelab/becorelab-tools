"""
월별 매출 정산 마스터 스크립트
전 채널 원본 파일 → 자동 처리 → 종합 매핑 파일 생성

사용법: python 정산.py 2026-04
       python 정산.py 2026-04 --retry  (미매핑 수정 후 재실행)

폴더 구조:
  ★월별매출정산 with Claude Cowork/
    4월 매출 정산/
      04월 카페24/         ← 원본 파일 넣기
      04월 스마트스토어/     ← 원본 파일 넣기
      04월 로켓배송/        ← ...
      ...
"""
import sys, io, os, json, re, glob
from collections import defaultdict
from datetime import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(SCRIPT_DIR)

# ─── 공통 유틸 ───
def sf(v):
    if v is None or v == '' or v == 'None': return 0.0
    s = str(v).replace(',', '').replace('원', '').strip()
    try: return float(s)
    except: return 0.0

def si(v):
    if v is None or v == '' or v == 'None': return 0
    try: return int(float(str(v).replace(',', '')))
    except: return 0

def col_idx(headers, keyword):
    for i, h in enumerate(headers):
        if h and keyword in str(h):
            return i
    return None

def load_name_map(name):
    path = os.path.join(PARENT_DIR, f'{name}_name_map.json')
    if not os.path.exists(path):
        path = os.path.join(PARENT_DIR, f'{name}_name_map_temp.json')
    if os.path.exists(path):
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}

def read_file(path):
    """xls/xlsx/csv 읽어서 (headers, rows) 반환"""
    ext = os.path.splitext(path)[1].lower()
    if ext == '.csv':
        import csv
        for enc in ['utf-8-sig', 'utf-8', 'euc-kr', 'cp949']:
            try:
                with open(path, 'r', encoding=enc) as f:
                    reader = csv.reader(f)
                    headers = next(reader)
                    rows = list(reader)
                return headers, rows
            except: continue
    elif ext == '.xls':
        import xlrd
        try:
            wb = xlrd.open_workbook(path)
            ws = wb.sheet_by_index(0)
            headers = [ws.cell_value(0, c) for c in range(ws.ncols)]
            rows = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(1, ws.nrows)]
            return headers, rows
        except:
            # .xls 확장자지만 실제 xlsx인 경우
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb[wb.sheetnames[0]]
            headers = [cell.value for cell in ws[1]]
            rows = [list(r) for r in ws.iter_rows(min_row=2, values_only=True)]
            return headers, rows
    else:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = [cell.value for cell in ws[1]]
        rows = [list(r) for r in ws.iter_rows(min_row=2, values_only=True)]
        return headers, rows

def find_files(folder, *patterns):
    """폴더에서 패턴에 맞는 파일 찾기"""
    result = []
    if not os.path.exists(folder):
        return result
    for f in os.listdir(folder):
        fp = os.path.join(folder, f)
        if os.path.isfile(fp) and not f.startswith('~') and not f.startswith('.'):
            if any(f.endswith(ext) for ext in ['.xlsx', '.xls', '.csv']):
                if not patterns or any(p.lower() in f.lower() for p in patterns):
                    result.append(fp)
    return sorted(result)

def match_name(name, name_map):
    """품명리스트에서 매칭"""
    if name in name_map:
        return name_map[name]
    for mk, mv in name_map.items():
        if name and (name in mk or mk in name):
            return mv
    return None


# ─── 채널별 처리 함수 ───

def process_cafe24(folder, name_map):
    """카페24: 결제총액 + 쿠폰/앱할인 차감"""
    files = find_files(folder)
    if not files: return None
    headers, rows = read_file(files[0])
    IDX_NAME = col_idx(headers, '주문상품명(옵션포함)') or col_idx(headers, '주문상품명')
    IDX_OPT_PRICE = col_idx(headers, '옵션+판매가')
    IDX_QTY = col_idx(headers, '수량')
    IDX_SHIP = col_idx(headers, '총 배송비(KRW)')
    IDX_ORDER = col_idx(headers, '주문번호')
    IDX_COUPON = col_idx(headers, '주문서 쿠폰 할인금액')
    IDX_APP = col_idx(headers, '앱 상품할인 금액(최종)')
    IDX_REFUND = col_idx(headers, '실제 환불금액')

    pivot = defaultdict(lambda: {'qty': 0, 'amount': 0, 'shipping': 0})
    unmapped = []
    order_ship = set()
    coupon_orders = set()
    sum_coupon = sum_app = sum_refund = 0.0

    for row in rows:
        name = str(row[IDX_NAME] or '').strip()
        final = match_name(name, name_map)
        if not final:
            unmapped.append(name)
            continue
        qty = si(row[IDX_QTY])
        amount = sf(row[IDX_OPT_PRICE]) * qty
        order = str(row[IDX_ORDER] or '').strip()
        ship = 0.0
        if order and order not in order_ship:
            ship = sf(row[IDX_SHIP])
            order_ship.add(order)
        pivot[final]['qty'] += qty
        pivot[final]['amount'] += amount
        pivot[final]['shipping'] += ship
        if IDX_REFUND: sum_refund += sf(row[IDX_REFUND])
        if IDX_COUPON and order not in coupon_orders:
            sum_coupon += sf(row[IDX_COUPON])
            coupon_orders.add(order)
        if IDX_APP: sum_app += sf(row[IDX_APP])

    return {'pivot': dict(pivot), 'unmapped': unmapped,
            'deductions': {'환불': sum_refund, '쿠폰': sum_coupon, '앱할인': sum_app},
            'note': '⚠️ 환불금액 수동 확인 필요'}


def process_smartstore(folder, name_map):
    """스마트스토어: 정산기준금액 - 수수료"""
    settle_file = find_files(folder, '결제정산', '일별상세')
    confirm_file = find_files(folder, '구매확정')
    order_file = find_files(folder, '주문조회')
    if not settle_file: return None

    # 옵션 로드
    import openpyxl
    option_map = {}
    for fpath, label in [(confirm_file[0] if confirm_file else None, '구매확정'),
                          (order_file[0] if order_file else None, '주문조회')]:
        if not fpath: continue
        try:
            h, rows = read_file(fpath)
            idx_ino = col_idx(h, '상품주문번호')
            idx_nm = col_idx(h, '상품명')
            idx_opt = col_idx(h, '옵션정보')
            for row in rows:
                ino = str(row[idx_ino] or '').strip()
                nm = str(row[idx_nm] or '').strip()
                opt = str(row[idx_opt] or '').strip() if idx_opt else ''
                if ino and nm:
                    option_map[ino] = f"{nm}{opt}" if opt else nm
        except: pass

    headers, rows = read_file(settle_file[0])
    IDX_TYPE = col_idx(headers, '구분')
    IDX_ITEM = col_idx(headers, '상품주문번호')
    IDX_NAME = col_idx(headers, '상품명')
    IDX_SETTLE = col_idx(headers, '정산기준금액')
    IDX_NPAY = col_idx(headers, 'Npay 수수료')
    IDX_SALES = col_idx(headers, '매출연동 수수료')
    IDX_INSTALL = col_idx(headers, '무이자할부 수수료')

    pivot = defaultdict(lambda: {'qty': 0, 'amount': 0, 'shipping': 0})
    unmapped = []
    ship_total = 0

    for row in rows:
        rtype = str(row[IDX_TYPE] or '').strip()
        net = sf(row[IDX_SETTLE]) + sf(row[IDX_NPAY]) + sf(row[IDX_SALES])
        if IDX_INSTALL: net += sf(row[IDX_INSTALL])
        if rtype == '배송비':
            ship_total += net
            continue
        if rtype != '상품주문': continue

        name = str(row[IDX_NAME] or '').strip()
        item_no = str(row[IDX_ITEM] or '').strip()
        final = match_name(name, name_map)
        if not final and item_no in option_map:
            final = match_name(option_map[item_no], name_map)
        if not final:
            unmapped.append(name)
            continue
        pivot[final]['qty'] += 1
        pivot[final]['amount'] += net

    # 배송비를 별도로
    if ship_total > 0:
        pivot['배송비'] = {'qty': 0, 'amount': ship_total, 'shipping': 0}

    return {'pivot': dict(pivot), 'unmapped': unmapped}


def process_gmarket_auction(folder, name_map, channel_name):
    """지마켓/옥션: 판매자 최종정산금 + 배송비"""
    sales_file = find_files(folder, '매출기준_상품판매', '상세내역')
    ship_file = find_files(folder, '배송비')
    confirm_file = find_files(folder, '구매결정', '구매확정')
    if not sales_file: return None

    headers, rows = read_file(sales_file[0])
    IDX_NAME = col_idx(headers, '상품명')
    IDX_QTY = col_idx(headers, '주문수량')
    IDX_SETTLE = col_idx(headers, '판매자 최종정산금')
    if IDX_NAME is None or IDX_SETTLE is None: return None

    # 배송비
    total_ship = 0
    if ship_file:
        h_s, rows_s = read_file(ship_file[0])
        ship_idx = None
        for i, c in enumerate(h_s):
            if c and '정산' in str(c) and '배송' in str(c):
                ship_idx = i; break
        if ship_idx:
            for row in rows_s:
                total_ship += sf(row[ship_idx])

    pivot = defaultdict(lambda: {'qty': 0, 'amount': 0, 'shipping': 0})
    unmapped = []
    for row in rows:
        name = str(row[IDX_NAME] or '').strip()
        if not name: continue
        final = match_name(name, name_map)
        if not final:
            unmapped.append(name)
            continue
        pivot[final]['qty'] += si(row[IDX_QTY])
        pivot[final]['amount'] += sf(row[IDX_SETTLE])

    return {'pivot': dict(pivot), 'unmapped': unmapped,
            'extra': {'배송비 합계': total_ship},
            'note': '⚠️ 배송비는 상품별 배분 안 됨, 합계만 제공'}


def process_11st(folder, name_map):
    """11번가: 정산금액 + 배송비 + 후불광고비 (헤더 Row 6)"""
    files = find_files(folder)
    if not files: return None

    ext = os.path.splitext(files[0])[1].lower()
    if ext == '.xls':
        import xlrd
        try:
            wb = xlrd.open_workbook(files[0])
        except:
            import openpyxl
            wb = openpyxl.load_workbook(files[0], data_only=True)
            ws = wb[wb.sheetnames[0]]
            header_row = 6
            for r in range(1, 11):
                if ws.cell(row=r, column=1).value == 'NO':
                    header_row = r; break
            headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column+1)]
            rows = []
            for r in range(header_row+1, ws.max_row+1):
                rows.append([ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)])
            # continue below
            IDX_NAME = col_idx(headers, '상품명')
            IDX_QTY = col_idx(headers, '수량')
            IDX_SETTLE = col_idx(headers, '정산금액')
            IDX_SHIP = col_idx(headers, '선결제배송비')
            IDX_AD = col_idx(headers, '후불광고비')

            pivot = defaultdict(lambda: {'qty': 0, 'amount': 0, 'shipping': 0})
            unmapped = []
            total_ad = 0
            for row in rows:
                name = str(row[IDX_NAME] or '').strip()
                if not name: continue
                final = match_name(name, name_map)
                if not final:
                    unmapped.append(name)
                    continue
                pivot[final]['qty'] += si(row[IDX_QTY])
                pivot[final]['amount'] += sf(row[IDX_SETTLE])
                pivot[final]['shipping'] += sf(row[IDX_SHIP]) if IDX_SHIP else 0
                if IDX_AD: total_ad += sf(row[IDX_AD])
            return {'pivot': dict(pivot), 'unmapped': unmapped, 'extra': {'후불광고비': total_ad}}

        ws = wb.sheet_by_index(0)
        header_row = 5
        for r in range(10):
            if ws.cell_value(r, 0) == 'NO':
                header_row = r; break
        headers = [ws.cell_value(header_row, c) for c in range(ws.ncols)]
        rows = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(header_row+1, ws.nrows)]
    else:
        import openpyxl
        wb = openpyxl.load_workbook(files[0], data_only=True)
        ws = wb[wb.sheetnames[0]]
        header_row = 6
        for r in range(1, 11):
            if ws.cell(row=r, column=1).value == 'NO':
                header_row = r; break
        headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column+1)]
        rows = []
        for r in range(header_row+1, ws.max_row+1):
            rows.append([ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)])

    IDX_NAME = col_idx(headers, '상품명')
    IDX_QTY = col_idx(headers, '수량')
    IDX_SETTLE = col_idx(headers, '정산금액')
    IDX_SHIP = col_idx(headers, '선결제배송비')
    IDX_AD = col_idx(headers, '후불광고비')

    pivot = defaultdict(lambda: {'qty': 0, 'amount': 0, 'shipping': 0})
    unmapped = []
    total_ad = 0
    for row in rows:
        name = str(row[IDX_NAME] or '').strip()
        if not name: continue
        final = match_name(name, name_map)
        if not final:
            unmapped.append(name)
            continue
        pivot[final]['qty'] += si(row[IDX_QTY])
        pivot[final]['amount'] += sf(row[IDX_SETTLE])
        pivot[final]['shipping'] += sf(row[IDX_SHIP]) if IDX_SHIP else 0
        if IDX_AD: total_ad += sf(row[IDX_AD])

    return {'pivot': dict(pivot), 'unmapped': unmapped, 'extra': {'후불광고비': total_ad}}


def process_simple_settle(folder, name_map, settle_col='판매정산금액'):
    """카카오선물하기/오늘의집 등 단순 정산"""
    files = find_files(folder)
    if not files: return None
    headers, rows = read_file(files[0])
    # 헤더가 None이면 xls→xlsx 변환 시도
    if all(h is None for h in headers):
        return None
    IDX_NAME = col_idx(headers, '상품명')
    IDX_QTY = col_idx(headers, '수량') or col_idx(headers, '개수')
    IDX_SETTLE = col_idx(headers, settle_col) or col_idx(headers, '주문 정산금액') or col_idx(headers, '정산금액')
    if IDX_NAME is None or IDX_SETTLE is None: return None

    pivot = defaultdict(lambda: {'qty': 0, 'amount': 0, 'shipping': 0})
    unmapped = []
    for row in rows:
        name = str(row[IDX_NAME] or '').strip()
        if not name: continue
        qty = si(row[IDX_QTY])
        settle = sf(row[IDX_SETTLE])
        final = match_name(name, name_map)
        if not final:
            if '배송비' in name: final = '배송비'
            elif '할인' in name: continue  # 할인 항목 스킵
            else:
                unmapped.append(name)
                continue
        pivot[final]['qty'] += qty
        pivot[final]['amount'] += settle

    return {'pivot': dict(pivot), 'unmapped': unmapped}


def process_ssg(folder, name_map):
    """신세계: 정산금액(VAT포함) + 배송비(VAT포함)"""
    files = find_files(folder)
    if not files: return None
    headers, rows = read_file(files[0])
    IDX_NAME = col_idx(headers, '상품명')
    IDX_QTY = col_idx(headers, '수량')
    IDX_SETTLE = col_idx(headers, '정산금액(VAT포함)') or col_idx(headers, '정산금액')
    if IDX_NAME is None or IDX_SETTLE is None: return None
    IDX_SHIP = None
    for i, h in enumerate(headers):
        if h and '배송비' in str(h) and 'VAT' in str(h):
            IDX_SHIP = i; break
    if IDX_SHIP is None:
        IDX_SHIP = col_idx(headers, '배송비')

    pivot = defaultdict(lambda: {'qty': 0, 'amount': 0, 'shipping': 0})
    unmapped = []
    for row in rows:
        name = str(row[IDX_NAME] or '').strip()
        if not name: continue
        final = match_name(name, name_map)
        if not final:
            unmapped.append(name)
            continue
        pivot[final]['qty'] += si(row[IDX_QTY])
        pivot[final]['amount'] += sf(row[IDX_SETTLE])
        pivot[final]['shipping'] += sf(row[IDX_SHIP]) if IDX_SHIP else 0

    return {'pivot': dict(pivot), 'unmapped': unmapped}


def process_rocket(folder, name_map):
    """로켓배송: 총단가 기준"""
    data_file = find_files(folder, 'Coupang', 'Stocked', 'Data')
    ad_file = find_files(folder, '광고보고서')
    if not data_file: return None

    import openpyxl
    wb = openpyxl.load_workbook(data_file[0], data_only=True)
    ws = wb['data'] if 'data' in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [cell.value for cell in ws[1]]
    IDX_SKU = col_idx(headers, 'SKU명')
    IDX_QTY = col_idx(headers, '수량')
    IDX_PRICE = col_idx(headers, '총단가')

    pivot = defaultdict(lambda: {'qty': 0, 'amount': 0, 'shipping': 0})
    unmapped = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None: continue
        sku = str(row[IDX_SKU] or '').replace('\xa0', ' ').strip()
        if not sku: continue
        final = match_name(sku, name_map)
        if not final:
            unmapped.append(sku)
            continue
        pivot[final]['qty'] += si(row[IDX_QTY])
        pivot[final]['amount'] += sf(row[IDX_PRICE])

    # 광고비 (별도 정보)
    total_ad = 0
    if ad_file:
        try:
            wb_ad = openpyxl.load_workbook(ad_file[0], data_only=True)
            ws_ad = wb_ad[wb_ad.sheetnames[0]]
            ad_h = [cell.value for cell in ws_ad[1]]
            idx_cost = col_idx(ad_h, '광고비')
            for row in ws_ad.iter_rows(min_row=2, values_only=True):
                total_ad += sf(row[idx_cost])
        except: pass

    extra = {'광고비': total_ad} if total_ad > 0 else {}
    return {'pivot': dict(pivot), 'unmapped': unmapped, 'extra': extra}


def process_dispatch(folder, name_map, channel_filter, cost_map):
    """에드가/두버/지엠홀딩스: 발송내역에서 필터 + 공급가 적용"""
    import openpyxl, xlrd
    SHIP_BASE = os.path.join(SCRIPT_DIR, '3월 수동 발주', '3월 발송 내역')
    # 동적 경로: 월 기반으로 변경 가능하도록
    # 현재는 발송내역 폴더를 직접 지정

    # 실제로는 folder 안에 발송내역이 있거나, 별도 경로를 참조
    # 여기서는 기존 로직 유지
    all_items = []
    if not os.path.exists(SHIP_BASE):
        return None

    month_num = 3  # TODO: 파라미터로 받기

    for fname in sorted(os.listdir(SHIP_BASE)):
        date_match = re.search(r'_(\d{4})', fname)
        if date_match and int(date_match.group(1)[:2]) != month_num:
            continue
        fpath = os.path.join(SHIP_BASE, fname)
        ext = os.path.splitext(fname)[1].lower()
        try:
            if ext == '.xlsx':
                wb = openpyxl.load_workbook(fpath, data_only=True)
                ws = wb[wb.sheetnames[0]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if channel_filter.lower() in str(row[0] or '').lower():
                        all_items.append({
                            'order': str(row[1] or '').strip(),
                            'name': str(row[7] or '').strip(),
                            'option': str(row[8] or '').strip(),
                            'qty': si(row[9])
                        })
            elif ext == '.xls':
                wb = xlrd.open_workbook(fpath)
                ws = wb.sheet_by_index(0)
                for r in range(1, ws.nrows):
                    if channel_filter.lower() in str(ws.cell_value(r, 0)).lower():
                        all_items.append({
                            'order': str(ws.cell_value(r, 1)).strip(),
                            'name': str(ws.cell_value(r, 7)).strip(),
                            'option': str(ws.cell_value(r, 8)).strip(),
                            'qty': si(ws.cell_value(r, 9))
                        })
        except: pass

    if not all_items:
        return None

    pivot = defaultdict(lambda: {'qty': 0, 'amount': 0, 'shipping': 0})
    unmapped = []
    shipped_orders = set()

    for item in all_items:
        full = f"{item['name']}{item['option']}" if item['option'] else item['name']
        final = match_name(full, name_map) or match_name(item['name'], name_map)
        if not final:
            unmapped.append(f"{item['name']} | {item['option']}")
            continue

        # 공급가
        supply = cost_map.get(final, 0)
        if not supply:
            m = re.match(r'^(.+?)\s*(\d+)개$', final)
            if m:
                supply = cost_map.get(f"{m.group(1)}*{m.group(2)}", 0)
                if not supply:
                    supply = cost_map.get(f"{m.group(1)} 믹스 {m.group(2)}개", 0)

        pivot[final]['qty'] += item['qty']
        pivot[final]['amount'] += supply * item['qty']

    return {'pivot': dict(pivot), 'unmapped': unmapped,
            'note': '⚠️ 배송비는 묶음배송 규칙으로 수동 확인 필요'}


# ─── 메인 실행 ───
def main():
    if len(sys.argv) < 2:
        print("사용법: python 정산.py 2026-04")
        print("       python 정산.py 2026-04 --retry")
        sys.exit(1)

    month_str = sys.argv[1]  # 예: 2026-04
    year, month = month_str.split('-')
    month_num = int(month)
    month_kr = f"{month_num}월"
    base_dir = os.path.join(SCRIPT_DIR, f'{month_kr} 매출 정산')

    print(f"{'='*60}")
    print(f"  📊 {year}년 {month_kr} 매출 정산 시작")
    print(f"{'='*60}")
    import openpyxl
    print(f"  경로: {base_dir}\n")

    if not os.path.exists(base_dir):
        os.makedirs(base_dir)
        print(f"  폴더 생성됨. 채널별 하위 폴더에 원본 파일을 넣어주세요!")
        sys.exit(0)

    # 채널별 결과 엑셀이 이미 있으면 그걸 읽고, 없으면 원본에서 처리
    mm = f"{month_num:02d}"

    # 1단계: 이미 처리된 결과 파일 읽기
    all_results = {}
    all_unmapped = {}

    channel_folders = [
        ('카페24', f'{mm}월 카페24'),
        ('스마트스토어', f'{mm}월 스마트스토어'),
        ('카카오선물하기', f'{mm}월 카카오 선물하기'),
        ('로켓배송', f'{mm}월 로켓배송'),
        ('옥션', f'{mm}월 옥션'),
        ('지마켓', f'{mm}월 지마켓'),
        ('11번가', f'{mm}월 11번가'),
        ('신세계', f'{mm}월 신세계'),
        ('오늘의집', f'{mm}월 오늘의집'),
        ('에드가', f'{mm}월 에드가'),
        ('두버', f'{mm}월 두버'),
        ('지엠홀딩스', f'{mm}월 지엠홀딩스'),
    ]

    for ch_name, folder_name in channel_folders:
        folder = os.path.join(base_dir, folder_name)
        if not os.path.exists(folder):
            print(f"  ⏭️  {ch_name}: 폴더 없음")
            continue

        # 결과 파일 찾기 (03월 XXX.xlsx 패턴)
        result_file = None
        for f in os.listdir(folder):
            if f.startswith(f'{mm}월') and f.endswith('.xlsx') and '정산결과' not in f and '매핑' not in f and not f.startswith('~'):
                result_file = os.path.join(folder, f)
                break
        # 정산결과 파일도 체크
        if not result_file:
            for f in os.listdir(folder):
                if f.endswith('.xlsx') and ('정산결과' in f) and not f.startswith('~'):
                    result_file = os.path.join(folder, f)
                    break

        if result_file:
            try:
                wb = openpyxl.load_workbook(result_file, data_only=True)
                ws = wb['피벗'] if '피벗' in wb.sheetnames else wb[wb.sheetnames[0]]
                pivot = {}
                skip = {'총합계', '(비어 있음)', '행 레이블', None, ''}
                for row in ws.iter_rows(min_row=2, values_only=True):
                    name = row[0]
                    if name and str(name).strip() not in skip:
                        try:
                            name = str(name).strip()
                            # 컬럼 위치는 파일마다 다를 수 있으므로 헤더 확인
                            qty = int(row[1]) if row[1] and not isinstance(row[1], str) else 0
                            settle = float(row[2]) if row[2] and not isinstance(row[2], str) else 0
                            ship = float(row[3]) if len(row) > 3 and row[3] and not isinstance(row[3], str) else 0
                            pivot[name] = {'qty': qty, 'amount': round(settle), 'shipping': round(ship)}
                        except (ValueError, TypeError):
                            continue
                if pivot:
                    all_results[ch_name] = {'pivot': pivot, 'unmapped': []}
                    total = sum(d['amount'] for d in pivot.values())
                    print(f"  ✅ {len(pivot)}개 상품, {total:,.0f}원  {ch_name} (결과 파일)")
                else:
                    print(f"  ⚠️  {ch_name}: 결과 파일 비어있음")
            except Exception as e:
                print(f"  ❌ {ch_name}: 결과 파일 읽기 실패 - {e}")
        else:
            # 2단계: 원본에서 처리 시도
            map_names = {'카페24': 'cafe24', '스마트스토어': 'smartstore', '로켓배송': 'rocket',
                         '카카오선물하기': 'kakao_gift', '옥션': 'auction', '지마켓': 'gmarket',
                         '11번가': '11st', '신세계': 'ssg', '오늘의집': 'ohouse'}
            processors = {
                '카페24': process_cafe24,
                '스마트스토어': process_smartstore,
                '로켓배송': process_rocket,
                '신세계': process_ssg,
            }
            map_name = map_names.get(ch_name)
            processor = processors.get(ch_name)
            if processor and map_name:
                name_map = load_name_map(map_name)
                if name_map:
                    try:
                        result = processor(folder, name_map)
                        if result:
                            all_results[ch_name] = result
                            total = sum(d.get('amount', 0) for d in result['pivot'].values())
                            print(f"  ✅ {len(result['pivot'])}개 상품, {total:,.0f}원  {ch_name} (원본 처리)")
                        else:
                            print(f"  ⏭️  {ch_name}: 처리 결과 없음")
                    except Exception as e:
                        print(f"  ❌ {ch_name}: {e}")
                else:
                    print(f"  ⚠️  {ch_name}: 품명리스트 없음")
            else:
                print(f"  ⏭️  {ch_name}: 파일 없음 (원본 또는 결과)")

    # 종합 매핑 파일 생성
    print(f"\n{'='*60}")
    print(f"  📋 종합 매핑 파일 생성")
    print(f"{'='*60}")

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    HF = PatternFill(start_color='1B2A4A', end_color='1B2A4A', fill_type='solid')
    HFont = Font(bold=True, size=11, color='FFFFFF')
    CH_FILL = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
    CH_FONT = Font(bold=True, size=12, color='FFFFFF')
    MATCH_FILL = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    WARN_FILL = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
    NUM_FONT = Font(size=10, name='Consolas')
    BOLD_NUM = Font(bold=True, size=11, name='Consolas', color='1565C0')
    thin = Border(bottom=Side(style='thin', color='E0E0E0'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{month_kr} 정산 종합'

    headers_out = ['채널', '상품명', '판매수량', '매출액/정산금', '배송비', '비고']
    for c, h in enumerate(headers_out, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HFont; cell.fill = HF; cell.alignment = Alignment(horizontal='center')

    r = 2
    for ch_name, result in all_results.items():
        # 채널 헤더
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1, value=f"📦 {ch_name}").font = CH_FONT
        ws.cell(row=r, column=1).fill = CH_FILL
        r += 1

        pivot = result.get('pivot', {})
        note = result.get('note', '')
        extra = result.get('extra', {})

        for name, data in sorted(pivot.items()):
            ws.cell(row=r, column=1, value=ch_name)
            ws.cell(row=r, column=2, value=name)
            ws.cell(row=r, column=3, value=data.get('qty', 0)).number_format = '#,##0'
            ws.cell(row=r, column=4, value=round(data.get('amount', 0))).number_format = '#,##0'
            ship = data.get('shipping', 0)
            if ship: ws.cell(row=r, column=5, value=round(ship)).number_format = '#,##0'
            for c2 in range(1, 7):
                ws.cell(row=r, column=c2).fill = MATCH_FILL
                ws.cell(row=r, column=c2).border = thin
            r += 1

        # 채널 소계
        total_amt = sum(d.get('amount', 0) for d in pivot.values())
        total_ship = sum(d.get('shipping', 0) for d in pivot.values())
        ws.cell(row=r, column=2, value='소계').font = Font(bold=True)
        ws.cell(row=r, column=4, value=round(total_amt)).font = BOLD_NUM
        ws.cell(row=r, column=4).number_format = '#,##0'
        if total_ship:
            ws.cell(row=r, column=5, value=round(total_ship)).font = BOLD_NUM
            ws.cell(row=r, column=5).number_format = '#,##0'

        # 비고/추가정보
        notes = []
        if note: notes.append(note)
        for k, v in extra.items():
            notes.append(f"{k}: {v:,.0f}")
        if notes:
            ws.cell(row=r, column=6, value=' / '.join(notes)).font = Font(size=9, color='E53935')
        r += 2

    # 미매핑 시트
    if all_unmapped:
        ws_um = wb.create_sheet('미매핑')
        ws_um.cell(row=1, column=1, value='채널').font = HFont
        ws_um.cell(row=1, column=1).fill = HF
        ws_um.cell(row=1, column=2, value='미매핑 상품명').font = HFont
        ws_um.cell(row=1, column=2).fill = HF
        ws_um.cell(row=1, column=3, value='매핑 입력').font = HFont
        ws_um.cell(row=1, column=3).fill = HF
        ur = 2
        for ch, names in all_unmapped.items():
            for name in names:
                ws_um.cell(row=ur, column=1, value=ch)
                ws_um.cell(row=ur, column=2, value=name)
                ws_um.cell(row=ur, column=2).fill = WARN_FILL
                ur += 1
        ws_um.column_dimensions['A'].width = 18
        ws_um.column_dimensions['B'].width = 70
        ws_um.column_dimensions['C'].width = 30

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 40
    ws.freeze_panes = 'A2'

    outpath = os.path.join(base_dir, f'★{month_kr} 정산 종합.xlsx')
    wb.save(outpath)
    print(f"\n  💾 저장: {outpath}")

    if all_unmapped:
        total_um = sum(len(v) for v in all_unmapped.values())
        print(f"\n  ⚠️ 미매핑 {total_um}개 — '미매핑' 시트에서 확인 후 품명리스트 업데이트 → 재실행")
    else:
        print(f"\n  ✅ 전 채널 매핑 완료!")

    print(f"\n{'='*60}")
    print(f"  완료! 종합 파일을 확인하고 정산시트에 복붙해주세요")
    print(f"{'='*60}\n")


if __name__ == '__main__':
    main()
