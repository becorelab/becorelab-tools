"""
에드가 매출 정산 스크립트
바이피엘 출고요청 파일 합치기 → 에드가 필터 → 공급가 적용 → 피벗

사용법: python edgar_settlement.py
"""
import sys, io, os, json, re
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(SCRIPT_DIR)

def sf(v):
    if v is None or v == '': return 0.0
    return float(str(v).replace(',', ''))

def si(v):
    if v is None or v == '': return 0
    return int(float(str(v).replace(',', '')))

# 1. 공급가 로드
import openpyxl
wb_cost = openpyxl.load_workbook('N:/개인/Becorelab/15. 물류/04. 일일발송내역/★거래처공급가(공구).xlsx', data_only=True)
ws_cost = wb_cost[wb_cost.sheetnames[0]]
cost_map = {}
for r in range(195, 214):
    name = str(ws_cost.cell(row=r, column=3).value or '').strip()
    supply = sf(ws_cost.cell(row=r, column=5).value)
    ship = sf(ws_cost.cell(row=r, column=6).value)
    if name and supply > 0:
        cost_map[name] = {'price': supply, 'ship': ship}
print(f"[1/5] 공급가 로드: {len(cost_map)}개")

# 2. 품명리스트 로드
edgar_map_path = os.path.join(PARENT_DIR, 'edgar_name_map.json')
if os.path.exists(edgar_map_path):
    with open(edgar_map_path, 'r', encoding='utf-8') as f:
        edgar_map = json.load(f)
else:
    wb_ref = openpyxl.load_workbook('N:/개인/Becorelab/03. 영업/20. 월별 매출정산/2026.02/02월 에드가.xlsx', data_only=True)
    edgar_map = {}
    ws_nm = wb_ref['품명리스트']
    for row in ws_nm.iter_rows(min_row=2, values_only=True):
        orig = str(row[1] or '').strip()
        mapped = str(row[2] or '').strip()
        if orig and mapped:
            edgar_map[orig] = mapped
    with open(edgar_map_path, 'w', encoding='utf-8') as f:
        json.dump(edgar_map, f, ensure_ascii=False, indent=2)
print(f"[2/5] 품명리스트: {len(edgar_map)}개")

# 3. 발송 파일 합치기
BASE = os.path.join(SCRIPT_DIR, '3월 수동 발주', '3월 발송 내역')
import xlrd

all_edgar = []
files = sorted(os.listdir(BASE))
print(f"[3/5] 발송 파일 스캔: {len(files)}개")

for fname in files:
    fpath = os.path.join(BASE, fname)
    ext = os.path.splitext(fname)[1].lower()

    # 3월만 (파일명에서 03xx 또는 날짜 확인)
    date_match = re.search(r'_(\d{4})', fname)
    if date_match:
        mmdd = date_match.group(1)
        month = int(mmdd[:2])
        if month != 3:
            continue

    try:
        if ext == '.xlsx':
            wb = openpyxl.load_workbook(fpath, data_only=True)
            ws = wb[wb.sheetnames[0]]
            h = [cell.value for cell in ws[1]]
            idx_ch = next((i for i, c in enumerate(h) if c and '채널' in str(c)), 0)
            idx_nm = next((i for i, c in enumerate(h) if c and '상품명' in str(c)), 7)
            idx_op = next((i for i, c in enumerate(h) if c and '옵션' in str(c)), 8)
            idx_qt = next((i for i, c in enumerate(h) if c and '수량' in str(c)), 9)
            for row in ws.iter_rows(min_row=2, values_only=True):
                ch = str(row[idx_ch] or '').strip()
                if '에드가' in ch:
                    all_edgar.append({
                        'file': fname,
                        'name': str(row[idx_nm] or '').strip(),
                        'option': str(row[idx_op] or '').strip(),
                        'qty': si(row[idx_qt]),
                    })
        elif ext == '.xls':
            wb = xlrd.open_workbook(fpath)
            ws = wb.sheet_by_index(0)
            h = [ws.cell_value(0, c) for c in range(ws.ncols)]
            idx_ch = next((i for i, c in enumerate(h) if '채널' in str(c)), 0)
            idx_nm = next((i for i, c in enumerate(h) if '상품명' in str(c)), 7)
            idx_op = next((i for i, c in enumerate(h) if '옵션' in str(c)), 8)
            idx_qt = next((i for i, c in enumerate(h) if '수량' in str(c)), 9)
            for r in range(1, ws.nrows):
                ch = str(ws.cell_value(r, idx_ch)).strip()
                if '에드가' in ch:
                    all_edgar.append({
                        'file': fname,
                        'name': str(ws.cell_value(r, idx_nm)).strip(),
                        'option': str(ws.cell_value(r, idx_op)).strip(),
                        'qty': si(ws.cell_value(r, idx_qt)),
                    })
    except Exception as e:
        print(f"  ⚠️ {fname}: {e}")

print(f"  에드가 3월 발송: {len(all_edgar)}건")

# 4. 매핑 + 공급가 적용
print(f"[4/5] 매핑 + 공급가 적용...")
pivot = defaultdict(lambda: {'qty': 0, 'supply': 0, 'ship': 0})
unmapped = []
proc_rows = []

for item in all_edgar:
    name = item['name']
    option = item['option']
    qty = item['qty']
    full = f"{name}{option}" if option else name

    final = edgar_map.get(full) or edgar_map.get(name)
    if not final:
        for mk, mv in edgar_map.items():
            if name and (name in mk or mk in name):
                final = mv
                break

    if not final:
        unmapped.append(f"{name} | {option}")
        proc_rows.append({'file': item['file'], 'name': name, 'option': option, 'final': '', 'qty': qty, 'supply': 0, 'ship': 0})
        continue

    # 공급가표 키 변환: "건조기시트 2개" → "건조기시트*2"
    cost = cost_map.get(final, None)
    if not cost:
        import re
        m = re.match(r'^(.+?)\s*(\d+)개$', final)
        if m:
            cost_key = f"{m.group(1)}*{m.group(2)}"
            cost = cost_map.get(cost_key, None)
    supply = cost['price'] if cost else 0
    ship = cost['ship'] if cost else 0

    pivot[final]['qty'] += qty
    pivot[final]['supply'] += supply * qty
    pivot[final]['ship'] += ship
    proc_rows.append({'file': item['file'], 'name': name, 'option': option, 'final': final, 'qty': qty, 'supply': supply * qty, 'ship': ship})

# 5. 결과 출력
print(f"[5/5] 결과\n")
sorted_items = sorted(pivot.items())
tq = ts = tsh = 0
print(f"{'#':>3}  {'상품명':<25} {'수량':>6} {'공급가합계':>14} {'배송비':>10}")
print(f"{'─'*3}  {'─'*25} {'─'*6} {'─'*14} {'─'*10}")
for i, (name, data) in enumerate(sorted_items, 1):
    tq += data['qty']; ts += data['supply']; tsh += data['ship']
    print(f"{i:>3}  {name:<25} {data['qty']:>6,} {data['supply']:>14,.0f} {data['ship']:>10,.0f}")
print(f"{'─'*3}  {'─'*25} {'─'*6} {'─'*14} {'─'*10}")
print(f"{'':>3}  {'총합계':<25} {tq:>6,} {ts:>14,.0f} {tsh:>10,.0f}")
print(f"\n  공급가+배송비 = {ts + tsh:,.0f}")

if unmapped:
    unique = sorted(set(unmapped))
    print(f"\n⚠️ 미매핑 {len(unique)}개:")
    for u in unique:
        print(f"  - {u[:60]} ({unmapped.count(u)}건)")
else:
    print(f"\n✅ 미매핑 없음")

# 엑셀 출력
from openpyxl.styles import Font, PatternFill, Border, Side
HF = PatternFill(start_color='2D4A7A', end_color='2D4A7A', fill_type='solid')
HFont = Font(bold=True, size=11, color='FFFFFF')
SF = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
TF = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
BN = Font(bold=True, size=11, name='Consolas')
tb = Border(top=Side(style='medium', color='1B2A4A'), bottom=Side(style='medium', color='1B2A4A'))

wbo = openpyxl.Workbook()
wsp = wbo.active
wsp.title = '피벗'
wsp.merge_cells('A1:D1')
wsp['A1'] = '에드가 매출 피벗 (3월)'
wsp['A1'].font = Font(bold=True, size=14, color='1B2A4A')
for c, hd in enumerate(['행 레이블', '수량', '공급가합계', '배송비'], 1):
    cell = wsp.cell(row=3, column=c, value=hd)
    cell.font = HFont
    cell.fill = HF
r = 4
for name, data in sorted_items:
    wsp.cell(row=r, column=1, value=name)
    wsp.cell(row=r, column=2, value=data['qty']).number_format = '#,##0'
    wsp.cell(row=r, column=3, value=round(data['supply'])).number_format = '#,##0'
    wsp.cell(row=r, column=4, value=round(data['ship'])).number_format = '#,##0'
    r += 1
for c2 in range(1, 5):
    wsp.cell(row=r, column=c2).fill = SF
    wsp.cell(row=r, column=c2).border = tb
wsp.cell(row=r, column=1, value='총합계').font = Font(bold=True)
wsp.cell(row=r, column=2, value=tq).font = BN
wsp.cell(row=r, column=2).number_format = '#,##0'
wsp.cell(row=r, column=3, value=round(ts)).font = BN
wsp.cell(row=r, column=3).number_format = '#,##0'
wsp.cell(row=r, column=4, value=round(tsh)).font = BN
wsp.cell(row=r, column=4).number_format = '#,##0'
r += 2
for c2 in range(3, 5):
    wsp.cell(row=r, column=c2).fill = TF
    wsp.cell(row=r, column=c2).border = tb
wsp.cell(row=r, column=3, value='공급가+배송비').font = Font(bold=True, size=10, color='1565C0')
wsp.cell(row=r, column=4, value=round(ts + tsh)).font = Font(bold=True, size=12, color='1565C0', name='Consolas')
wsp.cell(row=r, column=4).number_format = '#,##0'
wsp.column_dimensions['A'].width = 30
wsp.column_dimensions['B'].width = 10
wsp.column_dimensions['C'].width = 16
wsp.column_dimensions['D'].width = 14

wp = wbo.create_sheet('가공')
for c, hd in enumerate(['파일', '상품명', '옵션', '최종상품명', '수량', '공급가', '배송비'], 1):
    cell = wp.cell(row=1, column=c, value=hd)
    cell.font = HFont
    cell.fill = HF
for i, pr in enumerate(proc_rows, 2):
    wp.cell(row=i, column=1, value=pr['file'])
    wp.cell(row=i, column=2, value=pr['name'])
    wp.cell(row=i, column=3, value=pr['option'])
    wp.cell(row=i, column=4, value=pr['final'])
    wp.cell(row=i, column=5, value=pr['qty']).number_format = '#,##0'
    wp.cell(row=i, column=6, value=round(pr['supply'])).number_format = '#,##0'
    wp.cell(row=i, column=7, value=round(pr['ship']) if pr['ship'] > 0 else None)
wp.column_dimensions['A'].width = 35
wp.column_dimensions['B'].width = 40
wp.column_dimensions['C'].width = 25
wp.column_dimensions['D'].width = 20
wp.column_dimensions['E'].width = 8
wp.column_dimensions['F'].width = 12
wp.column_dimensions['G'].width = 10

outdir = os.path.join(SCRIPT_DIR, '3월 매출 정산', '03월 에드가')
os.makedirs(outdir, exist_ok=True)
outpath = os.path.join(outdir, '03월 에드가.xlsx')
wbo.save(outpath)
print(f"\n엑셀 저장: {outpath}")
