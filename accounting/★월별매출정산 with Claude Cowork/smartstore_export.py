"""
스마트스토어 정산 결과 → 엑셀 출력 (피벗 + 가공 시트)

사용법: python smartstore_export.py <결제정산.xlsx> [구매확정.xlsx] [주문조회.xlsx]
"""
import sys, io, os, json, openpyxl
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(SCRIPT_DIR)
NAME_MAP_PATH = os.path.join(PARENT_DIR, 'smartstore_name_map.json')

with open(NAME_MAP_PATH, 'r', encoding='utf-8') as f:
    NAME_MAP = json.load(f)

def safe_float(val):
    if val is None or val == '' or val == 'None': return 0.0
    return float(str(val).replace(',', ''))

def col_idx(headers, keyword):
    for i, h in enumerate(headers):
        if h and keyword in str(h):
            return i
    return None

settle_path = sys.argv[1]
confirm_path = sys.argv[2] if len(sys.argv) > 2 else None
order_path = sys.argv[3] if len(sys.argv) > 3 else None

# 옵션 로드
option_map = {}
for path, label in [(confirm_path, '구매확정'), (order_path, '주문조회')]:
    if not path or not os.path.exists(path):
        continue
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    h = [cell.value for cell in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        ino = str(row[h.index('상품주문번호')] or '').strip()
        name = str(row[h.index('상품명')] or '').strip()
        opt = str(row[h.index('옵션정보')] or '').strip()
        if ino and name:
            option_map[ino] = f"{name}{opt}" if opt else name

# 결제정산 처리
wb_in = openpyxl.load_workbook(settle_path, data_only=True)
ws_in = wb_in[wb_in.sheetnames[0]]
headers = [cell.value for cell in ws_in[1]]
IDX_ORDER = col_idx(headers, '주문번호')
IDX_ITEM = col_idx(headers, '상품주문번호')
IDX_TYPE = col_idx(headers, '구분')
IDX_NAME = col_idx(headers, '상품명')
IDX_BUYER = col_idx(headers, '구매자명')
IDX_SETTLE = col_idx(headers, '정산기준금액')
IDX_NPAY = col_idx(headers, 'Npay 수수료')
IDX_SALES = col_idx(headers, '매출연동 수수료')
IDX_INSTALL = col_idx(headers, '무이자할부 수수료')
IDX_STATUS = col_idx(headers, '정산상태')

pivot = defaultdict(lambda: {'count': 0, 'net': 0})
proc_rows = []
shipping_total = 0

for row in ws_in.iter_rows(min_row=2, values_only=True):
    row_type = str(row[IDX_TYPE] or '').strip()
    item_no = str(row[IDX_ITEM] or '').strip()
    product_name = str(row[IDX_NAME] or '').strip()
    settle_amt = safe_float(row[IDX_SETTLE])
    npay = safe_float(row[IDX_NPAY])
    sales = safe_float(row[IDX_SALES])
    install = safe_float(row[IDX_INSTALL]) if IDX_INSTALL else 0
    net = settle_amt + npay + sales + install

    if row_type == '배송비':
        shipping_total += net
        proc_rows.append({
            'order': str(row[IDX_ORDER] or ''),
            'item': item_no, 'type': row_type,
            'name': '', 'final': '', 'settle': settle_amt,
            'npay': npay, 'sales': sales, 'net': net,
        })
        continue
    if row_type != '상품주문':
        continue

    final = NAME_MAP.get(product_name, None)
    if not final and item_no in option_map:
        final = NAME_MAP.get(option_map[item_no], None)
    if not final:
        for mk, mv in NAME_MAP.items():
            if product_name and product_name in mk:
                final = mv
                break
    if final:
        pivot[final]['count'] += 1
        pivot[final]['net'] += net

    proc_rows.append({
        'order': str(row[IDX_ORDER] or ''),
        'item': item_no, 'type': row_type,
        'name': product_name, 'final': final or '',
        'settle': settle_amt, 'npay': npay, 'sales': sales, 'net': net,
    })

# === 스타일 ===
HEADER_FILL = PatternFill(start_color='2D4A7A', end_color='2D4A7A', fill_type='solid')
HEADER_FONT = Font(bold=True, size=11, color='FFFFFF')
LIGHT_ROW = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
SUMMARY_FILL = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
TOTAL_FILL = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
TITLE_FONT = Font(bold=True, size=14, color='1B2A4A')
NUM_FONT = Font(size=10, name='Consolas')
BOLD_NUM = Font(bold=True, size=11, name='Consolas')
thin_border = Border(bottom=Side(style='thin', color='E0E0E0'))
thick_border = Border(top=Side(style='medium', color='1B2A4A'), bottom=Side(style='medium', color='1B2A4A'))

wb = openpyxl.Workbook()

# === 피벗 시트 ===
ws_pv = wb.active
ws_pv.title = '피벗'
ws_pv.sheet_properties.tabColor = '4CAF50'
ws_pv.merge_cells('A1:C1')
ws_pv['A1'] = '스마트스토어 매출 피벗'
ws_pv['A1'].font = TITLE_FONT

for c, h in enumerate(['행 레이블', '수량', '최종정산금'], 1):
    cell = ws_pv.cell(row=3, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center')

sorted_items = sorted(pivot.items(), key=lambda x: x[0])
total_count = total_net = 0
r = 4
for name, data in sorted_items:
    total_count += data['count']
    total_net += data['net']
    ws_pv.cell(row=r, column=1, value=name).font = Font(size=10)
    ws_pv.cell(row=r, column=2, value=data['count']).font = NUM_FONT
    ws_pv.cell(row=r, column=2).number_format = '#,##0'
    ws_pv.cell(row=r, column=2).alignment = Alignment(horizontal='right')
    ws_pv.cell(row=r, column=3, value=round(data['net'])).font = NUM_FONT
    ws_pv.cell(row=r, column=3).number_format = '#,##0'
    ws_pv.cell(row=r, column=3).alignment = Alignment(horizontal='right')
    if (r - 4) % 2 == 1:
        for c in range(1, 4):
            ws_pv.cell(row=r, column=c).fill = LIGHT_ROW
    for c in range(1, 4):
        ws_pv.cell(row=r, column=c).border = thin_border
    r += 1

for c in range(1, 4):
    ws_pv.cell(row=r, column=c).fill = SUMMARY_FILL
    ws_pv.cell(row=r, column=c).border = thick_border
ws_pv.cell(row=r, column=1, value='총합계').font = Font(bold=True, size=11)
ws_pv.cell(row=r, column=2, value=total_count).font = BOLD_NUM
ws_pv.cell(row=r, column=2).number_format = '#,##0'
ws_pv.cell(row=r, column=3, value=round(total_net)).font = BOLD_NUM
ws_pv.cell(row=r, column=3).number_format = '#,##0'
r += 2

ws_pv.cell(row=r, column=2, value='배송비 정산금').font = Font(bold=True, size=10)
ws_pv.cell(row=r, column=3, value=round(shipping_total)).font = BOLD_NUM
ws_pv.cell(row=r, column=3).number_format = '#,##0'
r += 1
for c in range(2, 4):
    ws_pv.cell(row=r, column=c).fill = TOTAL_FILL
    ws_pv.cell(row=r, column=c).border = thick_border
ws_pv.cell(row=r, column=2, value='상품 + 배송비 합계').font = Font(bold=True, size=10, color='1565C0')
ws_pv.cell(row=r, column=3, value=round(total_net + shipping_total)).font = Font(bold=True, size=12, color='1565C0', name='Consolas')
ws_pv.cell(row=r, column=3).number_format = '#,##0'

ws_pv.column_dimensions['A'].width = 40
ws_pv.column_dimensions['B'].width = 12
ws_pv.column_dimensions['C'].width = 18

# === 가공 시트 ===
ws_proc = wb.create_sheet('가공')
ws_proc.sheet_properties.tabColor = '2196F3'
proc_h = ['주문번호', '상품주문번호', '구분', '상품명', '최종상품명', '정산기준금액', 'Npay수수료', '매출연동수수료', '최종정산금']
for c, h in enumerate(proc_h, 1):
    cell = ws_proc.cell(row=1, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal='center')

for i, pr in enumerate(proc_rows, 2):
    ws_proc.cell(row=i, column=1, value=pr['order'])
    ws_proc.cell(row=i, column=2, value=pr['item'])
    ws_proc.cell(row=i, column=3, value=pr['type'])
    ws_proc.cell(row=i, column=4, value=pr['name'])
    ws_proc.cell(row=i, column=5, value=pr['final'])
    ws_proc.cell(row=i, column=6, value=round(pr['settle'])).number_format = '#,##0'
    ws_proc.cell(row=i, column=7, value=round(pr['npay'])).number_format = '#,##0'
    ws_proc.cell(row=i, column=8, value=round(pr['sales'])).number_format = '#,##0'
    ws_proc.cell(row=i, column=9, value=round(pr['net'])).number_format = '#,##0'

ws_proc.column_dimensions['A'].width = 20
ws_proc.column_dimensions['B'].width = 22
ws_proc.column_dimensions['C'].width = 10
ws_proc.column_dimensions['D'].width = 55
ws_proc.column_dimensions['E'].width = 30
ws_proc.column_dimensions['F'].width = 15
ws_proc.column_dimensions['G'].width = 13
ws_proc.column_dimensions['H'].width = 13
ws_proc.column_dimensions['I'].width = 13
ws_proc.auto_filter.ref = f"A1:I{len(proc_rows)+1}"

outpath = os.path.join(os.path.dirname(settle_path), '03월 스마트스토어.xlsx')
wb.save(outpath)
print(f"저장 완료: {outpath}")
print(f"피벗: {len(pivot)}개 상품, 합계: {round(total_net + shipping_total):,}원")
