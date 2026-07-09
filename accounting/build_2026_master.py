# -*- coding: utf-8 -*-
"""2026 통합 마스터 — 채널 × 월(1~6) 매출·이익 합본 (규격화, 삽입 0) 2026-07-10"""
import openpyxl, pickle
from collections import OrderedDict, defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
B="/Users/macmini_ky/Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브/Claude AI work space/02. 매출 정산"
ch15=pickle.load(open('/tmp/ch15.pkl','rb'))   # {월:{채널:(매출총,이익)}}
# 6월 채널 총계 (검증값, build_june_consolidated v5)
JUN={'카페24':(25190876,17453129),'스마트스토어':(49467040,30111052),'쿠팡(로켓배송)':(46974511,24885484),
'쿠팡(로켓그로스)':(24845425,8605278),'지마켓':(1586186,1015684),'옥션':(425275,269497),'11번가':(1351612,834519),
'SSG':(96600,74583),'GS샵':(30160,20223),'오늘의집':(58380,37430),'에이블리':(45333,29649),
'카카오선물하기':(101760,72021),'카카오쇼핑하기':(113949,88501),'라이플로우':(19578280,3701498),
'PG컴퍼니':(977990,143166),'두버':(1069460,349893),'에드가':(45690,19556)}
def norm(s): return s.replace(' ','').replace('/','').replace('\n','')
# 정규화 매핑 (legacy→canonical)
ALIAS={'카페24네이버페이':'카페24','신세계몰(에스에스지닷컴)':'SSG','쿠팡':'쿠팡윙'}
def canon(name):
    n=norm(name); return ALIAS.get(n,name.strip().replace('\n',' '))
MONTHS=['1월','2월','3월','4월','5월','6월']
# 채널 그룹·순서
GROUPS=OrderedDict([("자사몰",["카페24"]),("네이버",["스마트스토어"]),
 ("쿠팡",["쿠팡(로켓배송)","쿠팡(로켓그로스)","쿠팡윙"]),("오픈마켓",["지마켓","옥션","11번가"]),
 ("종합몰",["SSG","GS샵","오늘의집","에이블리","카카오선물하기","카카오쇼핑하기","카카오메이커스","카카오쇼핑하기"]),
 ("수동발주",["라이플로우","PG컴퍼니","두버","에드가","굿트리","지엠홀딩스","고그로우"])])
# 데이터 통합: M[채널][월]=(매출,이익)
M=defaultdict(lambda: {mn:(0,0) for mn in MONTHS})
for mn in ['1월','2월','3월','4월','5월']:
    for ch,(rv,pf) in ch15[mn].items():
        M[canon(ch)][mn]=(rv,pf)
for ch,(rv,pf) in JUN.items():
    M[ch]['6월']=(rv,pf)
# 그룹에 없는 채널 → 수동발주/기타 편입
known=set(c for chs in GROUPS.values() for c in chs)
for ch in list(M):
    if ch not in known: GROUPS["수동발주"].append(ch)
# 활동 있는 채널만
def active(ch): return any(M[ch][mn][0] or M[ch][mn][1] for mn in MONTHS)

HF=PatternFill('solid',fgColor='1B2A4A'); HFONT=Font(color='FFFFFF',bold=True,size=10)
GF=PatternFill('solid',fgColor='4CAF50'); GFONT=Font(color='FFFFFF',bold=True,size=11)
SUB=PatternFill('solid',fgColor='DCEDC8'); TOT=PatternFill('solid',fgColor='FFF2CC')
bd=Border(bottom=Side(style='thin',color='DDDDDD'))
def mm(c): c.number_format='#,##0'
wb=openpyxl.Workbook()

def matrix_sheet(title, idx):
    """idx=0 매출, 1 이익"""
    ws=wb.create_sheet(title)
    ws['A1']=f'2026 채널별 월별 {"매출" if idx==0 else "이익"} (통합 마스터)'; ws['A1'].font=Font(bold=True,size=13)
    hdr=['그룹','채널']+MONTHS+['합계']
    ws.append([]); ws.append(hdr)
    for c in range(1,len(hdr)+1): ws.cell(row=3,column=c).fill=HF; ws.cell(row=3,column=c).font=HFONT
    r=4; gtot={mn:0 for mn in MONTHS}
    for g,chs in GROUPS.items():
        chs=[c for c in dict.fromkeys(chs) if c in M and active(c)]
        if not chs: continue
        gs=r; sub={mn:0 for mn in MONTHS}
        for ch in chs:
            vals=[M[ch][mn][idx] for mn in MONTHS]
            ws.append([g if ch==chs[0] else '',ch]+vals+[sum(vals)])
            for c in range(3,len(hdr)+1): mm(ws.cell(row=r,column=c))
            for c in range(1,len(hdr)+1): ws.cell(row=r,column=c).border=bd
            for i,mn in enumerate(MONTHS): sub[mn]+=vals[i]
            r+=1
        if len(chs)>1: ws.merge_cells(start_row=gs,start_column=1,end_row=r-1,end_column=1)
        ws.cell(row=gs,column=1).alignment=Alignment(vertical='center',horizontal='center'); ws.cell(row=gs,column=1).font=Font(bold=True)
        ws.append(['',g+' 소계']+[sub[mn] for mn in MONTHS]+[sum(sub.values())])
        for c in range(1,len(hdr)+1): ws.cell(row=r,column=c).fill=SUB; ws.cell(row=r,column=c).font=Font(bold=True)
        for c in range(3,len(hdr)+1): mm(ws.cell(row=r,column=c))
        for mn in MONTHS: gtot[mn]+=sub[mn]
        r+=1
    ws.append(['','■ 종합']+[gtot[mn] for mn in MONTHS]+[sum(gtot.values())])
    for c in range(1,len(hdr)+1): ws.cell(row=r,column=c).fill=TOT; ws.cell(row=r,column=c).font=Font(bold=True,size=11)
    for c in range(3,len(hdr)+1): mm(ws.cell(row=r,column=c))
    for col,w in zip('ABCDEFGHIJ',[9,20,12,12,12,12,12,12,12]): ws.column_dimensions[col].width=w
    return ws,r  # 종합행 r

ws_rev,rev_tot_row=matrix_sheet('채널별 매출(월별)',0)
ws_prof,_=matrix_sheet('채널별 이익(월별)',1)

# 대시보드: 그룹×월 매출 + 라인차트
ws1=wb.create_sheet('대시보드',0); ws1['A1']='2026 매출·이익 추이 (그룹별)'; ws1['A1'].font=Font(bold=True,size=14)
ws1.append([]); ws1.append(['그룹']+MONTHS+['합계'])
for c in range(1,9): ws1.cell(row=3,column=c).fill=HF; ws1.cell(row=3,column=c).font=HFONT
r=4; gt={mn:0 for mn in MONTHS}
for g,chs in GROUPS.items():
    chs=[c for c in dict.fromkeys(chs) if c in M and active(c)]
    if not chs: continue
    row=[sum(M[c][mn][0] for c in chs) for mn in MONTHS]
    ws1.append([g]+row+[sum(row)])
    for c in range(2,9): mm(ws1.cell(row=r,column=c))
    for i,mn in enumerate(MONTHS): gt[mn]+=row[i]
    r+=1
ws1.append(['■ 종합']+[gt[mn] for mn in MONTHS]+[sum(gt.values())])
for c in range(1,9): ws1.cell(row=r,column=c).fill=TOT; ws1.cell(row=r,column=c).font=Font(bold=True)
for c in range(2,9): mm(ws1.cell(row=r,column=c))
dr=r
# 종합 매출 추이 라인
lc=LineChart(); lc.title='월별 종합 매출 추이'; lc.height=8; lc.width=18; lc.style=12; lc.y_axis.numFmt='#,##0'
lc.add_data(Reference(ws1,min_col=2,max_col=7,min_row=r,max_row=r)); lc.set_categories(Reference(ws1,min_col=2,max_col=7,min_row=3))
ws1.add_chart(lc,'A'+str(dr+3))
# 그룹별 매출 막대(6월)
bc=BarChart(); bc.type='bar'; bc.title='그룹별 6월 매출'; bc.height=8; bc.width=14; bc.legend=None
bc.add_data(Reference(ws1,min_col=7,max_col=7,min_row=3,max_row=dr-1),titles_from_data=True)
bc.set_categories(Reference(ws1,min_col=1,min_row=4,max_row=dr-1)); ws1.add_chart(bc,'J'+str(dr+3))
for col,w in zip('ABCDEFGH',[14,12,12,12,12,12,12,12]): ws1.column_dimensions[col].width=w

# 성장률: 월별 종합 MoM
ws2=wb.create_sheet('성장률'); ws2['A1']='월별 성장률 (전월 대비)'; ws2['A1'].font=Font(bold=True,size=13)
ws2.append([]); ws2.append(['월','매출','전월대비','이익','이익률'])
for c in range(1,6): ws2.cell(row=3,column=c).fill=HF; ws2.cell(row=3,column=c).font=HFONT
prev=None; r=4
for mn in MONTHS:
    rv=sum(M[c][mn][0] for c in M); pf=sum(M[c][mn][1] for c in M)
    mom=((rv-prev)/prev) if prev else 0
    ws2.append([mn,rv,mom,pf,(pf/rv if rv else 0)])
    mm(ws2.cell(row=r,column=2)); mm(ws2.cell(row=r,column=4))
    ws2.cell(row=r,column=3).number_format='0.0%'; ws2.cell(row=r,column=5).number_format='0.0%'
    prev=rv; r+=1
for col,w in zip('ABCDE',[8,14,10,14,8]): ws2.column_dimensions[col].width=w
ws2.cell(row=r+1,column=1,value='※ 매출=매출+배송(총). 5월 이익은 시트3 기준(79.48M); 마감 확정 77.24M(별도 보고서). 6월 그로스 물류 반영.')

OUT=f"{B}/[마스터] 2026 온라인 매출정산_통합.xlsx"
wb.save(OUT)
print("✅ 마스터 저장")
# 검증
for mn in MONTHS:
    print(f"{mn}: 매출 {sum(M[c][mn][0] for c in M):,.0f} / 이익 {sum(M[c][mn][1] for c in M):,.0f}")
