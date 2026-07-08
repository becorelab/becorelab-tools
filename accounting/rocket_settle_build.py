# -*- coding: utf-8 -*-
"""쿠팡 로켓배송 월 정산 파일 생성 (Stocked_Data_List → 품명매핑 → 원가·이익 + VAT별도 손익)
사용: python3 rocket_settle_build.py 2026-06 [광고비VAT별도]"""
import openpyxl, json, re, sys, os, glob
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

MONTH = sys.argv[1] if len(sys.argv) > 1 else "2026-06"
AD = float(sys.argv[2]) if len(sys.argv) > 2 else 0  # 쿠팡 광고비 (VAT별도, 확정 시 입력)
MM = MONTH.split("-")[1]
BASE = f"/Users/macmini_ky/Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브/Claude AI work space/02. 매출 정산/26.{MM}/{MM}월 쿠팡 로켓배송"

NAME_MAP = json.load(open('rocket_name_map.json'))
NM = {re.sub(r'\s+',' ',k.replace('\xa0',' ')).strip(): v for k,v in NAME_MAP.items()}
cost = json.load(open('cost_master_2026.json'))
norm = lambda s: re.sub(r'\s+','',str(s))
cost_n = {norm(k):v for k,v in cost.items()}
get_cost = lambda p: cost.get(p) or cost_n.get(norm(p))
f = lambda v: 0.0 if v in (None,'','None') else float(str(v).replace(',',''))

src = glob.glob(f"{BASE}/Coupang_Stocked_Data_List*.xlsx")
assert len(src)==1, f"원본 1개여야 함: {src}"
wb = openpyxl.load_workbook(src[0], data_only=True)
ws = wb['data'] if 'data' in wb.sheetnames else wb[wb.sheetnames[0]]
hdr = [c.value for c in ws[1]]
ci = lambda k: next(i for i,h in enumerate(hdr) if h and k in str(h))
i_sku, i_qty, i_amt = ci('SKU명'), ci('수량'), ci('총단가')
prod = defaultdict(lambda:[0,0]); unmapped = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    sku = re.sub(r'\s+',' ',str(row[i_sku] or '').replace('\xa0',' ')).strip()
    if not sku or sku=='None': continue
    std = NM.get(sku)
    if not std: unmapped.add(sku); continue
    prod[std][0] += f(row[i_qty]); prod[std][1] += f(row[i_amt])
wb.close()
if unmapped:
    print("⚠️ 미매핑:", *unmapped, sep="\n  "); sys.exit(1)

wbo = openpyxl.Workbook(); wso = wbo.active; wso.title = f'쿠팡 로켓 {int(MM)}월'
thin=Side(style='thin',color='CCCCCC'); bd=Border(thin,thin,thin,thin)
hf=PatternFill('solid',fgColor='4472C4'); hfont=Font(color='FFFFFF',bold=True,size=10); tf=PatternFill('solid',fgColor='FFF2CC')
wso.append(['품명','수량','정산금액','원가','이익'])
for c in range(1,6):
    cell=wso.cell(row=1,column=c); cell.fill=hf; cell.font=hfont; cell.alignment=Alignment(horizontal='center'); cell.border=bd
T=[0,0,0,0]; r=2; miss=[]
for nm,(q,a) in sorted(prod.items(), key=lambda x:-x[1][1]):
    u=get_cost(nm); cst=u*q if u is not None else 0
    if u is None: miss.append(nm)
    wso.append([nm,q,a,cst,a-cst]); T[0]+=q; T[1]+=a; T[2]+=cst; T[3]+=a-cst; r+=1
wso.append(['합계',T[0],T[1],T[2],T[3]])
for c in range(1,6): wso.cell(row=r,column=c).fill=tf; wso.cell(row=r,column=c).font=Font(bold=True)
for row in wso.iter_rows(min_row=2,max_row=r,min_col=2,max_col=5):
    for cell in row:
        if isinstance(cell.value,(int,float)): cell.number_format='#,##0'
for col,w in zip('ABCDE',[32,8,13,12,13]): wso.column_dimensions[col].width=w

# VAT별도 손익
w2=wbo.create_sheet('VAT별도 손익')
w2['A1']=f'쿠팡 로켓 {int(MM)}월 손익 — 전항목 VAT별도 / 노란셀=가정'; w2['A1'].font=Font(bold=True,size=12)
ASSUME=PatternFill('solid',fgColor='FFF9C4')
w2.append([]); w2.append(['항목','금액','비고'])
for c in range(1,4): w2.cell(row=3,column=c).fill=hf; w2.cell(row=3,column=c).font=hfont
w2.append(['매출(공급가) = 매입가 ÷ 1.1', round(T[1]/1.1), f'{T[1]:,.0f} ÷ 1.1'])
w2.append(['원가 (VAT별도)', -round(T[2]), 'cost_master'])
w2.append(['쿠팡 광고비 (VAT별도)', -round(AD), '미입력 시 0 — 광고 XLSX 확정값 입력'])
if not AD: w2.cell(row=6,column=2).fill=ASSUME
w2.append(['영업이익','=SUM(B4:B6)',''])
w2.cell(row=7,column=1).font=Font(bold=True); w2.cell(row=7,column=2).font=Font(bold=True)
for rr in range(4,8): w2.cell(row=rr,column=2).number_format='#,##0'
w2.cell(row=9,column=1,value='※ 물류·수수료는 쿠팡 부담(직매입)이라 없음.')
for col,w in zip('ABC',[30,14,36]): w2.column_dimensions[col].width=w

out=f"{BASE}/{MM}월 쿠팡로켓 정산.xlsx"
wbo.save(out)
print(f"✅ {out}")
print(f"수량 {T[0]:,.0f} / 정산금액 {T[1]:,.0f} / 원가 {T[2]:,.0f} / 이익 {T[3]:,.0f} / 원가미상 {miss or '없음'}")
