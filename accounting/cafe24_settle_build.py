# -*- coding: utf-8 -*-
"""카페24 월 정산 파일 생성 (품명별 + 배송비 + 차감 6종: 환불/쿠폰/앱할인/적립금/회원등급/상품별)
사용: python3 cafe24_settle_build.py 2026-06  (기본 2026-06)"""
import openpyxl, json, re, sys

MONTH = sys.argv[1] if len(sys.argv) > 1 else "2026-06"
MM = MONTH.split("-")[1]
BASE = f"/Users/macmini_ky/Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브/Claude AI work space/02. 매출 정산/26.{MM}/{MM}월 카페24"
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAME_MAP = json.load(open('cafe24_name_map_temp.json'))
cost = json.load(open('cost_master_2026.json'))
# 혼합세트 원가 추가: 캡슐세제 4개(3448*4) + 캡슐표백제 2개(2769*2 or cost_master)
cap2 = cost.get('캡슐표백제 2개') or (cost.get('캡슐표백제',2769)*2)
cost['캡슐세제 4개+캡슐표백제 2개'] = cost.get('캡슐세제',3448)*4 + cap2

def norm(s): return re.sub(r'\s+','',str(s))
cost_n = {norm(k):v for k,v in cost.items()}
def get_cost(p):
    if p in cost: return cost[p]
    if norm(p) in cost_n: return cost_n[norm(p)]
    return None

raw=f"{BASE}/{MM}월 카페24 로우데이터.xlsx"
wb=openpyxl.load_workbook(raw, data_only=True, read_only=True); ws=wb[wb.sheetnames[0]]
hdr=[c.value for c in next(ws.iter_rows(max_row=1))]
def ci(n):
    for i,h in enumerate(hdr):
        if h and n in str(h): return i
    return None
i_nm=ci('주문상품명(옵션포함)'); i_opt=ci('옵션+판매가'); i_qty=ci('수량'); i_ship=ci('총 배송비')
i_ord=ci('주문번호'); i_ref=ci('실제 환불금액'); i_cpn=ci('주문서 쿠폰 할인금액')
i_app=ci('앱 상품할인 금액(최종)'); i_mil=ci('사용한 적립금액(최종)')
# 2026-07-06 추가: 회원등급/상품별 추가할인 (5월 성락님 임직원구매 67,900 누락 발견 → 대표님 B안: 6월분부터 차감)
i_grd=ci('회원등급 추가할인금액'); i_itm=ci('상품별 추가할인금액')
# 2026-07-14 추가: 취소·반품 행 선제외 (성락님 주간정산 방식 채택 — '환불금액' 컬럼은 취소분을 전부 기록하지 않아 매출 과대 원인이었음)
i_st=ci('주문 상태')
def f(v):
    if v in (None,'','None'): return 0.0
    try: return float(str(v).replace(',',''))
    except: return 0.0

prod=defaultdict(lambda:[0,0,0])  # 품명 -> [수량, 매출(결제금액), 배송비]
oship=set(); oref=set(); ocpn=set(); omil=set(); ogrd=set()
ref=cpn=app=mil=grd=itm=0.0; unmapped=[]
cancel_qty=0; cancel_amt=0.0
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[i_qty] in (None,'') and row[i_opt] in (None,''): continue
    # 취소·반품 행 선제외 (매출·원가·배송·차감 전부 스킵 — 주간정산과 동일 방식)
    if i_st is not None:
        st=str(row[i_st] or '')
        if '취소' in st or '반품' in st:
            cancel_qty+=int(f(row[i_qty])); cancel_amt+=f(row[i_opt])*f(row[i_qty]); continue
    rawname=str(row[i_nm] or '').strip()
    std=NAME_MAP.get(rawname)
    o=str(row[i_ord] or '').strip()
    if std:
        qty=int(f(row[i_qty])); amt=f(row[i_opt])*qty
        prod[std][0]+=qty; prod[std][1]+=amt
        if o and o not in oship: prod[std][2]+=f(row[i_ship]); oship.add(o)
    else:
        unmapped.append(rawname)
    # 차감 (주문/품목 단위 중복제거)
    if o and o not in oref:
        rv=f(row[i_ref])
        if rv>0: ref+=rv; oref.add(o)
    if o and o not in ocpn: cpn+=f(row[i_cpn]); ocpn.add(o)
    app+=f(row[i_app])
    if o and o not in omil:
        mv=f(row[i_mil])
        if mv>0: mil+=mv; omil.add(o)
    if i_grd is not None and o and o not in ogrd:  # 주문 내 행마다 동일값 반복 → 주문당 1회
        gv=f(row[i_grd])
        if gv>0: grd+=gv; ogrd.add(o)
    if i_itm is not None: itm+=f(row[i_itm])  # 상품별 할인은 행별 합산

# 출력 파일
out=f"{BASE}/{MM}월 카페24 정산.xlsx" 
wbo=openpyxl.Workbook(); wso=wbo.active; wso.title=f'카페24 {int(MM)}월'
thin=Side(style='thin',color='CCCCCC'); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
hf=PatternFill('solid',fgColor='4472C4'); hfont=Font(color='FFFFFF',bold=True,size=10)
tf=PatternFill('solid',fgColor='FFF2CC'); mf=PatternFill('solid',fgColor='FCE4D6')
wso['A1']=f'카페24(자사몰) {int(MM)}월 정산'; wso['A1'].font=Font(bold=True,size=13)
wso.append([]); wso.append(['품명','수량','매출액','배송비','원가','이익'])
for c in range(1,7):
    cell=wso.cell(row=3,column=c); cell.fill=hf; cell.font=hfont; cell.alignment=Alignment(horizontal='center'); cell.border=bd
r=4; t_qty=t_sales=t_ship=t_cost=t_profit=0; miss_cost=[]
for nm,(q,s,sh) in sorted(prod.items(), key=lambda x:-x[1][1]):
    unit=get_cost(nm); cst=unit*q if unit is not None else None
    if cst is None: miss_cost.append(nm); cst=0
    profit=s+sh-cst
    wso.append([nm,q,s,sh,cst if unit is not None else '원가미상',profit])
    t_qty+=q; t_sales+=s; t_ship+=sh; t_cost+=cst; t_profit+=profit; r+=1
# 소계
wso.append(['상품 소계',t_qty,t_sales,t_ship,t_cost,t_profit])
for c in range(1,7): wso.cell(row=r,column=c).fill=tf; wso.cell(row=r,column=c).font=Font(bold=True)
r+=1
# 차감 행
for label,amt in [('환불 (차감)',ref),('쿠폰 (차감)',cpn),('앱 상품할인 (차감)',app),('사용 적립금 (차감)',mil),('회원등급할인 (차감)',grd),('상품별 추가할인 (차감)',itm)]:
    wso.append([label,'',-amt,'','',-amt]);
    for c in range(1,7): wso.cell(row=r,column=c).fill=mf
    r+=1
# 최종
final_sales=t_sales+t_ship-ref-cpn-app-mil-grd-itm
final_profit=t_profit-ref-cpn-app-mil-grd-itm
wso.append(['최종 (정산)','',final_sales,'',t_cost,final_profit])
for c in range(1,7): wso.cell(row=r,column=c).fill=tf; wso.cell(row=r,column=c).font=Font(bold=True,size=11)
for row in wso.iter_rows(min_row=4,max_row=r,min_col=2,max_col=6):
    for cell in row:
        if isinstance(cell.value,(int,float)): cell.number_format='#,##0'
for col,w in zip('ABCDEF',[40,7,12,10,11,13]): wso.column_dimensions[col].width=w
wso.cell(row=r+2,column=1,value='※ 매출액=결제금액(옵션+판매가×수량), 이익=매출+배송비−원가. 취소·반품 행은 선제외(2026-07-14~), 환불 차감행은 제외 안 된 행의 부분환불만. 쿠폰/앱할인/적립금/회원등급/상품별할인은 매출·이익 모두 차감.')
# ── 시트2: 자사몰 손익 (VAT별도 통일 + 택배·PG·광고비) 2026-07-07 신설 ──
w2=wbo.create_sheet('자사몰 손익(VAT별도)')
w2['A1']=f'카페24 {int(MM)}월 손익 — VAT별도 통일 / 노란 셀은 가정값(수정 시 자동 재계산)'
w2['A1'].font=Font(bold=True,size=12)
n_orders=len(oship)  # 배송 집계에 잡힌 unique 주문수
rows2=[
 ('매출(공급가) = 최종정산 ÷ 1.1', round(final_sales/1.1), ''),
 ('원가 (VAT별도)', -t_cost, ''),
 ('출고 택배비 = 주문수 × 단가', None, f'주문 {n_orders}건'),
 ('PG 수수료 = 매출(포함) × 요율', None, ''),
 ('메타 광고비 (VAT별도, 수동입력)', 0, '메타 인사이트 지출값 입력'),
 ('영업이익', None, ''),
]
w2.append([]); w2.append(['항목','금액','비고'])
for c in range(1,4):
    cell=w2.cell(row=3,column=c); cell.fill=hf; cell.font=hfont; cell.border=bd
ASSUME=PatternFill('solid',fgColor='FFF9C4')
w2['E3']='가정값'; w2['E4']=3000; w2['F4']='택배 단가(원/건)'
w2['E5']=0.034; w2['F5']='PG 요율'
w2['E4'].fill=ASSUME; w2['E5'].fill=ASSUME
w2['B4']=round(final_sales/1.1); w2['C4']=''
w2.cell(row=4,column=1,value='매출(공급가)')
w2.cell(row=5,column=1,value='원가(VAT별도)'); w2['B5']=-t_cost
w2.cell(row=6,column=1,value=f'출고 택배비 ({n_orders}건 × 단가)'); w2['B6']=f'=-{n_orders}*E4'
w2.cell(row=7,column=1,value='PG 수수료 (결제액 × 요율)'); w2['B7']=f'=-ROUND({final_sales}*E5,0)'
w2.cell(row=8,column=1,value='메타 광고비 (수동입력)'); w2['B8']=0; w2['B8'].fill=ASSUME
w2.cell(row=9,column=1,value='영업이익'); w2['B9']='=SUM(B4:B8)'
w2.cell(row=9,column=1).font=Font(bold=True); w2['B9'].font=Font(bold=True)
for rr in range(4,10): w2.cell(row=rr,column=2).number_format='#,##0'
w2['E4'].number_format='#,##0'; w2['E5'].number_format='0.0%'
for col,w in zip('ABCDEF',[30,13,20,3,10,14]): w2.column_dimensions[col].width=w
wbo.save(out)
print(f'✅ 저장: {out}')
print(f'   상품 {len(prod)}종 / 결제총액 {t_sales:,.0f} + 배송 {t_ship:,.0f}')
print(f'   차감: 환불 {ref:,.0f} / 쿠폰 {cpn:,.0f} / 앱할인 {app:,.0f} / 적립금 {mil:,.0f} / 등급할인 {grd:,.0f} / 상품별할인 {itm:,.0f}')
print(f'   최종 매출 {final_sales:,.0f} / 원가 {t_cost:,.0f} / 최종 이익 {final_profit:,.0f}')
print(f'   미매핑 {len(set(unmapped))} / 원가미상 {sorted(set(miss_cost))}')
