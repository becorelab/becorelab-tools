"""
월간 매출 보고서 자동 생성 (Obsidian용)
정산 엑셀 → 월별 마크다운 보고서 생성
채널 범위를 B열에서 동적으로 감지

사용법: python generate_monthly_reports.py
"""
import sys, io, os
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OBSIDIAN = r'C:\Users\User\Documents\비코어랩\01. Becorelab AI Agent Team\📊 Sales Report\월간'
os.makedirs(OBSIDIAN, exist_ok=True)

def sf(v):
    if v is None or v == '' or v == '#DIV/0!' or v == ' ':
        return 0.0
    try: return float(v)
    except: return 0.0

def si(v): return int(sf(v))

def fmt(n):
    if n == 0: return '-'
    return f"₩{int(round(n)):,}"

def pct(n):
    if n == 0 or n is None: return '-'
    if isinstance(n, float) and abs(n) < 10:
        return f"{n*100:.1f}%" if abs(n) < 1 else f"{n:.1f}%"
    return f"{n:.1f}%"

def mom(cur, prev):
    if prev == 0 or cur == 0: return '-'
    return f"{(cur/prev-1)*100:+.1f}%"

# 채널명 정규화
CHANNEL_NORMALIZE = {
    '카페24 / 네이버페이': '카페24', '카페24 / 네이버스토어': '카페24',
    '카카오선물하기': '카카오선물', '카카오쇼핑하기': '카카오쇼핑',
    '쿠팡(로켓배송)': '쿠팡 로켓배송', '쿠팡': '쿠팡(일반)',
    '신세계몰\n(에스에스지닷컴)': '신세계', '신세계몰(에스에스지닷컴)': '신세계',
}

def detect_channels(ws):
    """B열 스캔하여 채널별 행 범위 동적 감지"""
    channels = {}
    current_ch = None
    start_row = None

    for r in range(7, ws.max_row + 1):
        b_val = ws.cell(row=r, column=2).value
        if b_val:
            ch_name = str(b_val).strip().replace('\n', '')
            ch_name = CHANNEL_NORMALIZE.get(ch_name, ch_name)

            if ch_name != current_ch:
                if current_ch and start_row:
                    channels[current_ch] = (start_row, r - 1)
                current_ch = ch_name
                start_row = r

    if current_ch and start_row:
        channels[current_ch] = (start_row, ws.max_row)

    return channels


def extract_month(ws, month, channels):
    """한 달 데이터 추출"""
    MONTH_COLS = {
        1: 6, 2: 15, 3: 24, 4: 33, 5: 42, 6: 51,
        7: 60, 8: 69, 9: 78, 10: 87, 11: 96, 12: 105
    }
    sc = MONTH_COLS[month]
    result = {}
    products_all = []

    for ch_name, (row_start, row_end) in channels.items():
        ch_data = {'qty': 0, 'revenue': 0, 'ship': 0, 'total': 0, 'profit': 0, 'products': []}

        for r in range(row_start, row_end + 1):
            product = ws.cell(row=r, column=5).value
            if not product: continue
            product = str(product).strip()

            qty = si(ws.cell(row=r, column=sc).value)
            revenue = sf(ws.cell(row=r, column=sc + 1).value)
            ship = sf(ws.cell(row=r, column=sc + 2).value)
            total = sf(ws.cell(row=r, column=sc + 3).value)
            profit = sf(ws.cell(row=r, column=sc + 5).value)
            profit_rate = sf(ws.cell(row=r, column=sc + 6).value)

            if qty == 0 and revenue == 0 and total == 0:
                continue

            ch_data['qty'] += qty
            ch_data['revenue'] += revenue
            ch_data['ship'] += ship
            ch_data['total'] += total
            ch_data['profit'] += profit

            p = {'name': product, 'qty': qty, 'revenue': revenue,
                 'ship': ship, 'total': total, 'profit': profit,
                 'profit_rate': profit_rate, 'channel': ch_name}
            ch_data['products'].append(p)
            products_all.append(p)

        if ch_data['total'] != 0 or ch_data['qty'] != 0:
            ch_data['profit_rate'] = ch_data['profit'] / ch_data['total'] if ch_data['total'] > 0 else 0
            result[ch_name] = ch_data

    return result, products_all


def extract_ad_costs(wb, month):
    """광고비적용 시트에서 채널별 광고비 추출"""
    ad_costs = {}
    for sn in wb.sheetnames:
        if '광고비적용' in sn or '광고비 적용' in sn:
            ws_ad = wb[sn]
            # 헤더 찾기: Row 9~10 근처
            header_row = None
            for r in range(1, 15):
                v = ws_ad.cell(row=r, column=3).value
                if v and '매출' in str(v):
                    header_row = r
                    break
            if not header_row: break

            # 월 컬럼 찾기 (4열씩: 매출, 이익, 광고비, 이익률)
            ad_col = None
            for c in range(3, ws_ad.max_column + 1):
                month_header = ws_ad.cell(row=header_row - 1, column=c).value
                col_header = ws_ad.cell(row=header_row, column=c).value
                if col_header and '광고비' in str(col_header):
                    m_str = str(month_header or '')
                    if f'{month}월' in m_str:
                        ad_col = c
                        break

            if not ad_col: break

            for r in range(header_row + 1, ws_ad.max_row + 1):
                ch = ws_ad.cell(row=r, column=2).value
                if not ch: continue
                ch = str(ch).strip().replace('\n', '')
                ch = CHANNEL_NORMALIZE.get(ch, ch)
                ad = sf(ws_ad.cell(row=r, column=ad_col).value)
                if ad > 0:
                    ad_costs[ch] = ad
            break
    return ad_costs


def generate_report(year, month, channels, products, prev_channels=None, yoy_channels=None, ad_costs=None):
    """마크다운 보고서 생성"""
    if not channels: return None

    total_revenue = sum(c['total'] for c in channels.values())
    total_profit = sum(c['profit'] for c in channels.values())
    total_ad = sum(ad_costs.values()) if ad_costs else 0
    total_net = total_profit - total_ad
    total_rate = total_profit / total_revenue if total_revenue > 0 else 0
    net_rate = total_net / total_revenue if total_revenue > 0 else 0

    prev_total = sum(c['total'] for c in prev_channels.values()) if prev_channels else 0
    prev_profit = sum(c['profit'] for c in prev_channels.values()) if prev_channels else 0
    yoy_total = sum(c['total'] for c in yoy_channels.values()) if yoy_channels else 0
    yoy_profit = sum(c['profit'] for c in yoy_channels.values()) if yoy_channels else 0

    sorted_channels = sorted(channels.items(), key=lambda x: -x[1]['total'])
    top_products = sorted(products, key=lambda x: -abs(x['total']))[:10]

    # 카테고리 분류
    cat_map = {'건조기': '건조기 시트', '식세기': '식기세척기 세제', '하트식세기': '식기세척기 세제',
               '캡슐': '캡슐세제', '섬유탈취': '섬유탈취제', '수세미': '수세미',
               '얼룩': '얼룩제거제', '이염': '이염방지시트', '세탁세제': '세탁세제', '다목적': '다목적세정제'}
    categories = {}
    for p in products:
        cat = '기타'
        for kw, cn in cat_map.items():
            if kw in p['name']: cat = cn; break
        if cat not in categories: categories[cat] = {'revenue': 0, 'profit': 0, 'qty': 0}
        categories[cat]['revenue'] += p['total']
        categories[cat]['profit'] += p['profit']
        categories[cat]['qty'] += p['qty']

    warning_products = [p for p in products if p['profit_rate'] < 0 and p['qty'] > 0]

    lines = []
    L = lines.append

    # 프론트매터
    L(f"---")
    L(f"tags: [매출, 월간, {year}]")
    L(f"date: {year}-{month:02d}-28")
    L(f"---")
    L(f"")
    L(f"# 📋 {year}년 {month}월 월간 매출 보고서")
    L(f"")
    L(f"> 기준: {year}. 12 온라인 매출정산")
    L(f"")
    L(f"---")
    L(f"")

    # 총괄
    L(f"## 💰 월간 총괄")
    L(f"")
    L(f"> [!summary] 핵심 수치")

    if prev_total > 0 or yoy_total > 0:
        L(f"> | 지표 | {month}월 실적 | {'전월' if prev_total > 0 else ''} | {'MoM' if prev_total > 0 else ''} | {'전년동월' if yoy_total > 0 else ''} | {'YoY' if yoy_total > 0 else ''} |")
        L(f"> | :--- | ---: | {'---:' if prev_total > 0 else ''} | {':---:' if prev_total > 0 else ''} | {'---:' if yoy_total > 0 else ''} | {':---:' if yoy_total > 0 else ''} |")
        prev_str = f" {fmt(prev_total)} | {mom(total_revenue, prev_total)} |" if prev_total > 0 else " |  |"
        yoy_str = f" {fmt(yoy_total)} | {mom(total_revenue, yoy_total)} |" if yoy_total > 0 else " |  |"
        L(f"> | **총 매출** | `{fmt(total_revenue)}` |{prev_str}{yoy_str}")
        prev_p = f" {fmt(prev_profit)} | {mom(total_profit, prev_profit)} |" if prev_profit > 0 else " |  |"
        yoy_p = f" {fmt(yoy_profit)} | {mom(total_profit, yoy_profit)} |" if yoy_profit > 0 else " |  |"
        L(f"> | **총 이익** | `{fmt(total_profit)}` — 이익률 =={pct(total_rate)}== |{prev_p}{yoy_p}")
    else:
        L(f"> - **총 매출** `{fmt(total_revenue)}`")
        L(f"> - **총 이익** `{fmt(total_profit)}` — 이익률 =={pct(total_rate)}==")

    if total_ad > 0:
        L(f"> - **총 광고비** `{fmt(total_ad)}` — 매출 대비 =={pct(total_ad/total_revenue if total_revenue else 0)}==")
        L(f"> - **순이익** `{fmt(total_net)}` — 순이익률 =={pct(net_rate)}==")
    L(f"> - 활성 채널 **{len(channels)}개**")
    L(f"")
    L(f"---")
    L(f"")

    # 채널별 성과
    L(f"## 📊 채널별 성과")
    L(f"")
    if total_ad > 0:
        L(f"| 순위 | 채널 | 매출 | 비중 | 이익률 | 광고비 | 순이익률 |")
        L(f"| :--: | :--- | ---: | :--: | :---: | ---: | :---: |")
    else:
        L(f"| 순위 | 채널 | 매출 | 비중 | 이익 | 이익률 |")
        L(f"| :--: | :--- | ---: | :--: | ---: | :---: |")

    for i, (name, data) in enumerate(sorted_channels, 1):
        share = data['total'] / total_revenue * 100 if total_revenue > 0 else 0
        rate = pct(data['profit_rate'])
        emoji = ' 🔥' if data['profit_rate'] > 0.6 else ''
        if total_ad > 0:
            ad = ad_costs.get(name, 0)
            ad_str = fmt(ad) if ad > 0 else '-'
            net = data['profit'] - ad
            net_r = pct(net / data['total']) if data['total'] > 0 and ad > 0 else rate
            warn = ' ⚠️' if ad > 0 and net / data['total'] < 0.25 else emoji
            L(f"| {i} | **{name}** | {fmt(data['total'])} | {share:.1f}% | {rate} | {ad_str} | {net_r}{warn} |")
        else:
            L(f"| {i} | **{name}** | {fmt(data['total'])} | {share:.1f}% | {fmt(data['profit'])} | {rate}{emoji} |")
    L(f"")
    L(f"---")
    L(f"")

    # 상품 TOP 10
    L(f"## 🏆 상품 TOP 10")
    L(f"")
    L(f"| 순위 | 상품 | 채널 | 수량 | 매출 | 이익률 |")
    L(f"| :--: | :--- | :---: | ---: | ---: | :---: |")
    for i, p in enumerate(top_products, 1):
        rate = pct(p['profit_rate'])
        emoji = ' 🔥' if p['profit_rate'] > 0.6 else (' ⚠️' if p['profit_rate'] < 0.4 and p['profit_rate'] > 0 else '')
        ch_short = p['channel'][:8]
        L(f"| {i} | **{p['name'][:30]}** | {ch_short} | {p['qty']:,} | {fmt(p['total'])} | {rate}{emoji} |")
    L(f"")

    # 카테고리
    if categories:
        L(f"### 카테고리 요약")
        L(f"")
        L(f"| 카테고리 | 매출 | 비중 | 이익률 |")
        L(f"| :--- | ---: | :---: | :---: |")
        for cat, data in sorted(categories.items(), key=lambda x: -x[1]['revenue']):
            if data['revenue'] == 0: continue
            share = data['revenue'] / total_revenue * 100 if total_revenue > 0 else 0
            rate = pct(data['profit'] / data['revenue']) if data['revenue'] > 0 else '-'
            L(f"| **{cat}** | {fmt(data['revenue'])} | {share:.1f}% | {rate} |")
        L(f"")

    # 적자 상품
    if warning_products:
        L(f"### ⚠️ 적자/저마진 주의")
        L(f"")
        L(f"| 상품 | 채널 | 이익률 | 수량 |")
        L(f"| :--- | :--- | :---: | ---: |")
        for p in sorted(warning_products, key=lambda x: x['profit_rate'])[:5]:
            L(f"| {p['name'][:25]} | {p['channel'][:8]} | =={pct(p['profit_rate'])}== | {p['qty']:,} |")
        L(f"")

    L(f"---")
    L(f"")

    # 광고비 (있을 때만)
    if total_ad > 0:
        L(f"## 📢 광고 효율")
        L(f"")
        L(f"| 채널 | 광고비 | 광고비/매출 | 순이익 | 순이익률 |")
        L(f"| :--- | ---: | :---: | ---: | :---: |")
        for name, ad in sorted(ad_costs.items(), key=lambda x: -x[1]):
            ch = channels.get(name)
            if not ch: continue
            ad_ratio = pct(ad / ch['total']) if ch['total'] > 0 else '-'
            net = ch['profit'] - ad
            net_r = pct(net / ch['total']) if ch['total'] > 0 else '-'
            warn = ' ⚠️' if ch['total'] > 0 and ad / ch['total'] > 0.35 else ''
            L(f"| **{name}** | {fmt(ad)} | {ad_ratio}{warn} | {fmt(net)} | {net_r} |")
        L(f"| **합계** | **{fmt(total_ad)}** | | | |")
        L(f"")
        L(f"---")
        L(f"")

    # 인사이트
    L(f"## 💡 인사이트")
    L(f"")
    best_ch = sorted_channels[0] if sorted_channels else None
    best_margin = max(channels.items(), key=lambda x: x[1]['profit_rate']) if channels else None

    if best_ch:
        L(f"> [!note] 채널")
        share_pct = best_ch[1]['total'] / total_revenue * 100 if total_revenue > 0 else 0
        L(f"> - **{best_ch[0]}** 매출 1위 ({fmt(best_ch[1]['total'])}, 비중 {share_pct:.0f}%)")
        if best_margin and best_margin[0] != best_ch[0]:
            L(f"> - **{best_margin[0]}** 이익률 최고 ({pct(best_margin[1]['profit_rate'])})")
        L(f"")

    if top_products:
        L(f"> [!note] 상품")
        L(f"> - **{top_products[0]['name'][:25]}** 매출 1위 ({fmt(top_products[0]['total'])})")
        high_margin = [p for p in products if p['profit_rate'] > 0.7 and p['qty'] >= 5]
        if high_margin:
            best_hm = max(high_margin, key=lambda x: x['qty'])
            L(f"> - **{best_hm['name'][:25]}** 고마진 ({pct(best_hm['profit_rate'])}, {best_hm['qty']}건)")
        L(f"")

    if total_ad > 0 and total_revenue > 0:
        L(f"> [!warning] 광고")
        L(f"> - 광고비 {fmt(total_ad)} = 매출의 {total_ad/total_revenue*100:.1f}%, 이익의 {total_ad/total_profit*100:.0f}% 소진")
        worst_ad = max(ad_costs.items(), key=lambda x: x[1] / channels[x[0]]['total'] if x[0] in channels and channels[x[0]]['total'] > 0 else 0)
        if worst_ad[0] in channels and channels[worst_ad[0]]['total'] > 0:
            L(f"> - **{worst_ad[0]}** 광고비/매출 {worst_ad[1]/channels[worst_ad[0]]['total']*100:.0f}%로 가장 비효율")
        L(f"")

    L(f"---")
    L(f"")

    # 이전/다음 링크
    prev_m = month - 1 if month > 1 else 12
    prev_y = year if month > 1 else year - 1
    next_m = month + 1 if month < 12 else 1
    next_y = year if month < 12 else year + 1
    L(f"*📁 데이터 소스: `{year}. 12 온라인 매출정산.xlsx`*")
    L(f"*⬅️ [[{prev_y}-{prev_m:02d}]] | ➡️ [[{next_y}-{next_m:02d}]]*")

    return '\n'.join(lines)


# ─── 메인 실행 ───
FILES = {
    2023: os.path.join(SCRIPT_DIR, '2023. 12 온라인 매출정산.xlsx'),
    2024: os.path.join(SCRIPT_DIR, '2024. 12 온라인 매출정산 -원가수정.xlsx'),
    2025: os.path.join(SCRIPT_DIR, '2025. 12 온라인 매출정산.xlsx'),
    2026: os.path.join(SCRIPT_DIR, '2026. 02 온라인 매출정산.xlsx'),
}

# 2026-01, 02는 수동 작성본 유지 (덮어쓰지 않음)
SKIP = {'2026-01', '2026-02'}

print("=" * 50)
print("  📊 월간 매출 보고서 자동 생성")
print("=" * 50)

# 모든 파일의 월별 데이터를 미리 추출 (전월/전년 비교용)
all_data = {}  # {(year, month): {channel_data}}

for year, filepath in sorted(FILES.items()):
    if not os.path.exists(filepath):
        print(f"\n  ⚠️ {year}년: 파일 없음")
        continue

    print(f"\n  📂 {year}년: {os.path.basename(filepath)}")
    wb = openpyxl.load_workbook(filepath, data_only=True)

    ws_main = None
    for sn in wb.sheetnames:
        if '월별 온라인 매출정산' in sn:
            ws_main = wb[sn]; break
    if not ws_main and len(wb.sheetnames) >= 2:
        ws_main = wb[wb.sheetnames[1]]
    if not ws_main:
        print(f"    ❌ 매출정산 시트 없음")
        continue

    # 채널 범위 동적 감지
    channels_range = detect_channels(ws_main)
    print(f"    채널 {len(channels_range)}개 감지")

    max_month = 12 if year <= 2025 else 2

    for month in range(1, max_month + 1):
        ch_data, products = extract_month(ws_main, month, channels_range)
        if ch_data:
            # 광고비 추출
            ad = extract_ad_costs(wb, month)
            all_data[(year, month)] = {'channels': ch_data, 'products': products, 'ad': ad}

# 보고서 생성
total_generated = 0
for (year, month), data in sorted(all_data.items()):
    doc_id = f"{year}-{month:02d}"
    if doc_id in SKIP:
        print(f"    ⏭️  {doc_id}: 수동 작성본 유지")
        continue

    ch = data['channels']
    products = data['products']
    ad = data['ad']

    # 전월 데이터
    prev_m = month - 1 if month > 1 else 12
    prev_y = year if month > 1 else year - 1
    prev = all_data.get((prev_y, prev_m), {}).get('channels', {})

    # 전년동월
    yoy = all_data.get((year - 1, month), {}).get('channels', {})

    report = generate_report(year, month, ch, products, prev, yoy, ad)
    if report:
        outpath = os.path.join(OBSIDIAN, f'{doc_id}.md')
        with open(outpath, 'w', encoding='utf-8') as f:
            f.write(report)
        total_rev = sum(c['total'] for c in ch.values())
        total_ad_cost = sum(ad.values()) if ad else 0
        ad_str = f", 광고비 {fmt(total_ad_cost)}" if total_ad_cost > 0 else ""
        print(f"    ✅ {doc_id}: {len(ch)}개 채널, {fmt(total_rev)}{ad_str}")
        total_generated += 1

print(f"\n{'=' * 50}")
print(f"  완료! {total_generated}개 보고서 생성 (수동 유지 {len(SKIP)}개)")
print(f"  📁 {OBSIDIAN}")
print(f"{'=' * 50}")
