# -*- coding: utf-8 -*-
"""5/18~31 물류비 반영 정상기간 손익 시트 추가 + 비교"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 계산 결과 (gross_normal_period 분석값)
SALES=11758450; SETTLE=10456526; COST=4196475; LOGI=1255172; AD=1311450
PROFIT=SETTLE-COST-LOGI-AD  # 3,693,429

path='/Users/macmini_ky/Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브/앱/매출 정산/26.05/26.05 로켓그로스 정산_채움컴퍼니.xlsx'
wb=load_workbook(path)
if '정상기간 손익(5.18~31)' in wb.sheetnames: del wb['정상기간 손익(5.18~31)']
ws=wb.create_sheet('정상기간 손익(5.18~31)',1)
thin=Side(style='thin',color='CCCCCC'); border=Border(left=thin,right=thin,top=thin,bottom=thin)
hdrfill=PatternFill('solid',fgColor='C00000'); hdrfont=Font(color='FFFFFF',bold=True,size=10)
totfill=PatternFill('solid',fgColor='FFF2CC'); pf=PatternFill('solid',fgColor='E2EFDA')

ws['A1']='로켓그로스 — 5/18~31 정상기간 손익 (비용제로 종료, 물류비 본격 반영)'
ws['A1'].font=Font(bold=True,size=13)
ws['A2']='= 앞으로의 기준선 마진. 정산주기 5/24+5/31 (매출인식일 5/18~31)'
ws['A2'].font=Font(size=9,italic=True)

# 손익 구조
ws.append([]); ws.append(['항목','금액','판매액 대비'])
for c in range(1,4):
    cell=ws.cell(row=4,column=c); cell.fill=hdrfill; cell.font=hdrfont
    cell.alignment=Alignment(horizontal='center'); cell.border=border
items=[
    ('판매액', SALES, SALES/SALES),
    ('정산대상액', SETTLE, SETTLE/SALES),
    ('− 원가', COST, COST/SALES),
    ('− 물류비', LOGI, LOGI/SALES),
    ('− 광고비', AD, AD/SALES),
    ('= 이익', PROFIT, PROFIT/SALES),
]
r=5
for name,val,rate in items:
    ws.append([name,val,rate])
    if '이익' in name:
        for c in range(1,4): ws.cell(row=r,column=c).fill=pf; ws.cell(row=r,column=c).font=Font(bold=True)
    r+=1
ws.append(['이익률(정산대상 기준)','', PROFIT/SETTLE])
r+=1
for row in ws.iter_rows(min_row=5,max_row=r,min_col=2,max_col=2):
    for cell in row:
        if isinstance(cell.value,(int,float)): cell.number_format='#,##0'
for row in ws.iter_rows(min_row=5,max_row=r,min_col=3,max_col=3):
    for cell in row:
        if isinstance(cell.value,(int,float)): cell.number_format='0.0%'

# 기간 비교
r+=1
ws.cell(row=r,column=1,value='[ 이익률 추이 비교 ]').font=Font(bold=True,size=11); r+=1
ws.append(['기간','이익률(판매)','물류비 상태']);
for c in range(1,4):
    cell=ws.cell(row=r,column=c); cell.fill=hdrfill; cell.font=hdrfont; cell.alignment=Alignment(horizontal='center')
r+=1
comp=[('3~4월',0.40,'0원 (비용제로 90일)'),
      ('5월 전체',0.335,'1,255,172 (일부 반영)'),
      ('5/18~31 정상',0.314,'반영 (기준선)')]
for nm,rt,st in comp:
    ws.append([nm,rt,st])
    if '정상' in nm:
        for c in range(1,4): ws.cell(row=r,column=c).fill=totfill; ws.cell(row=r,column=c).font=Font(bold=True)
    r+=1
for row in ws.iter_rows(min_row=r-3,max_row=r,min_col=2,max_col=2):
    for cell in row:
        if isinstance(cell.value,(int,float)): cell.number_format='0.0%'

r+=1
for note in ['※ 광고비 = 빠른정산 화면 실제 차감액 (5/18~31).',
             '※ 물류비 10.7% / 광고비 11.2%가 양대 변동비. 비용제로 종료로 이익률 40%→31% 하락.',
             '※ 물류비는 앞으로 증가 방향 → 광고비·쿠폰 관리가 이익률 방어 핵심.']:
    ws.cell(row=r,column=1,value=note).font=Font(size=9); r+=1

ws.column_dimensions['A'].width=24
ws.column_dimensions['B'].width=14
ws.column_dimensions['C'].width=22
wb.save(path)
print('정상기간 손익 시트 추가 완료')
print('시트 순서:', wb.sheetnames)
print(f'\n5/18~31 이익 {PROFIT:,.0f}원 / 이익률(판매) {PROFIT/SALES*100:.1f}% / 이익률(정산) {PROFIT/SETTLE*100:.1f}%')
