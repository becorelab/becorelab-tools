#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
쏘핑네흰집 X 일비아 공동구매 트래킹 (2026-06-22 ~ 06-25)

스마트스토어 '전체주문발주발송관리' 암호화 엑셀(.xlsx)을 입력하면:
  1) 주문을 '상품주문번호'로 마스터(orders.json)에 누적 병합
     → 발송완료로 목록에서 빠진 주문도 보존됨
  2) 결제일 기준 일별·시간대별 추이 집계
  3) 세트 구성을 풀어 제품 단위로 환산 (재고 차감용)
  4) 매출/정산예정 합계

사용법:
  python3 track.py "<암호화엑셀경로>" [비번(기본 1111)]
  python3 track.py --report          # 새 파일 없이 마스터로 리포트만 다시 출력

풀코스 기획세트의 '세제 2개'는 캡슐세제로 확정 환산 (대표님 확인 2026-06-22).
"""
import sys, os, re, io, json, csv
from collections import defaultdict
from datetime import datetime

import msoffcrypto
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
ORDERS_JSON = os.path.join(BASE, "orders.json")        # 마스터 (주문번호 기준 누적)
DAY_CSV = os.path.join(BASE, "일별_매출.csv")
DAYPROD_CSV = os.path.join(BASE, "일별_제품수량.csv")
HOUR_CSV = os.path.join(BASE, "시간대별_매출.csv")

# 윈도우 PC에서도 보도록 구글드라이브에 통합 엑셀 자동 저장
GDRIVE_DIR = os.path.expanduser(
    "~/Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브/Claude AI work space/02. 매출 정산/공동구매_쏘핑네흰집"
)
XLSX_OUT = os.path.join(GDRIVE_DIR, "공구_트래킹_쏘핑네흰집.xlsx")

PRODUCT_ORDER = [
    "캡슐세제", "하트 식세기세제", "캡슐표백제",
    "건조기시트(코튼블루)", "건조기시트(바이올렛)", "건조기시트(베이비크림)",
    "얼룩제거제", "섬유탈취제", "샘플키트",
]

# 과거 자사몰 공구 실적 (고객결제 GMV 기준, 정산서에서 산출 — 고정값)
# (주문수, 세트수량, 매출GMV)
PAST_GONGU = {
    "25년10월 (자사몰, 4일)": {"기간": 4, "전체": (445, 1064, 19516400), "Day1": (164, 403, 7284600)},
    "26년02월 (자사몰, 5일)": {"기간": 5, "전체": (368, 779, 16676400), "Day1": (145, 337, 6939300)},
}
SOPING_FEE = 0.28  # 쏘핑 인플루언서 공구 수수료 (자사몰·스토어 동일)


def compute_compare(master):
    """스토어 누적/Day1 메트릭을 산출해 과거 공구와 합친 비교용 dict 반환."""
    recs = list(master.values())
    days = sorted({(r["결제일"] or "")[:10] for r in recs if r["결제일"]})
    d1 = days[0] if days else ""

    def metrics(rs):
        gmv = sum(r["주문금액"] for r in rs)
        orders = len({r.get("주문번호") for r in rs})
        sets = sum(r["수량"] for r in rs)
        return (orders, sets, int(gmv))

    day1_recs = [r for r in recs if (r["결제일"] or "")[:10] == d1]
    store_label = f"26년06월 (스토어, {len(days)}일·진행중)"
    return {
        "store_label": store_label,
        "전체": {**{k: v["전체"] for k, v in PAST_GONGU.items()}, store_label: metrics(recs)},
        "Day1": {**{k: v["Day1"] for k, v in PAST_GONGU.items()}, store_label: metrics(day1_recs)},
    }


def decrypt(path, pw):
    with open(path, "rb") as fh:
        off = msoffcrypto.OfficeFile(fh)
        off.load_key(password=pw)
        out = io.BytesIO()
        off.decrypt(out)
    out.seek(0)
    return out


def explode_products(name, opt, q):
    """옵션 구성을 풀어 {제품: 개수}를 q배로 반환. 미분류는 note로."""
    prod = defaultdict(int)
    note = None

    def add(p, n):
        prod[p] += n * q

    if "단독 공구" in name:
        if "캡슐 세제" in opt and "2+2" in opt:
            add("캡슐세제", 4)
        elif "캡슐 세제" in opt and "1+1" in opt:
            add("캡슐세제", 2)
        elif "식기세척기" in opt and "3개" in opt:
            add("하트 식세기세제", 3)
        elif "표백제" in opt and "2+2" in opt:
            add("캡슐표백제", 4)
        elif "표백제" in opt and "1+1" in opt:
            add("캡슐표백제", 2)
        elif "풀코스" in opt:  # 건조3종 + 캡슐세제2 + 표백제2 + 얼룩1
            add("건조기시트(코튼블루)", 1); add("건조기시트(바이올렛)", 1); add("건조기시트(베이비크림)", 1)
            add("캡슐세제", 2); add("캡슐표백제", 2); add("얼룩제거제", 1)
        elif "건조기 시트" in opt:
            comp = opt.split("구성 선택:")[-1]
            if "코튼1개+바이올렛1개+베이비1개" in comp:
                add("건조기시트(코튼블루)", 1); add("건조기시트(바이올렛)", 1); add("건조기시트(베이비크림)", 1)
            elif "코튼2개+바이올렛2개+베이비2개" in comp:
                add("건조기시트(코튼블루)", 2); add("건조기시트(바이올렛)", 2); add("건조기시트(베이비크림)", 2)
            elif "코튼" in comp and "6개" in comp:
                add("건조기시트(코튼블루)", 6)
            elif "코튼" in comp and "3개" in comp:
                add("건조기시트(코튼블루)", 3)
            elif "바이올렛" in comp and "6개" in comp:
                add("건조기시트(바이올렛)", 6)
            elif "바이올렛" in comp and "3개" in comp:
                add("건조기시트(바이올렛)", 3)
            elif "베이비" in comp and "6개" in comp:
                add("건조기시트(베이비크림)", 6)
            elif "베이비" in comp and "3개" in comp:
                add("건조기시트(베이비크림)", 3)
            else:
                note = "미분류 건조기: " + comp[:40]
        else:
            note = "미분류 공구옵션: " + opt[:50]
    else:  # 단품 (구체적 조건부터 — 묶음 수량은 상품명/옵션에서 파싱)
        # 샘플류 (식세기 6정 무료/테스트, 건조기 2매 등 — 종류 무관 합산)
        if "샘플 키트" in name or "샘플키트" in name or "테스트샘플" in name or "맛보기용 샘플" in opt:
            add("샘플키트", 1)
        # 식세기세제 선물세트 (1세트 = 하트 식세기세제 2개, 대표님 확인 2026-06-22)
        elif "선물세트" in name:
            add("하트 식세기세제", 2)
        # 식세기세제 3개 묶음 (옵션 "하트 식세기세제x3개")
        elif "식세기세제x3개" in opt or ("식기세척기" in name and "60정 x" in name):
            add("하트 식세기세제", 3)
        elif "하트 식세기 세제 60" in name:
            add("하트 식세기세제", 1)
        # 식세기세제 60개입 단품 — 상품명 "60개입, N개" 수량 파싱 (없으면 1)
        elif ("식기세척기" in name or "식세기" in name) and "60개입" in name:
            m = re.search(r"60개입,?\s*(\d+)개", name)
            add("하트 식세기세제", int(m.group(1)) if m else 1)
        elif "식기세척기" in name and "2개" in name:
            add("하트 식세기세제", 2)
        elif "식기세척기" in name and "1개" in name:
            add("하트 식세기세제", 1)
        # 캡슐 세탁세제 초고농축 — 옵션의 "캡슐세제 N개" 수량 파싱
        elif "초고농축" in name:
            m = re.search(r"캡슐세제\s*(\d+)개", opt)
            add("캡슐세제", int(m.group(1)) if m else 1)
        elif "캡슐 세제 30" in name or "캡슐세제 30" in name:
            add("캡슐세제", 1)
        elif "캡슐 표백제" in name or "캡슐표백제" in name:
            add("캡슐표백제", 1)
        elif "얼룩제거제" in name:
            add("얼룩제거제", 1)
        elif "섬유탈취제" in name:
            add("섬유탈취제", 1)
        # 퍼퓸 건조기시트 단품 — 수량: 상품명 "40매, N개" 또는 옵션 "N개입" (향은 상품명 기준)
        elif ("퍼퓸 건조기 시트" in name or "향기" in name) and "40매" in name:
            m = re.search(r"40매,\s*(\d+)개", name) or re.search(r"(\d+)개입", opt)
            cnt = int(m.group(1)) if m else 1
            if "코튼" in name:
                add("건조기시트(코튼블루)", cnt)
            elif "바이올렛" in name:
                add("건조기시트(바이올렛)", cnt)
            elif "베이비" in name:
                add("건조기시트(베이비크림)", cnt)
            else:
                note = "미분류 퍼퓸시트: " + name[:40]
        elif "베이비크림 1박스" in name:
            add("건조기시트(베이비크림)", 1)
        elif "코튼블루 1박스" in name:
            add("건조기시트(코튼블루)", 1)
        elif "바이올렛 1박스" in name:
            add("건조기시트(바이올렛)", 1)
        else:
            note = "미분류 단품: " + name[:50]
    return dict(prod), note


def load_master():
    if os.path.exists(ORDERS_JSON):
        with open(ORDERS_JSON, encoding="utf-8") as f:
            return json.load(f)
    return {}


def ingest(path, pw, master):
    """엑셀을 읽어 master(주문번호->주문) 에 병합. (신규건수, 갱신건수) 반환."""
    wb = openpyxl.load_workbook(decrypt(path, pw))
    ws = wb["발주발송관리"]
    H = {ws.cell(2, c).value: c for c in range(1, ws.max_column + 1)}
    new_cnt = upd_cnt = 0
    for r in range(3, ws.max_row + 1):
        oid = ws.cell(r, H["상품주문번호"]).value
        name = ws.cell(r, H["상품명"]).value
        if oid is None or name is None:
            continue
        oid = str(oid)
        pay = ws.cell(r, H["결제일"]).value
        if isinstance(pay, datetime):
            pay = pay.strftime("%Y-%m-%d %H:%M:%S")
        rec = {
            "주문번호": str(ws.cell(r, H["주문번호"]).value),
            "결제일": pay,
            "상품명": name,
            "옵션": ws.cell(r, H["옵션정보"]).value or "(옵션없음)",
            "수량": ws.cell(r, H["수량"]).value or 0,
            "주문금액": ws.cell(r, H["최종 상품별 총 주문금액"]).value or 0,
            "정산예정": ws.cell(r, H["정산예정금액"]).value or 0,
            "주문상태": ws.cell(r, H["주문상태"]).value,
        }
        if oid in master:
            upd_cnt += 1
        else:
            new_cnt += 1
        master[oid] = rec
    return new_cnt, upd_cnt


def build_report(master):
    """master 전체에서 일별/시간대별/제품별 집계."""
    day = defaultdict(lambda: dict(cnt=0, qty=0, amt=0, settle=0, prod=defaultdict(int)))
    hour = defaultdict(lambda: dict(cnt=0, qty=0, amt=0, settle=0))
    sku = defaultdict(lambda: [0, 0, 0, 0])  # (상품,옵션) -> [건수,수량,주문금액,정산예정]
    prod_total = defaultdict(int)
    tot = dict(cnt=0, qty=0, amt=0, settle=0)
    notes = set()

    for oid, rec in master.items():
        pay = rec["결제일"] or ""
        d = pay[:10] if pay else "(미상)"
        h = pay[11:13] if len(pay) >= 13 else "??"
        q = rec["수량"]; amt = rec["주문금액"]; st = rec["정산예정"]

        for bucket, key in ((day, d), (hour, f"{d} {h}시")):
            bucket[key]["cnt"] += 1; bucket[key]["qty"] += q
            bucket[key]["amt"] += amt; bucket[key]["settle"] += st
        s = sku[(rec["상품명"], rec["옵션"])]
        s[0] += 1; s[1] += q; s[2] += amt; s[3] += st
        tot["cnt"] += 1; tot["qty"] += q; tot["amt"] += amt; tot["settle"] += st

        pmap, note = explode_products(rec["상품명"], rec["옵션"], q)
        if note:
            notes.add(note)
        for p, n in pmap.items():
            prod_total[p] += n
            day[d]["prod"][p] += n

    return day, hour, sku, prod_total, tot, notes


def write_csvs(day, hour, prod_total):
    days = sorted(day)
    with open(DAY_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["일자", "주문건수", "수량(세트)", "주문금액", "정산예정"])
        for d in days:
            v = day[d]
            w.writerow([d, v["cnt"], v["qty"], v["amt"], v["settle"]])
    with open(DAYPROD_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["일자"] + PRODUCT_ORDER + ["건조기시트합계"])
        for d in days:
            pr = day[d]["prod"]
            dryer = sum(pr.get(p, 0) for p in PRODUCT_ORDER if "건조기시트" in p)
            w.writerow([d] + [pr.get(p, 0) for p in PRODUCT_ORDER] + [dryer])
    with open(HOUR_CSV, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["시간대", "주문건수", "수량(세트)", "주문금액", "정산예정"])
        for h in sorted(hour):
            v = hour[h]
            w.writerow([h, v["cnt"], v["qty"], v["amt"], v["settle"]])


def write_xlsx(day, hour, sku, prod_total, tot, cmp):
    """구글드라이브에 통합 엑셀 1개(시트 6개) 저장 — 윈도우 PC 열람용."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        return None
    if not os.path.isdir(os.path.dirname(XLSX_OUT)):
        try:
            os.makedirs(os.path.dirname(XLSX_OUT), exist_ok=True)
        except OSError:
            return None

    hdr_fill = PatternFill("solid", fgColor="DCE6F1")
    hdr_font = Font(bold=True)
    left = Alignment(horizontal="left")
    won = "#,##0"

    wb = Workbook()

    def sheet(title, headers, rows, num_cols=()):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(1, c)
            cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = left
        for row in rows:
            ws.append(row)
        for c in num_cols:
            for r in range(2, ws.max_row + 1):
                ws.cell(r, c).number_format = won
        for col in ws.columns:
            w = max((len(str(x.value)) for x in col if x.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 10), 50)
        for r in range(1, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).alignment = left
        return ws

    # 1) 요약
    dryer = sum(prod_total.get(p, 0) for p in PRODUCT_ORDER if "건조기시트" in p)
    sheet("요약", ["항목", "값"], [
        ["집계 시각", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["누적 주문건수", tot["cnt"]],
        ["총 수량(세트)", tot["qty"]],
        ["총 주문금액", tot["amt"]],
        ["정산예정금액", tot["settle"]],
        ["건조기시트 합계(개)", dryer],
    ], num_cols=[2])

    # 1.5) 공구비교 (Day1 + 전체/누적, 과거 자사몰 vs 현재 스토어)
    cs = wb.create_sheet("공구비교")

    def cmp_block(title, data):
        cs.append([title])
        cs.cell(cs.max_row, 1).font = Font(bold=True)
        cols = list(data.keys())
        cs.append(["지표"] + cols)
        for c in range(1, len(cols) + 2):
            cell = cs.cell(cs.max_row, c)
            cell.fill = hdr_fill; cell.font = hdr_font
        rowdefs = [
            ("주문(결제) 수", lambda m: m[0], False),
            ("세트 수량", lambda m: m[1], False),
            ("매출(고객결제)", lambda m: m[2], True),
            (f"쏘핑 {int(SOPING_FEE*100)}% 차감 후", lambda m: int(m[2] * (1 - SOPING_FEE)), True),
            ("객단가(주문당)", lambda m: int(m[2] / m[0]) if m[0] else 0, True),
            ("세트당 단가", lambda m: int(m[2] / m[1]) if m[1] else 0, True),
            ("주문당 세트수", lambda m: round(m[1] / m[0], 2) if m[0] else 0, False),
        ]
        for label, fn, money in rowdefs:
            cs.append([label] + [fn(data[c]) for c in cols])
            if money:
                for ci in range(2, len(cols) + 2):
                    cs.cell(cs.max_row, ci).number_format = won
        cs.append([])

    cmp_block("◆ 1일차(Day1) 비교", cmp["Day1"])
    cmp_block("◆ 전체/누적 비교", cmp["전체"])
    for col in cs.columns:
        w = max((len(str(x.value)) for x in col if x.value is not None), default=10)
        cs.column_dimensions[col[0].column_letter].width = min(max(w + 2, 12), 28)
    for row in cs.iter_rows():
        for c in row:
            c.alignment = left

    # 2) 일별 매출
    sheet("일별매출", ["일자", "주문건수", "수량(세트)", "주문금액", "정산예정"],
          [[d, day[d]["cnt"], day[d]["qty"], day[d]["amt"], day[d]["settle"]] for d in sorted(day)],
          num_cols=[4, 5])

    # 3) 일별 제품수량 (재고용)
    sheet("일별제품수량", ["일자"] + PRODUCT_ORDER + ["건조기시트합계"],
          [[d] + [day[d]["prod"].get(p, 0) for p in PRODUCT_ORDER]
           + [sum(day[d]["prod"].get(p, 0) for p in PRODUCT_ORDER if "건조기시트" in p)]
           for d in sorted(day)])

    # 4) 시간대별
    sheet("시간대별", ["시간대", "주문건수", "수량(세트)", "주문금액", "정산예정"],
          [[h, hour[h]["cnt"], hour[h]["qty"], hour[h]["amt"], hour[h]["settle"]] for h in sorted(hour)],
          num_cols=[4, 5])

    # 5) SKU별
    sheet("SKU별", ["상품", "옵션", "건수", "수량", "주문금액", "정산예정"],
          [[k[0], k[1], v[0], v[1], v[2], v[3]]
           for k, v in sorted(sku.items(), key=lambda x: -x[1][2])],
          num_cols=[5, 6])

    wb.remove(wb["Sheet"])
    wb.save(XLSX_OUT)
    return XLSX_OUT


def fmt(n):
    return f"{n:,}"


def report(master):
    day, hour, sku, prod_total, tot, notes = build_report(master)
    cmp = compute_compare(master)
    write_csvs(day, hour, prod_total)
    xlsx = write_xlsx(day, hour, sku, prod_total, tot, cmp)

    print("\n" + "=" * 52)
    print(f"📊 쏘핑네흰집 X 일비아 공구 트래킹 (누적 주문 {tot['cnt']}건)")
    print("=" * 52)
    print(f"총 수량 {tot['qty']}세트 / 주문금액 {fmt(tot['amt'])}원 / 정산예정 {fmt(tot['settle'])}원")

    print("\n[일별 추이]")
    print(f"  {'일자':<12}{'건수':>5}{'수량':>6}{'주문금액':>12}{'정산예정':>12}")
    for d in sorted(day):
        v = day[d]
        print(f"  {d:<12}{v['cnt']:>5}{v['qty']:>6}{fmt(v['amt']):>13}{fmt(v['settle']):>13}")

    print("\n[제품별 누적 판매 — 재고 차감용]")
    for p in PRODUCT_ORDER:
        if prod_total.get(p):
            print(f"  {p:22s} {prod_total[p]:>4d}개")
    dryer = sum(prod_total.get(p, 0) for p in PRODUCT_ORDER if "건조기시트" in p)
    print(f"  └ 건조기시트 합계         {dryer:>4d}개")

    if notes:
        print("\n⚠️ 미분류:", *notes, sep="\n  - ")
    print(f"\n💾 CSV: 일별_매출 / 일별_제품수량 / 시간대별_매출  (폴더: {BASE})")
    if xlsx:
        print(f"☁️ 구글드라이브 엑셀: {xlsx}")
    else:
        print("⚠️ 구글드라이브 엑셀 저장 실패 (경로/동기화 확인 필요)")


def main():
    args = sys.argv[1:]
    master = load_master()

    if args and args[0] == "--report":
        report(master)
        return

    if not args:
        print("사용법: python3 track.py <암호화엑셀경로> [비번]  |  --report")
        sys.exit(1)

    path = args[0]
    pw = args[1] if len(args) > 1 else "1111"
    new_cnt, upd_cnt = ingest(path, pw, master)
    with open(ORDERS_JSON, "w", encoding="utf-8") as f:
        json.dump(master, f, ensure_ascii=False, indent=1)
    print(f"📥 {os.path.basename(path)} 반영: 신규 {new_cnt}건 / 갱신 {upd_cnt}건 → 마스터 총 {len(master)}건")
    report(master)


if __name__ == "__main__":
    main()
