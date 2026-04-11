"""
카페24 14개월 검증 스크립트 v2 (최종)
- 옛날 담당자 방식: 취소/반품 사전 제외 (배송완료/구매확정/배송중/배송준비중만)
- 주간 정산 방식: 행마다 모든 할인 직접 차감
- 월별 품명리스트 + 월별 원가 시트 사용 (정확도 ↑)
- 월별 정산 파일의 카페24 매출/이익과 비교
"""
import sys, io, os, re, shutil, tempfile
import openpyxl
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

VERIFY = 'N:/개인/Becorelab/03. 영업/20. 월별 매출정산/기타/자사몰 검증'
MONTHLY_2025 = 'N:/개인/Becorelab/03. 영업/20. 월별 매출정산/2025년'
MONTHLY_2026 = 'N:/개인/Becorelab/03. 영업/20. 월별 매출정산'

# 포함할 주문 상태 (보수적: 배송 흐름에 있는 것만)
INCLUDED_STATUS = {
    '배송 완료', '구매 확정', '배송중', '배송 준비중'
}

def sf(v):
    if v is None or v == '': return 0.0
    try: return float(str(v).replace(',', ''))
    except: return 0.0

def safe_copy_load(src):
    """N드라이브 파일을 임시로 복사 후 로드 (권한 문제 회피)"""
    fd, dst = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)
    shutil.copy(src, dst)
    wb = openpyxl.load_workbook(dst, data_only=True)
    return wb, dst

# 월별 파일 경로 매핑
MONTH_FILES = {}
# 2025
for month in range(1, 13):
    mm = f'{month:02d}'
    folder = f'{MONTHLY_2025}/2025.{mm}'
    if os.path.exists(folder):
        for f in os.listdir(folder):
            fl = f.lower()
            if ('카페24' in f or '자사몰' in f) and (fl.endswith('.xlsx') or fl.endswith('.xls')):
                if not any(k in f for k in ['적립금', '복사본', '안나', '성소미', '쏘핑', '채아맘', '회송', '온라인']):
                    MONTH_FILES[f'2025-{mm}'] = {
                        'mapping': f'{folder}/{f}',
                    }
                    break
# 2026
for month, fname in [(1, '01월 카페24.xlsx'), (2, '02월 카페24 (version 1).xlsx')]:
    fp = f'{MONTHLY_2026}/2026.{month:02d}/{fname}'
    if os.path.exists(fp):
        MONTH_FILES[f'2026-{month:02d}'] = {'mapping': fp}

# 원본 + 정산 파일 추가
for label in MONTH_FILES:
    yr, mn = label.split('-')
    yr_short = yr[-2:]
    raw_file = f'{VERIFY}/{yr_short}. {mn} 카페24 원본 데이터.xlsx'
    if os.path.exists(raw_file):
        MONTH_FILES[label]['raw'] = raw_file

    # 정산 파일 — 검증 폴더 우선
    pattern = f'{yr}. {mn} 온라인 매출정산'
    found = False
    for f in os.listdir(VERIFY):
        if f.startswith(pattern) and f.endswith('.xlsx'):
            MONTH_FILES[label]['settlement'] = f'{VERIFY}/{f}'
            found = True
            break
    # 검증 폴더에 없으면 월별 폴더에서 찾기
    if not found:
        if yr == '2025':
            month_folder = f'{MONTHLY_2025}/2025.{mn}'
        else:
            month_folder = f'{MONTHLY_2026}/2026.{mn}'
        if os.path.exists(month_folder):
            for f in os.listdir(month_folder):
                if f.startswith(pattern) and f.endswith('.xlsx'):
                    MONTH_FILES[label]['settlement'] = f'{month_folder}/{f}'
                    break

print(f'[1/4] 월별 파일 매핑 완료: {len(MONTH_FILES)}개월')
for label in sorted(MONTH_FILES):
    files = MONTH_FILES[label]
    print(f'  {label}: 매핑={"O" if "mapping" in files else "X"}, 원본={"O" if "raw" in files else "X"}, 정산={"O" if "settlement" in files else "X"}')


def load_name_map(filepath):
    """월별 카페24 파일의 품명리스트 시트 로드"""
    wb, dst = safe_copy_load(filepath)
    name_map = {}
    if '품명리스트' in wb.sheetnames:
        ws = wb['품명리스트']
        for row in ws.iter_rows(min_row=2, values_only=True):
            orig = str(row[0] or '').strip() if row[0] else ''
            mapped = str(row[1] or '').strip() if len(row) > 1 and row[1] else ''
            if orig and mapped:
                name_map[orig] = mapped
    os.remove(dst)
    return name_map


def load_cost_map(filepath):
    """월별 정산 파일의 원가 시트 로드"""
    wb, dst = safe_copy_load(filepath)
    cost_map = {}
    cost_norm = {}
    if '원가' in wb.sheetnames:
        ws = wb['원가']
        # Row 2~4 특수 항목
        for r in [2, 3, 4]:
            name = str(ws.cell(row=r, column=1).value or '').strip()
            cost = ws.cell(row=r, column=3).value
            if name and cost:
                try:
                    c = float(cost)
                    cost_map[name] = c
                    cost_norm[re.sub(r'\s+', '', name)] = c
                except: pass
        # Row 7부터
        for row in ws.iter_rows(min_row=7, values_only=True):
            name = str(row[1] or '').strip() if row[1] else ''
            cost = row[4] if len(row) > 4 else None
            if name and cost:
                try:
                    c = float(cost)
                    cost_map[name] = c
                    cost_norm[re.sub(r'\s+', '', name)] = c
                except: pass
    os.remove(dst)
    return cost_map, cost_norm


def get_settlement_cafe24(filepath, month):
    """월별 정산 파일의 채널별 매출 이익 시트에서 카페24 매출/이익 추출"""
    wb, dst = safe_copy_load(filepath)
    if '채널별 매출 이익' not in wb.sheetnames:
        os.remove(dst)
        return None
    ws = wb['채널별 매출 이익']

    # 월 헤더 행 찾기 (Row 12 근처)
    month_header_row = None
    col_header_row = None
    for r in range(8, 20):
        for c in range(2, 50):
            v = ws.cell(row=r, column=c).value
            if v and f'{month}월' == str(v).strip():
                month_header_row = r
                col_header_row = r + 1
                month_start_col = c
                break
        if month_header_row: break

    if not month_header_row:
        os.remove(dst)
        return None

    # 매출/이익 컬럼 찾기 (월별 3컬럼: 매출/이익/이익률)
    rev_col = profit_col = None
    for c in range(month_start_col, month_start_col + 3):
        h = ws.cell(row=col_header_row, column=c).value
        if h and '매출' in str(h) and rev_col is None:
            rev_col = c
        elif h and '이익' in str(h) and '률' not in str(h) and profit_col is None:
            profit_col = c

    if not rev_col or not profit_col:
        os.remove(dst)
        return None

    # 카페24 행 찾기 (B열 = "카페24 / 네이버페이")
    cafe_revenue = cafe_profit = 0
    for r in range(col_header_row + 1, ws.max_row + 1):
        b = ws.cell(row=r, column=2).value
        if b and '카페24' in str(b):
            rev = ws.cell(row=r, column=rev_col).value
            profit = ws.cell(row=r, column=profit_col).value
            if isinstance(rev, (int, float)): cafe_revenue += rev
            if isinstance(profit, (int, float)): cafe_profit += profit
            break  # 카페24는 한 행

    os.remove(dst)
    return {
        'revenue': cafe_revenue,
        'profit': cafe_profit,
        'rev_with_ship': cafe_revenue,  # 채널별 시트는 매출+배송 통합값
    }


def process_month(label, files):
    """한 달 카페24 데이터 처리 (옛날 담당자 + 주간 정산 방식)"""
    if 'raw' not in files or 'mapping' not in files:
        return None

    # 1. 품명 매핑 로드
    name_map = load_name_map(files['mapping'])

    # 2. 원가 매핑 로드
    cost_map = {}
    cost_norm = {}
    if 'settlement' in files:
        cost_map, cost_norm = load_cost_map(files['settlement'])

    # 3. 원본 처리
    wb, dst = safe_copy_load(files['raw'])
    ws = wb['원본']
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]

    def col_idx(kw):
        for i, h in enumerate(headers):
            if h and kw in str(h): return i
        return None

    IDX_ORDER = col_idx('주문번호')
    IDX_QTY = col_idx('수량')
    IDX_OPT_PRICE = col_idx('옵션+판매가')
    IDX_NAME = col_idx('주문상품명(옵션포함)')
    IDX_SHIP = col_idx('총 배송비(KRW)')
    IDX_COUPON = col_idx('쿠폰 할인금액(최초)')
    IDX_POINT = col_idx('사용한 적립금액(최종)')
    IDX_RANK = col_idx('회원등급 추가할인금액')
    IDX_APP = col_idx('앱 상품할인 금액(최종)')
    IDX_EXTRA = col_idx('상품별 추가할인금액')
    IDX_STATUS = col_idx('주문 상태')

    orders_seen = set()
    sum_gross = sum_ship = sum_coupon = sum_point = sum_rank = sum_app = sum_extra = 0
    total_payment = total_cost = total_qty = 0
    excluded_qty = 0
    excluded_count = 0
    no_name_map = 0
    no_cost = 0
    no_cost_value = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[IDX_ORDER]: continue

        # 주문 상태 필터 (옛날 담당자 방식 — 보수적)
        status = str(row[IDX_STATUS] or '').strip()
        if status not in INCLUDED_STATUS:
            excluded_count += 1
            qty_ex = int(sf(row[IDX_QTY]))
            opt_ex = sf(row[IDX_OPT_PRICE])
            excluded_qty += qty_ex
            continue

        order = str(row[IDX_ORDER]).strip()
        qty = int(sf(row[IDX_QTY]))
        opt_price = sf(row[IDX_OPT_PRICE])
        gross = qty * opt_price

        is_first = order not in orders_seen
        if is_first: orders_seen.add(order)

        # 주문 단위: 첫 행에만 카운트
        coupon = sf(row[IDX_COUPON]) if is_first and IDX_COUPON is not None else 0
        point = sf(row[IDX_POINT]) if is_first and IDX_POINT is not None else 0
        ship = sf(row[IDX_SHIP]) if is_first and IDX_SHIP is not None else 0

        # 행 단위
        rank = sf(row[IDX_RANK]) if IDX_RANK is not None else 0
        app = sf(row[IDX_APP]) if IDX_APP is not None else 0
        extra = sf(row[IDX_EXTRA]) if IDX_EXTRA is not None else 0

        # 최종결제금액 (주간 정산 방식)
        payment = gross - coupon - point - rank - app - extra

        # 원가 계산
        raw_name = str(row[IDX_NAME] or '').strip()
        final = name_map.get(raw_name, '')
        unit_cost = 0
        if final:
            unit_cost = cost_map.get(final, 0)
            if unit_cost == 0:
                unit_cost = cost_norm.get(re.sub(r'\s+', '', final), 0)
            if unit_cost == 0:
                no_cost += 1
                no_cost_value += gross
        else:
            no_name_map += 1

        cost = unit_cost * qty

        total_payment += payment
        total_cost += cost
        total_qty += qty

        sum_gross += gross
        sum_coupon += coupon
        sum_point += point
        sum_rank += rank
        sum_app += app
        sum_extra += extra
        sum_ship += ship

    os.remove(dst)

    return {
        'qty': total_qty,
        'gross': sum_gross,
        'shipping': sum_ship,
        'coupon': sum_coupon,
        'point': sum_point,
        'rank': sum_rank,
        'app': sum_app,
        'extra': sum_extra,
        'payment': total_payment,
        'rev_with_ship': total_payment + sum_ship,
        'cost': total_cost,
        'profit': total_payment - total_cost,
        'margin_rate': (total_payment - total_cost) / total_payment if total_payment > 0 else 0,
        'excluded_count': excluded_count,
        'excluded_qty': excluded_qty,
        'no_name_map': no_name_map,
        'no_cost': no_cost,
        'no_cost_value': no_cost_value,
        'name_map_size': len(name_map),
        'cost_map_size': len(cost_map),
    }


# === 메인 실행 ===
print(f'\n[2/4] 14개월 처리 시작...')
results = {}
settlement_results = {}

for label in sorted(MONTH_FILES):
    print(f'  {label} 처리 중...')
    files = MONTH_FILES[label]

    # 새 방식 계산
    r = process_month(label, files)
    if r:
        results[label] = r

    # 정산 파일에서 카페24 매출 추출
    if 'settlement' in files:
        month_num = int(label.split('-')[1])
        s = get_settlement_cafe24(files['settlement'], month_num)
        if s:
            settlement_results[label] = s

# === 결과 출력 ===
print(f'\n[3/4] 결과 출력\n')
print('=' * 140)
print(f'  카페24 14개월 검증 — 새 방식 vs 종전 정산')
print('=' * 140)

print(f"\n{'월':<10} {'[종전 정산]':<35} {'[새 방식 (옛담당자+주간)]':<45} {'차이':<25}")
print(f"{'':<10} {'매출+배송':>13} {'이익':>12} {'마진율':>7} | {'매출+배송':>13} {'이익':>12} {'마진율':>7} | {'매출차이':>12} {'이익차이':>12}")
print('-' * 140)

total_old_rev = total_old_pr = 0
total_new_rev = total_new_pr = 0

for label in sorted(results.keys()):
    new = results[label]
    old = settlement_results.get(label, {})

    new_rev = new['rev_with_ship']
    new_pr = new['profit']
    new_mr = new['margin_rate'] * 100

    old_rev = old.get('rev_with_ship', 0)
    old_pr = old.get('profit', 0)
    old_mr = (old_pr / old_rev * 100) if old_rev > 0 else 0

    rev_diff = new_rev - old_rev
    pr_diff = new_pr - old_pr

    total_old_rev += old_rev
    total_old_pr += old_pr
    total_new_rev += new_rev
    total_new_pr += new_pr

    print(f"{label:<10} {old_rev:>13,.0f} {old_pr:>12,.0f} {old_mr:>6.1f}% | {new_rev:>13,.0f} {new_pr:>12,.0f} {new_mr:>6.1f}% | {rev_diff:>+12,.0f} {pr_diff:>+12,.0f}")

print('-' * 140)
total_old_mr = total_old_pr / total_old_rev * 100 if total_old_rev > 0 else 0
total_new_mr = total_new_pr / total_new_rev * 100 if total_new_rev > 0 else 0
total_rev_diff = total_new_rev - total_old_rev
total_pr_diff = total_new_pr - total_old_pr
print(f"{'합계':<10} {total_old_rev:>13,.0f} {total_old_pr:>12,.0f} {total_old_mr:>6.1f}% | {total_new_rev:>13,.0f} {total_new_pr:>12,.0f} {total_new_mr:>6.1f}% | {total_rev_diff:>+12,.0f} {total_pr_diff:>+12,.0f}")

print(f'\n[4/4] 데이터 품질 점검')
for label in sorted(results.keys()):
    r = results[label]
    issues = []
    if r['no_name_map'] > 0:
        issues.append(f"미매핑 {r['no_name_map']}")
    if r['no_cost'] > 0:
        issues.append(f"원가없음 {r['no_cost']} (₩{r['no_cost_value']:,.0f})")
    if r['excluded_count'] > 0:
        issues.append(f"제외 {r['excluded_count']}건")
    if issues:
        print(f"  {label}: {', '.join(issues)} | 매핑={r['name_map_size']}, 원가={r['cost_map_size']}")
