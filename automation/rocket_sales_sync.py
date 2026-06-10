"""쿠팡 로켓배송(1P 직매입) 매출 → ERP sales 반영

로켓 1P는 쿠팡 직매입이라 "발주(공급가액)" = 우리 매출.
소스: 서플라이허브 입고상세내역 엑셀 (Coupang_Stocked_Data_List)
  - 구분 '발주' = 매출(+), '반출' = 차감(-)
  - 총공급가액(col12) = 공급가, 총세액(col13) = 세액
사용: python3 rocket_sales_sync.py <엑셀경로>
※ 자동화: supplyhub_scraper로 입고상세 엑셀 다운로드(Akamai 세션재활용) 후 이 스크립트 호출.
"""
import openpyxl, sqlite3, sys, re
from collections import defaultdict

ERP_DB = "/Users/macmini_ky/ClaudeAITeam/erp/erp.db"
CHANNEL = "비코어랩 쿠팡 로켓배송"

def num(v):
    if v is None: return 0
    if isinstance(v,(int,float)): return v
    try: return float(str(v).replace(',','').strip())
    except Exception: return 0

def parse(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['data'] if 'data' in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0])
    def col(key):
        for i,h in enumerate(hdr):
            if h and key in str(h): return i
        return None
    c_time, c_supply, c_tax = col('입고/반출시각'), col('총공급가'), col('총세액')
    daily = defaultdict(lambda: {'supply':0.0,'tax':0.0})
    for r in rows[1:]:
        if not r[0]: continue
        d = str(r[c_time] or '')[:10].replace('/','-')
        if not re.match(r'\d{4}-\d{2}-\d{2}', d): continue
        sign = 1 if r[0]=='발주' else -1
        daily[d]['supply'] += num(r[c_supply]) * sign
        daily[d]['tax'] += num(r[c_tax]) * sign
    return daily

def upsert(daily):
    conn = sqlite3.connect(ERP_DB, timeout=30); n=0
    for d, v in daily.items():
        supply, tax = round(v['supply']), round(v['tax'])
        amount = supply + tax
        if amount == 0: continue
        conn.execute("DELETE FROM sales WHERE channel=? AND sale_date=?", (CHANNEL, d))
        conn.execute("""INSERT INTO sales
            (sale_date,partner_id,channel,channel_order_no,total_supply,total_tax,total_amount,status,source)
            VALUES (?,?,?,?,?,?,?, 'confirmed','supplyhub')""",
            (d, None, CHANNEL, f"ROCKET-{d}", supply, tax, amount))
        n+=1
    conn.commit(); conn.close(); return n

if __name__ == "__main__":
    path = sys.argv[1]
    daily = parse(path)
    n = upsert(daily)
    tot = sum(round(v['supply'])+round(v['tax']) for v in daily.values())
    print(f"✅ 로켓 1P {n}일 → ERP sales 반영. 총 {tot:,}원 (공급가+세액)")
