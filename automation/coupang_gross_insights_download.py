#!/usr/bin/python3
"""쿠팡 Wing 판매분석의 상품별 판매 리포트를 다운로드하고 구조를 검증한다.

기본 대상은 어제 1일이다. 화면의 12:30 안내는 트래픽·전환율 그래프
표시 시각이며 판매량·매출 리포트 확정 시각이 아니다.

사용:
  /usr/bin/python3 automation/coupang_gross_insights_download.py
  /usr/bin/python3 automation/coupang_gross_insights_download.py --validate FILE
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import Locator, Page, sync_playwright

BASE = Path(__file__).resolve().parent
CDP_URL = "http://127.0.0.1:9222"
SALES_URL = "https://wing.coupang.com/tenants/business-insight/sales-analysis"
TRANSFER_DIR = (
    Path.home()
    / "Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브"
    / "Claude AI work space/mac window file transfer"
)
ACCOUNT_ID = "A00940134"
ACCOUNT_NAME = "채움컴퍼니"
REPORT_LABEL = "상품별 판매 리포트"
EXPECTED_SHEET = "vendor item metrics"
REQUIRED_HEADERS = {
    "옵션 ID", "옵션명", "상품명", "등록상품ID", "판매방식",
    "매출(원)", "주문", "판매량", "총 매출(원)", "총 판매수",
    "총 취소 금액(원)", "총 취소된 상품수",
}


def _number(value, cell: str) -> int:
    try:
        if value is None or str(value).strip() == "":
            raise ValueError("빈 값")
        number = float(str(value).replace(",", "").strip())
        if not number.is_integer():
            raise ValueError("정수가 아님")
        return int(number)
    except (TypeError, ValueError):
        raise ValueError(f"숫자 셀 손상: {cell}={value!r}")


def _empty_totals() -> dict:
    return {"rows": 0, "net_sales": 0, "orders": 0, "units": 0,
            "gross_sales": 0, "gross_units": 0, "cancel_sales": 0,
            "cancel_units": 0}


def validate_report(path: Path) -> tuple[bool, str, dict]:
    """상품별 판매 리포트의 열·행·합계를 검증한다."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        if ws.title.strip().lower() != EXPECTED_SHEET:
            wb.close()
            return False, f"시트명이 예상과 다름: {ws.title}", {}
        rows = ws.iter_rows(values_only=True)
        first = next(rows, ())
        headers = [str(v).strip() if v is not None else "" for v in first]
        duplicates = sorted({h for h in headers if h and headers.count(h) > 1})
        if duplicates:
            wb.close()
            return False, f"중복 헤더: {', '.join(duplicates)}", {}
        index = {name: i for i, name in enumerate(headers)}
        missing = sorted(REQUIRED_HEADERS - set(headers))
        if missing:
            wb.close()
            return False, f"필수열 누락: {', '.join(missing)}", {}

        totals = {"all": _empty_totals(), "rocket_gross": _empty_totals(),
                  "seller_delivery": _empty_totals(), "other": _empty_totals()}
        methods = set()
        option_ids = set()
        for excel_row, row in enumerate(rows, start=2):
            if not any(v not in (None, "") for v in row):
                continue
            option_id = str(row[index["옵션 ID"]]).strip()
            if not option_id or option_id == "None":
                raise ValueError(f"옵션 ID 누락: {excel_row}행")
            if option_id in option_ids:
                raise ValueError(f"옵션 ID 중복: {option_id} ({excel_row}행)")
            option_ids.add(option_id)
            method = str(row[index["판매방식"]]).strip()
            methods.add(method)
            values = {
                "net_sales": _number(row[index["매출(원)"]], f"G{excel_row}"),
                "orders": _number(row[index["주문"]], f"H{excel_row}"),
                "units": _number(row[index["판매량"]], f"I{excel_row}"),
                "gross_sales": _number(row[index["총 매출(원)"]], f"O{excel_row}"),
                "gross_units": _number(row[index["총 판매수"]], f"P{excel_row}"),
                "cancel_sales": _number(row[index["총 취소 금액(원)"]], f"Q{excel_row}"),
                "cancel_units": _number(row[index["총 취소된 상품수"]], f"R{excel_row}"),
            }
            if values["cancel_sales"] > 0 or values["cancel_units"] > 0:
                raise ValueError(f"취소값 부호 오류: {excel_row}행")
            if values["gross_sales"] + values["cancel_sales"] != values["net_sales"]:
                raise ValueError(f"매출 취소반영식 불일치: {excel_row}행")
            if values["gross_units"] + values["cancel_units"] != values["units"]:
                raise ValueError(f"판매수 취소반영식 불일치: {excel_row}행")
            bucket = ("rocket_gross" if method == "로켓그로스" else
                      "seller_delivery" if method == "판매자배송" else "other")
            for key in ("all", bucket):
                totals[key]["rows"] += 1
                for metric, value in values.items():
                    totals[key][metric] += value
        wb.close()
        if totals["all"]["rows"] == 0:
            return False, "데이터 행 0개", {}
        if totals["rocket_gross"]["rows"] == 0:
            return False, f"판매방식이 로켓그로스가 아님: {sorted(methods)}", {}
        summary = {
            **totals,
            "sales_methods": sorted(methods),
        }
        gross = totals["rocket_gross"]
        extra = totals["seller_delivery"]
        reason = (f"로켓그로스 {gross['rows']}개 옵션 / 순매출 {gross['net_sales']:,}원 / "
                  f"순판매 {gross['units']:,}개")
        if extra["rows"]:
            reason += (f" + 판매자배송 별도 {extra['rows']}개 / "
                       f"{extra['net_sales']:,}원 / {extra['units']:,}개")
        return True, reason, summary
    except Exception as exc:
        return False, f"XLSX 검증 실패: {type(exc).__name__}: {exc}", {}


def _logged_in(page: Page) -> bool:
    url = page.url.lower()
    if "xauth" in url or "login" in url:
        return False
    try:
        return "access denied" not in page.locator("body").inner_text(timeout=5000).lower()
    except Exception:
        return False


def _open_sales_page(context) -> Page:
    """현재 CDP 세션으로 판매분석 페이지를 열고 로그인 상태를 검증한다."""
    page = context.new_page()
    page.goto(SALES_URL, wait_until="domcontentloaded", timeout=60000)
    if _logged_in(page):
        return page
    page.close()

    raise RuntimeError("Wing 세션 만료")


def _ensure_session() -> None:
    """세션을 먼저 검사하고 만료면 별도 브라우저 사이클에서 브릿지를 한 번만 시도한다."""
    alive = False
    with sync_playwright() as playwright:
        browser = playwright.chromium.connect_over_cdp(CDP_URL, timeout=30000)
        context = browser.contexts[0]
        page = context.new_page()
        try:
            page.goto(SALES_URL, wait_until="domcontentloaded", timeout=60000)
            alive = _logged_in(page)
        finally:
            page.close()
    if alive:
        return

    sys.path.insert(0, str(BASE))
    try:
        from gross_cookie_bridge import refresh_from_user_chrome
        bridged = refresh_from_user_chrome(verbose=True)
    except Exception as exc:
        raise RuntimeError(f"일반 Chrome 쿠키 브릿지 실패: {exc}") from exc
    if not bridged:
        raise RuntimeError("Wing 세션 만료 — 일반 Chrome에서 로그인 1회 필요")


def _last_visible(locator: Locator) -> Locator | None:
    for i in range(locator.count() - 1, -1, -1):
        item = locator.nth(i)
        try:
            if item.is_visible():
                return item
        except Exception:
            pass
    return None


def _select_and_verify_yesterday(page: Page, data_date: date) -> None:
    """어제 퀵필터를 직접 누르고, 실제 조회 요청에도 대상일이 들어갔는지 확인한다."""
    requests = []
    def capture(request):
        if "vi-detail-search" in request.url and request.method == "POST":
            requests.append(request.post_data or "")
    page.on("request", capture)
    choices = page.get_by_text("어제", exact=True)
    choice = _last_visible(choices)
    if choice is None:
        raise RuntimeError("판매분석의 '어제' 기간 버튼을 찾지 못함")
    # 이미 '어제'가 선택돼 있어 클릭 이벤트가 생략되는 UI도 있어 한 번 다른
    # 퀵필터로 이동한 뒤 돌아온다. 최종 조회 payload만 검증한다.
    other = _last_visible(page.get_by_text("최근 7일", exact=True))
    if other is not None:
        other.click()
        try:
            page.wait_for_load_state("networkidle", timeout=15000)
        except Exception:
            pass
        requests.clear()
    choice.click()
    try:
        page.wait_for_load_state("networkidle", timeout=15000)
    except Exception:
        pass
    target = data_date.isoformat()
    exact_date_seen = False
    for payload in requests:
        try:
            request_body = json.loads(payload)
            if (request_body.get("startDate") == target
                    and request_body.get("endDate") == target):
                exact_date_seen = True
                break
        except (TypeError, ValueError, json.JSONDecodeError):
            continue
    if not exact_date_seen:
        raise RuntimeError("'어제' 선택 후 실제 조회 요청의 시작일·종료일을 검증하지 못함")


def _verify_account(page: Page) -> None:
    # storage/cookie에는 과거에 열었던 다른 계정도 남을 수 있으므로 증거로 쓰지
    # 않는다. 현재 화면 헤더에 실제로 보이는 선택 업체명만 허용한다.
    visible_account = _last_visible(page.get_by_text(
        re.compile(rf"^\s*{re.escape(ACCOUNT_NAME)}(?:\s|$)")
    ))
    if visible_account is None:
        raise RuntimeError(f"현재 선택 Wing 계정이 {ACCOUNT_NAME}({ACCOUNT_ID})인지 확인하지 못함")


def _click_download(page: Page):
    """텍스트 기반으로 드롭다운을 열고 상품별 판매 리포트를 다운로드한다."""
    option_title = page.get_by_text(re.compile(r"옵션목록\s*\(총\s*\d+\)"))
    if option_title.count():
        option_title.last.scroll_into_view_if_needed()
    else:
        page.mouse.wheel(0, 1600)
    time.sleep(1)

    download_button = _last_visible(page.get_by_text("엑셀 다운로드", exact=True))
    if download_button is None:
        download_button = _last_visible(page.get_by_role("button", name=re.compile("엑셀 다운로드")))
    if download_button is None:
        raise RuntimeError("'엑셀 다운로드' 버튼을 찾지 못함 — Wing 화면 구조 확인 필요")
    download_button.click()

    report_item = page.get_by_text(REPORT_LABEL, exact=True)
    report_item.last.wait_for(state="visible", timeout=10000)
    with page.expect_download(timeout=60000) as info:
        report_item.last.click()
    download = info.value
    if "SELLER_INSIGHTS_VENDOR_ITEM_METRICS" not in download.suggested_filename.upper():
        raise RuntimeError(f"예상하지 않은 원본 파일명: {download.suggested_filename}")
    return download


def _target_path(data_date: date) -> Path:
    return TRANSFER_DIR / f"{ACCOUNT_ID}_seller_insights_vendor_item_{data_date:%Y%m%d}.xlsx"


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _trusted_existing(target: Path, data_date: date) -> bool:
    meta_path = target.with_suffix(".json")
    try:
        meta = json.loads(meta_path.read_text(encoding="utf-8"))
        return (meta.get("validator_version") == 2
                and meta.get("account_id") == ACCOUNT_ID
                and meta.get("data_date") == data_date.isoformat()
                and meta.get("sha256") == _sha256(target)
                and validate_report(target)[0])
    except Exception:
        return False


def download_report_in_page(page: Page, data_date: date, force: bool = False) -> Path:
    """이미 채움컴퍼니로 인증된 브라우저 페이지에서 상품별 판매 리포트를 받는다."""
    if data_date != date.today() - timedelta(days=1):
        raise ValueError("현재 자동 다운로드는 Wing 기본값인 '어제' 1일만 지원")
    captured_at = datetime.now()
    TRANSFER_DIR.mkdir(parents=True, exist_ok=True)
    target = _target_path(data_date)
    if target.exists() and not force and _trusted_existing(target, data_date):
        valid, reason, _ = validate_report(target)
        if valid:
            print(f"✅ 기존 검증본 사용: {target.name} — {reason}")
            return target

    temp = target.with_suffix(".download.xlsx")
    if temp.exists():
        temp.unlink()

    page.goto(SALES_URL, wait_until="domcontentloaded", timeout=60000)
    if not _logged_in(page):
        raise RuntimeError("광고 로그인 세션으로 Wing 판매분석 진입 실패")
    try:
        page.wait_for_load_state("networkidle", timeout=20000)
    except Exception:
        pass
    _verify_account(page)
    _select_and_verify_yesterday(page, data_date)
    body = page.locator("body").inner_text(timeout=10000)
    if "옵션목록" not in body:
        raise RuntimeError("판매분석의 옵션목록 화면을 확인하지 못함")
    download = _click_download(page)
    download.save_as(str(temp))

    valid, reason, summary = validate_report(temp)
    if not valid:
        temp.unlink(missing_ok=True)
        raise RuntimeError(f"다운로드 파일 검증 실패: {reason}")
    os.replace(temp, target)

    meta = {
        "validator_version": 2,
        "account_id": ACCOUNT_ID,
        "data_date": data_date.isoformat(),
        "captured_at": captured_at.isoformat(timespec="seconds"),
        "status": "downloaded",
        "note": "화면의 12:30 안내는 트래픽·전환율 그래프 표시 시각이며 판매 리포트 제한이 아님",
        "source": "Wing > 비즈니스 인사이트 > 판매분석 > 상품별 판매 리포트",
        "file": str(target),
        "original_filename": download.suggested_filename,
        "sha256": _sha256(target),
        "summary": summary,
    }
    target.with_suffix(".json").write_text(
        json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"✅ 다운로드: {target.name} — {reason}")
    return target


def download_report(data_date: date, force: bool = False) -> Path:
    """상주 CDP 세션을 검증·복원한 뒤 상품별 판매 리포트를 받는다."""
    _ensure_session()
    with sync_playwright() as playwright:
        browser = playwright.chromium.connect_over_cdp(CDP_URL, timeout=30000)
        context = browser.contexts[0]
        page = context.new_page()
        try:
            return download_report_in_page(page, data_date, force=force)
        finally:
            page.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="그로스 상품별 판매 리포트 자동 다운로드")
    parser.add_argument("--data-date", help="YYYY-MM-DD, 기본=어제")
    parser.add_argument("--force", action="store_true", help="기존 파일이 있어도 다시 다운로드")
    parser.add_argument("--validate", type=Path, help="브라우저 없이 기존 XLSX만 검증")
    args = parser.parse_args()

    if args.validate:
        valid, reason, summary = validate_report(args.validate)
        print(("✅ " if valid else "❌ ") + reason)
        if summary:
            print(json.dumps(summary, ensure_ascii=False, indent=2))
        return 0 if valid else 1

    data_date = (
        datetime.strptime(args.data_date, "%Y-%m-%d").date()
        if args.data_date
        else date.today() - timedelta(days=1)
    )
    try:
        download_report(data_date, force=args.force)
        return 0
    except Exception as exc:
        print(f"❌ {type(exc).__name__}: {exc}")
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
