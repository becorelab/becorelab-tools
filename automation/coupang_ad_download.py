"""쿠팡 광고센터 보고서 자동 다운로드 + JSON 변환

흐름:
1. stealth + Chrome으로 xauth.coupang.com 로그인
2. 광고센터 보고서 페이지 접속
3. 보고서 목록 스캔 → 다운로드
4. 엑셀 → JSON 변환
"""
import os
import sys
import time
import json
import glob
from datetime import datetime, timedelta
from playwright.sync_api import sync_playwright
from playwright_stealth import Stealth

sys.path.insert(0, os.path.dirname(__file__))
from coupang_ad_config import ACCOUNTS, AD_CENTER_URL, DOWNLOAD_DIR, DATA_DIR

WING_AUTH_URL = "https://advertising.coupang.com/user/login?_cap_client=WING&returnUrl=%2Fdashboard"
REPORT_PAGE_URL = f"{AD_CENTER_URL}/marketing-reporting/billboard/reports/pa"
KEYWORD_REQUIRED_HEADERS = {
    "날짜", "캠페인명", "노출수", "클릭수", "광고비", "총 전환매출액(14일)",
    "광고그룹", "광고집행 상품명", "광고집행 옵션ID",
    "광고 노출 지면", "키워드",
}
KEYWORD_STRUCTURE_LABEL = "캠페인 > 광고그룹 > 상품 > 키워드"
LAST_GROSS_INSIGHTS_PATH = None
LAST_GROSS_INSIGHTS_ERROR = None


def validate_keyword_report(xlsx_path):
    """키워드 분석용 보고서인지 파일명과 실제 열을 함께 검증한다."""
    filename = os.path.basename(xlsx_path)
    if "_pa_daily_keyword_" not in filename:
        return False, f"파일명이 키워드 보고서가 아님: {filename}"
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = {str(v).strip() for v in first_row if v is not None}
        wb.close()
    except Exception as e:
        return False, f"XLSX 열 검증 실패: {e}"
    missing = sorted(KEYWORD_REQUIRED_HEADERS - headers)
    if missing:
        return False, f"키워드 필수열 누락: {', '.join(missing)}"
    return True, f"키워드 필수열 {len(KEYWORD_REQUIRED_HEADERS)}개 확인"


def login_and_download_all(account_key="chaewoom", headless=True, max_reports=10,
                           create_first=False, create_date_from=None, create_date_to=None):
    """로그인 → (옵션: 보고서 생성) → 보고서 목록 스캔 → 전체 다운로드
    create_first=True: 같은 로그인 세션으로 어제자(또는 지정기간) 보고서를 먼저 생성 (논스톱 파이프라인)"""
    global LAST_GROSS_INSIGHTS_PATH, LAST_GROSS_INSIGHTS_ERROR
    LAST_GROSS_INSIGHTS_PATH = None
    LAST_GROSS_INSIGHTS_ERROR = None
    acct = ACCOUNTS[account_key]
    print(f"[{acct['name']}] 쿠팡 광고 보고서 자동 다운로드")
    print(f"  headless={headless}, max_reports={max_reports}")
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=headless,
            channel="chrome",
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
                "--disable-dev-shm-usage",
            ]
        )
        context = browser.new_context(
            viewport={"width": 1920, "height": 1080},
            user_agent="Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/148.0.0.0 Safari/537.36",
            locale="ko-KR",
            accept_downloads=True,
        )
        stealth = Stealth()
        page = context.new_page()
        stealth.apply_stealth_sync(page)

        # 1) 로그인
        print("\n[1/3] Wing 로그인...")
        page.goto(WING_AUTH_URL, wait_until="domcontentloaded", timeout=60000)
        time.sleep(3)

        body_text = page.inner_text('body')
        if "Access Denied" in body_text:
            print("  ❌ Akamai Access Denied")
            browser.close()
            return []

        try:
            page.wait_for_selector('input[name="username"]', timeout=30000)
        except Exception:
            print("  ⚠️ 로그인 폼 없음 — 1회 재시도")
            page.reload(wait_until="domcontentloaded", timeout=30000)
            time.sleep(5)
            try:
                page.wait_for_selector('input[name="username"]', timeout=30000)
            except Exception:
                print("  ❌ 로그인 폼 없음 (재시도 실패)")
                browser.close()
                return []

        page.fill('input[name="username"]', acct["id"])
        page.fill('input[name="password"]', acct["pw"])
        time.sleep(1)

        login_success = False
        try:
            with page.expect_navigation(timeout=30000, wait_until="load"):
                page.click('input[name="login"]')
            time.sleep(3)
        except Exception:
            time.sleep(3)

        if "xauth" in page.url and "advertising.coupang.com" not in page.url:
            body_text = page.inner_text('body')[:100]
            if "Access Denied" in body_text:
                print("  ⚠️ Akamai 차단 — 추가 로그인 없이 즉시 중단")
            else:
                print("  ⚠️ 로그인 실패 — 추가 로그인 없이 즉시 중단")
        else:
            login_success = True

        if not login_success:
            print("  ❌ 로그인 실패 — 최소 30분 쿨다운 후 재실행 필요")
            browser.close()
            return []

        print(f"  ✅ 로그인 성공 → {page.url}")

        # 1.5) (옵션) 보고서 생성 — 같은 세션 재활용 (2026-07-16, 생성→다운→분석 논스톱)
        target_from = None
        target_to = None
        if create_first:
            from coupang_ad_create import create_in_page
            from datetime import date as _date, timedelta as _td
            _yday = (_date.today() - _td(days=1)).isoformat()
            _from = create_date_from or _yday
            _to = create_date_to or _yday
            target_from, target_to = _from, _to
            print(f"\n[1.5/3] 보고서 생성: {_from} ~ {_to} (일별/전체캠페인)")
            ok, msg = create_in_page(page, _from, _to, wait_done=True)
            print(("  ✅ " if ok else "  ⚠️ ") + msg)
            if not ok:
                print("  ❌ 생성 검증 실패 — 과거/다른 구조 보고서 다운로드를 막고 중단")
                browser.close()
                return []

        # 2) 보고서 페이지
        print(f"\n[2/3] 보고서 페이지...")
        page.goto(REPORT_PAGE_URL, wait_until="domcontentloaded", timeout=60000)
        time.sleep(10)

        # 보고서 행 스캔
        rows = page.query_selector_all('div[role="row"][row-id]')
        print(f"  보고서 {len(rows)}개 발견")

        reports = []
        for row in rows[:max_reports]:
            row_id = row.get_attribute("row-id")
            text = row.inner_text().strip().replace('\n', ' | ')
            reports.append({"row_id": row_id, "text": text})
            print(f"    [{row_id}] {text[:120]}")

        # 3) 다운로드
        print(f"\n[3/3] 다운로드 시작...")
        if create_first:
            expected_range = f"{target_from} ~ {target_to}"
            expected_structure = f"[일별] {KEYWORD_STRUCTURE_LABEL}"
            target_rows = []
            for row in rows:
                text = row.inner_text().replace("\n", " ")
                if (expected_range in text and expected_structure in text
                        and "생성 완료" in text):
                    target_rows.append(row)
            if not target_rows:
                print(f"  ❌ 대상 보고서 행 없음: {expected_range} / {expected_structure}")
                browser.close()
                return []
            dl_buttons = []
            for row in target_rows[:1]:
                btn = row.query_selector('button:has-text("다운로드")')
                if btn:
                    dl_buttons.append(btn)
            if not dl_buttons:
                print("  ❌ 대상 키워드 보고서 행에서 다운로드 버튼을 찾지 못함")
                browser.close()
                return []
        else:
            dl_buttons = page.query_selector_all('button:has-text("다운로드")')[:max_reports]
        downloaded = []

        for i, btn in enumerate(dl_buttons):
            # 모달이 열려 있으면 닫기
            modal_close = page.query_selector('.ant-modal-wrap .ant-modal-close, .ant-modal-wrap button:has-text("닫기")')
            if modal_close:
                try:
                    modal_close.click()
                    time.sleep(1)
                except Exception:
                    page.keyboard.press("Escape")
                    time.sleep(1)

            try:
                with page.expect_download(timeout=30000) as dl_info:
                    btn.click()
                download = dl_info.value
                filename = download.suggested_filename
                save_path = os.path.join(DOWNLOAD_DIR, filename)
                download.save_as(save_path)
                size = os.path.getsize(save_path)
                valid, reason = validate_keyword_report(save_path)
                if valid:
                    print(f"  ✅ [{i+1}/{len(dl_buttons)}] {filename} ({size:,} bytes) — {reason}")
                    downloaded.append(save_path)
                else:
                    print(f"  ⚠️ [{i+1}/{len(dl_buttons)}] 분석 제외: {filename} — {reason}")
                time.sleep(2)
            except Exception as e:
                # 모달 닫기 재시도
                page.keyboard.press("Escape")
                time.sleep(1)
                try:
                    with page.expect_download(timeout=30000) as dl_info:
                        btn.click()
                    download = dl_info.value
                    filename = download.suggested_filename
                    save_path = os.path.join(DOWNLOAD_DIR, filename)
                    download.save_as(save_path)
                    size = os.path.getsize(save_path)
                    valid, reason = validate_keyword_report(save_path)
                    if valid:
                        print(f"  ✅ [{i+1}/{len(dl_buttons)}] {filename} ({size:,} bytes) (재시도) — {reason}")
                        downloaded.append(save_path)
                    else:
                        print(f"  ⚠️ [{i+1}/{len(dl_buttons)}] 분석 제외: {filename} — {reason}")
                    time.sleep(2)
                except Exception as e2:
                    print(f"  ❌ [{i+1}/{len(dl_buttons)}] 실패: {e2}")

        # 4) 채움컴퍼니는 같은 인증 세션으로 그로스 상품별 판매 리포트까지 수집한다.
        #    별도 로그인/쿠키 브릿지를 만들지 않아 Akamai 접촉을 최소화한다.
        if account_key == "chaewoom":
            print("\n[4/4] 그로스 오가닉 포함 전체판매 리포트...")
            try:
                from coupang_gross_insights_download import download_report_in_page
                data_date = datetime.now().date() - timedelta(days=1)
                LAST_GROSS_INSIGHTS_PATH = str(
                    download_report_in_page(page, data_date, force=True)
                )
                print(f"  ✅ {os.path.basename(LAST_GROSS_INSIGHTS_PATH)}")
            except Exception as e:
                LAST_GROSS_INSIGHTS_ERROR = f"{type(e).__name__}: {e}"
                print(f"  ❌ 그로스 전체판매 다운로드 실패: {LAST_GROSS_INSIGHTS_ERROR}")

        # 쿠키 저장
        cookies = context.cookies()
        with open(os.path.join(DOWNLOAD_DIR, "cookies.json"), "w") as f:
            json.dump(cookies, f, indent=2)

        browser.close()
        print(f"\n완료! {len(downloaded)}개 다운로드됨")
        return downloaded


def excel_to_json(xlsx_path, output_dir=None):
    """엑셀 보고서 → JSON 변환"""
    valid, reason = validate_keyword_report(xlsx_path)
    if not valid:
        print(f"  변환 중단: {reason}")
        return None
    try:
        import openpyxl
    except ImportError:
        print("openpyxl 설치 필요: pip3 install openpyxl")
        return None

    if output_dir is None:
        output_dir = DATA_DIR
    os.makedirs(output_dir, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print(f"  빈 파일: {xlsx_path}")
        return None

    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    data = []
    for row in rows[1:]:
        record = {}
        for j, val in enumerate(row):
            if j < len(headers):
                if isinstance(val, datetime):
                    val = val.strftime("%Y-%m-%d")
                record[headers[j]] = val
        data.append(record)

    basename = os.path.splitext(os.path.basename(xlsx_path))[0]
    json_path = os.path.join(output_dir, f"{basename}.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"  {os.path.basename(xlsx_path)} → {os.path.basename(json_path)} ({len(data)}행)")
    wb.close()
    return json_path


if __name__ == "__main__":
    headless = "--headed" not in sys.argv
    convert = "--convert" in sys.argv or "--json" in sys.argv
    create_first = "--create" in sys.argv  # 어제자 보고서 생성 후 다운로드 (논스톱)
    account = "chaewoom"
    for arg in sys.argv[1:]:
        if not arg.startswith("-"):
            account = arg
            break

    print(f"=== 쿠팡 광고 보고서 자동화 ===")
    print(f"    계정: {account}, headless: {headless}, create: {create_first}")
    print(f"    시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()

    downloaded = login_and_download_all(account, headless=headless, create_first=create_first)

    if downloaded and convert:
        print(f"\n=== JSON 변환 ===")
        for path in downloaded:
            if path.endswith('.xlsx'):
                excel_to_json(path)

    if not downloaded:
        sys.exit(1)
    if account == "chaewoom" and LAST_GROSS_INSIGHTS_ERROR:
        # 광고 XLSX/JSON은 보존하되 전체판매 누락을 성공으로 숨기지 않는다.
        sys.exit(2)
