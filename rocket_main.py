#!/usr/bin/env python3
"""
로켓배송 자동화 메인 스크립트
- 매일 오전 9시: 쿠팡 발주 확인 + 확정 + PO SKU LIST 다운로드 + 양식 변환
- 매일 오후 3시: 이지어드민에 변환 파일 자동 업로드
"""

import os
import re
import glob
import json
import time
import subprocess
import schedule
import logging
from datetime import datetime, date
from pathlib import Path
import xlwt
import openpyxl
from playwright.sync_api import sync_playwright

from rocket_config import (
    COUPANG_ID, COUPANG_PW,
    EZADMIN_DOMAIN, EZADMIN_ID, EZADMIN_PW,
    WORK_DIR, CENTER_MAP, EXPIRY_FILE,
    UPLOAD_HOUR, UPLOAD_MINUTE,
    CHECK_HOUR, CHECK_MINUTE,
    CHECK2_HOUR, CHECK2_MINUTE,
)

# 로그 설정
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(os.path.join(WORK_DIR, 'rocket_auto.log'), encoding='utf-8'),
        logging.StreamHandler()
    ]
)
log = logging.getLogger(__name__)


# ============================================================
# 브라우저 공통 설정
# ============================================================
def get_browser(p):
    browser = p.chromium.launch(
        headless=False,
        args=['--disable-blink-features=AutomationControlled']
    )
    context = browser.new_context(
        user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
        viewport={'width': 1440, 'height': 900}
    )
    page = context.new_page()
    page.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    return browser, page


# ============================================================
# 쿠팡 로그인
# ============================================================
def coupang_login(page):
    log.info("쿠팡 로그인 중...")
    page.goto('https://supplier.coupang.com/login', timeout=30000)
    page.wait_for_timeout(2000)
    page.fill('input[name="username"]', COUPANG_ID)
    page.fill('input[name="password"]', COUPANG_PW)
    page.click('input[type="submit"], button[type="submit"]')
    page.wait_for_timeout(5000)
    log.info(f"쿠팡 로그인 완료: {page.url}")


# ============================================================
# 쿠팡 발주 확인 + 확정
# ============================================================
def coupang_confirm_orders():
    log.info("=" * 50)
    log.info("쿠팡 발주 확인 시작")
    today = date.today().strftime('%Y-%m-%d')

    with sync_playwright() as p:
        browser, page = get_browser(p)
        try:
            coupang_login(page)

            # 발주리스트 이동
            page.click('text=물류')
            page.wait_for_timeout(1000)
            page.click('text=발주리스트')
            page.wait_for_timeout(5000)

            # 오늘 발주 필터링 (어제~오늘)
            try:
                page.click('text=최근 1일')
                page.wait_for_timeout(1000)
                page.click('text=검색')
                page.wait_for_timeout(3000)
            except:
                pass

            # 발주 확인 및 확정 링크 클릭
            try:
                page.click('text=발주 확인 및 확정하기', timeout=5000)
                page.wait_for_timeout(3000)
                log.info("발주 확인 및 확정 페이지 이동 완료")
            except:
                log.info("발주 확정 버튼 없음 (이미 처리됨)")

            # 미확정 발주 확정 처리
            confirmed_count = 0
            while True:
                try:
                    # 미확정 상태 체크박스 선택 후 확정
                    checkboxes = page.query_selector_all('input[type="checkbox"]:not(:disabled)')
                    if not checkboxes:
                        break
                    for cb in checkboxes:
                        if not cb.is_checked():
                            cb.check()

                    # 확정 버튼 클릭
                    confirm_btn = page.query_selector('button:has-text("확정"), button:has-text("발주확정")')
                    if confirm_btn:
                        confirm_btn.click()
                        page.wait_for_timeout(2000)
                        confirmed_count += 1
                    else:
                        break
                except:
                    break

            log.info(f"발주 확정 완료: {confirmed_count}건")

            # PO SKU LIST 다운로드
            sku_file = download_po_sku_list(page, today)

            browser.close()
            return sku_file

        except Exception as e:
            log.error(f"쿠팡 작업 오류: {e}")
            browser.close()
            return None


# ============================================================
# PO SKU LIST 다운로드
# ============================================================
def download_po_sku_list(page, today):
    log.info("PO SKU LIST 다운로드 중...")
    try:
        page.click('text=물류')
        page.wait_for_timeout(500)
        page.click('text=발주 SKU 리스트')
        page.wait_for_timeout(4000)

        # 오늘 날짜로 필터
        page.click('text=오늘')
        page.wait_for_timeout(500)
        page.click('text=검색')
        page.wait_for_timeout(3000)

        # 다운로드
        with page.expect_download() as dl_info:
            page.click('text=전체 다운로드 하기')
        download = dl_info.value

        # 저장 경로
        filename = f'PO_SKU_LIST_{today.replace("-", "")}.xlsx'
        save_path = os.path.join(WORK_DIR, filename)
        download.save_as(save_path)
        log.info(f"PO SKU LIST 저장: {save_path}")
        return save_path

    except Exception as e:
        log.error(f"PO SKU LIST 다운로드 실패: {e}")
        # 폴더에서 최신 PO 파일 찾기
        return find_latest_po_file(today)


def find_latest_po_file(today=None):
    """폴더에서 오늘 날짜의 PO SKU LIST 파일 찾기"""
    today_str = (today or date.today().strftime('%Y%m%d')).replace('-', '')
    patterns = [
        os.path.join(WORK_DIR, f'*PO_SKU_LIST*{today_str}*.xlsx'),
        os.path.join(WORK_DIR, f'*PO_SKU*{today_str}*.xlsx'),
        os.path.join(WORK_DIR, f'*po_sku*{today_str}*.xlsx'),
        os.path.join(WORK_DIR, '*PO_SKU_LIST*.xlsx'),  # 오늘 날짜 없으면 최신파일
    ]
    for pat in patterns:
        files = glob.glob(pat)
        if files:
            latest = max(files, key=os.path.getmtime)
            log.info(f"PO 파일 발견: {latest}")
            return latest
    return None


# ============================================================
# 유통기한 데이터 로드 (제조일자 시트 첫 번째 시트)
# ============================================================
def load_expiry_dates():
    """
    ●쿠팡물류주소_제조일자 리스트.xlsx 의 '제조일자' 시트에서
    상품별 최신 제조일자 + 유통기한 로드.
    시트 상단 = 가장 최신 데이터이므로 첫 번째 출현 우선 사용.
    반환: {정규화된상품명: {'제조일자': '2026.02.10', '유통기한': '2029.02.09', '원본': ...}}
    """
    if not os.path.exists(EXPIRY_FILE):
        log.warning(f"유통기한 파일 없음: {EXPIRY_FILE}")
        return {}
    try:
        wb = openpyxl.load_workbook(EXPIRY_FILE, read_only=True, data_only=True)
        sheet = wb.worksheets[0]
        rows = list(sheet.iter_rows(values_only=True))

        result = {}
        skip_values = {'상품명', '제조일자', '유통기한', '', None}

        for row in rows:
            name_raw = row[3]
            mfg_raw  = str(row[4] or '').strip()
            exp_raw  = str(row[5] or '').strip()

            if not name_raw:
                continue
            name_raw = str(name_raw).strip()

            # 헤더/구분행 제외
            if name_raw in skip_values or 'orporation' in name_raw:
                continue
            # "XX.XX.XX 기준" 같은 구분행 제외
            if '기준' in name_raw or '예정' in name_raw or '입고일자' in name_raw:
                continue
            # 날짜 형식처럼 보이는 행 제외 (숫자만 있는 행)
            if re.match(r'^\d+\.?\d*$', name_raw):
                continue
            # 유통기한이 X이거나 없는 경우 (얼룩제거제 등) → None으로 저장
            if exp_raw in ('X', 'x', '', 'None') or mfg_raw in ('X', 'x', '', 'None'):
                exp_full = None
                mfg_full = None
            else:
                # YY.MM.DD → 20YY.MM.DD 변환
                def expand_year(d):
                    parts = d.split('.')
                    if len(parts) == 3 and len(parts[0]) == 2:
                        return f"20{parts[0]}.{parts[1]}.{parts[2]}"
                    return d
                mfg_full = expand_year(mfg_raw)
                exp_full = expand_year(exp_raw)

            # 상품명 정규화 (특수문자·이모지 제거, 공백 정리)
            clean_name = re.sub(r'[♥●★▲▶◆■♦○□]', '', name_raw).strip()

            if clean_name not in result:  # 첫 번째 출현 = 가장 최신
                result[clean_name] = {
                    '제조일자': mfg_full,
                    '유통기한': exp_full,
                    '원본': name_raw,
                }

        log.info(f"유통기한 데이터 로드 완료: {len(result)}개 제품")
        return result
    except Exception as e:
        log.error(f"유통기한 파일 로드 오류: {e}")
        return {}


def match_expiry(sku_name, expiry_map):
    """
    SKU 이름으로 유통기한 데이터 매칭.
    단어 겹침 점수가 가장 높은 제품 반환.
    반환: {'제조일자': ..., '유통기한': ...} 또는 None
    """
    if not expiry_map or not sku_name:
        return None

    sku_words = set(sku_name.replace('일비아', '').split())
    best_score = 1  # 최소 2단어 이상 겹쳐야 매칭
    best = None

    for prod_name, dates in expiry_map.items():
        prod_words = set(prod_name.split())
        score = len(sku_words & prod_words)
        if score > best_score:
            best_score = score
            best = dates

    if best:
        log.info(f"  유통기한 매칭: '{sku_name[:20]}...' → 유통기한 {best['유통기한']}")
    return best


# ============================================================
# PO SKU LIST → 이지어드민 양식 변환
# ============================================================
def convert_po_to_ezadmin(po_file_path, today=None):
    today = today or date.today()
    today_str = today.strftime('%m%d') if isinstance(today, date) else today

    log.info(f"PO 파일 변환 중: {po_file_path}")

    # 유통기한 데이터 로드
    expiry_map = load_expiry_dates()

    wb = openpyxl.load_workbook(po_file_path, read_only=True, data_only=True)

    # 두 번째 시트 사용 ('바이피엘' 컬럼 있는 시트)
    sheet = wb.worksheets[1] if len(wb.worksheets) > 1 else wb.active
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        log.error("PO 파일이 비어있습니다")
        return None

    # 헤더 매핑
    header = rows[0]
    col = {str(v): i for i, v in enumerate(header) if v}

    # 결과 행 생성
    order_num_count = {}  # 중복 발주번호 카운터
    result_rows = []

    for row in rows[1:]:
        if not row[col.get('발주번호', 0)]:
            continue

        status = str(row[col.get('발주현황', 2)] or '')
        if '발주확정' not in status:
            continue

        po_num = row[col.get('발주번호', 0)]
        sku_name = str(row[col.get('SKU 이름', 4)] or '').strip()
        center = str(row[col.get('물류센터', 6)] or '').strip()
        qty = row[col.get('확정수량', 10)] or 0
        amount = row[col.get('총발주 매입금', 20)] or 0

        # 물류센터 정보 조회
        center_info = CENTER_MAP.get(center)
        if not center_info:
            log.warning(f"⚠️ 알 수 없는 물류센터: '{center}' → rocket_config.py에 추가 필요!")
            center_info = {'phone': '070-0000-0000', 'addr': f'[주소확인필요] {center}'}

        # 중복 발주번호 처리
        key = str(po_num)
        if key in order_num_count:
            order_num_count[key] += 1
            display_num = f"{key}_{str(order_num_count[key]-1).zfill(2)}"
        else:
            order_num_count[key] = 1
            display_num = key

        # 유통기한 매칭
        expiry = match_expiry(sku_name, expiry_map)
        mfg_date = expiry['제조일자'] if expiry else ''
        exp_date = expiry['유통기한'] if expiry else ''

        result_rows.append([
            '쿠팡 로켓배송',          # 채널명
            display_num,              # 주문번호
            f'쿠팡물류센터 {center}', # 받는분성함
            center_info['phone'],     # 받는분전화번호
            '',                       # 기타연락처
            '',                       # 우편번호
            center_info['addr'],      # 받는분주소
            sku_name,                 # 상품명
            '',                       # 옵션
            qty,                      # 수량
            amount,                   # 판매가격
            amount,                   # 정산금액
            '',                       # 송장번호
            '',                       # 배송요청사항
            mfg_date,                 # [내부용] 제조일자
            exp_date,                 # [내부용] 유통기한
        ])

    if not result_rows:
        log.warning("변환할 발주확정 건이 없습니다")
        return None

    # XLS 파일로 저장
    output_filename = f'바이피엘 출고요청_{today_str}(로켓배송).xls'
    output_path = os.path.join(WORK_DIR, output_filename)

    wbout = xlwt.Workbook(encoding='utf-8')
    ws = wbout.add_sheet('비코어랩 업로드 양식')

    # 이지어드민 업로드용 헤더 (14개) — 유통기한은 내부용이라 XLS에 미포함
    upload_headers = ['채널명', '주문번호', '받는분성함', '받는분전화번호', '기타연락처',
                      '우편번호', '받는분주소', '상품명', '옵션', '수량',
                      '판매가격', '정산금액', '송장번호', '배송요청사항']
    for j, h in enumerate(upload_headers):
        ws.write(0, j, h)

    # 데이터 (업로드 14컬럼만)
    for i, row in enumerate(result_rows):
        for j, val in enumerate(row[:14]):
            ws.write(i + 1, j, val)

    wbout.save(output_path)
    log.info(f"✅ 변환 완료: {output_path} ({len(result_rows)}건)")

    # 브라우저 미리보기 자동 오픈 (유통기한 포함)
    open_preview(result_rows, upload_headers, output_filename)

    return output_path


# ============================================================
# 브라우저 미리보기 자동 생성 + 오픈
# ============================================================
def open_preview(result_rows, headers, filename):
    """변환 결과를 HTML로 생성하고 브라우저에서 자동으로 열기"""
    total_qty = sum(int(r[9]) for r in result_rows if r[9])
    total_amt = sum(int(r[10]) for r in result_rows if r[10])
    centers = list({r[2] for r in result_rows})
    warnings = [r for r in result_rows if '[주소확인필요]' in str(r[6])]

    rows_json = json.dumps(result_rows, ensure_ascii=False)
    headers_json = json.dumps(headers, ensure_ascii=False)
    warnings_html = ''
    if warnings:
        items = ''.join(f'<li>주문번호 {r[1]} — {r[2]} → 물류센터 주소 매핑 없음 (rocket_config.py에 추가 필요)</li>' for r in warnings)
        warnings_html = f'<div class="warnings"><strong>⚠️ 주소 확인 필요</strong><ul>{items}</ul></div>'

    html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<title>로켓배송 발주서 미리보기</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Apple SD Gothic Neo', 'Noto Sans KR', sans-serif; background: #0f1117; color: #e2e8f0; padding: 28px; }}
  h1 {{ font-size: 20px; font-weight: 700; color: #fff; margin-bottom: 4px; }}
  .sub {{ font-size: 13px; color: #64748b; margin-bottom: 24px; }}
  .summary {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 12px; margin-bottom: 20px; }}
  .card {{ background: #1e2330; border-radius: 10px; padding: 16px 20px; border: 1px solid #2d3748; }}
  .card-label {{ font-size: 11px; color: #64748b; margin-bottom: 6px; text-transform: uppercase; letter-spacing: 0.05em; }}
  .card-value {{ font-size: 22px; font-weight: 700; color: #fff; }}
  .orange {{ color: #e8542a; }} .blue {{ color: #60a5fa; }}
  .warnings {{ background: #3b1a0a; border: 1px solid #7c2d12; border-radius: 8px; padding: 12px 16px; margin-bottom: 20px; font-size: 13px; color: #fca5a5; }}
  .warnings strong {{ color: #f87171; display: block; margin-bottom: 6px; }}
  .warnings ul {{ padding-left: 18px; }}
  .table-wrap {{ background: #1e2330; border-radius: 10px; border: 1px solid #2d3748; overflow: auto; max-height: 68vh; }}
  .toolbar {{ display: flex; justify-content: space-between; align-items: center; margin-bottom: 12px; }}
  .toolbar-title {{ font-size: 14px; color: #94a3b8; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
  thead th {{ position: sticky; top: 0; background: #161b27; color: #64748b; font-weight: 600; font-size: 11px; text-transform: uppercase; letter-spacing: 0.04em; padding: 10px 14px; text-align: left; border-bottom: 1px solid #2d3748; white-space: nowrap; }}
  tbody tr {{ border-bottom: 1px solid #1a202c; }}
  tbody tr:hover {{ background: #252d3d; }}
  tbody td {{ padding: 10px 14px; vertical-align: middle; white-space: nowrap; }}
  .badge {{ display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 11px; font-weight: 600; }}
  .badge-blue {{ background: #1e3a5f; color: #60a5fa; }}
  .badge-red {{ background: #3b1a0a; color: #f87171; }}
  .num {{ text-align: right; font-variant-numeric: tabular-nums; }}
  .dim {{ color: #475569; }}
  .addr {{ max-width: 220px; overflow: hidden; text-overflow: ellipsis; }}
  .prod {{ max-width: 260px; overflow: hidden; text-overflow: ellipsis; }}
</style>
</head>
<body>
<h1>로켓배송 발주서 미리보기</h1>
<p class="sub">📄 {filename} &nbsp;·&nbsp; 생성: {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
<div class="summary">
  <div class="card"><div class="card-label">총 발주 건수</div><div class="card-value blue">{len(result_rows)}건</div></div>
  <div class="card"><div class="card-label">물류센터 수</div><div class="card-value">{len(centers)}개</div></div>
  <div class="card"><div class="card-label">총 수량</div><div class="card-value">{total_qty:,}개</div></div>
  <div class="card"><div class="card-label">총 매입금액</div><div class="card-value orange">₩{total_amt:,}</div></div>
</div>
{warnings_html}
<div class="toolbar"><div class="toolbar-title">발주 상세 내역</div></div>
<div class="table-wrap"><table>
<thead><tr>
  <th>주문번호</th><th>물류센터</th><th>전화번호</th><th>주소</th><th>상품명</th><th>수량</th><th>매입금액</th><th>제조일자 / 유통기한</th>
</tr></thead>
<tbody id="tbody"></tbody>
</table></div>
<script>
const rows = {rows_json};
const tbody = document.getElementById('tbody');
rows.forEach(r => {{
  const isWarn = String(r[6]).includes('[주소확인필요]');
  const badge = isWarn
    ? `<span class="badge badge-red">⚠️ ${{r[2]}}</span>`
    : `<span class="badge badge-blue">${{r[2]}}</span>`;
  const mfg = r[14] || '';
  const exp = r[15] || '';
  const expiryCell = exp
    ? `<span style="color:#4ade80;font-size:12px">${{mfg}}</span><br><span style="color:#f87171;font-size:11px">까지: ${{exp}}</span>`
    : `<span style="color:#475569;font-size:12px">입력 불필요</span>`;
  tbody.innerHTML += `<tr>
    <td>${{r[1]}}</td>
    <td>${{badge}}</td>
    <td>${{r[3]}}</td>
    <td class="addr" title="${{r[6]}}">${{r[6]}}</td>
    <td class="prod" title="${{r[7]}}">${{r[7]}}</td>
    <td class="num">${{Number(r[9]).toLocaleString()}}</td>
    <td class="num">₩${{Number(r[10]).toLocaleString()}}</td>
    <td style="white-space:nowrap">${{expiryCell}}</td>
  </tr>`;
}});
</script>
</body>
</html>"""

    preview_path = os.path.join(WORK_DIR, 'rocket_preview_latest.html')
    with open(preview_path, 'w', encoding='utf-8') as f:
        f.write(html)

    subprocess.Popen(['open', preview_path])
    log.info(f"🌐 미리보기 브라우저 오픈: {preview_path}")


# ============================================================
# 이지어드민 업로드
# ============================================================
def ezadmin_upload(file_path):
    if not file_path or not os.path.exists(file_path):
        log.error(f"업로드할 파일 없음: {file_path}")
        # 오늘 날짜 파일 자동 탐색
        today_str = date.today().strftime('%m%d')
        pattern = os.path.join(WORK_DIR, f'*출고요청*{today_str}*(로켓배송)*.xls')
        files = glob.glob(pattern)
        if not files:
            log.error("오늘 변환 파일을 찾을 수 없습니다")
            return False
        file_path = max(files, key=os.path.getmtime)
        log.info(f"자동 탐색된 파일: {file_path}")

    log.info(f"이지어드민 업로드 시작: {file_path}")

    with sync_playwright() as p:
        browser, page = get_browser(p)
        try:
            # 로그인
            page.goto('https://www.ezadmin.co.kr/index.html', timeout=30000)
            page.wait_for_timeout(3000)
            page.evaluate(f"document.getElementById('login-domain').value = '{EZADMIN_DOMAIN}'")
            page.evaluate(f"document.getElementById('login-id').value = '{EZADMIN_ID}'")
            page.evaluate(f"document.getElementById('login-pwd').value = '{EZADMIN_PW}'")
            page.evaluate("document.querySelector('input[type=\"button\"]').click()")
            page.wait_for_timeout(6000)
            log.info("이지어드민 로그인 완료")

            # 오버레이 제거
            page.evaluate("document.querySelectorAll('.dim').forEach(el => el.remove())")
            page.wait_for_timeout(500)

            # 주문배송관리 > 발주 탭 이동
            page.evaluate("move_page35('DC10')")
            page.wait_for_timeout(4000)
            page.evaluate("document.querySelectorAll('.dim').forEach(el => el.remove())")
            page.wait_for_timeout(500)
            page.locator("text=발주").first.click(force=True)
            page.wait_for_timeout(3000)
            log.info("발주 탭 이동 완료")

            # '비코어랩 로켓배송' 행의 업로드 버튼 클릭
            page.evaluate("""
                let rows = document.querySelectorAll('tr');
                for(let row of rows) {
                    if(row.innerText.includes('로켓배송') && row.innerText.includes('업로드')) {
                        let uploadBtns = row.querySelectorAll('span, button, a');
                        for(let btn of uploadBtns) {
                            if(btn.innerText.trim() === '업로드') {
                                btn.click();
                                break;
                            }
                        }
                        break;
                    }
                }
            """)
            page.wait_for_timeout(2000)
            log.info("업로드 버튼 클릭")

            # 파일 선택
            with page.expect_file_chooser() as fc_info:
                page.evaluate("""
                    let fileInput = document.querySelector('input[type="file"]');
                    if(fileInput) fileInput.click();
                """)
            file_chooser = fc_info.value
            file_chooser.set_files(file_path)
            page.wait_for_timeout(1000)

            # 업로드 확인 버튼 클릭
            page.evaluate("""
                let btns = document.querySelectorAll('button, input[type="button"]');
                for(let btn of btns) {
                    if(btn.innerText && btn.innerText.trim() === '업로드') {
                        btn.click();
                        break;
                    }
                }
            """)
            page.wait_for_timeout(3000)

            # 결과 확인
            page.screenshot(path=os.path.join(WORK_DIR, f'upload_result_{date.today().strftime("%Y%m%d")}.png'))
            log.info("✅ 이지어드민 업로드 완료!")

            browser.close()
            return True

        except Exception as e:
            log.error(f"이지어드민 업로드 오류: {e}")
            browser.close()
            return False


# ============================================================
# 일일 작업 함수
# ============================================================
def morning_job():
    """오전 작업: 쿠팡 발주 확인 + 다운로드 + 변환"""
    log.info("🌅 오전 작업 시작")
    today = date.today()

    # 1. 쿠팡 발주 확인 + 다운로드
    po_file = coupang_confirm_orders()

    # 2. 파일 못 받으면 폴더에서 찾기
    if not po_file:
        po_file = find_latest_po_file()

    if not po_file:
        log.warning("⚠️ PO SKU LIST 파일을 찾을 수 없습니다. 수동으로 파일을 넣어주세요.")
        return

    # 3. 이지어드민 양식으로 변환
    convert_po_to_ezadmin(po_file, today)
    log.info("🌅 오전 작업 완료")


def afternoon_job():
    """오후 3시 작업: 이지어드민 업로드"""
    log.info("☀️ 오후 3시 이지어드민 업로드 시작")
    today_str = date.today().strftime('%m%d')

    # 오늘 변환된 파일 찾기
    pattern = os.path.join(WORK_DIR, f'*출고요청*{today_str}*(로켓배송)*.xls')
    files = glob.glob(pattern)

    if not files:
        # 혹은 수동으로 넣어둔 파일
        pattern2 = os.path.join(WORK_DIR, f'*{today_str}*로켓*.xls*')
        files = glob.glob(pattern2)

    file_path = max(files, key=os.path.getmtime) if files else None
    ezadmin_upload(file_path)
    log.info("☀️ 오후 작업 완료")


# ============================================================
# 수동 실행 (테스트용)
# ============================================================
def run_now(task='all'):
    """즉시 실행 (테스트/수동)"""
    if task in ('morning', 'all'):
        morning_job()
    if task in ('afternoon', 'all'):
        afternoon_job()


# ============================================================
# 스케줄러
# ============================================================
def start_scheduler():
    log.info("🤖 로켓배송 자동화 스케줄러 시작")
    log.info(f"  - {CHECK_HOUR:02d}:{CHECK_MINUTE:02d} → 쿠팡 발주 확인 + 변환 (1차)")
    log.info(f"  - {CHECK2_HOUR:02d}:{CHECK2_MINUTE:02d} → 쿠팡 발주 확인 + 변환 (2차)")
    log.info(f"  - {UPLOAD_HOUR:02d}:{UPLOAD_MINUTE:02d} → 이지어드민 업로드")

    schedule.every().day.at(f"{CHECK_HOUR:02d}:{CHECK_MINUTE:02d}").do(morning_job)
    schedule.every().day.at(f"{CHECK2_HOUR:02d}:{CHECK2_MINUTE:02d}").do(morning_job)
    schedule.every().day.at(f"{UPLOAD_HOUR:02d}:{UPLOAD_MINUTE:02d}").do(afternoon_job)

    while True:
        schedule.run_pending()
        time.sleep(30)


if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1:
        cmd = sys.argv[1]
        if cmd == 'morning':
            morning_job()
        elif cmd == 'afternoon':
            afternoon_job()
        elif cmd == 'convert':
            # 폴더의 최신 PO 파일 변환만
            po = find_latest_po_file()
            if po:
                convert_po_to_ezadmin(po)
            else:
                print("PO SKU LIST 파일을 찾을 수 없습니다")
        elif cmd == 'upload':
            afternoon_job()
        else:
            print("사용법: python3 rocket_main.py [morning|afternoon|convert|upload]")
    else:
        # 스케줄러 실행
        start_scheduler()
