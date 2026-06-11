#!/usr/bin/env python3
"""
Extract monthly channel sales data from settlement Excel files.
"""

import openpyxl
import json
from datetime import datetime

# Month column indices (0-based from col B=index 1):
# 1월 매출 = col C = index 2
# 2월 매출 = col F = index 5
# ... each month adds 3 columns
MONTH_COLS = {
    1: 3,   # col C (1-indexed)
    2: 6,
    3: 9,
    4: 12,
    5: 15,
    6: 18,
    7: 21,
    8: 24,
    9: 27,
    10: 30,
    11: 33,
    12: 36,
}

# Channels to exclude (these are totals or aggregates, not actual channels)
SKIP_NAMES = {'총합(부가세포함)', '총합(부가세별도)', '채널', '-', None, ''}

# Also skip rows where column B starts with common aggregate patterns
def is_skip_row(channel_name):
    if channel_name is None:
        return True
    name = str(channel_name).strip()
    if not name or name == '-':
        return True
    if '총합' in name:
        return True
    return False


def get_month_value(ws, row, month_num):
    """Get the 매출 value for a given month (1-12) from a row."""
    col = MONTH_COLS[month_num]
    val = ws.cell(row=row, column=col).value
    if val is None:
        return 0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0


def extract_year_data(ws, data_start_row, year):
    """
    Extract all channel/month data from a sheet.
    Returns dict: {month_num: {channel: sales_value}}
    """
    result = {m: {} for m in range(1, 13)}

    for row in range(data_start_row, ws.max_row + 1):
        channel = ws.cell(row=row, column=2).value
        if is_skip_row(channel):
            continue

        channel_name = str(channel).strip().replace('\n', ' ')

        for month_num in range(1, 13):
            val = get_month_value(ws, row, month_num)
            if val and val > 0:
                if channel_name in result[month_num]:
                    result[month_num][channel_name] += val
                else:
                    result[month_num][channel_name] = val

    return result


def load_sheet(path):
    """Load workbook and find the correct channel sheet."""
    wb = openpyxl.load_workbook(path, data_only=True)

    # Try sheet names in priority order
    for name in wb.sheetnames:
        if name == '채널별 매출 이익':
            return wb[name]

    for name in wb.sheetnames:
        if '채널별' in name and '이익' in name and '광고비' not in name:
            return wb[name]

    for name in wb.sheetnames:
        if name == '일비아 매출':
            return wb[name]

    for name in wb.sheetnames:
        if '채널별' in name:
            return wb[name]

    # Fall back to 3rd sheet
    if len(wb.sheetnames) >= 3:
        return wb[wb.sheetnames[2]]

    return wb[wb.sheetnames[0]]


def find_data_start(ws):
    """Find the row where channel data starts (after headers)."""
    for row in range(1, 25):
        for col in range(1, 5):
            val = ws.cell(row=row, column=col).value
            if val and '채널' in str(val) and row < 20:
                # Next row should be sub-header or data
                # Check if next row has '매출' as sub-header
                next_row_col2 = ws.cell(row=row+1, column=2).value
                next_row_col3 = ws.cell(row=row+1, column=3).value
                if next_row_col3 and '매출' in str(next_row_col3):
                    return row + 2  # data starts 2 rows after header
                elif next_row_col2 is None or str(next_row_col2).strip() in ('', '-'):
                    return row + 2
                else:
                    return row + 1
    return 5  # default


def main():
    files = {
        "2023": {
            "path": "/Users/macmini_ky/Library/CloudStorage/MYBOX-igimylife/개인 폴더/Becorelab/03. 영업/20. 월별 매출정산/2023년/2023. 12/2023. 12 온라인 매출정산.xlsx",
            "sheet_override": "일비아 매출",
            "data_start": 5,
        },
        "2024": {
            "path": "/Users/macmini_ky/Library/CloudStorage/MYBOX-igimylife/개인 폴더/Becorelab/03. 영업/20. 월별 매출정산/2024년/2024. 12/2024. 12 온라인 매출정산.xlsx",
            "sheet_override": "채널별 매출 이익",
            "data_start": 14,
        },
        "2025": {
            "path": "/Users/macmini_ky/Library/CloudStorage/MYBOX-igimylife/개인 폴더/Becorelab/03. 영업/20. 월별 매출정산/2025년/2025.12/2025. 12 온라인 매출정산_자사몰 수정.xlsx",
            "sheet_override": "채널별 매출 이익",
            "data_start": 14,
        },
        "2026": {
            "path": "/Users/macmini_ky/Library/CloudStorage/GoogleDrive-cky2833@gmail.com/내 드라이브/(주)비코어랩/Claude-Setup/매출 정산/2026.04/2026. 04 온라인 매출정산.xlsx",
            "sheet_override": "채널별 매출 이익",
            "data_start": 14,
        },
    }

    # Build months output
    months_data = {}

    for year_str, cfg in files.items():
        year = int(year_str)
        path = cfg["path"]

        print(f"\nProcessing {year}...")
        print(f"  File: {path}")

        wb = openpyxl.load_workbook(path, data_only=True)
        sheet_name = cfg["sheet_override"]
        ws = wb[sheet_name]
        print(f"  Sheet: {sheet_name}")

        data_start = cfg["data_start"]
        year_data = extract_year_data(ws, data_start, year)

        # Determine which months to include
        if year == 2023:
            start_month = 3  # From March 2023
            end_month = 12
        elif year == 2026:
            start_month = 1
            end_month = 4  # Only through April 2026
        else:
            start_month = 1
            end_month = 12

        for month_num in range(start_month, end_month + 1):
            month_key = f"{year}-{month_num:02d}"
            channels = year_data[month_num]

            if not channels:
                print(f"  {month_key}: No data, skipping")
                continue

            total = sum(channels.values())
            if total <= 0:
                print(f"  {month_key}: Total is 0 or negative, skipping")
                continue

            # Round all values to int
            channels_int = {k: round(v) for k, v in channels.items() if v > 0}
            total_int = sum(channels_int.values())

            months_data[month_key] = {
                "channels": channels_int,
                "total": total_int,
            }
            print(f"  {month_key}: {len(channels_int)} channels, total={total_int:,}")

    # Build final output
    output = {
        "extracted_at": datetime.now().isoformat(),
        "months": months_data,
    }

    output_path = "/Users/macmini_ky/ClaudeAITeam/erp/settlement_history.json"
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\nSaved to: {output_path}")
    print(f"Total months extracted: {len(months_data)}")
    print(f"Month range: {min(months_data.keys())} to {max(months_data.keys())}")


if __name__ == "__main__":
    main()
