"""
쿠팡 리뷰 전체 다운로더
- 방법 1 (기본): 소싱앱 wing 브라우저 활용 (Akamai 우회, 가장 안정적)
- 방법 2 (대안): 독립 Playwright + stealth (wing 미실행 시)
- JSON + Excel 저장
"""

import sys, os, re, json, time, argparse
from datetime import datetime

SOURCING_API = 'http://localhost:8090'
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'review_output')
PROFILE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.review_profile')


def extract_product_id(url_or_id: str) -> str:
    if url_or_id.isdigit():
        return url_or_id
    m = re.search(r'products/(\d+)', url_or_id)
    if m:
        return m.group(1)
    m = re.search(r'itemId=(\d+)', url_or_id)
    if m:
        return m.group(1)
    return url_or_id.strip()


def _wing_available() -> bool:
    import requests
    try:
        r = requests.get(f'{SOURCING_API}/api/wing/status', timeout=3)
        data = r.json()
        return data.get('has_browser', False)
    except:
        return False


def _collect_via_wing(pid: str, product_url: str, max_reviews: int = 9999) -> dict:
    """소싱앱 API를 통해 전체 리뷰 수집 (Wing 브라우저 경유, Akamai 우회)"""
    import requests

    print('[INFO] Wing 브라우저로 전체 리뷰 수집 시작...')
    print(f'[INFO] (소싱앱 API 경유 → Wing 브라우저 → 쿠팡 내부 API)')

    try:
        r = requests.post(f'{SOURCING_API}/api/reviews/download', json={
            'url': product_url,
            'max_reviews': max_reviews,
        }, timeout=600)
        resp = r.json()
        if resp.get('success'):
            return resp.get('data', {})
        else:
            print(f'[ERROR] API 응답 에러: {resp.get("error", "unknown")}')
            return {}
    except requests.exceptions.Timeout:
        print('[ERROR] 타임아웃 (600초)')
        return {}
    except Exception as e:
        print(f'[ERROR] API 호출 실패: {e}')
        return {}


def _collect_via_standalone(pid: str, product_url: str, max_reviews: int = 9999) -> dict:
    """독립 Playwright 브라우저로 리뷰 수집 (wing 미실행 시)"""
    from playwright.sync_api import sync_playwright
    try:
        from playwright_stealth import Stealth
        stealth = Stealth()
    except ImportError:
        stealth = None

    os.makedirs(PROFILE_DIR, exist_ok=True)

    pw = sync_playwright().start()
    reviews = []
    product_title = ''
    total_count = 0
    rating_summary = {}

    try:
        ctx = pw.chromium.launch_persistent_context(
            user_data_dir=PROFILE_DIR,
            headless=False,
            args=[
                '--disable-blink-features=AutomationControlled',
                '--window-position=-32000,-32000',
                '--window-size=1,1',
                '--no-focus-on-launch',
            ],
            viewport={'width': 1280, 'height': 900},
            user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
        )

        page = ctx.new_page()
        if stealth:
            stealth.apply_stealth_sync(page)

        print('[INFO] 쿠팡 홈 접속 → 쿠키 확보...')
        page.goto('https://www.coupang.com/', wait_until='domcontentloaded', timeout=20000)
        page.wait_for_timeout(2000)
        page.evaluate('window.scrollTo(0, 300)')
        page.wait_for_timeout(1000)

        print('[INFO] 상품 페이지 접속 중...')
        page.goto(product_url, wait_until='domcontentloaded', timeout=20000)
        page.wait_for_timeout(3000)

        title_check = page.title()
        if 'Access Denied' in title_check:
            print('[WARN] Access Denied — 홈 경유 재시도')
            page.goto('https://www.coupang.com/', wait_until='domcontentloaded', timeout=15000)
            page.wait_for_timeout(3000)
            page.goto(product_url, wait_until='domcontentloaded', timeout=20000)
            page.wait_for_timeout(3000)
            title_check = page.title()

        if 'Access Denied' in title_check:
            print('[ERROR] Akamai 차단. 소싱앱(8090)을 실행하면 Wing 브라우저로 우회됩니다.')
            page.close()
            ctx.close()
            return {'error': 'access_denied'}

        product_title = page.evaluate("""() => {
            const selectors = ['h2.prod-buy-header__title', 'h1.prod-buy-header__title', '.prod-buy-header__title', 'h1', 'h2'];
            for (const s of selectors) {
                const el = document.querySelector(s);
                if (el && el.textContent.trim().length > 3) return el.textContent.trim();
            }
            return '';
        }""")

        reviews, total_count, rating_summary = _paginate_reviews(page, pid, max_reviews)
        page.close()
        ctx.close()
    finally:
        pw.stop()

    return {
        'product_title': product_title,
        'total_count': total_count,
        'rating_summary': rating_summary,
        'reviews': reviews,
    }


def _paginate_reviews(page, pid, max_reviews, page_size=20):
    """쿠팡 내부 API로 리뷰 페이지네이션"""
    reviews = []
    page_num = 1
    total_count = 0
    rating_summary = {}
    max_pages = (max_reviews // page_size) + 1

    while page_num <= max_pages:
        result = page.evaluate(f'''() => {{
            return new Promise(resolve => {{
                fetch('/next-api/review?productId={pid}&page={page_num}&size={page_size}&sortBy=DATE_DESC&ratingSummary=true', {{
                    credentials: 'include',
                    headers: {{ 'accept': 'application/json' }}
                }})
                .then(r => r.json())
                .then(d => resolve(d))
                .catch(e => resolve({{error: e.toString()}}));
            }});
        }}''')

        if result.get('error'):
            print(f'\n[ERROR] API 에러: {result["error"]}')
            break

        data = result.get('data', {})

        if page_num == 1:
            total_count = data.get('totalCount', 0)
            rating_summary = data.get('ratingSummary', {})
            print(f'[INFO] 총 리뷰 수: {total_count}')
            if rating_summary:
                avg = rating_summary.get('averageRating', '?')
                print(f'[INFO] 평균 평점: {avg}')

        review_list = data.get('reviews', [])
        if not review_list:
            break

        for r in review_list:
            reviews.append({
                'rating': r.get('rating', 0),
                'headline': r.get('headline', ''),
                'content': r.get('content', ''),
                'created_at': r.get('createdAt', ''),
                'helpful_count': r.get('helpfulCount', 0),
                'user_name': r.get('userName', r.get('nickName', '')),
                'option': r.get('productOptionName', ''),
                'photos': [p.get('url', '') for p in r.get('photos', [])],
                'photo_count': len(r.get('photos', [])),
                'answer': r.get('answer', {}).get('content', '') if r.get('answer') else '',
            })

        print(f'  페이지 {page_num}: +{len(review_list)}건 (누적 {len(reviews)}/{total_count})', end='\r')

        if len(review_list) < page_size:
            break

        page_num += 1
        delay = 0.5 + (0.5 * (page_num % 5 == 0))
        page.wait_for_timeout(int(delay * 1000))

    print()
    return reviews, total_count, rating_summary


def download_reviews(url_or_id: str, max_reviews: int = 9999) -> dict:
    pid = extract_product_id(url_or_id)
    if not pid.isdigit():
        print(f'[ERROR] 유효한 상품 ID를 찾을 수 없습니다: {url_or_id}')
        return {'error': 'invalid_product_id'}

    product_url = f'https://www.coupang.com/vp/products/{pid}'
    print(f'[INFO] 상품 ID: {pid}')

    reviews = []
    product_title = ''
    total_count = 0
    rating_summary = {}

    if _wing_available():
        print('[INFO] Wing 브라우저 감지 → Wing 경유 수집')
        wing_result = _collect_via_wing(pid, product_url, max_reviews)
        if wing_result and not wing_result.get('error'):
            reviews = wing_result.get('reviews', [])
            product_title = wing_result.get('product_title', '')
            total_count = wing_result.get('total_count', len(reviews))
            rating_summary = wing_result.get('rating_summary', {})
    else:
        print('[INFO] Wing 미감지 → 독립 브라우저 수집')
        result = _collect_via_standalone(pid, product_url, max_reviews)
        if result.get('error'):
            return result
        reviews = result.get('reviews', [])
        product_title = result.get('product_title', '')
        total_count = result.get('total_count', 0)
        rating_summary = result.get('rating_summary', {})

    # wing 경유 시 상품명이 없으므로 첫 리뷰의 정보에서 추정 or 별도 API
    if not product_title and reviews:
        product_title = f'상품 {pid}'

    print(f'[INFO] 수집 완료: {len(reviews)}건')

    return {
        'product_id': pid,
        'product_title': product_title,
        'product_url': product_url,
        'total_count': total_count or len(reviews),
        'collected_count': len(reviews),
        'downloaded_at': datetime.now().isoformat(),
        'rating_summary': rating_summary,
        'reviews': reviews,
    }


def save_json(data: dict, filepath: str):
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f'[SAVE] JSON: {filepath}')


def save_excel(data: dict, filepath: str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        print('[WARN] openpyxl 미설치 — pip3 install openpyxl')
        return

    wb = Workbook()

    # 요약 시트
    ws_summary = wb.active
    ws_summary.title = '요약'
    ws_summary.append(['항목', '값'])
    ws_summary.append(['상품명', data.get('product_title', '')])
    ws_summary.append(['상품 URL', data.get('product_url', '')])
    ws_summary.append(['총 리뷰 수', data.get('total_count', 0)])
    ws_summary.append(['수집 건수', data.get('collected_count', 0)])
    ws_summary.append(['다운로드 시각', data.get('downloaded_at', '')])

    rs = data.get('rating_summary', {})
    if rs:
        ws_summary.append(['평균 평점', rs.get('averageRating', '')])
        dist = rs.get('ratingDistribution', {})
        for star in ['5', '4', '3', '2', '1']:
            ws_summary.append([f'★{star} 개수', dist.get(star, dist.get(int(star), 0))])

    ws_summary.column_dimensions['A'].width = 15
    ws_summary.column_dimensions['B'].width = 60

    # 리뷰 시트
    ws = wb.create_sheet('리뷰')
    headers = ['번호', '별점', '제목', '내용', '작성일', '작성자', '옵션', '도움됨', '사진수', '판매자답변']
    header_fill = PatternFill(start_color='2D3A5C', end_color='2D3A5C', fill_type='solid')
    header_font = Font(bold=True, color='E8EAF0')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for i, r in enumerate(data.get('reviews', []), 1):
        ws.append([
            i,
            r.get('rating', 0),
            r.get('headline', ''),
            r.get('content', ''),
            r.get('created_at', '')[:10] if r.get('created_at') else '',
            r.get('user_name', ''),
            r.get('option', ''),
            r.get('helpful_count', 0),
            r.get('photo_count', 0),
            r.get('answer', ''),
        ])

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 80
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 8
    ws.column_dimensions['I'].width = 8
    ws.column_dimensions['J'].width = 40

    ws.auto_filter.ref = f'A1:J{len(data.get("reviews", [])) + 1}'
    ws.freeze_panes = 'A2'

    wb.save(filepath)
    print(f'[SAVE] Excel: {filepath}')


def main():
    parser = argparse.ArgumentParser(description='쿠팡 리뷰 다운로더')
    parser.add_argument('url', help='쿠팡 상품 URL 또는 상품 ID')
    parser.add_argument('--max', type=int, default=9999, help='최대 리뷰 수 (기본: 전체)')
    parser.add_argument('--json-only', action='store_true', help='JSON만 저장')
    parser.add_argument('--excel-only', action='store_true', help='Excel만 저장')
    args = parser.parse_args()

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    data = download_reviews(args.url, max_reviews=args.max)
    if data.get('error'):
        sys.exit(1)

    pid = data['product_id']
    title = data.get('product_title', '').strip()
    safe_title = re.sub(r'[^\w가-힣\s\-]', '', title)[:40].strip() if title else pid
    folder_name = f'{pid}_{safe_title}' if safe_title != pid else pid
    product_dir = os.path.join(OUTPUT_DIR, folder_name)
    os.makedirs(product_dir, exist_ok=True)

    ts = datetime.now().strftime('%Y%m%d_%H%M')
    base = os.path.join(product_dir, f'reviews_{ts}')

    if not args.excel_only:
        save_json(data, f'{base}.json')
    if not args.json_only:
        save_excel(data, f'{base}.xlsx')

    print(f'\n✅ 완료! {data["collected_count"]}건 리뷰 저장됨')
    return data


if __name__ == '__main__':
    main()
